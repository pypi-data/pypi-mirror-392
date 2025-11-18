from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
from .merge_file import clean_data
from .config import *
import sqlite3, pandas as pd, openpyxl, os, subprocess

def shorten_filename(filename: str) -> str:
    """
    ตัดชื่อไฟล์ให้เหลือคำแรกและคำสุดท้าย (ไม่รวม .xlsx หรือสกุลไฟล์อื่น)
    เช่น:
        "สรุปบัญชี คูปอง รอบ ต.ค.68 (7Delivery TMW Prefixแบบใหม่) มูก้า-namm.xlsx"
        -> "สรุปบัญชีมูก้า-namm"
    """
    name, _ = os.path.splitext(filename)
    parts = name.split()

    if not parts:
        return name  # กรณีไม่มีคำเลย

    if len(parts) >= 2:
        short_name = parts[0] + parts[-1]
    else:
        short_name = parts[0]

    return short_name

# with open('instance/config.json', 'r',encoding="utf-8", errors="replace") as file:
#     Format_Design = json.load(file)


def sanitize_sheet_name(name: str) -> str:
    """Replace invalid Excel sheet characters and trim to 31 chars."""
    for ch in ['\\', '/', '*', '?', ':', '[', ']',')','(',' ']:
        name = name.replace(ch, '')
    return name[:30]

def format_and_save_excel(export_dict:dict,sheetss =False):
    """
    Generate formatted Excel file with multiple sheets grouped by OPTIMAL_DATE and WORKSHEET.
    Each sheet name = OPTIMAL_DATE_WORKSHEET.
    """
    version =export_dict["Version"][:5]
    sub_version = export_dict["Version"][6:]
    conn = sqlite3.connect(export_dict.get("Database Path"))
    query = """
        SELECT promotion_code AS "PROMOTION CODE",promotion_name AS "PROMOTION NAME",active_from AS "ACTIVE FROM",active_to AS "ACTIVE TO",
        redemption_limit_per_transaction AS "LIMIT",bucketid AS "BUCKET",trigger_value AS "TRIGGER",attachmentmode AS "ATTACHMENTMODE",
        entity_code AS "ENTITY CODE",entity_name AS "ENTITY NAME",barcode AS "BARCODE",barcode_code39 AS "BARCODE CODE39",
        reward_type AS "REWARD TYPE",reward_value AS "REWARD VALUE",notes AS "NOTES",coupon_id AS "COUPON ID",coupon_code39 AS "COUPON CODE39",
        reward_ma_id AS "REWARD MA ID",reward_ma_name AS "REWARD MA NAME",
        sheet AS "SHEET",worksheet AS "WORKSHEET",optimal_date AS "OPTIMAL_DATE"
        FROM promotion_data 
        WHERE version = ? AND round = ?
    """
    df = pd.read_sql_query(query, conn, params=(version, sub_version,))
    conn.close()
    df = clean_data(df)

    # เตรียม workbook
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)  # ลบ sheet เปล่าที่ระบบสร้างมา

    # Define styles
    thin_border = Border(
        left=Side(style=TABEL_LINE_LEFT),
        right=Side(style=TABEL_LINE_RIGHT),
        top=Side(style=TABEL_LINE_TOP),
        bottom=Side(style=TABEL_LINE_BOTTOM)
    )
    thick_border = Border(top=Side(style='thick'),
                        left=Side(style=TABEL_LINE_LEFT),
                        right=Side(style=TABEL_LINE_RIGHT),
                        bottom=Side(style=TABEL_LINE_BOTTOM))
    barcode_font = Font(name=FONT_BARCODE, size=HEARDER_FONT_SIZE)
    default_font = Font(name=FONT_NOMAL, size=BACKGROUND_FONT_SIZE)
    white_bold_font = Font(
        name=BACKGROUND_FONT_NOMAL,
        color=BACKGROUND_COLOR,
        bold=True,
        size=BACKGROUND_FONT_SIZE)
    light_grey_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    grey_fill = PatternFill(
        start_color=HEARDER_COLOR,
        end_color=HEARDER_COLOR,
        fill_type=SOLID
    )
    column_widths = WIDTH_COLUMN
    barcode_columns = COLUMN_BARCODE
    # แยกข้อมูลตาม OPTIMAL_DATE และ WORKSHEET
    for (opt_date, ws_name), group_df in df.groupby(["OPTIMAL_DATE", "WORKSHEET"]):
        if sheetss:
            group_df = df
        try:
            formatted = datetime.strptime(opt_date, "%d/%m/%Y").strftime("%d%m%y")
        except:
            formatted=""
        ws_name = shorten_filename(ws_name)
        sheet_name = sanitize_sheet_name(f"{formatted}_{ws_name}")  # จำกัดชื่อไม่เกิน 31 ตัวอักษร
        worksheet = workbook.create_sheet(title=sheet_name)

        headers = list(group_df.columns)
        worksheet.append(headers)


        for row in group_df.itertuples(index=False):
            worksheet.append(list(row))


        # Apply style
        for r_idx ,row in enumerate(worksheet.iter_rows(min_row=2), start=2):
            entity_code_col = headers.index("ATTACHMENTMODE") + 1
            entity_value = worksheet.cell(row=r_idx, column=entity_code_col).value
            curr_value = worksheet.cell(row=r_idx, column=1).value
            prev_value = worksheet.cell(row=r_idx-1, column=1).value

            is_entity_one = str(entity_value) not in ["Include"]



            for cell in row:
                cell.font = default_font
                if prev_value != curr_value:
                    cell.border = thick_border
                else:
                    cell.border = thin_border
                cell.alignment = Alignment(
                    wrap_text=BACKGROUND_WRAP_TEXT,
                    horizontal=BACKGROUND_HORIZONTAL,
                    vertical=BACKGROUND_VERTICAL
                )
                if is_entity_one:
                    cell.fill = light_grey_fill 
        # Apply barcode font
        for col in barcode_columns:
            for cell in worksheet[col]:
                cell.font = barcode_font

        # Apply header style & column width
        for col, width in zip(range(1, worksheet.max_column + 1), column_widths):
            worksheet.cell(row=1, column=col).alignment = Alignment(
                                                         wrap_text=BACKGROUND_WRAP_TEXT,
                                                         horizontal=BACKGROUND_HORIZONTAL,
                                                         vertical=BACKGROUND_VERTICAL)
            worksheet.cell(row=1, column=col).border = thin_border
            worksheet.cell(row=1, column=col).fill = grey_fill
            worksheet.cell(row=1, column=col).font = white_bold_font
            worksheet.column_dimensions[get_column_letter(col)].width = width

        # ถ้าคอลัมน์เกินที่ตั้งไว้
        if len(column_widths) < worksheet.max_column:
            last_width = column_widths[-1]
            for col in range(len(column_widths) + 1, worksheet.max_column + 1):
                worksheet.column_dimensions[get_column_letter(col)].width = last_width

        worksheet.row_dimensions[1].height = 105

        # Page setup
        worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A3
        worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE
        worksheet.page_margins = openpyxl.worksheet.page.PageMargins(
            left=PRINT_SIZE_LEFT,
            right=PRINT_SIZE_RIGHT,
            top=PRINT_SIZE_TOP,
            bottom=PRINT_SIZE_BOTTOM,
            header=PRINT_SIZE_HEADER,
            footer=PRINT_SIZE_FOOTER
        )
        worksheet.page_setup.fitToWidth = PRINT_WIDTH
        worksheet.page_setup.fitToHeight = PRINT_HEIGT
        worksheet.page_setup.scale = PRINT_SCALE
        worksheet.print_title_rows = "1:1"

        worksheet.oddHeader.left.text = PRINT_TEXT_HEARDER_RIGHT
        worksheet.oddHeader.center.text = PRINT_TEXT_HEARDER_CENTER
        worksheet.oddHeader.right.text = str(export_dict.get("Version"))
        if sheetss:
            break
    # Save or return workbook
    if export_dict["save_to_file"]:
        full_name = os.path.join(export_dict["Export File Path"], f'Promotion_{export_dict["Version"]}.xlsx')
        workbook.save(full_name)
        subprocess.run(['start', 'excel', str(full_name)], shell=True)
    else:
        return workbook
