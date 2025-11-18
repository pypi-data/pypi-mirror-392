
import oracledb
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from config import ALL_TABEL_MAPPING
from model import DataMapItem,Dict
from icecream import ic

class OracleDB:
    DATAMAP:list[DataMapItem] = ALL_TABEL_MAPPING
    VALUE: Dict = {}
    DUMMY =["CS_CODE","VENDER_ID","SERVICE_ID"]
    CONFIG={"Operational Metadata":"SELECT COMP_NAME,VERSION,STATUS,MODIFIED,NAMESPACE,CONTROL,SCHEMA,SYSTIMESTAMP FROM dba_registry"
            ,"Baseline Operation":"SELECT * FROM nls_database_parameters"}
    VENDOR_CODE =None
    VENDOR_ID =None
    VENDOR_NAME =None
    SERVICE_ID =None
    USER = 'CS_SUPPORT'
    PASSWORD ='1234'
    HOST='csonldbqa01.counterservice.co.th'
    PORT='1521'
    DATABASE ='ONLPRD'
    def __init__(self,URL=None, **kwargs):
        self._set_url(URL)
    def _set_url(self,URL):
        
        if not URL is None:
            ic(URL)
            self.HOST = URL
        self.__connect()

    def __connect(self):
        try:
            self.conn = oracledb.connect(
                user=self.USER,
                password=self.PASSWORD,
                dsn=f'{self.HOST}:{self.PORT}/{self.DATABASE}'
            )
            ic(f"Connected to Oracle Database")
        except Exception as e:
            self.conn = None
            ic("Connection failed:", e)
    def _get_vendor(self,query):
        if query == '' or self.conn is None:
            return {'':''}
        cur = self.conn.cursor()
        select =f"SELECT VENDOR_CODE ,VENDOR_NAME ,VENDOR_ID,SERVICE_ID FROM ONLSTD.WS_CLIENT_CONFIG WHERE {query} "
        ic(select)
        cur.execute(select)
        cols = [c[0] for c in cur.description] # type: ignore
        data = cur.fetchone()
        if not data:
            data= ['','','','']
        return  dict(zip(cols, data))
    def __set_config_return(self):
        self.VENDOR_CODE = self.item.get('VENDOR_CODE',None)
        self.SERVICE_ID = self.item.get('SERVICE_ID',None)
        self.VENDOR_ID = self.item.get('VENDOR_ID',None)
        self.VENDOR_NAME =  self.item.get('VENDOR_NAME',None)
    def __set_config(self, client: str, service_id: str):
        where_config : str  = ''
        column = 'VENDOR_CODE'
        logi = 'AND'
        if len(client) >= 7:
            column = 'VENDOR_ID'
        if len(service_id) == 0:
            logi = 'OR'
        where_config += f" {column} = '{client}' {logi} SERVICE_ID = '{service_id}' "
        # ic(where_config)
        if not self.VENDOR_ID:
            self.item = self._get_vendor(where_config) # type: ignore
            self.__set_config_return()
            for value in self.DATAMAP:
                where_config = ''.join(f"AND {v['RESULT']} = '{self.item.get(k)}' "  for k, v in value['RULE_MAP'].items() if v.get('STATUS', 'S') == 'A')
                self.CONFIG[value['TABEL']] = f"SELECT * FROM {value['SCHEMA']}.{value['TABEL']} WHERE 1=1 {where_config} "
                # ic(self.CONFIG[value['TABEL']])
            return self.CONFIG
    def _run_query(self, query):
        """รัน query (SELECT / UPDATE)"""
        rows_cols ={"Column":[],"Row":[]}
        try:
            if self.conn is None:
                return rows_cols , pd.DataFrame(rows_cols['Row'], columns=rows_cols['Column'])
            cur = self.conn.cursor()
            cur.execute(str(query))
            rows_cols['Column'] = [c[0] for c in cur.description] # type: ignore
            rows_cols['Row'] = [[str(v).replace("\n", " ").replace("\r", " ").strip().encode('latin1').decode('tis-620',errors='ignore') 
                        if v is not None else "" 
                    for v in r]
                for r in cur.fetchall()]
            ic(rows_cols['Row'])
        except Exception as e:
            ic(f"Error: {e}")
        return rows_cols , pd.DataFrame(rows_cols['Row'], columns=rows_cols['Column'])
    def export_file(self,**kg):
        wb = Workbook()
        ws = wb.active
        ws.title = "config" # pyright: ignore[reportOptionalMemberAccess]
        header_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        column_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
        border = Border(left=Side(style="thin"), right=Side(style="thin"),
                        top=Side(style="thin"), bottom=Side(style="thin"))
        row = 1
        # data = self.fetch_tables(self.VENDOR_CODE,self.SERVICE_ID) # pyright: ignore[reportAssignmentType]
        # ic(data)
        for table_name, df in self.VALUE.items(): # type: ignore
            # ic(df.get('Row'),df.get('Column'))
            df = pd.DataFrame(df.get('Row'), columns=df.get('Column'))
            ws.cell(row=row, column=1, value=table_name) # pyright: ignore[reportCallIssue, reportOptionalMemberAccess]
            ws.cell(row=row, column=1).fill = header_fill # pyright: ignore[reportOptionalMemberAccess]
            ws.cell(row=row, column=1).font = Font(bold=True) # pyright: ignore[reportOptionalMemberAccess]
            row += 2
            if df is not None and not df.empty:
                for col_num, col_name in enumerate(df.columns, 1):
                    cell = ws.cell(row=row, column=col_num, value=col_name) # pyright: ignore[reportOptionalMemberAccess]
                    cell.fill = column_fill
                    cell.font = Font(bold=True)
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center")
                for r in df.itertuples(index=False):
                    row += 1
                    for col_num, val in enumerate(r, 1):
                        cell = ws.cell(row=row, column=col_num, value=val) # pyright: ignore[reportOptionalMemberAccess]
                        cell.border = border
                row += 3
            else:
                ws.cell(row=row, column=1, value="ไม่พบการ Set ที่ Tabel นี้") # pyright: ignore[reportOptionalMemberAccess]
                row += 3

        # Auto adjust column width
        for col in ws.columns: # pyright: ignore[reportOptionalMemberAccess]
            max_length = 0
            col_letter = col[0].column_letter # pyright: ignore[reportAttributeAccessIssue]
            for cell in col:
                
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2 # pyright: ignore[reportOptionalMemberAccess]
        wb.save(f"QCM-COU-Test Case&Result_CS Online_{self.VENDOR_CODE}_SV{str(self.SERVICE_ID).zfill(2)}_{self.VENDOR_NAME}-R1.xlsx")
        return True
    def fetch_tables(self,client = None, service= None,include=[],URL=None):
        """
        ดึงข้อมูลจากหลาย table + แสดงผลใน Jupyter
        export_excel: ถ้าใส่ชื่อไฟล์ เช่น "output.xlsx" → export เป็น Excel ด้วย
        """
        self._set_url(URL)
        tabel = [ i.get('TABEL') for i in self.DATAMAP if  len(include)==0 or i.get('TABEL') in include ]

        self.__set_config(str(client),str(service))
        # self.VALUE = { ky: result[0] 
        #               for ky, vl in self.CONFIG.items() 
        #             if ky in tabel 
        #         for result in [self._run_query(vl)]}
        
        self.VALUE = {ky: self._run_query(vl)[0] for ky, vl in self.CONFIG.items()}
        filtered = {ky: self.VALUE[ky] for ky in tabel if ky in self.VALUE}
        ic(filtered)
        return  filtered

if __name__ =="__main__":
    base = OracleDB()
    base.fetch_tables('062','00','WS_STD_SERVICE_LOOKUP')
    base.export_file()