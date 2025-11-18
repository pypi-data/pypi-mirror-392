#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from __future__ import annotations
import os
import tempfile
from abc import ABC, abstractmethod
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet._read_only import ReadOnlyWorksheet
from pandas.io.parsers import TextFileReader
from io import BytesIO, StringIO
import pandas as pd
import csv
from soup_files import File, ProgressBarAdapter
from sheet_stream.utils import VoidAdapter
from sheet_stream.erros import *
from sheet_stream.type_utils import (
    ListString, HeadCell, HeadValues, MetaDataFile,
    MetaDataItem, LibSheet,
)


class ABCSheetReader(ABC):
    """
        Classe abstrata para leitura de planilhas CSV, ODS e EXCEL.
    """

    def __init__(
            self,
            file: File | bytes | BytesIO, *,
            pbar: ProgressBarAdapter = VoidAdapter()
    ):
        self.pbar: ProgressBarAdapter = pbar
        self.file: File = file
        self.df = None
        self.isLoading: bool = False
        self._temp_path = None
        self.metadata_file: MetaDataFile = MetaDataFile.create_metadata(self.file)
        self.lib_sheet: LibSheet = LibSheet.NOT_IMPLEMENTED

    @property
    def is_running(self) -> bool:
        return self.isLoading

    @property
    def file_bytes(self) -> bytes:
        if isinstance(self.file, bytes):
            return self.file
        elif isinstance(self.file, BytesIO):
            self.file.seek(0)
            return self.file.getvalue()
        elif isinstance(self.file, File):
            try:
                return self.file.path.read_bytes()
            except FileNotFoundError:
                raise FileNotFoundError()
        else:
            raise InvalidFileSheetError('[!] ... O arquivo/planilha de ser File(), bytes, ou BytesIO()')

    @property
    def file_bytes_io(self) -> BytesIO:
        if isinstance(self.file, bytes):
            return BytesIO(self.file)
        elif isinstance(self.file, BytesIO):
            self.file.seek(0)
            return self.file
        elif isinstance(self.file, File):
            try:
                return BytesIO(self.file.path.read_bytes())
            except FileNotFoundError:
                raise FileNotFoundError()
        else:
            raise InvalidFileSheetError('[!] ... O arquivo/planilha de ser File(), bytes, ou BytesIO()')

    @property
    def file_path(self) -> File:
        """
        Retorna o caminho absoluto de leitura.
        - Se for File → caminho real
        - Se for bytes/BytesIO → salva num arquivo temporário e retorna seu path
        """
        if isinstance(self.file, File):
            return self.file

        # Cria arquivo temporário (extensão .ods)
        _tmp_sheet_file = tempfile.NamedTemporaryFile(delete=False, suffix='.ods')
        if isinstance(self.file, BytesIO):
            _tmp_sheet_file.write(self.file.getvalue())
        elif isinstance(self.file, bytes):
            _tmp_sheet_file.write(self.file)
        _tmp_sheet_file.flush()
        _tmp_sheet_file.close()
        self._temp_path = _tmp_sheet_file.name
        return File(self._temp_path)

    def _cleanup_temp(self) -> None:
        """Remove o arquivo temporário, se criado."""
        if self._temp_path is None:
            return
        if self._temp_path and os.path.exists(self._temp_path):
            try:
                os.remove(self._temp_path)
            except Exception as err:
                print(f"Aviso: falha ao remover arquivo temporário {self._temp_path}: {err}")
            finally:
                self._temp_path = None

    @abstractmethod
    def read(self, sheet_name: str = None) -> None:
        pass

    @abstractmethod
    def get_sheet_names(self) -> ListString:
        pass

    @abstractmethod
    def get_total_rows(self, sheet_name: str = None) -> int:
        pass

    @abstractmethod
    def get_total_columns(self, sheet_name: str = None) -> int:
        pass

    @abstractmethod
    def get_columns(self, sheet_name: str = None) -> HeadValues:
        pass

    @abstractmethod
    def get_dataframe(self, sheet_name: str = None) -> pd.DataFrame:
        pass


class ReaderCsv(ABCSheetReader):

    def __init__(
                self, file: File | bytes | BytesIO, *,
                pbar: ProgressBarAdapter = VoidAdapter(),
                separator: str = '\t'
            ):
        super().__init__(file, pbar=pbar)
        self.separator: str = separator
        self.df = pd.DataFrame()
        self.lib_sheet: LibSheet = LibSheet.CSV
        self.metadata_file.extension = MetaDataItem('.csv')

    def get_sheet_names(self) -> ListString:
        if isinstance(self.file, File):
            return ListString([self.file.name()])
        return ListString(['Sheet'])

    def get_total_rows(self, sheet_name: str = None) -> int:
        """
        Conta o número total de linhas (excluindo o cabeçalho) a partir dos bytes do CSV.
        """
        try:
            # Converte bytes -> texto (usando UTF-8)
            text_stream = StringIO(self.file_bytes.decode('utf-8'))
            reader = csv.reader(text_stream, delimiter=self.separator)
            # Conta todas as linhas do CSV
            total_lines = sum(1 for _ in reader)
            # Subtrai 1 do cabeçalho (se houver pelo menos uma linha)
            return max(total_lines - 1, 0)
        except Exception as e:
            self.pbar.update_text(f'Erro ao contar linhas do CSV: {e}')
            return 0

    def get_total_columns(self, sheet_name: str = None) -> int:
        try:
            text_stream = StringIO(self.file_bytes.decode('utf-8'))
            reader = csv.reader(text_stream, delimiter=self.separator)
            # Lê a primeira linha do arquivo
            _first_line = next(reader)
            # O número de colunas é o tamanho da primeira linha
            num_cols = len(_first_line)
        except StopIteration:
            self.pbar.update_text(f"Erro: O arquivo CSV está vazio.")
            return 0
        except Exception as e:
            self.pbar.update_text(f"Ocorreu um erro: {e}")
            return 0
        else:
            return num_cols

    def get_columns(self, sheet_name: str = None) -> HeadValues:
        try:
            text_stream = StringIO(self.file_bytes.decode('utf-8'))
            reader = csv.reader(text_stream, delimiter=self.separator)
            head_csv = next(reader)  # lê a primeira linha
        except Exception as err:
            self.pbar.update_text(f'{__class__.__name__}: {err}')
            return HeadValues([])
        else:
            return HeadValues([HeadCell(x) for x in head_csv])

    def read(self, sheet_name: str = None) -> None:
        self.isLoading = True
        self.pbar.start()
        self.pbar.update(0, "Iniciando leitura do CSV")

        total_lines = self.get_total_rows()
        chunks_text: list[TextFileReader] = []
        for num, chunk in enumerate(
                    pd.read_csv(self.file.absolute(), chunksize=1000, sep=self.separator),
                ):
            #percent: float = (((num + 1) * 1000) / total_lines) * 100
            #self.pbar.update(percent, f"Lendo CSV [{self.file.basename()}]")
            chunks_text.append(chunk)

        if len(chunks_text) > 0:
            self.df = pd.concat(chunks_text, ignore_index=True)
        self.pbar.update(100, "Leitura finalizada!", )
        self.isLoading = False
        self.pbar.stop()

    def get_dataframe(self, sheet_name: str = None) -> pd.DataFrame:
        if self.df.empty:
            self.read()
        return self.df


class ReaderExcel(ABCSheetReader):

    def __init__(self, file: File | bytes | BytesIO, *, pbar: ProgressBarAdapter = VoidAdapter()):
        super().__init__(file, pbar=pbar)
        self.df = pd.DataFrame()
        self.lib_sheet: LibSheet = LibSheet.EXCEL
        self.metadata_file.extension = MetaDataItem('.xlsx')

    def get_columns(self, sheet_name: str = None) -> HeadValues:
        """Cabeçalho original da planilha"""
        ws = self.get_work_sheet(sheet_name)
        if ws is None:
            return HeadValues([])
        __head = list(next(ws.iter_rows(values_only=True)))
        __head_values = HeadValues([])
        for x in __head:
            __head_values.append(x)
        return __head_values

    def get_total_rows(self, sheet_name: str = None) -> int:
        return self.get_work_sheet(sheet_name).max_row

    def get_total_columns(self, sheet_name: str = None) -> int:
        return self.get_work_sheet(sheet_name).max_column

    def get_sheet_names(self) -> ListString:
        try:
            return ListString(self.get_workbook().sheetnames)
        except Exception as e:
            print(e)
            return ListString([])

    def get_workbook(self) -> Workbook | None:
        self.pbar.start()
        self.pbar.update_text(f'Lendo Excel')
        try:
            excel_wb: Workbook = load_workbook(self.file_bytes_io, read_only=True)
        except Exception as e:
            self.pbar.update_text(f'{__class__.__name__} -> {e}\n')
            self.pbar.stop()
            return None
        else:
            self.pbar.stop()
            return excel_wb

    def get_active_sheet(self) -> ReadOnlyWorksheet | None:
        return self.get_workbook().active

    def get_work_sheet(self, sheet_name: str = None) -> ReadOnlyWorksheet | None:
        file_workbook = self.get_workbook()
        if file_workbook is None:
            return None

        if sheet_name is None:
            return file_workbook.active
        else:
            _names = self.get_sheet_names()
            if not sheet_name in _names:
                return None
            return file_workbook[sheet_name]

    def read(self, sheet_name: str = None) -> None:
        return self._read_yes_progress(sheet_name)

    def _read_yes_progress(self, sheet_name: str = None) -> None:
        """
        Lê o Excel inteiro diretamente em um DataFrame usando pandas.read_excel(),
        sem iterar linha a linha. Muito mais rápido para planilhas grandes.
        """
        self.isLoading = True
        self.pbar.start()

        try:
            # Determina qual aba será lida
            sheet_to_read = sheet_name or self.get_sheet_names()[0]
            self.pbar.update(0, f"Iniciando leitura do Excel {sheet_to_read}")
            # Lê direto em DataFrame
            self.df = pd.read_excel(self.file_bytes_io, sheet_name=sheet_to_read, engine='openpyxl')
            # Garante que o cabeçalho esteja correto
            if (self.df.columns is None) or (len(self.df.columns) == 0):
                ws = self.get_work_sheet(sheet_to_read)
                if ws is not None:
                    self.df.columns = list(next(ws.iter_rows(values_only=True)))
            self.pbar.update(100, 'Operação finalizada!')
        except Exception as e:
            self.pbar.update_text(f'Erro ao ler Excel: {e}')
        finally:
            self.isLoading = False
            self.pbar.stop()

    def get_dataframe(self, sheet_name: str = None) -> pd.DataFrame:
        if self.df.empty:
            self.read(sheet_name)
        return self.df


class ReaderOds(ABCSheetReader):

    def __init__(self, file: File | bytes | BytesIO, *, pbar: ProgressBarAdapter = VoidAdapter()):
        super().__init__(file, pbar=pbar)
        self.df = pd.DataFrame()
        self.lib_sheet: LibSheet = LibSheet.ODS
        self.metadata_file.extension = MetaDataItem('.ods')

    def get_columns(self, sheet_name: str = None) -> HeadValues:
        self.read(sheet_name)
        try:
            return HeadValues(self.df.columns.to_list())
        except Exception as err:
            print(f'Erro ao ler cabeçalho: {err}')
            return HeadValues([])

    def get_total_rows(self, sheet_name: str = None) -> int:
        self.read(sheet_name)
        try:
            return self.df.shape[0]
        except Exception as err:
            print(f'Erro ao contar linhas: {err}')
            return 0

    def get_total_columns(self, sheet_name: str = None) -> int:
        cols = self.get_columns(sheet_name)
        return len(cols)

    def get_sheet_names(self) -> ListString:
        try:
            xls = pd.ExcelFile(self.file_bytes_io, engine='odf')
            return ListString(xls.sheet_names)
        except Exception as err:
            print(f'Erro ao obter nomes das abas: {err}')
            return ListString([])

    def read(self, sheet_name: str = None) -> None:
        if not self.df.empty:
            return

        self.isLoading = True
        self.pbar.start()
        self.pbar.update(0, f'Iniciando leitura do ODS')

        # 1 - Tentar a leitura com pandas
        self.read_with_pandas(sheet_name)
        if not self.df.empty:
            return

        # 2 - Tentar a leitura com odfpy.
        self._read_with_odfpy(sheet_name)
        if not self.df.empty:
            return

        # 3 - Tentar a leitura com XML.
        self._read_xml_raw(sheet_name)

        self.pbar.update(100, 'Operação finalizada!')
        self.pbar.stop()
        self._cleanup_temp()
        self.isLoading = False

    def read_with_pandas(self, sheet_name: str = None) -> None:
        self.pbar.update(0, f'Lendo com Pandas')
        try:
            result = pd.read_excel(
                self.file_bytes_io,
                sheet_name=sheet_name,
                engine='odf',
                dtype=str,
                parse_dates=False,
                na_values=[],
                keep_default_na=False
            )

            # Se o resultado for dict, pega a primeira aba
            if isinstance(result, dict):
                first_key = next(iter(result))
                self.df = result[first_key]
                print(f"DEBUG: Várias abas detectadas, usando '{first_key}'")
            else:
                self.df = result

        except Exception as err_pandas:
            self.pbar.update_text(f"Aviso: Leitura com Pandas falhou: {err_pandas}")
            print(f"DEBUG CRÍTICO (Pandas): {err_pandas}")

    def _read_with_odfpy(self, sheet_name: str = None) -> None:
        """
        Modo seguro de leitura ODS usando odfpy diretamente.
        Compatível com versões antigas e novas do odfpy.
        """
        self.pbar.update(0, f'Lendo com OdfPy')
        try:
            import odf.opendocument
            from odf.table import Table, TableRow, TableCell
            from odf.text import P
            try:
                # Nova versão
                from odf.text import teletype
            except ImportError:
                # Compatível com versões antigas
                from odf import teletype

            doc = odf.opendocument.load(self.file_bytes_io)
            sheets = [t for t in doc.spreadsheet.getElementsByType(Table)]

            if not sheets:
                self.df = pd.DataFrame()
                return

            # Seleciona a aba desejada
            if sheet_name:
                sheet_to_read = next((s for s in sheets if s.getAttribute("name") == sheet_name), None)
                if sheet_to_read is None:
                    print(f"DEBUG: Aba '{sheet_name}' não encontrada, usando a primeira.")
                    sheet_to_read = sheets[0]
            else:
                sheet_to_read = sheets[0]

            data_rows = []
            for row in sheet_to_read.getElementsByType(TableRow):
                values = []
                for cell in row.getElementsByType(TableCell):
                    paragraphs = cell.getElementsByType(P)
                    text_content = ' '.join(teletype.extractText(p) for p in paragraphs)
                    values.append(text_content.strip())
                data_rows.append(values)

            if not data_rows:
                self.df = pd.DataFrame()
                return

            columns = data_rows[0]
            data = data_rows[1:]
            self.df = pd.DataFrame(data, columns=columns, dtype=str)

        except Exception as err:
            self.pbar.update_text(f'Erro Pandas/odfpy (leitura bruta): {err}')
            print(f"DEBUG: Falha no modo seguro de leitura ODS: {err}")

    def _read_xml_raw(self, sheet_name: str = None) -> None:
        """
        Leitura bruta do ODS via XML (modo ultra seguro).
        Extrai apenas texto visível das células.
        Agora suporta seleção de aba (sheet_name).
        """
        import zipfile
        from xml.etree import ElementTree as ET

        self.pbar.update(0, f'Lendo com XML')
        try:
            with zipfile.ZipFile(self.file_bytes_io) as zf:
                content = zf.read("content.xml")

            xml_root = ET.fromstring(content)

            ns = {
                'office': 'urn:oasis:names:tc:opendocument:xmlns:office:1.0',
                'table': 'urn:oasis:names:tc:opendocument:xmlns:table:1.0',
                'text': 'urn:oasis:names:tc:opendocument:xmlns:text:1.0'
            }

            # Obtém todas as tabelas (abas)
            sheets = xml_root.findall('.//table:table', ns)
            if not sheets:
                print("DEBUG: Nenhuma aba encontrada no ODS.")
                return

            # Seleciona aba específica (ou a primeira)
            if sheet_name:
                sheet = next(
                    (s for s in sheets if s.get('{urn:oasis:names:tc:opendocument:xmlns:table:1.0}name') == sheet_name),
                    None
                )
                if sheet is None:
                    print(f"DEBUG: Aba '{sheet_name}' não encontrada, usando a primeira disponível.")
                    sheet = sheets[0]
            else:
                sheet = sheets[0]

            rows = []
            for row in sheet.findall('table:table-row', ns):
                row_data = []
                for cell in row.findall('table:table-cell', ns):
                    # Captura textos (pode haver mais de um <text:p> por célula)
                    texts = [t.text or '' for t in cell.findall('.//text:p', ns)]
                    cell_text = ' '.join(texts).strip()
                    # Lida com células repetidas (número de repetições)
                    repeat = cell.get('{urn:oasis:names:tc:opendocument:xmlns:table:1.0}number-columns-repeated}')
                    repeat = int(repeat) if repeat else 1
                    row_data.extend([cell_text] * repeat)
                if any(row_data):
                    rows.append(row_data)

            if not rows:
                print("DEBUG: Nenhuma linha encontrada na aba.")
                return

            columns = rows[0]
            data = rows[1:]
            self.df = pd.DataFrame(data, columns=columns, dtype=str)
        except Exception as err:
            print(f"DEBUG: Falha ao ler XML bruto ODS: {err}")

    def get_dataframe(self, sheet_name: str = None) -> pd.DataFrame:
        self.read(sheet_name)
        return self.df


class ReadFileSheet(object):

    def __init__(
            self, file: File | bytes | BytesIO, *,
            separator: str = '\t',
            lib_sheet: LibSheet = LibSheet.NOT_IMPLEMENTED,
            pbar: ProgressBarAdapter = VoidAdapter()
    ):
        if isinstance(file, str):
            file = File(file)

        if isinstance(file, File):
            if file.is_csv():
                self.sheet_reader: ABCSheetReader = ReaderCsv(file, pbar=pbar, separator=separator)
            elif file.is_excel():
                self.sheet_reader: ABCSheetReader = ReaderExcel(file, pbar=pbar)
            elif file.is_ods():
                self.sheet_reader: ABCSheetReader = ReaderOds(file, pbar=pbar)
            else:
                raise InvalidFileSheetError(f'Tipo de planilha desconhecida: {file.basename()}')
        else:
            if lib_sheet == LibSheet.CSV:
                self.sheet_reader: ABCSheetReader = ReaderCsv(file, pbar=pbar, separator=separator)
            elif lib_sheet == LibSheet.EXCEL:
                self.sheet_reader: ABCSheetReader = ReaderExcel(file, pbar=pbar)
            elif lib_sheet == LibSheet.ODS:
                self.sheet_reader: ABCSheetReader = ReaderOds(file, pbar=pbar)
            else:
                raise InvalidFileSheetError(f'Tipo de planilha desconhecida')

    @property
    def lib_sheet(self) -> LibSheet:
        return self.sheet_reader.lib_sheet

    def get_metadata(self) -> MetaDataFile:
        return self.sheet_reader.metadata_file

    def set_metadata(self, new: MetaDataFile):
        self.sheet_reader.metadata_file = new

    def read(self, sheet_name: str = None) -> None:
        self.sheet_reader.read(sheet_name)

    def get_sheet_names(self) -> ListString:
        return self.sheet_reader.get_sheet_names()

    def get_total_rows(self, sheet_name: str = None) -> int:
        return self.sheet_reader.get_total_rows(sheet_name)

    def get_total_columns(self, sheet_name: str = None) -> int:
        return self.sheet_reader.get_total_columns(sheet_name)

    def get_columns(self, sheet_name: str = None) -> HeadValues:
        return self.sheet_reader.get_columns(sheet_name)

    def get_dataframe(self, sheet_name: str = None) -> pd.DataFrame:
        return self.sheet_reader.get_dataframe(sheet_name)
