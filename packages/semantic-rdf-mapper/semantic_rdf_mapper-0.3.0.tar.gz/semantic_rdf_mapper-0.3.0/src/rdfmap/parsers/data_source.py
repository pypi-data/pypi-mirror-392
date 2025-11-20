"""High-performance data source parsers using Polars for big data processing."""

from abc import ABC, abstractmethod
from pathlib import Path
from typing import Generator, List, Optional, Any, Dict
import json
import xml.etree.ElementTree as ET

import polars as pl


class DataSourceParser(ABC):
    """Abstract base class for data source parsers using Polars."""

    @abstractmethod
    def parse(
        self, chunk_size: Optional[int] = None
    ) -> Generator[pl.DataFrame, None, None]:
        """Parse data source and yield Polars DataFrames in chunks."""
        pass

    @abstractmethod
    def get_column_names(self) -> List[str]:
        """Get list of column names."""
        pass


class CSVParser(DataSourceParser):
    """High-performance CSV parser using Polars."""

    def __init__(
        self,
        file_path: Path,
        delimiter: str = ",",
        has_header: bool = True,
        encoding: str = "utf8",
    ):
        """Initialize CSV parser.

        Args:
            file_path: Path to CSV file
            delimiter: Column delimiter
            has_header: Whether first row is header
            encoding: File encoding
        """
        self.file_path = file_path
        self.delimiter = delimiter
        self.has_header = has_header
        self.encoding = encoding

        if not self.file_path.exists():
            raise FileNotFoundError(f"CSV file not found: {self.file_path}")

    def parse(
        self, chunk_size: Optional[int] = None
    ) -> Generator[pl.DataFrame, None, None]:
        """Parse CSV file and yield Polars DataFrames.

        Args:
            chunk_size: Number of rows per chunk. If None, load entire file.

        Yields:
            Polars DataFrames containing parsed data
        """
        if chunk_size:
            # Get column names from header once
            if self.has_header:
                header_df = pl.read_csv(
                    self.file_path,
                    separator=self.delimiter,
                    has_header=True,
                    n_rows=0,  # Just get column names
                )
                column_names = header_df.columns
            else:
                column_names = None

            # True chunked processing without pre-loading entire dataset
            offset = 0
            skip_rows_start = 1 if self.has_header else 0  # Skip header row

            while True:
                # Read chunk directly without pre-calculating total rows
                try:
                    # Calculate how many rows to skip (header + previously read data rows)
                    actual_skip = skip_rows_start + offset

                    chunk = pl.read_csv(
                        self.file_path,
                        separator=self.delimiter,
                        has_header=False,  # Don't treat any row as header
                        encoding=self.encoding if self.encoding in ['utf8', 'utf8-lossy'] else 'utf8',
                        null_values=[""],
                        ignore_errors=True,
                        skip_rows=actual_skip,
                        n_rows=chunk_size,
                    )

                    if len(chunk) == 0:
                        break

                    # Apply column names if we have header
                    if column_names:
                        chunk = chunk.rename({f"column_{i+1}": name for i, name in enumerate(column_names)})

                    yield chunk
                    offset += len(chunk)

                    # If we got fewer rows than requested, we've reached the end
                    if len(chunk) < chunk_size:
                        break

                except Exception as e:
                    # Silent break on empty CSV or EOF
                    break
        else:
            # Load entire file using lazy evaluation
            df = pl.read_csv(
                self.file_path,
                separator=self.delimiter,
                has_header=self.has_header,
                encoding=self.encoding if self.encoding in ['utf8', 'utf8-lossy'] else 'utf8',
                null_values=[""],
                ignore_errors=True,
            )
            yield df

    def get_column_names(self) -> List[str]:
        """Get list of column names from CSV."""
        if not self.has_header:
            # Read first row to determine number of columns
            df_sample = pl.read_csv(
                self.file_path,
                separator=self.delimiter,
                n_rows=1,
                has_header=False,
            )
            return [f"Column_{i}" for i in range(df_sample.width)]

        # Use lazy scan to get column names efficiently
        lazy_df = pl.scan_csv(
            self.file_path,
            separator=self.delimiter,
            has_header=self.has_header,
        )
        return lazy_df.collect_schema().names()


class XLSXParser(DataSourceParser):
    """XLSX parser using Polars with openpyxl backend."""

    def __init__(
        self,
        file_path: Path,
        sheet_name: Optional[str] = None,
        has_header: bool = True,
    ):
        """Initialize XLSX parser.

        Args:
            file_path: Path to XLSX file
            sheet_name: Name of sheet to read. If None, reads first sheet.
            has_header: Whether first row is header
        """
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.has_header = has_header

        if not self.file_path.exists():
            raise FileNotFoundError(f"XLSX file not found: {self.file_path}")

    def parse(
        self, chunk_size: Optional[int] = None
    ) -> Generator[pl.DataFrame, None, None]:
        """Parse XLSX file and yield Polars DataFrames.

        Args:
            chunk_size: Number of rows per chunk. If None, load entire sheet.

        Yields:
            Polars DataFrames containing parsed data
        """
        try:
            # Use Polars Excel reading
            df = pl.read_excel(
                self.file_path,
                sheet_name=self.sheet_name,
                has_header=self.has_header,
            )
        except (ImportError, AttributeError):
            # Use openpyxl manually if Polars Excel support not available
            from openpyxl import load_workbook

            wb = load_workbook(self.file_path, read_only=True)
            ws = wb[self.sheet_name] if self.sheet_name else wb.active

            # Extract data as list of lists
            data = []
            for row in ws.iter_rows(values_only=True):
                data.append(list(row))

            if not data:
                return

            # Create DataFrame
            if self.has_header and data:
                columns = [str(col) if col is not None else f"Column_{i}" for i, col in enumerate(data[0])]
                df = pl.DataFrame(data[1:], schema=columns)
            else:
                df = pl.DataFrame(data)

        if chunk_size and len(df) > chunk_size:
            # Yield in chunks
            for i in range(0, len(df), chunk_size):
                yield df.slice(i, chunk_size)
        else:
            # Yield entire DataFrame
            yield df

    def get_column_names(self) -> List[str]:
        """Get list of column names from XLSX."""
        if not self.has_header:
            # Read first row to determine columns
            try:
                df_sample = pl.read_excel(
                    self.file_path,
                    sheet_name=self.sheet_name,
                    has_header=False,
                    n_rows=1,
                )
                return [f"Column_{i}" for i in range(df_sample.width)]
            except (ImportError, AttributeError):
                # Fallback to manual reading
                from openpyxl import load_workbook
                wb = load_workbook(self.file_path, read_only=True)
                ws = wb[self.sheet_name] if self.sheet_name else wb.active
                first_row = next(ws.iter_rows(values_only=True), [])
                return [f"Column_{i}" for i in range(len(first_row))]

        # Get column names
        try:
            df_sample = pl.read_excel(
                self.file_path,
                sheet_name=self.sheet_name,
                has_header=True,
                n_rows=0,
            )
            return df_sample.columns
        except (ImportError, AttributeError):
            # Fallback to manual reading
            from openpyxl import load_workbook
            wb = load_workbook(self.file_path, read_only=True)
            ws = wb[self.sheet_name] if self.sheet_name else wb.active
            first_row = next(ws.iter_rows(values_only=True), [])
            return [str(col) if col is not None else f"Column_{i}" for i, col in enumerate(first_row)]

    def list_sheets(self) -> List[str]:
        """List all sheet names in the workbook."""
        from openpyxl import load_workbook
        wb = load_workbook(self.file_path, read_only=True)
        return wb.sheetnames


class JSONParser(DataSourceParser):
    """JSON parser with array expansion using Polars."""

    def __init__(self, file_path: Path, encoding: str = "utf-8"):
        """Initialize JSON parser.

        Args:
            file_path: Path to JSON file
            encoding: File encoding
        """
        self.file_path = file_path
        self.encoding = encoding

        if not self.file_path.exists():
            raise FileNotFoundError(f"JSON file not found: {self.file_path}")

    def parse(
        self, chunk_size: Optional[int] = None
    ) -> Generator[pl.DataFrame, None, None]:
        """Parse JSON file and yield Polars DataFrames.

        Args:
            chunk_size: Number of rows per chunk. If None, process all data.

        Yields:
            Polars DataFrames containing flattened JSON data
        """
        with open(self.file_path, 'r', encoding=self.encoding) as f:
            data = json.load(f)

        # Flatten and expand JSON data
        flattened_data = self._flatten_json_data(data)

        if not flattened_data:
            return

        # Convert to Polars DataFrame
        df = pl.DataFrame(flattened_data)

        if chunk_size and len(df) > chunk_size:
            # Yield in chunks
            for i in range(0, len(df), chunk_size):
                yield df.slice(i, chunk_size)
        else:
            yield df

    def _flatten_json_data(self, data: Any, prefix: str = "") -> List[Dict[str, Any]]:
        """Flatten nested JSON data with array expansion.

        Args:
            data: JSON data to flatten
            prefix: Prefix for nested keys

        Returns:
            List of flattened dictionaries
        """
        if isinstance(data, list):
            # Handle top-level arrays
            if not data:
                return []

            # Check if it's an array of simple objects (no nesting)
            # If all items are dicts with same keys, treat as tabular data
            if all(isinstance(item, dict) for item in data):
                # Check if this is a simple flat structure
                sample_keys = set()
                for item in data[:5]:  # Sample first 5 items
                    sample_keys.update(item.keys())

                # Check if any values are complex (nested)
                has_nested = False
                for item in data[:5]:
                    for value in item.values():
                        if isinstance(value, (dict, list)):
                            has_nested = True
                            break
                    if has_nested:
                        break

                if not has_nested:
                    # Simple flat array of objects - return as-is
                    result = []
                    for item in data:
                        flattened = {}
                        for key, value in item.items():
                            new_key = f"{prefix}.{key}" if prefix else key
                            flattened[new_key] = value
                        result.append(flattened)
                    return result
                else:
                    # Has nested structures - need to expand
                    result = []
                    for item in data:
                        sub_rows = self._flatten_json_data(item, prefix)
                        result.extend(sub_rows)
                    return result
            else:
                # Mixed types or simple values - use array indexing
                result = []
                for i, item in enumerate(data):
                    if isinstance(item, (dict, list)):
                        sub_rows = self._flatten_json_data(item, f"{prefix}[{i}]")
                        result.extend(sub_rows)
                    else:
                        result.append({f"{prefix}[{i}]": item})
                return result

        elif isinstance(data, dict):
            # Handle dictionary - flatten keys
            result = []
            flattened = {}
            arrays_to_expand = {}

            for key, value in data.items():
                new_key = f"{prefix}.{key}" if prefix else key

                if isinstance(value, list):
                    # Handle arrays
                    if not value:  # Empty array
                        flattened[new_key] = None
                    elif all(isinstance(item, dict) for item in value):
                        # Array of objects - expand into multiple rows
                        arrays_to_expand[new_key] = value
                    elif all(not isinstance(item, (dict, list)) for item in value):
                        # Array of simple values - could convert to string or expand
                        # For now, store as None and expand below
                        arrays_to_expand[new_key] = value
                    else:
                        # Mixed array - expand with indexing
                        for i, item in enumerate(value):
                            if isinstance(item, (dict, list)):
                                sub_rows = self._flatten_json_data(item, f"{new_key}[{i}]")
                                result.extend(sub_rows)
                            else:
                                flattened[f"{new_key}[{i}]"] = item
                elif isinstance(value, dict):
                    # Recursively flatten nested dictionaries
                    nested_flat = self._flatten_dict(value, new_key)
                    flattened.update(nested_flat)
                else:
                    # Simple value
                    flattened[new_key] = value

            # Handle array expansion
            if arrays_to_expand:
                for array_key, array_values in arrays_to_expand.items():
                    for item in array_values:
                        row = flattened.copy()
                        if isinstance(item, dict):
                            # Flatten the dict and merge
                            item_flat = self._flatten_dict(item, array_key)
                            row.update(item_flat)
                        else:
                            # Simple value
                            row[array_key] = item
                        result.append(row)

            if result:
                return result
            else:
                return [flattened] if flattened else []

        else:
            # Simple value at root
            return [{"value": data}]

    def _flatten_dict(self, data: Dict[str, Any], prefix: str = "") -> Dict[str, Any]:
        """Flatten a dictionary without array expansion.

        Args:
            data: Dictionary to flatten
            prefix: Prefix for keys

        Returns:
            Flattened dictionary
        """
        result = {}
        for key, value in data.items():
            new_key = f"{prefix}.{key}" if prefix else key

            if isinstance(value, dict):
                # Recursively flatten nested dict
                nested = self._flatten_dict(value, new_key)
                result.update(nested)
            elif isinstance(value, list):
                # For lists in dict flattening, just store None or join strings
                if value and all(isinstance(item, str) for item in value):
                    result[new_key] = ", ".join(value)
                else:
                    result[new_key] = None  # Will be expanded in main logic
            else:
                result[new_key] = value

        return result

    def get_column_names(self) -> List[str]:
        """Get list of column names from JSON."""
        # Parse a sample to determine columns
        with open(self.file_path, 'r', encoding=self.encoding) as f:
            data = json.load(f)

        flattened_data = self._flatten_json_data(data)
        if flattened_data:
            # Get all unique keys across all flattened rows
            all_keys = set()
            for row in flattened_data[:100]:  # Sample first 100 rows
                all_keys.update(row.keys())
            return sorted(list(all_keys))
        return []


class XMLParser(DataSourceParser):
    """XML parser using Polars."""

    def __init__(
        self,
        file_path: Path,
        row_xpath: str = "./*",
        encoding: str = "utf-8"
    ):
        """Initialize XML parser.

        Args:
            file_path: Path to XML file
            row_xpath: XPath expression for row elements
            encoding: File encoding
        """
        self.file_path = file_path
        self.row_xpath = row_xpath
        self.encoding = encoding

        if not self.file_path.exists():
            raise FileNotFoundError(f"XML file not found: {self.file_path}")

    def parse(
        self, chunk_size: Optional[int] = None
    ) -> Generator[pl.DataFrame, None, None]:
        """Parse XML file and yield Polars DataFrames.

        Args:
            chunk_size: Number of rows per chunk. If None, process all data.

        Yields:
            Polars DataFrames containing XML data
        """
        tree = ET.parse(self.file_path)
        root = tree.getroot()

        # Find all row elements using XPath
        row_elements = root.findall(self.row_xpath)

        # Convert XML elements to dictionaries
        rows_data = []
        for element in row_elements:
            row_dict = self._xml_element_to_dict(element)
            rows_data.append(row_dict)

        if not rows_data:
            return

        # Convert to Polars DataFrame
        df = pl.DataFrame(rows_data)

        if chunk_size and len(df) > chunk_size:
            # Yield in chunks
            for i in range(0, len(df), chunk_size):
                yield df.slice(i, chunk_size)
        else:
            yield df

    def _xml_element_to_dict(self, element: ET.Element) -> Dict[str, Any]:
        """Convert XML element to dictionary.

        Args:
            element: XML element

        Returns:
            Dictionary representation of the element
        """
        result = {}

        # Add attributes
        if element.attrib:
            for key, value in element.attrib.items():
                result[f"@{key}"] = value

        # Add text content
        if element.text and element.text.strip():
            result["text"] = element.text.strip()

        # Add child elements
        for child in element:
            child_dict = self._xml_element_to_dict(child)

            if child.tag in result:
                # Handle multiple elements with same tag
                if not isinstance(result[child.tag], list):
                    result[child.tag] = [result[child.tag]]
                result[child.tag].append(child_dict)
            else:
                result[child.tag] = child_dict

        return result

    def get_column_names(self) -> List[str]:
        """Get list of column names from XML."""
        tree = ET.parse(self.file_path)
        root = tree.getroot()

        # Sample first few elements to determine columns
        row_elements = root.findall(self.row_xpath)[:10]

        all_keys = set()
        for element in row_elements:
            row_dict = self._xml_element_to_dict(element)
            all_keys.update(self._flatten_dict_keys(row_dict))

        return sorted(list(all_keys))

    def _flatten_dict_keys(self, d: Dict[str, Any], prefix: str = "") -> List[str]:
        """Flatten dictionary keys recursively.

        Args:
            d: Dictionary to flatten
            prefix: Prefix for nested keys

        Returns:
            List of flattened keys
        """
        keys = []
        for key, value in d.items():
            new_key = f"{prefix}.{key}" if prefix else key

            if isinstance(value, dict):
                keys.extend(self._flatten_dict_keys(value, new_key))
            elif isinstance(value, list) and value and isinstance(value[0], dict):
                # Handle list of dictionaries
                keys.extend(self._flatten_dict_keys(value[0], f"{new_key}[0]"))
            else:
                keys.append(new_key)

        return keys


def create_parser(
    file_path: Path,
    delimiter: str = ",",
    has_header: bool = True,
    encoding: str = "utf-8",
    sheet_name: Optional[str] = None,
    row_xpath: str = "./*",
) -> DataSourceParser:
    """Create appropriate parser based on file extension.

    Args:
        file_path: Path to data file
        delimiter: CSV delimiter
        has_header: Whether file has header row
        encoding: File encoding
        sheet_name: Excel sheet name
        row_xpath: XPath for XML row elements

    Returns:
        Appropriate parser instance

    Raises:
        ValueError: If file type is not supported
    """
    suffix = file_path.suffix.lower()

    if suffix in [".csv", ".tsv", ".txt"]:
        if suffix == ".tsv":
            delimiter = "\t"
        return CSVParser(file_path, delimiter, has_header, encoding)
    elif suffix in [".xlsx", ".xls"]:
        return XLSXParser(file_path, sheet_name, has_header)
    elif suffix == ".json":
        return JSONParser(file_path, encoding)
    elif suffix == ".xml":
        return XMLParser(file_path, row_xpath, encoding)
    else:
        raise ValueError(f"Unsupported file type: {suffix}")
