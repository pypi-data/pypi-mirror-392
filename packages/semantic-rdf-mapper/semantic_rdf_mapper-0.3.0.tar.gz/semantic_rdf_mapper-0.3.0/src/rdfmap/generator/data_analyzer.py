"""Enhanced data source analyzer supporting CSV, XLSX, JSON, and XML formats."""

from typing import Dict, List, Optional, Any
from pathlib import Path
import polars as pl
import re
import json
import xml.etree.ElementTree as ET

# Import Polars helper functions
from .polars_helpers import (_infer_polars_type, _suggest_xsd_datatype_polars,
                           _is_likely_identifier_polars, _detect_pattern_polars)


class DataFieldAnalysis:
    """Analysis results for a single data field (column/attribute/key)."""

    def __init__(self, name: str, path: str = None):
        self.name = name
        self.path = path or name  # JSON path or XML XPath for nested data
        self.sample_values: List[Any] = []
        self.null_count: int = 0
        self.total_count: int = 0
        self.inferred_type: Optional[str] = None
        self.is_identifier: bool = False
        self.is_unique: bool = False
        self.suggested_datatype: Optional[str] = None
        self.pattern: Optional[str] = None
        self.is_nested: bool = False
        self.nested_fields: List['DataFieldAnalysis'] = []
        self.parent_path: Optional[str] = None

    @property
    def null_percentage(self) -> float:
        """Calculate percentage of null values."""
        if self.total_count == 0:
            return 0.0
        return (self.null_count / self.total_count) * 100

    @property
    def is_required(self) -> bool:
        """Determine if field is required (low null percentage)."""
        return self.null_percentage < 20.0

    @property
    def uniqueness(self) -> float:
        """Calculate uniqueness ratio."""
        if self.total_count == 0:
            return 0.0
        unique_count = len(set(str(v) for v in self.sample_values))
        return unique_count / self.total_count


class DataSourceAnalyzer:
    """Enhanced analyzer for multiple data source formats (CSV, XLSX, JSON, XML)."""

    def __init__(self, file_path: str):
        """Initialize analyzer with data file.

        Args:
            file_path: Path to data file (CSV, XLSX, JSON, or XML)
        """
        self.file_path = Path(file_path)
        self.file_extension = self.file_path.suffix.lower()
        self.data_format = self._detect_format()

        # Analysis results
        self.field_analyses: Dict[str, DataFieldAnalysis] = {}
        self.total_rows: int = 0
        self.sample_data: List[Dict[str, Any]] = []
        self.nested_structure: Dict[str, Any] = {}
        self.has_multiple_sheets: bool = False  # For Excel files
        self.sheet_count: int = 0  # For Excel files

        # Perform analysis
        self._analyze()

    def _detect_format(self) -> str:
        """Detect data format from file extension."""
        format_map = {
            '.csv': 'csv',
            '.tsv': 'csv',
            '.xlsx': 'excel',
            '.xls': 'excel',
            '.json': 'json',
            '.jsonl': 'json',
            '.xml': 'xml',
        }

        format_type = format_map.get(self.file_extension)
        if not format_type:
            raise ValueError(f"Unsupported data format: {self.file_extension}")

        return format_type

    def _analyze(self) -> None:
        """Perform format-specific analysis."""
        if self.data_format == 'csv':
            self._analyze_csv()
        elif self.data_format == 'excel':
            self._analyze_excel()
        elif self.data_format == 'json':
            self._analyze_json()
        elif self.data_format == 'xml':
            self._analyze_xml()

    def _analyze_csv(self) -> None:
        """Analyze CSV/TSV data."""
        try:
            # Read with Polars for high performance
            df = pl.read_csv(self.file_path, n_rows=100)  # Sample first 100 rows
            self.total_rows = len(df)

            # Convert to list of dictionaries for uniform processing
            self.sample_data = df.to_dicts()

            # Analyze each column
            for col_name in df.columns:
                analysis = DataFieldAnalysis(col_name)
                column = df[col_name]

                # Basic statistics
                analysis.total_count = len(df)
                analysis.null_count = column.null_count()
                analysis.sample_values = column.drop_nulls().head(10).to_list()

                # Infer type and datatype using Polars helpers
                analysis.inferred_type = _infer_polars_type(column)
                analysis.suggested_datatype = _suggest_xsd_datatype_polars(column)

                # Check uniqueness and identifier potential
                non_null = column.drop_nulls()
                analysis.is_unique = non_null.n_unique() == len(non_null)
                analysis.is_identifier = _is_likely_identifier_polars(col_name, column)

                # Pattern detection
                analysis.pattern = _detect_pattern_polars(column)

                self.field_analyses[col_name] = analysis

        except Exception as e:
            raise ValueError(f"Failed to analyze CSV file: {e}")

    def _analyze_excel(self) -> None:
        """Analyze Excel data."""
        try:
            # First, check if file has multiple sheets
            try:
                from openpyxl import load_workbook
                wb = load_workbook(self.file_path, read_only=True)
                self.sheet_count = len(wb.sheetnames)
                self.has_multiple_sheets = self.sheet_count > 1
                wb.close()
            except ImportError:
                # If openpyxl not available, assume single sheet
                self.sheet_count = 1
                self.has_multiple_sheets = False

            # Read Excel file using Polars - use first sheet for now
            try:
                # Use sheet_id instead of sheet_name for compatibility
                df = pl.read_excel(self.file_path, sheet_id=1)
                df = df.head(100)  # Sample first 100 rows
            except (ImportError, AttributeError, ValueError) as e:
                # Fallback to openpyxl for Excel reading
                from openpyxl import load_workbook
                wb = load_workbook(self.file_path, read_only=True)
                ws = wb.active

                # Extract data as list of lists
                data = []
                for row in ws.iter_rows(values_only=True, max_row=101):  # Header + 100 rows
                    data.append(list(row))

                wb.close()

                if data:
                    columns = [str(col) if col is not None else f"Column_{i}" for i, col in enumerate(data[0])]
                    df = pl.DataFrame(data[1:], schema=columns, strict=False)
                else:
                    return

            self.total_rows = len(df)

            # Convert to list of dictionaries
            self.sample_data = df.to_dicts()

            # Analyze each column (same as CSV)
            for col_name in df.columns:
                analysis = DataFieldAnalysis(col_name)
                column = df[col_name]

                analysis.total_count = len(df)
                analysis.null_count = column.null_count()
                analysis.sample_values = column.drop_nulls().head(10).to_list()
                analysis.inferred_type = _infer_polars_type(column)
                analysis.suggested_datatype = _suggest_xsd_datatype_polars(column)
                non_null = column.drop_nulls()
                analysis.is_unique = non_null.n_unique() == len(non_null)
                analysis.is_identifier = _is_likely_identifier_polars(col_name, column)
                analysis.pattern = _detect_pattern_polars(column)

                self.field_analyses[col_name] = analysis

        except Exception as e:
            raise ValueError(f"Failed to analyze Excel file: {e}")

    def _analyze_json(self) -> None:
        """Analyze JSON data with nested structure support."""
        try:
            with open(self.file_path, 'r', encoding='utf-8') as f:
                if self.file_extension == '.jsonl':
                    # JSON Lines format
                    data = []
                    for line in f:
                        if line.strip():
                            data.append(json.loads(line))
                else:
                    # Standard JSON
                    content = json.load(f)
                    # Handle different JSON structures
                    if isinstance(content, list):
                        data = content
                    elif isinstance(content, dict):
                        # Look for array field or treat as single record
                        array_fields = [k for k, v in content.items() if isinstance(v, list)]
                        if array_fields:
                            data = content[array_fields[0]]  # Use first array field
                        else:
                            data = [content]  # Single record
                    else:
                        raise ValueError("JSON must contain object or array")

            self.total_rows = len(data)
            self.sample_data = data[:100]

            # Expand arrays and flatten nested structure
            expanded_data = []
            for record in self.sample_data:
                expanded_records = self._expand_arrays(record)
                for expanded_record in expanded_records:
                    flattened = self._flatten_json(expanded_record)
                    expanded_data.append(flattened)

            # Create DataFrame from expanded and flattened data
            if expanded_data:
                df = pl.DataFrame(expanded_data)

                for col_name in df.columns:
                    analysis = DataFieldAnalysis(col_name, path=col_name)
                    column = df[col_name]

                    analysis.total_count = len(df)
                    analysis.null_count = column.null_count()
                    analysis.sample_values = column.drop_nulls().head(10).to_list()
                    analysis.inferred_type = _infer_polars_type(column)
                    analysis.suggested_datatype = _suggest_xsd_datatype_polars(column)
                    analysis.is_unique = column.n_unique() == len(column.drop_nulls())
                    analysis.is_identifier = _is_likely_identifier_polars(col_name, column)
                    analysis.pattern = _detect_pattern_polars(column)
                    analysis.is_nested = '.' in col_name  # Nested if contains dot notation

                    self.field_analyses[col_name] = analysis

            # Store nested structure info
            if self.sample_data:
                self.nested_structure = self._analyze_json_structure(self.sample_data[0])

        except Exception as e:
            raise ValueError(f"Failed to analyze JSON file: {e}")

    def _analyze_xml(self) -> None:
        """Analyze XML data with nested structure support."""
        try:
            tree = ET.parse(self.file_path)
            root = tree.getroot()

            # Find repeating elements (likely records)
            record_elements = self._find_xml_records(root)

            self.total_rows = len(record_elements)

            # Convert XML records to dictionaries
            xml_data = []
            for elem in record_elements[:100]:  # Sample first 100
                record_dict = self._xml_to_dict(elem)
                xml_data.append(record_dict)

            self.sample_data = xml_data

            # Flatten and analyze like JSON
            flattened_data = []
            for record in xml_data:
                flattened = self._flatten_json(record, separator='.')
                flattened_data.append(flattened)

            if flattened_data:
                df = pl.DataFrame(flattened_data)

                for col_name in df.columns:
                    analysis = DataFieldAnalysis(col_name, path=col_name)
                    column = df[col_name]

                    analysis.total_count = len(df)
                    analysis.null_count = column.null_count()
                    analysis.sample_values = column.drop_nulls().head(10).to_list()
                    analysis.inferred_type = _infer_polars_type(column)
                    analysis.suggested_datatype = _suggest_xsd_datatype_polars(column)
                    analysis.is_unique = column.n_unique() == len(column.drop_nulls())
                    analysis.is_identifier = _is_likely_identifier_polars(col_name, column)
                    analysis.pattern = _detect_pattern_polars(column)
                    analysis.is_nested = '.' in col_name

                    self.field_analyses[col_name] = analysis

            # Store XML structure info
            if record_elements:
                self.nested_structure = self._analyze_xml_structure(record_elements[0])

        except Exception as e:
            raise ValueError(f"Failed to analyze XML file: {e}")

    def _expand_arrays(self, obj: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Expand arrays in JSON objects to create separate records for each array item.

        This handles cases like a student with multiple courses by creating
        separate records for each course while preserving the parent data.

        Args:
            obj: JSON object that may contain arrays

        Returns:
            List of expanded objects, one for each combination of array items
        """
        # Find all array fields in the object
        array_fields = []

        def find_arrays(data, path=""):
            if isinstance(data, dict):
                for key, value in data.items():
                    current_path = f"{path}.{key}" if path else key
                    if isinstance(value, list) and value and isinstance(value[0], dict):
                        # Found an array of objects
                        array_fields.append((current_path, value))
                    elif isinstance(value, dict):
                        find_arrays(value, current_path)

        find_arrays(obj)

        if not array_fields:
            # No arrays found, return original object
            return [obj]

        # For now, handle the first array field found
        # More complex expansion (multiple arrays) can be added later
        array_path, array_items = array_fields[0]

        expanded_records = []
        for item in array_items:
            # Create a new record for each array item
            expanded_record = self._deep_copy_object(obj)

            # Replace the array with the single item
            self._set_nested_value(expanded_record, array_path, item)

            expanded_records.append(expanded_record)

        return expanded_records

    def _deep_copy_object(self, obj: Any) -> Any:
        """Create a deep copy of an object."""
        if isinstance(obj, dict):
            return {key: self._deep_copy_object(value) for key, value in obj.items()}
        elif isinstance(obj, list):
            return [self._deep_copy_object(item) for item in obj]
        else:
            return obj

    def _set_nested_value(self, obj: Dict[str, Any], path: str, value: Any) -> None:
        """Set a value in a nested object using dot notation path."""
        parts = path.split('.')
        current = obj

        # Navigate to the parent of the target
        for part in parts[:-1]:
            if part not in current:
                current[part] = {}
            current = current[part]

        # Set the final value
        current[parts[-1]] = value

    def _flatten_json(self, obj: Any, parent_key: str = '', separator: str = '.') -> Dict[str, Any]:
        """Flatten nested JSON object."""
        items = []

        if isinstance(obj, dict):
            for key, value in obj.items():
                new_key = f"{parent_key}{separator}{key}" if parent_key else key

                if isinstance(value, dict):
                    items.extend(self._flatten_json(value, new_key, separator).items())
                elif isinstance(value, list) and value and isinstance(value[0], dict):
                    # Handle array of objects - flatten first object
                    items.extend(self._flatten_json(value[0], new_key, separator).items())
                else:
                    items.append((new_key, value))
        else:
            items.append((parent_key, obj))

        return dict(items)

    def _xml_to_dict(self, element: ET.Element) -> Dict[str, Any]:
        """Convert XML element to dictionary."""
        result = {}

        # Add attributes
        if element.attrib:
            for key, value in element.attrib.items():
                result[f"@{key}"] = value

        # Add text content
        if element.text and element.text.strip():
            if len(element) == 0:  # Leaf element
                result = element.text.strip()
            else:
                result['#text'] = element.text.strip()

        # Add child elements
        for child in element:
            child_data = self._xml_to_dict(child)

            if child.tag in result:
                # Multiple elements with same tag - convert to array
                if not isinstance(result[child.tag], list):
                    result[child.tag] = [result[child.tag]]
                result[child.tag].append(child_data)
            else:
                result[child.tag] = child_data

        return result if isinstance(result, dict) and result else element.text or ""

    def _find_xml_records(self, root: ET.Element) -> List[ET.Element]:
        """Find repeating XML elements that represent records."""
        # Find elements with multiple children of the same tag
        tag_counts = {}

        def count_tags(elem):
            for child in elem:
                tag_counts[child.tag] = tag_counts.get(child.tag, 0) + 1
                count_tags(child)

        count_tags(root)

        # Find the most common tag (likely records)
        if tag_counts:
            most_common_tag = max(tag_counts.items(), key=lambda x: x[1])[0]
            return root.findall(f".//{most_common_tag}")

        # Fallback: return direct children
        return list(root)

    def _analyze_json_structure(self, sample_obj: Dict[str, Any]) -> Dict[str, Any]:
        """Analyze JSON structure for nested relationships."""
        structure = {}

        def analyze_obj(obj, path=""):
            if isinstance(obj, dict):
                for key, value in obj.items():
                    current_path = f"{path}.{key}" if path else key

                    if isinstance(value, dict):
                        structure[current_path] = {"type": "object", "properties": list(value.keys())}
                        analyze_obj(value, current_path)
                    elif isinstance(value, list) and value:
                        if isinstance(value[0], dict):
                            structure[current_path] = {"type": "array", "item_properties": list(value[0].keys())}
                            analyze_obj(value[0], current_path)
                        else:
                            structure[current_path] = {"type": "array", "item_type": type(value[0]).__name__}
                    else:
                        structure[current_path] = {"type": type(value).__name__}

        analyze_obj(sample_obj)
        return structure

    def _analyze_xml_structure(self, sample_elem: ET.Element) -> Dict[str, Any]:
        """Analyze XML structure for nested relationships."""
        structure = {}

        def analyze_elem(elem, path=""):
            current_path = f"{path}.{elem.tag}" if path else elem.tag

            if len(elem) > 0:
                # Has children
                structure[current_path] = {
                    "type": "element",
                    "children": [child.tag for child in elem],
                    "attributes": list(elem.attrib.keys())
                }

                for child in elem:
                    analyze_elem(child, current_path)
            else:
                # Leaf element
                structure[current_path] = {
                    "type": "leaf",
                    "attributes": list(elem.attrib.keys())
                }

        analyze_elem(sample_elem)
        return structure

    # Public interface methods
    def get_column_names(self) -> List[str]:
        """Get list of all field/column names."""
        return list(self.field_analyses.keys())

    def get_analysis(self, field_name: str) -> DataFieldAnalysis:
        """Get analysis for a specific field."""
        return self.field_analyses.get(field_name)

    def suggest_iri_template_columns(self) -> List[str]:
        """Suggest columns that could be used for IRI generation."""
        candidates = []

        for name, analysis in self.field_analyses.items():
            if analysis.is_identifier or analysis.is_unique:
                candidates.append(name)

        # Sort by preference (explicit IDs first, then unique fields)
        def sort_key(name):
            analysis = self.field_analyses[name]
            score = 0
            if analysis.is_identifier:
                score += 10
            if analysis.is_unique:
                score += 5
            if 'id' in name.lower():
                score += 3
            return score

        candidates.sort(key=sort_key, reverse=True)
        return candidates[:1]  # Return only the best candidate for IRI template

    def get_nested_fields(self) -> Dict[str, List[str]]:
        """Get nested field relationships for JSON/XML data."""
        nested = {}

        for name, analysis in self.field_analyses.items():
            if analysis.is_nested and '.' in name:
                parent = name.split('.')[0]
                if parent not in nested:
                    nested[parent] = []
                nested[parent].append(name)

        return nested

    def get_structure_info(self) -> Dict[str, Any]:
        """Get information about data structure (for nested formats)."""
        return {
            'format': self.data_format,
            'total_rows': self.total_rows,
            'field_count': len(self.field_analyses),
            'nested_structure': self.nested_structure,
            'has_nested_data': any(analysis.is_nested for analysis in self.field_analyses.values())
        }


# Maintain backward compatibility
class SpreadsheetAnalyzer(DataSourceAnalyzer):
    """Backward compatibility alias for DataSourceAnalyzer."""
    pass


# Legacy ColumnAnalysis alias
class ColumnAnalysis(DataFieldAnalysis):
    """Backward compatibility alias for DataFieldAnalysis."""
    pass
