"""Multi-sheet analysis for Excel workbooks.

Detects and analyzes relationships between sheets in Excel workbooks,
enabling automatic generation of cross-sheet mappings and joins.
"""

from typing import Dict, List, Optional, Any
from pathlib import Path
import polars as pl
from dataclasses import dataclass, field


@dataclass
class SheetInfo:
    """Information about a single sheet."""
    name: str
    row_count: int
    column_names: List[str]
    identifier_columns: List[str] = field(default_factory=list)  # Columns that look like IDs
    foreign_key_candidates: Dict[str, str] = field(default_factory=dict)  # col -> referenced_sheet
    sample_data: Optional[pl.DataFrame] = None


@dataclass
class SheetRelationship:
    """Detected relationship between two sheets."""
    source_sheet: str
    target_sheet: str
    source_column: str
    target_column: str
    relationship_type: str  # "one-to-many", "many-to-one", "many-to-many"
    confidence: float  # 0.0 to 1.0
    cardinality_analysis: Optional[Dict[str, Any]] = None


class MultiSheetAnalyzer:
    """Analyzer for Excel workbooks with multiple sheets."""

    def __init__(self, file_path: str):
        """Initialize the multi-sheet analyzer.

        Args:
            file_path: Path to Excel workbook
        """
        self.file_path = Path(file_path)
        self.sheets: Dict[str, SheetInfo] = {}
        self.relationships: List[SheetRelationship] = []

        if not self.file_path.suffix.lower() in ['.xlsx', '.xls']:
            raise ValueError(f"Multi-sheet analysis only supports Excel files, got: {self.file_path.suffix}")

        self._load_sheets()

    def _load_sheets(self):
        """Load all sheets from the workbook."""
        try:
            from openpyxl import load_workbook

            # Load workbook
            wb = load_workbook(self.file_path, read_only=True, data_only=True)

            # Process each sheet
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                # Extract data
                data = []
                for row in ws.iter_rows(values_only=True):
                    data.append(list(row))

                if not data or len(data) < 2:
                    continue  # Skip empty sheets

                # First row is header
                columns = [str(col) if col is not None else f"Column_{i}"
                          for i, col in enumerate(data[0])]

                # Create DataFrame from remaining rows
                rows = data[1:]
                if rows:
                    try:
                        # Filter out rows that are all None
                        valid_rows = [row for row in rows if any(cell is not None for cell in row)]

                        if valid_rows:
                            df = pl.DataFrame(valid_rows, schema=columns, orient="row")
                            sheet_info = self._analyze_sheet(sheet_name, df, len(self.sheets))
                            self.sheets[sheet_name] = sheet_info
                    except Exception as e:
                        # If DataFrame creation fails, skip this sheet
                        print(f"Warning: Could not analyze sheet '{sheet_name}': {e}")
                        continue

            wb.close()

        except ImportError:
            raise ValueError("openpyxl is required for multi-sheet analysis. Install with: pip install openpyxl")
        except Exception as e:
            raise ValueError(f"Error loading workbook {self.file_path}: {e}")

        if not self.sheets:
            raise ValueError(f"No sheets found in workbook: {self.file_path}")

    def _analyze_sheet(self, sheet_name: str, df: pl.DataFrame, sheet_id: int) -> SheetInfo:
        """Analyze a single sheet.

        Args:
            sheet_name: Name of the sheet
            df: DataFrame containing sheet data
            sheet_id: Sheet index

        Returns:
            SheetInfo with analysis results
        """
        # Identify potential ID columns
        identifier_cols = []
        for col in df.columns:
            col_lower = col.lower()
            # Check if column name suggests it's an ID
            if any(pattern in col_lower for pattern in ['id', '_id', 'key', 'code', 'number']):
                # Check if values are unique or mostly unique
                unique_ratio = df[col].n_unique() / len(df) if len(df) > 0 else 0
                if unique_ratio > 0.9:  # 90%+ unique
                    identifier_cols.append(col)

        # Identify potential foreign keys
        fk_candidates = {}
        for col in df.columns:
            col_lower = col.lower()
            # Look for patterns like "CustomerID", "OrderID", etc.
            if col_lower.endswith('id') and col not in identifier_cols:
                # Extract the referenced entity name
                # E.g., "CustomerID" -> "Customer", "customer_id" -> "customer"
                if col.endswith('ID'):
                    # CamelCase: CustomerID -> Customer
                    entity_name = col[:-2]
                elif col_lower.endswith('_id'):
                    # snake_case: customer_id -> customer
                    entity_name = col[:-3]
                else:
                    # lowercase: customerid -> customer
                    entity_name = col[:-2]
                fk_candidates[col] = entity_name

        return SheetInfo(
            name=sheet_name,
            row_count=len(df),
            column_names=df.columns,
            identifier_columns=identifier_cols,
            foreign_key_candidates=fk_candidates,
            sample_data=df.head(100) if len(df) > 0 else None
        )

    def detect_relationships(self) -> List[SheetRelationship]:
        """Detect relationships between sheets.

        Returns:
            List of detected relationships
        """
        self.relationships = []

        # For each sheet, look for foreign key relationships
        for source_name, source_sheet in self.sheets.items():
            for fk_col, referenced_entity in source_sheet.foreign_key_candidates.items():
                # Try to find a matching sheet
                target_sheet = self._find_matching_sheet(referenced_entity)

                if target_sheet:
                    # Found a potential relationship
                    target_info = self.sheets[target_sheet]

                    # Find the primary key column in target sheet
                    pk_col = self._find_primary_key(target_info, referenced_entity)

                    if pk_col:
                        # Analyze the relationship
                        relationship = self._analyze_relationship(
                            source_name, target_sheet,
                            fk_col, pk_col,
                            source_sheet, target_info
                        )

                        if relationship:
                            self.relationships.append(relationship)

        return self.relationships

    def _find_matching_sheet(self, entity_name: str) -> Optional[str]:
        """Find a sheet that matches an entity name.

        Args:
            entity_name: Name of the entity (e.g., "Customer", "Product")

        Returns:
            Matching sheet name or None
        """
        entity_lower = entity_name.lower()

        # Try exact match first
        for sheet_name in self.sheets.keys():
            if sheet_name.lower() == entity_lower:
                return sheet_name

        # Try plural forms
        plural_forms = [
            entity_lower + 's',      # customer -> customers
            entity_lower + 'es',     # address -> addresses
            entity_lower + 'ies',    # category -> categories (handled separately)
        ]

        # Handle special pluralization (e.g., "y" -> "ies")
        if entity_lower.endswith('y') and len(entity_lower) > 1:
            plural_forms.append(entity_lower[:-1] + 'ies')

        for sheet_name in self.sheets.keys():
            sheet_lower = sheet_name.lower()
            # Check if sheet name matches any plural form
            if sheet_lower in plural_forms:
                return sheet_name
            # Check if any plural form is contained in sheet name
            if any(form in sheet_lower for form in plural_forms):
                return sheet_name

        # Try if entity name is contained in sheet name
        for sheet_name in self.sheets.keys():
            if entity_lower in sheet_name.lower():
                return sheet_name

        return None

    def _find_primary_key(self, sheet_info: SheetInfo, entity_name: str) -> Optional[str]:
        """Find the primary key column in a sheet.

        Args:
            sheet_info: Sheet information
            entity_name: Entity name for hints

        Returns:
            Primary key column name or None
        """
        # First, check identifier columns
        if sheet_info.identifier_columns:
            # Prefer column that matches entity name + "ID"
            preferred_name = entity_name + "ID"
            for col in sheet_info.identifier_columns:
                if col == preferred_name or col.lower() == preferred_name.lower():
                    return col

            # Otherwise, return first identifier
            return sheet_info.identifier_columns[0]

        # Fallback: look for any column with "ID" in name
        for col in sheet_info.column_names:
            if 'id' in col.lower():
                return col

        return None

    def _analyze_relationship(
        self,
        source_sheet: str,
        target_sheet: str,
        source_col: str,
        target_col: str,
        source_info: SheetInfo,
        target_info: SheetInfo
    ) -> Optional[SheetRelationship]:
        """Analyze a potential relationship between sheets.

        Args:
            source_sheet: Source sheet name
            target_sheet: Target sheet name
            source_col: Foreign key column in source
            target_col: Primary key column in target
            source_info: Source sheet info
            target_info: Target sheet info

        Returns:
            SheetRelationship or None if relationship is not valid
        """
        if source_info.sample_data is None or target_info.sample_data is None:
            return None

        try:
            # Get sample values  (we know sample_data is not None from check above)
            source_df = source_info.sample_data
            target_df = target_info.sample_data

            source_values = source_df[source_col].drop_nulls()
            target_values = target_df[target_col].drop_nulls()

            if len(source_values) == 0 or len(target_values) == 0:
                return None

            # Check value overlap (referential integrity)
            source_set = set(source_values.to_list())
            target_set = set(target_values.to_list())

            overlap = len(source_set & target_set)
            overlap_ratio = overlap / len(source_set) if source_set else 0

            if overlap_ratio < 0.3:  # Less than 30% overlap, probably not a valid FK
                return None

            # Determine relationship type
            # Check cardinality
            source_unique_ratio = len(source_set) / len(source_values) if len(source_values) > 0 else 0
            target_unique_ratio = len(target_set) / len(target_values) if len(target_values) > 0 else 0

            if source_unique_ratio > 0.9 and target_unique_ratio > 0.9:
                rel_type = "one-to-one"
            elif source_unique_ratio < 0.5:
                rel_type = "many-to-one"  # Many rows in source reference same target
            else:
                rel_type = "one-to-many"  # Each source references different targets

            # Calculate confidence
            confidence = overlap_ratio * 0.7 + \
                        (1.0 if source_col.lower().endswith('id') else 0.5) * 0.3

            return SheetRelationship(
                source_sheet=source_sheet,
                target_sheet=target_sheet,
                source_column=source_col,
                target_column=target_col,
                relationship_type=rel_type,
                confidence=confidence,
                cardinality_analysis={
                    "overlap_ratio": overlap_ratio,
                    "overlap_count": overlap,
                    "source_unique_ratio": source_unique_ratio,
                    "target_unique_ratio": target_unique_ratio
                }
            )

        except Exception as e:
            # If analysis fails, return None
            # Log the error for debugging
            import warnings
            warnings.warn(f"Failed to analyze relationship {source_sheet}.{source_col} -> {target_sheet}.{target_col}: {e}")
            return None

    def get_sheet_names(self) -> List[str]:
        """Get all sheet names in the workbook."""
        return list(self.sheets.keys())

    def get_sheet_info(self, sheet_name: str) -> Optional[SheetInfo]:
        """Get information about a specific sheet."""
        return self.sheets.get(sheet_name)

    def get_primary_sheet(self) -> str:
        """Determine the primary (main) sheet in the workbook.

        Returns:
            Name of the sheet that appears to be the primary one
        """
        # Heuristics:
        # 1. Largest sheet (most rows)
        # 2. Sheet with most FK references from other sheets
        # 3. First sheet if no clear winner

        if not self.sheets:
            raise ValueError("No sheets available")

        # Count how many times each sheet is referenced
        reference_counts = {name: 0 for name in self.sheets.keys()}
        for rel in self.relationships:
            reference_counts[rel.target_sheet] += 1

        # Score each sheet
        scores = {}
        max_rows = max(sheet.row_count for sheet in self.sheets.values())

        for name, sheet in self.sheets.items():
            score = 0

            # Size score (normalized)
            score += (sheet.row_count / max_rows) * 50

            # Reference score
            score += reference_counts[name] * 30

            # First sheet bonus
            if list(self.sheets.keys())[0] == name:
                score += 20

            scores[name] = score

        # Return sheet with highest score
        return max(scores.items(), key=lambda x: x[1])[0]

    def generate_relationship_summary(self) -> str:
        """Generate a human-readable summary of detected relationships."""
        if not self.relationships:
            return "No relationships detected between sheets."

        lines = ["Detected Sheet Relationships:", ""]

        for i, rel in enumerate(self.relationships, 1):
            lines.append(f"{i}. {rel.source_sheet}.{rel.source_column} → "
                        f"{rel.target_sheet}.{rel.target_column}")
            lines.append(f"   Type: {rel.relationship_type}")
            lines.append(f"   Confidence: {rel.confidence:.0%}")

            if rel.cardinality_analysis:
                overlap = rel.cardinality_analysis['overlap_ratio']
                lines.append(f"   Referential integrity: {overlap:.0%}")

            lines.append("")

        return "\n".join(lines)

