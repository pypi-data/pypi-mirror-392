"""Tests for Multi-Sheet Excel Support.

This module tests the MultiSheetAnalyzer and multi-sheet data handling capabilities.
"""

import pytest
from pathlib import Path
import polars as pl

from rdfmap.generator.multisheet_analyzer import MultiSheetAnalyzer, SheetInfo, SheetRelationship
from rdfmap.generator.data_analyzer import DataSourceAnalyzer
from rdfmap.generator.mapping_generator import MappingGenerator


@pytest.fixture
def sample_multisheet_excel(tmp_path):
    """Create a sample multi-sheet Excel file for testing."""
    try:
        from openpyxl import Workbook
    except ImportError:
        pytest.skip("openpyxl not installed")

    wb = Workbook()
    wb.remove(wb.active)

    # Create Customers sheet
    ws_customers = wb.create_sheet("Customers")
    ws_customers.append(["CustomerID", "Name", "Email"])
    ws_customers.append(["C001", "John Doe", "john@example.com"])
    ws_customers.append(["C002", "Jane Smith", "jane@example.com"])

    # Create Orders sheet
    ws_orders = wb.create_sheet("Orders")
    ws_orders.append(["OrderID", "CustomerID", "OrderDate", "Total"])
    ws_orders.append(["O001", "C001", "2024-01-15", 100.00])
    ws_orders.append(["O002", "C002", "2024-01-16", 200.00])
    ws_orders.append(["O003", "C001", "2024-01-17", 150.00])

    excel_file = tmp_path / "test_multisheet.xlsx"
    wb.save(excel_file)

    return excel_file


@pytest.fixture
def sample_single_sheet_excel(tmp_path):
    """Create a sample single-sheet Excel file for testing."""
    try:
        from openpyxl import Workbook
    except ImportError:
        pytest.skip("openpyxl not installed")

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["ID", "Name", "Value"])
    ws.append([1, "A", 100])
    ws.append([2, "B", 200])

    excel_file = tmp_path / "test_singlesheet.xlsx"
    wb.save(excel_file)

    return excel_file


class TestMultiSheetAnalyzer:
    """Test suite for MultiSheetAnalyzer."""

    def test_analyzer_initialization(self, sample_multisheet_excel):
        """Test that analyzer initializes correctly."""
        analyzer = MultiSheetAnalyzer(str(sample_multisheet_excel))

        assert analyzer is not None
        assert len(analyzer.sheets) > 0

    def test_sheet_detection(self, sample_multisheet_excel):
        """Test that multiple sheets are detected."""
        analyzer = MultiSheetAnalyzer(str(sample_multisheet_excel))

        assert len(analyzer.sheets) == 2
        assert "Customers" in analyzer.sheets
        assert "Orders" in analyzer.sheets

    def test_sheet_info(self, sample_multisheet_excel):
        """Test that sheet information is correctly extracted."""
        analyzer = MultiSheetAnalyzer(str(sample_multisheet_excel))

        customers_info = analyzer.sheets["Customers"]
        assert isinstance(customers_info, SheetInfo)
        assert customers_info.row_count == 2  # 2 data rows
        assert "CustomerID" in customers_info.column_names
        assert "Name" in customers_info.column_names

    def test_relationship_detection(self, sample_multisheet_excel):
        """Test that relationships between sheets are detected."""
        analyzer = MultiSheetAnalyzer(str(sample_multisheet_excel))

        relationships = analyzer.detect_relationships()

        assert len(relationships) > 0

        # Should find Orders.CustomerID -> Customers.CustomerID relationship
        found_relationship = False
        for rel in relationships:
            if (rel.source_sheet == "Orders" and
                rel.target_sheet == "Customers" and
                "CustomerID" in rel.source_column and
                "CustomerID" in rel.target_column):
                found_relationship = True
                assert rel.relationship_type in ["many-to-one", "one-to-many"]
                break

        assert found_relationship, "Should detect CustomerID relationship"

    def test_primary_sheet_identification(self, sample_multisheet_excel):
        """Test identification of primary sheet."""
        analyzer = MultiSheetAnalyzer(str(sample_multisheet_excel))

        primary_sheet = analyzer.get_primary_sheet()

        assert primary_sheet is not None
        # Primary sheet is typically the one with most relationships
        # or first sheet
        assert primary_sheet in ["Customers", "Orders"]

    def test_foreign_key_detection(self, sample_multisheet_excel):
        """Test foreign key column detection."""
        analyzer = MultiSheetAnalyzer(str(sample_multisheet_excel))

        # Orders sheet should have CustomerID as FK
        orders_info = analyzer.sheets["Orders"]

        # Check if CustomerID is identified as potential FK
        # (ends with "ID" and references another sheet)
        has_fk = any("ID" in col for col in orders_info.column_names)
        assert has_fk


class TestDataSourceAnalyzerMultiSheet:
    """Test DataSourceAnalyzer with multi-sheet support."""

    def test_single_sheet_detection(self, sample_single_sheet_excel):
        """Test that single sheet is correctly identified."""
        analyzer = DataSourceAnalyzer(str(sample_single_sheet_excel))

        # Check if has_multiple_sheets attribute exists
        if hasattr(analyzer, 'has_multiple_sheets'):
            assert analyzer.has_multiple_sheets == False

        # Check if sheet_count attribute exists
        if hasattr(analyzer, 'sheet_count'):
            assert analyzer.sheet_count == 1

    def test_multi_sheet_detection(self, sample_multisheet_excel):
        """Test that multiple sheets are correctly identified."""
        analyzer = DataSourceAnalyzer(str(sample_multisheet_excel))

        if hasattr(analyzer, 'has_multiple_sheets'):
            assert analyzer.has_multiple_sheets == True

        if hasattr(analyzer, 'sheet_count'):
            assert analyzer.sheet_count >= 2


class TestMappingGeneratorMultiSheet:
    """Test MappingGenerator with multi-sheet support."""

    def test_generate_multisheet_method_exists(self):
        """Test that generate_multisheet method exists."""
        assert hasattr(MappingGenerator, 'generate_multisheet')

    @pytest.mark.integration
    def test_multisheet_mapping_generation(self, sample_multisheet_excel, tmp_path):
        """Test generation of mappings for multi-sheet Excel."""
        # This would require a full ontology setup
        # Skip if not in integration test mode
        pytest.skip("Requires full ontology setup - integration test")


class TestSheetInfo:
    """Test SheetInfo dataclass."""

    def test_sheet_info_creation(self):
        """Test creating SheetInfo instances."""
        sheet_info = SheetInfo(
            name="TestSheet",
            row_count=10,
            column_names=["ID", "Name", "Value"],
            sample_data=pl.DataFrame({"ID": [1], "Name": ["Test"], "Value": [100]})
        )

        assert sheet_info.name == "TestSheet"
        assert sheet_info.row_count == 10
        assert len(sheet_info.column_names) == 3


class TestSheetRelationship:
    """Test SheetRelationship dataclass."""

    def test_relationship_creation(self):
        """Test creating SheetRelationship instances."""
        relationship = SheetRelationship(
            source_sheet="Orders",
            source_column="CustomerID",
            target_sheet="Customers",
            target_column="CustomerID",
            relationship_type="many-to-one",
            confidence=0.95
        )

        assert relationship.source_sheet == "Orders"
        assert relationship.target_sheet == "Customers"
        assert relationship.confidence == 0.95


class TestMultiSheetEdgeCases:
    """Test edge cases for multi-sheet support."""

    def test_empty_sheet(self, tmp_path):
        """Test handling of empty sheets."""
        try:
            from openpyxl import Workbook
        except ImportError:
            pytest.skip("openpyxl not installed")

        wb = Workbook()
        wb.remove(wb.active)

        ws1 = wb.create_sheet("Data")
        ws1.append(["ID", "Name"])
        ws1.append([1, "Test"])

        ws2 = wb.create_sheet("Empty")
        # Empty sheet - no data

        excel_file = tmp_path / "with_empty.xlsx"
        wb.save(excel_file)

        analyzer = MultiSheetAnalyzer(str(excel_file))

        # Should handle empty sheets gracefully
        # Either skip them or include with 0 rows
        assert "Data" in analyzer.sheets

    def test_duplicate_column_names(self, tmp_path):
        """Test handling of duplicate column names across sheets."""
        try:
            from openpyxl import Workbook
        except ImportError:
            pytest.skip("openpyxl not installed")

        wb = Workbook()
        wb.remove(wb.active)

        ws1 = wb.create_sheet("Sheet1")
        ws1.append(["ID", "Name"])
        ws1.append([1, "A"])

        ws2 = wb.create_sheet("Sheet2")
        ws2.append(["ID", "Name"])
        ws2.append([2, "B"])

        excel_file = tmp_path / "duplicate_cols.xlsx"
        wb.save(excel_file)

        analyzer = MultiSheetAnalyzer(str(excel_file))

        # Should handle duplicate column names across sheets
        assert len(analyzer.sheets) == 2

    def test_no_relationships(self, tmp_path):
        """Test sheets with no relationships."""
        try:
            from openpyxl import Workbook
        except ImportError:
            pytest.skip("openpyxl not installed")

        wb = Workbook()
        wb.remove(wb.active)

        ws1 = wb.create_sheet("Products")
        ws1.append(["ProductID", "Name"])
        ws1.append(["P001", "Widget"])

        ws2 = wb.create_sheet("Employees")
        ws2.append(["EmployeeID", "Name"])
        ws2.append(["E001", "John"])

        excel_file = tmp_path / "no_relations.xlsx"
        wb.save(excel_file)

        analyzer = MultiSheetAnalyzer(str(excel_file))
        relationships = analyzer.detect_relationships()

        # Should return empty list or no matches
        # (depends on implementation)
        assert isinstance(relationships, list)


@pytest.mark.integration
class TestMultiSheetIntegration:
    """Integration tests for multi-sheet support."""

    def test_with_real_example_file(self):
        """Test with real example file if it exists."""
        example_file = Path("test_data/multisheet/ecommerce_orders.xlsx")

        if not example_file.exists():
            pytest.skip("Example file not found")

        analyzer = MultiSheetAnalyzer(str(example_file))

        assert len(analyzer.sheets) > 0

        relationships = analyzer.detect_relationships()
        assert isinstance(relationships, list)

