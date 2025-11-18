"""
Excel Report Generator for HVT6

This module generates professional Excel (.xlsx) reports with multiple sheets,
conditional formatting, and pivot-ready data structures for comprehensive
security audit analysis.
"""

from pathlib import Path
from typing import List, Dict, Optional
from datetime import datetime
from collections import defaultdict

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from loguru import logger

from hvt6.core.models import DeviceReport, CheckResult, CategoryScore
from hvt6.core.enums import Category, CheckStatus


class ExcelReportGenerator:
    """
    Generate multi-sheet Excel reports from device audit results.

    Features:
    - Summary sheet with executive overview and category aggregates
    - Devices sheet with detailed device information and scores
    - Check Results sheet with denormalized data (pivot-ready)
    - Conditional formatting (green/yellow/red scoring)
    - Auto-adjusted column widths
    - Professional formatting with headers and borders
    """

    def __init__(self, output_dir: Path):
        """
        Initialize Excel generator with output directory.

        Args:
            output_dir: Path to output directory for Excel files
        """
        self.output_dir = output_dir
        logger.debug(f"ExcelReportGenerator initialized with output: {output_dir}")

    def generate_excel(
        self,
        device_reports: List[DeviceReport],
        output_path: Path,
        customer: str = "Customer"
    ) -> None:
        """
        Generate comprehensive Excel report with multiple sheets.

        Args:
            device_reports: List of DeviceReport objects to include
            output_path: Path where Excel file will be saved
            customer: Customer name for report header
        """
        if not device_reports:
            logger.warning("No device reports provided for Excel generation")
            return

        logger.info(f"Generating Excel report for {len(device_reports)} devices")

        # Create DataFrames for each sheet
        summary_df = self._create_summary_dataframe(device_reports, customer)
        devices_df = self._create_devices_dataframe(device_reports)
        checks_df = self._create_checks_dataframe(device_reports)

        # Write to Excel with pandas
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Write sheets
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            devices_df.to_excel(writer, sheet_name='Devices', index=False)
            checks_df.to_excel(writer, sheet_name='Check Results', index=False)

            # Get workbook for formatting
            workbook = writer.book

            # Apply formatting to each sheet
            self._format_summary_sheet(workbook['Summary'], len(device_reports), customer)
            self._format_devices_sheet(workbook['Devices'], len(devices_df))
            self._format_checks_sheet(workbook['Check Results'], len(checks_df))

        logger.info(f"Excel report generated successfully: {output_path}")

    def _create_summary_dataframe(
        self,
        device_reports: List[DeviceReport],
        customer: str
    ) -> pd.DataFrame:
        """
        Create Summary sheet DataFrame with executive overview.

        Includes:
        - Report metadata (customer, date, device count)
        - Overall statistics (avg score, grade)
        - Device summary table
        - Category aggregates

        Args:
            device_reports: List of device reports
            customer: Customer name

        Returns:
            DataFrame for Summary sheet
        """
        # Calculate overall statistics
        total_achieved = sum(r.total_achieved for r in device_reports)
        total_max = sum(r.total_max_score for r in device_reports)
        avg_percentage = (total_achieved / total_max * 100) if total_max > 0 else 0
        overall_grade = self._calculate_grade(avg_percentage)

        # Build summary data rows
        summary_rows = []

        # Header section
        summary_rows.append(['SECURITY AUDIT EXECUTIVE SUMMARY', '', '', '', '', '', ''])
        summary_rows.append(['', '', '', '', '', '', ''])
        summary_rows.append(['Customer:', customer, '', '', '', '', ''])
        summary_rows.append(['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S'), '', '', '', '', ''])
        summary_rows.append(['Total Devices:', len(device_reports), '', '', '', '', ''])
        summary_rows.append(['Average Score:', f'{avg_percentage:.1f}%', '', '', '', '', ''])
        summary_rows.append(['Overall Grade:', overall_grade, '', '', '', '', ''])
        summary_rows.append(['', '', '', '', '', '', ''])

        # Device summary table
        summary_rows.append(['DEVICE SUMMARY', '', '', '', '', '', ''])
        summary_rows.append(['Hostname', 'Type', 'Model', 'OS', 'Version', 'Score %', 'Grade'])

        for report in device_reports:
            summary_rows.append([
                report.device_info.hostname,
                report.device_info.device_type.value.capitalize(),
                report.device_info.model,
                report.device_info.os,
                report.device_info.version,
                f'{report.total_percentage:.1f}%',
                self._calculate_grade(report.total_percentage)
            ])

        summary_rows.append(['', '', '', '', '', '', ''])

        # Category aggregates
        summary_rows.append(['CATEGORY PERFORMANCE', '', '', '', '', '', ''])
        summary_rows.append(['Category', 'Total Achieved', 'Total Max', 'Percentage', '', '', ''])

        # Aggregate scores by category
        category_totals = defaultdict(lambda: {'achieved': 0, 'max': 0})
        for report in device_reports:
            for category, score in report.category_scores.items():
                category_totals[category]['achieved'] += score.achieved
                category_totals[category]['max'] += score.max_score

        for category in Category:
            if category_totals[category]['max'] > 0:  # Only show categories with checks
                achieved = category_totals[category]['achieved']
                max_score = category_totals[category]['max']
                percentage = (achieved / max_score * 100) if max_score > 0 else 0
                summary_rows.append([
                    category.value.upper(),
                    achieved,
                    max_score,
                    f'{percentage:.1f}%',
                    '', '', ''
                ])

        return pd.DataFrame(summary_rows)

    def _create_devices_dataframe(self, device_reports: List[DeviceReport]) -> pd.DataFrame:
        """
        Create Devices sheet DataFrame with detailed device information.

        Args:
            device_reports: List of device reports

        Returns:
            DataFrame for Devices sheet
        """
        rows = []

        for report in device_reports:
            rows.append({
                'Hostname': report.device_info.hostname,
                'Device Type': report.device_info.device_type.value.capitalize(),
                'Model': report.device_info.model,
                'OS': report.device_info.os,
                'Version': report.device_info.version,
                'Serial Number': report.device_info.serial_number,
                'Score Achieved': report.total_achieved,
                'Max Score': report.total_max_score,
                'Percentage': report.total_percentage,
                'Grade': self._calculate_grade(report.total_percentage),
                'Passed Checks': report.passed_checks,
                'Total Checks': len(report.results),
                'Version Warning': 'Yes' if report.device_info.version_warning else 'No'
            })

        return pd.DataFrame(rows)

    def _create_checks_dataframe(self, device_reports: List[DeviceReport]) -> pd.DataFrame:
        """
        Create Check Results sheet DataFrame with denormalized data for pivot tables.

        One row per device-check combination, enabling flexible pivot table creation.

        Args:
            device_reports: List of device reports

        Returns:
            DataFrame for Check Results sheet
        """
        rows = []

        for report in device_reports:
            for check in report.results:
                rows.append({
                    'Hostname': report.device_info.hostname,
                    'Device Type': report.device_info.device_type.value.capitalize(),
                    'Check ID': check.check_id,
                    'Check Name': check.check_name,
                    'Category': check.category.value.upper(),
                    'Security Plane': check.security_plane.value.upper() if check.security_plane else 'N/A',
                    'Status': check.status.value.upper(),
                    'Score Achieved': check.achieved,
                    'Max Score': check.max_score,
                    'Percentage': check.percentage,
                    'Description': check.description,
                    'Recommendation': check.recommendation or 'N/A'
                })

        return pd.DataFrame(rows)

    def _format_summary_sheet(
        self,
        worksheet,
        device_count: int,
        customer: str
    ) -> None:
        """
        Apply formatting to Summary sheet.

        Args:
            worksheet: openpyxl worksheet object
            device_count: Number of devices in report
            customer: Customer name
        """
        # Format header (row 1)
        worksheet['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        worksheet['A1'].fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
        worksheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        worksheet.merge_cells('A1:G1')

        # Format metadata section (rows 3-7)
        for row in range(3, 8):
            worksheet[f'A{row}'].font = Font(bold=True)

        # Format "DEVICE SUMMARY" header (row 9)
        worksheet['A9'].font = Font(size=14, bold=True, color='FFFFFF')
        worksheet['A9'].fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
        worksheet.merge_cells('A9:G9')

        # Format device table header (row 10)
        for col in range(1, 8):
            cell = worksheet.cell(row=10, column=col)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                bottom=Side(style='thin', color='000000')
            )

        # Format "CATEGORY PERFORMANCE" header
        cat_header_row = 11 + device_count + 1
        worksheet[f'A{cat_header_row}'].font = Font(size=14, bold=True, color='FFFFFF')
        worksheet[f'A{cat_header_row}'].fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
        worksheet.merge_cells(f'A{cat_header_row}:G{cat_header_row}')

        # Format category table header
        cat_table_header = cat_header_row + 1
        for col in range(1, 5):
            cell = worksheet.cell(row=cat_table_header, column=col)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Auto-adjust column widths
        self._auto_adjust_columns(worksheet)

    def _format_devices_sheet(self, worksheet, row_count: int) -> None:
        """
        Apply formatting to Devices sheet with conditional formatting.

        Args:
            worksheet: openpyxl worksheet object
            row_count: Number of data rows
        """
        # Format header row
        for col in range(1, 14):  # 13 columns
            cell = worksheet.cell(row=1, column=col)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                bottom=Side(style='thin', color='000000')
            )

        # Add conditional formatting for Percentage column (column I)
        if row_count > 0:
            # Green: >= 80%
            worksheet.conditional_formatting.add(
                f'I2:I{row_count + 1}',
                CellIsRule(
                    operator='greaterThanOrEqual',
                    formula=['80'],
                    fill=PatternFill(start_color='10B981', end_color='10B981', fill_type='solid'),
                    font=Font(color='FFFFFF', bold=True)
                )
            )

            # Yellow: 60-79%
            worksheet.conditional_formatting.add(
                f'I2:I{row_count + 1}',
                CellIsRule(
                    operator='between',
                    formula=['60', '79.99'],
                    fill=PatternFill(start_color='F59E0B', end_color='F59E0B', fill_type='solid'),
                    font=Font(color='FFFFFF', bold=True)
                )
            )

            # Red: < 60%
            worksheet.conditional_formatting.add(
                f'I2:I{row_count + 1}',
                CellIsRule(
                    operator='lessThan',
                    formula=['60'],
                    fill=PatternFill(start_color='EF4444', end_color='EF4444', fill_type='solid'),
                    font=Font(color='FFFFFF', bold=True)
                )
            )

        # Auto-adjust column widths
        self._auto_adjust_columns(worksheet)

    def _format_checks_sheet(self, worksheet, row_count: int) -> None:
        """
        Apply formatting to Check Results sheet with status-based coloring.

        Args:
            worksheet: openpyxl worksheet object
            row_count: Number of data rows
        """
        # Format header row
        for col in range(1, 13):  # 12 columns
            cell = worksheet.cell(row=1, column=col)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(
                bottom=Side(style='thin', color='000000')
            )

        # Add conditional formatting for Status column (column G)
        if row_count > 0:
            # Green: PASS
            worksheet.conditional_formatting.add(
                f'G2:G{row_count + 1}',
                CellIsRule(
                    operator='equal',
                    formula=['"PASS"'],
                    fill=PatternFill(start_color='10B981', end_color='10B981', fill_type='solid'),
                    font=Font(color='FFFFFF', bold=True)
                )
            )

            # Red: FAIL
            worksheet.conditional_formatting.add(
                f'G2:G{row_count + 1}',
                CellIsRule(
                    operator='equal',
                    formula=['"FAIL"'],
                    fill=PatternFill(start_color='EF4444', end_color='EF4444', fill_type='solid'),
                    font=Font(color='FFFFFF', bold=True)
                )
            )

            # Yellow: WARNING
            worksheet.conditional_formatting.add(
                f'G2:G{row_count + 1}',
                CellIsRule(
                    operator='equal',
                    formula=['"WARNING"'],
                    fill=PatternFill(start_color='F59E0B', end_color='F59E0B', fill_type='solid'),
                    font=Font(color='FFFFFF', bold=True)
                )
            )

            # Gray: NOT_APPLICABLE
            worksheet.conditional_formatting.add(
                f'G2:G{row_count + 1}',
                CellIsRule(
                    operator='equal',
                    formula=['"NOT_APPLICABLE"'],
                    fill=PatternFill(start_color='9CA3AF', end_color='9CA3AF', fill_type='solid'),
                    font=Font(color='FFFFFF')
                )
            )

            # Add percentage conditional formatting (column J)
            # Green: >= 80%
            worksheet.conditional_formatting.add(
                f'J2:J{row_count + 1}',
                CellIsRule(
                    operator='greaterThanOrEqual',
                    formula=['80'],
                    fill=PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
                )
            )

            # Red: < 60%
            worksheet.conditional_formatting.add(
                f'J2:J{row_count + 1}',
                CellIsRule(
                    operator='lessThan',
                    formula=['60'],
                    fill=PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
                )
            )

        # Auto-adjust column widths
        self._auto_adjust_columns(worksheet)

    def _auto_adjust_columns(self, worksheet) -> None:
        """
        Auto-adjust column widths based on content.

        Args:
            worksheet: openpyxl worksheet object
        """
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass

            # Set column width with padding and max limit
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def _calculate_grade(self, percentage: float) -> str:
        """
        Calculate letter grade from percentage score.

        Args:
            percentage: Score percentage (0-100)

        Returns:
            Letter grade (A, B, C, D, F)
        """
        if percentage >= 90:
            return 'A'
        elif percentage >= 80:
            return 'B'
        elif percentage >= 70:
            return 'C'
        elif percentage >= 60:
            return 'D'
        else:
            return 'F'
