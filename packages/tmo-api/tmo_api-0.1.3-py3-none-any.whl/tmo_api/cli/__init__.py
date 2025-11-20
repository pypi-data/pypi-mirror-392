"""Command-line interface tools for The Mortgage Office API."""

import argparse
import configparser
import csv
import json
import os
import re
import sys
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional

from ..client import TMOClient
from ..environments import Environment
from ..exceptions import ValidationError

# Configuration file location
TMORC_PATH = Path.home() / ".tmorc"

# Demo credentials (fallback if no config file)
DEMO_TOKEN = "TMO"
DEMO_DATABASE = "API Sandbox"
DEMO_ENVIRONMENT = Environment.US


def load_config() -> configparser.ConfigParser:
    """Load configuration from ~/.tmorc file.

    Returns:
        ConfigParser instance with loaded configuration
    """
    config = configparser.ConfigParser()
    if TMORC_PATH.exists():
        config.read(TMORC_PATH)
    return config


def get_config_profiles(config: configparser.ConfigParser) -> List[str]:
    """Get list of available configuration profiles.

    Args:
        config: ConfigParser instance

    Returns:
        List of profile names
    """
    return list(config.sections())


def add_common_arguments(parser: argparse.ArgumentParser) -> None:  # pragma: no cover
    """Add common CLI arguments to a parser.

    Args:
        parser: ArgumentParser instance to add arguments to
    """
    config = load_config()
    available_profiles = get_config_profiles(config)
    profile_help = (
        f"Configuration profile to use (default: demo). Available: {', '.join(available_profiles)}"
        if available_profiles
        else "Configuration profile to use (default: demo). Run 'tmoapi init' to create ~/.tmorc"
    )

    parser.add_argument(
        "-P",
        "--profile",
        type=str,
        default="demo",
        help=profile_help,
    )
    parser.add_argument("--token", type=str, help="API token (overrides profile)")
    parser.add_argument("--database", type=str, help="Database name (overrides profile)")
    parser.add_argument(
        "--environment",
        type=str,
        choices=["us", "usa", "can", "aus"],
        help="API environment (overrides profile)",
    )
    parser.add_argument("--debug", action="store_true", help="Enable debug output")
    parser.add_argument("--user-agent", type=str, help="Override the default User-Agent header")


def resolve_config_values(args: argparse.Namespace) -> dict[str, Any]:  # pragma: no cover
    """Resolve configuration values from profile, command line args, and environment vars.

    Args:
        args: Parsed command-line arguments

    Returns:
        Dictionary with resolved configuration values

    Raises:
        ValidationError: If profile not found or required values missing
    """
    # Load config file
    config = load_config()

    # Start with defaults
    values = {"token": None, "database": None, "environment": "us", "timeout": 30}

    # Get profile name (defaults to "demo")
    profile = getattr(args, "profile", "demo")

    # Load from profile if it exists in config
    if config.has_section(profile):
        profile_section = config[profile]
        values.update(
            {
                "token": profile_section.get("token"),
                "database": profile_section.get("database"),
                "environment": profile_section.get("environment", "us"),
                "timeout": profile_section.getint("timeout", 30),
            }
        )
    elif profile == "demo":
        # Use built-in demo credentials if demo profile not in config
        values.update({"token": DEMO_TOKEN, "database": DEMO_DATABASE, "environment": "us"})
    else:
        # Profile specified but not found
        available = get_config_profiles(config)
        raise ValidationError(
            f"Profile '{profile}' not found in {TMORC_PATH}. "
            f"Available profiles: {', '.join(available) if available else 'none'}. "
            f"Run 'tmoapi init' to create the config file."
        )

    # Override with command line arguments if provided
    if hasattr(args, "token") and args.token:
        values["token"] = args.token
    if hasattr(args, "database") and args.database:
        values["database"] = args.database
    if hasattr(args, "environment") and args.environment:
        values["environment"] = args.environment

    # Override with environment variables if not already set
    if not values["token"]:
        values["token"] = os.getenv("TMO_API_TOKEN")
    if not values["database"]:
        values["database"] = os.getenv("TMO_DATABASE")

    # Validate required values
    if not values["token"]:
        raise ValidationError(
            "Token is required. Provide via --profile, --token, TMO_API_TOKEN env var, or run 'tmoapi init'."
        )
    if not values["database"]:
        raise ValidationError(
            "Database is required. Provide via --profile, --database, TMO_DATABASE env var, or run 'tmoapi init'."
        )

    return values


def create_client_from_args(args: argparse.Namespace) -> TMOClient:  # pragma: no cover
    """Create TMOClient from command-line arguments.

    Args:
        args: Parsed command-line arguments

    Returns:
        Configured TMOClient instance

    Raises:
        ValidationError: If required credentials are missing
    """
    # Resolve configuration values
    config_values = resolve_config_values(args)

    # Map environment string to enum
    env_map = {
        "us": Environment.US,
        "usa": Environment.US,
        "can": Environment.CANADA,
        "canada": Environment.CANADA,
        "aus": Environment.AUSTRALIA,
        "australia": Environment.AUSTRALIA,
    }

    env = env_map.get(config_values["environment"].lower(), Environment.US)
    user_agent = getattr(args, "user_agent", None) or os.getenv("TMO_USER_AGENT")

    return TMOClient(
        token=config_values["token"],
        database=config_values["database"],
        environment=env,
        timeout=config_values.get("timeout", 30),
        debug=getattr(args, "debug", False),
        user_agent=user_agent,
    )


def apply_default_date_ranges(args: argparse.Namespace) -> None:  # pragma: no cover
    """Apply default date ranges if not specified.

    Default logic:
    - No dates: last 31 days (31 days ago to today)
    - Only end date: 31 days before end date
    - Only start date: 31 days after start date
    - End date cannot exceed today

    Args:
        args: Parsed command-line arguments with start_date and end_date attributes
    """
    today = datetime.now()

    # If neither start nor end date provided, default to last 31 days
    if not getattr(args, "start_date", None) and not getattr(args, "end_date", None):
        args.end_date = today.strftime("%m/%d/%Y")
        args.start_date = (today - timedelta(days=31)).strftime("%m/%d/%Y")
        args._used_default_dates = True

    # If only end date provided, start date is 31 days before end date
    elif getattr(args, "end_date", None) and not getattr(args, "start_date", None):
        try:
            end_date = datetime.strptime(args.end_date, "%m/%d/%Y")
            # Ensure end date doesn't exceed today
            if end_date > today + timedelta(days=1):
                raise ValidationError(
                    f"End date cannot be later than today ({today.strftime('%m/%d/%Y')})"
                )
            args.start_date = (end_date - timedelta(days=31)).strftime("%m/%d/%Y")
        except ValueError:
            raise ValidationError("End date must be in MM/DD/YYYY format")

    # If only start date provided, end date is 31 days after start date
    elif getattr(args, "start_date", None) and not getattr(args, "end_date", None):
        try:
            start_date = datetime.strptime(args.start_date, "%m/%d/%Y")
            end_date = start_date + timedelta(days=31)
            # Ensure end date doesn't exceed today
            if end_date > today + timedelta(days=1):
                end_date = today
            args.end_date = end_date.strftime("%m/%d/%Y")
        except ValueError:
            raise ValidationError("Start date must be in MM/DD/YYYY format")

    # If both dates provided, validate end date doesn't exceed today
    else:
        try:
            end_date = datetime.strptime(args.end_date, "%m/%d/%Y")
            if end_date > today + timedelta(days=1):
                raise ValidationError(
                    f"End date cannot be later than today ({today.strftime('%m/%d/%Y')})"
                )
        except ValueError:
            raise ValidationError("End date must be in MM/DD/YYYY format")


def is_binary_field(field_name: str, field_value: Any) -> bool:
    """Check if a field contains binary data that should be hidden.

    Args:
        field_name: Name of the field
        field_value: Value of the field

    Returns:
        True if field appears to contain binary data
    """
    # Known binary/blob field names
    binary_field_names = [
        "Cert_TemplateFile",
        "TemplateFile",
        "FileContent",
        "BinaryData",
        "ImageData",
        "DocumentData",
        "AttachmentData",
        "FileData",
    ]

    # Check if field name indicates binary data
    field_name_lower = field_name.lower()
    if any(binary_name.lower() in field_name_lower for binary_name in binary_field_names):
        return True  # pragma: no cover - binary field detection

    # Check if value looks like binary data (base64 encoded strings over 100 chars)
    if isinstance(field_value, str) and len(field_value) > 100:  # pragma: no cover
        # Simple heuristic: if it's a long string with mostly base64-like characters
        if (
            re.match(r"^[A-Za-z0-9+/=\s]+$", field_value) and len(field_value) > 200
        ):  # pragma: no cover
            return True  # pragma: no cover

    # Check if it's a list/array with binary-looking data
    if isinstance(field_value, list) and field_value:  # pragma: no cover
        # If list contains long strings that look like binary
        first_item = field_value[0]  # pragma: no cover
        if isinstance(first_item, str) and len(first_item) > 100:  # pragma: no cover
            return True  # pragma: no cover

    return False


def format_output(data: Any, format_type: str = "text") -> str:
    """Format output data according to specified format.

    Args:
        data: Data to format
        format_type: Output format ("json" or "text")

    Returns:
        Formatted string
    """
    if format_type == "json":
        json_data: Any
        # Convert objects to dictionaries for JSON serialization
        if isinstance(data, list):
            json_data = []
            for item in data:
                if hasattr(item, "__dict__"):
                    json_data.append(
                        {
                            k: v
                            for k, v in item.__dict__.items()
                            if not k.startswith("_") and k != "raw_data"
                        }
                    )
                else:  # pragma: no cover - dict items handled in tests
                    json_data.append(item)
        elif hasattr(data, "__dict__"):  # pragma: no cover - single object
            json_data = {
                k: v for k, v in data.__dict__.items() if not k.startswith("_") and k != "raw_data"
            }
        else:  # pragma: no cover - primitive data
            json_data = data

        return json.dumps(json_data, indent=4, default=str)
    else:
        # Text format
        return format_table_output(data)


def format_table_output(data: Any) -> str:  # pragma: no cover
    """Format data as a readable table.

    Args:
        data: Data to format

    Returns:
        Formatted table string
    """
    if isinstance(data, dict):
        # Single object - display as key-value pairs
        lines = []
        for key, value in data.items():
            # Skip binary/blob fields
            if is_binary_field(key, value):
                lines.append(f"{key}: [BINARY DATA - {len(str(value))} bytes]")
                continue

            if isinstance(value, (dict, list)):
                value = json.dumps(value, default=str)
            lines.append(f"{key}: {value}")
        return "\n".join(lines)

    elif hasattr(data, "__dict__") and not isinstance(data, list):
        # Single object with attributes - convert to dict and display
        item_dict = {
            k: v
            for k, v in data.__dict__.items()
            if not k.startswith("_") and k != "raw_data" and v is not None
        }
        lines = []
        for key, value in item_dict.items():
            # Skip binary/blob fields
            if is_binary_field(key, value):
                lines.append(f"{key}: [BINARY DATA - {len(str(value))} bytes]")
                continue

            if isinstance(value, (dict, list)):
                value = json.dumps(value, default=str)
            lines.append(f"{key}: {value}")
        return "\n".join(lines)

    elif isinstance(data, list) and data:
        # Convert objects to dictionaries if needed
        dict_data = []
        for item in data:
            if hasattr(item, "__dict__"):
                # Convert object to dictionary, filtering out private attributes
                item_dict = {}
                for k, v in item.__dict__.items():
                    if not k.startswith("_") and k != "raw_data" and v is not None:
                        if is_binary_field(k, v):
                            item_dict[k] = f"[BINARY DATA - {len(str(v))} bytes]"
                        else:
                            item_dict[k] = v
                dict_data.append(item_dict)
            elif isinstance(item, dict):
                # Filter binary fields from dictionary items too
                filtered_item = {}
                for k, v in item.items():
                    if is_binary_field(k, v):
                        filtered_item[k] = f"[BINARY DATA - {len(str(v))} bytes]"
                    else:
                        filtered_item[k] = v
                dict_data.append(filtered_item)
            else:
                dict_data.append({"value": item})

        if dict_data and isinstance(dict_data[0], dict):
            # List of objects - display as table
            if not dict_data:
                return "No results found"

            # Get all unique keys from all objects
            all_keys: set[str] = set()
            for item in dict_data:
                if isinstance(item, dict):
                    all_keys.update(item.keys())

            headers = sorted(all_keys)

            # If more than 10 columns, use multi-line format
            if len(headers) > 10:
                return format_multiline_table(dict_data, headers)

            # Calculate column widths
            col_widths = {}
            for header in headers:
                col_widths[header] = len(str(header))
                for item in dict_data:
                    if isinstance(item, dict) and header in item:
                        value_str = str(item[header])
                        col_widths[header] = max(col_widths[header], len(value_str))

            # Build table
            lines = []

            # Header row
            header_row = " | ".join(header.ljust(col_widths[header]) for header in headers)
            lines.append(header_row)
            lines.append("-" * len(header_row))

            # Data rows
            for item in dict_data:
                if isinstance(item, dict):
                    row = " | ".join(
                        str(item.get(header, "")).ljust(col_widths[header]) for header in headers
                    )
                    lines.append(row)

            return "\n".join(lines)
        else:
            # List of simple values
            return "\n".join(str(item) for item in data)

    else:
        return str(data) if data else "No results found"


def format_multiline_table(data: list, headers: list) -> str:  # pragma: no cover
    """Format wide tables with multiple lines per record for better readability.

    Args:
        data: List of dictionaries to format
        headers: List of header names

    Returns:
        Formatted multi-line table string
    """
    lines = []

    # Calculate the maximum width for field names
    max_field_width = max(len(header) for header in headers)

    for i, item in enumerate(data):
        if not isinstance(item, dict):
            continue

        # Add record separator (except for first record)
        if i > 0:
            lines.append("")

        lines.append(f"Record {i + 1}:")
        lines.append("-" * (max_field_width + 50))

        # Display each field on its own line
        for header in headers:
            value = item.get(header, "")
            if isinstance(value, (dict, list)):
                value = json.dumps(value, default=str)

            # Handle very long values by wrapping them
            value_str = str(value)
            if len(value_str) > 80:
                # Wrap long values
                wrapped_lines = []
                for j in range(0, len(value_str), 80):
                    wrapped_lines.append(value_str[j : j + 80])
                value_display = "\n" + "\n".join(
                    f"{' ' * (max_field_width + 3)}{line}" for line in wrapped_lines
                )
            else:
                value_display = value_str

            lines.append(f"{header.ljust(max_field_width)} : {value_display}")

    return "\n".join(lines)


def is_name_value_array(data: Any) -> bool:
    """Check if an array contains name-value pair objects.

    Returns True if all items are dicts with 'Name' and 'Value' keys.

    Args:
        data: Data to check

    Returns:
        True if data is a name-value array
    """
    if not isinstance(data, list) or len(data) == 0:
        return False

    for item in data:
        if not isinstance(item, dict):
            return False
        if "Name" not in item or "Value" not in item:
            return False

    return True


def flatten_name_value_array(data: List[Dict[str, Any]], parent_key: str = "") -> Dict[str, Any]:
    """Flatten a name-value array into a dictionary using Name as keys and Value as values.

    Args:
        data: List of dictionaries with Name and Value keys
        parent_key: Parent key prefix for the flattened keys

    Returns:
        Flattened dictionary with Name fields as keys
    """
    items = {}
    for item in data:
        name = str(item.get("Name", "")).strip()
        value = item.get("Value", "")

        if name:  # Only add if name is not empty
            # Clean up the name to be a valid column name
            clean_name = name.replace(" ", "_").replace("-", "_").replace(".", "_")
            if parent_key:
                key = f"{parent_key}_{clean_name}"
            else:
                key = clean_name
            items[key] = value

    return items


def flatten_json(
    data: Any,
    separator: str = "_",
    max_levels: int = 2,
    current_level: int = 0,
    parent_key: str = "",
) -> Dict[str, Any]:
    """Flatten nested JSON data to a specified depth.

    Args:
        data: The data to flatten (dict, list, or primitive)
        separator: String to use for separating nested keys
        max_levels: Maximum levels to flatten
        current_level: Current nesting level (for recursion)
        parent_key: Parent key for building nested key names

    Returns:
        Flattened dictionary
    """
    items = {}

    if isinstance(data, dict):
        for key, value in data.items():
            new_key = f"{parent_key}{separator}{key}" if parent_key else key

            # If we've reached max levels, store as JSON string
            if current_level >= max_levels:
                if isinstance(value, (dict, list)):
                    items[new_key] = json.dumps(value, default=str)
                else:  # pragma: no cover - primitive at max level
                    items[new_key] = value
            else:
                # Continue flattening
                if isinstance(value, (dict, list)):
                    items.update(
                        flatten_json(value, separator, max_levels, current_level + 1, new_key)
                    )
                else:
                    items[new_key] = value

    elif isinstance(data, list):
        # Check if this is a name-value array
        if is_name_value_array(data):
            items.update(flatten_name_value_array(data, parent_key))
        else:
            # Regular array processing
            for i, value in enumerate(data):
                new_key = f"{parent_key}{separator}{i}" if parent_key else str(i)

                # If we've reached max levels, store as JSON string
                if current_level >= max_levels:  # pragma: no cover - array max level
                    if isinstance(value, (dict, list)):  # pragma: no cover
                        items[new_key] = json.dumps(value, default=str)  # pragma: no cover
                    else:  # pragma: no cover
                        items[new_key] = value  # pragma: no cover
                else:
                    # Continue flattening
                    if isinstance(value, (dict, list)):  # pragma: no cover - nested array
                        items.update(  # pragma: no cover
                            flatten_json(
                                value, separator, max_levels, current_level + 1, new_key
                            )  # pragma: no cover
                        )  # pragma: no cover
                    else:
                        items[new_key] = value
    else:
        # Primitive value
        if parent_key:
            items[parent_key] = data
        else:
            items["value"] = data

    return items


def prepare_data_for_flattening(data: Any) -> List[Dict[str, Any]]:
    """Prepare data for CSV/XLSX export by flattening and converting to list of dicts.

    Args:
        data: Input data (can be dict, list, or objects with __dict__)

    Returns:
        List of flattened dictionaries
    """
    # Fields to exclude from output
    excluded_fields = {"raw_data"}

    # Convert objects to dictionaries first
    if isinstance(data, list):
        dict_data = []
        for item in data:
            if hasattr(item, "__dict__"):
                dict_data.append(
                    {
                        k: v
                        for k, v in item.__dict__.items()
                        if not k.startswith("_") and k not in excluded_fields
                    }
                )
            elif isinstance(item, dict):
                dict_data.append({k: v for k, v in item.items() if k not in excluded_fields})
            else:  # pragma: no cover - simple list items
                dict_data.append({"value": item})
    elif hasattr(data, "__dict__"):
        dict_data = [
            {
                k: v
                for k, v in data.__dict__.items()
                if not k.startswith("_") and k not in excluded_fields
            }
        ]
    elif isinstance(data, dict):
        dict_data = [{k: v for k, v in data.items() if k not in excluded_fields}]
    else:  # pragma: no cover - primitive value
        dict_data = [{"value": data}]

    # Flatten each record
    flattened_records = []
    for record in dict_data:
        flattened = flatten_json(record, separator="_", max_levels=2)
        # Remove any excluded fields from flattened result as well
        flattened = {k: v for k, v in flattened.items() if k not in excluded_fields}
        flattened_records.append(flattened)

    return flattened_records


def write_to_csv(records: List[Dict[str, Any]], output_file: str) -> None:
    """Write records to CSV file.

    Args:
        records: List of flattened dictionaries
        output_file: Output file path
    """
    if not records:
        print("No records to write", file=sys.stderr)
        return

    # Get all unique fieldnames from all records
    fieldnames_set: set[str] = set()
    for record in records:
        fieldnames_set.update(record.keys())

    fieldnames = sorted(fieldnames_set)

    with open(output_file, "w", newline="", encoding="utf-8") as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()

        for record in records:
            # Ensure all values are strings and handle None values
            clean_record = {}
            for key in fieldnames:
                value = record.get(key)
                if value is None:
                    clean_record[key] = ""
                elif isinstance(value, (dict, list)):  # pragma: no cover - nested in CSV
                    clean_record[key] = json.dumps(value, default=str)  # pragma: no cover
                else:
                    clean_record[key] = str(value)
            writer.writerow(clean_record)


def write_to_xlsx(records: List[Dict[str, Any]], output_file: str) -> None:
    """Write records to XLSX file using openpyxl.

    Args:
        records: List of flattened dictionaries
        output_file: Output file path

    Raises:
        ImportError: If openpyxl is not installed
    """
    try:
        import openpyxl
        from openpyxl.worksheet.worksheet import Worksheet
    except ImportError:
        raise ImportError(
            "openpyxl is required for XLSX output. Install with: pip install tmo-api[xlsx]"
        )

    if not records:
        print("No records to write", file=sys.stderr)
        return

    # Get all unique fieldnames from all records
    fieldnames_set: set[str] = set()
    for record in records:
        fieldnames_set.update(record.keys())

    fieldnames = sorted(fieldnames_set)

    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:  # pragma: no cover - wb.active should always exist
        ws = wb.create_sheet("Data")  # pragma: no cover
    assert isinstance(ws, Worksheet)
    ws.title = "Data"

    # Write headers
    for col, fieldname in enumerate(fieldnames, 1):
        ws.cell(row=1, column=col, value=fieldname)

    # Write data
    for row_idx, record in enumerate(records, 2):
        for col_idx, fieldname in enumerate(fieldnames, 1):
            value = record.get(fieldname)
            if value is None:
                cell_value = ""
            elif isinstance(value, (dict, list)):  # pragma: no cover - nested in XLSX
                cell_value = json.dumps(value, default=str)  # pragma: no cover
            else:
                cell_value = str(value)

            ws.cell(row=row_idx, column=col_idx, value=cell_value)

    # Auto-adjust column widths
    for col_tuple in ws.columns:
        max_length = 0
        first_cell = col_tuple[0]
        if hasattr(first_cell, "column_letter"):
            column_letter = first_cell.column_letter
        else:  # pragma: no cover - merged cells edge case
            continue  # pragma: no cover

        for cell in col_tuple:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:  # pragma: no cover - cell value error
                pass  # pragma: no cover
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_file)


def handle_output(data: Any, output_path: Optional[str]) -> None:
    """Handle output formatting and writing based on file extension or stdout.

    Args:
        data: Data to output
        output_path: Output file path (None for stdout text format)

    Raises:
        ValueError: If output format cannot be determined
    """
    if output_path is None:
        # No output file specified - print as text to stdout
        output = format_output(data, "text")
        print(output)
        return

    # Determine format from file extension
    output_file = Path(output_path)
    extension = output_file.suffix.lower()

    if extension == ".json":
        # JSON format
        output = format_output(data, "json")
        output_file.write_text(output, encoding="utf-8")
        print(f"Wrote JSON output to {output_path}", file=sys.stderr)

    elif extension == ".csv":
        # CSV format - flatten and write
        records = prepare_data_for_flattening(data)
        write_to_csv(records, output_path)
        print(f"Wrote {len(records)} record(s) to {output_path}", file=sys.stderr)

    elif extension in [".xlsx", ".xls"]:
        # XLSX format - flatten and write
        records = prepare_data_for_flattening(data)
        write_to_xlsx(records, output_path)
        print(f"Wrote {len(records)} record(s) to {output_path}", file=sys.stderr)

    else:
        raise ValueError(
            f"Unsupported output format: {extension}. Supported formats: .json, .csv, .xlsx"
        )
