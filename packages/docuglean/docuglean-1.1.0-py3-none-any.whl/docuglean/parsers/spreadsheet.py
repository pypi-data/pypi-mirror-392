"""Spreadsheet parsing utilities using openpyxl."""

from __future__ import annotations


async def parse_spreadsheet(file_path: str) -> dict:
    if not file_path:
        return {"text": ""}
    
    from openpyxl import load_workbook
    
    workbook = load_workbook(filename=file_path, read_only=True)
    texts = []
    
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(values_only=True):
            row_text = "\t".join(str(cell) if cell is not None else "" for cell in row)
            texts.append(row_text)
    
    return {"text": "\n".join(texts)}
