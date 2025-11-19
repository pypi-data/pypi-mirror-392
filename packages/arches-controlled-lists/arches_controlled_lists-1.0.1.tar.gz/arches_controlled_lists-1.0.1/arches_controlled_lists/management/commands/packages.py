import os
import sys

import openpyxl
from django.db import transaction
from django.db.models import Q

from arches.management.commands.packages import Command as PackagesCommand
from arches.app.models import models
from arches_controlled_lists.models import List, ListItem, ListItemValue
from arches_controlled_lists.utils.skos import SKOSReader, SKOSWriter


class Command(PackagesCommand):

    def add_arguments(self, parser):
        super().add_arguments(parser)

        idx_of_operation_arg = [a.dest for a in parser._actions].index("operation")
        parser._actions[idx_of_operation_arg].choices.extend(
            ["import_controlled_lists", "export_controlled_lists"]
        )

        parser.add_argument(
            "-fn",
            "--file_name",
            type=str,
            dest="file_name",
            default="export_controlled_lists",
            help="The name of the file to export to. Default is export_controlled_lists",
        )

        parser.add_argument(
            "-cl",
            "--controlled_lists",
            type=str,
            dest="controlled_lists",
            help="A comma-separated list of controlled list names to export. If not provided, all controlled lists will be exported.",
        )

    def handle(self, *args, **options):
        super().handle(self, *args, **options)

        if options["operation"] == "import_controlled_lists":
            self.import_controlled_lists(options["source"], options["overwrite"])

        if options["operation"] == "export_controlled_lists":
            self.export_controlled_lists(
                options["dest_dir"],
                options["file_name"],
                options["controlled_lists"],
                options["format"],
            )

    def import_controlled_lists(self, source, overwrite_options):
        if source.lower().endswith(".xml"):
            skos = SKOSReader()
            rdf = skos.read_file(source)
            concepts = skos.save_controlled_lists_from_skos(
                rdf, overwrite_options=overwrite_options
            )

        elif source.lower().endswith(".xlsx"):
            created_instances_pks = []
            if os.path.exists(source):
                wb = openpyxl.load_workbook(source)
                with transaction.atomic():
                    for sheet in wb.sheetnames:
                        if sheet == "List":
                            created_instances_pks.extend(
                                self.import_sheet_to_model(wb[sheet], List)
                            )
                        elif sheet == "ListItem":
                            created_instances_pks.extend(
                                self.import_sheet_to_model(wb[sheet], ListItem)
                            )
                        elif sheet == "ListItemValue":
                            created_instances_pks.extend(
                                self.import_sheet_to_model(wb[sheet], ListItemValue)
                            )
                    # validate all data
                    for model in [
                        List,
                        ListItem,
                        ListItemValue,
                    ]:
                        for instance in model.objects.filter(
                            pk__in=created_instances_pks
                        ):
                            instance.full_clean()
                    self.stdout.write(
                        "Data imported successfully from {0}".format(source)
                    )
        else:
            self.stdout.write(
                "The source file does not exist or is not the correct format. Please rerun this command with a valid source file."
            )

    def import_sheet_to_model(self, sheet, model):
        fields = [
            {"name": field.name, "is_fk": field.get_internal_type() == "ForeignKey"}
            for field in model._meta.fields
        ]
        field_names = [field["name"] for field in fields]

        # Parse the sheet into a list of dictionaries
        import_table = []
        for imported_row in sheet.iter_rows(min_row=2, values_only=True):
            working_row = {}
            for field in field_names:
                working_row[field] = imported_row[field_names.index(field)]
            import_table.append(working_row)

        # Process row data and create instances of the model
        instances = []
        instance_pks = []
        list_items_with_parent = {}
        for row in import_table:
            instance = model()
            for field in fields:
                is_fk = field["is_fk"]
                field_name = field["name"]
                value = row[field_name] if row[field_name] else None  # might be ''
                if value and is_fk and model is ListItem:
                    if field_name == "list":
                        related_list = List.objects.get(id=value)
                        setattr(instance, field_name, related_list)
                    elif field_name == "parent":
                        # stash list items with parent relationships to create relationships after all list items have been created
                        # stashed object in the form of {child_list_item_instance : parent_list_item_pk, ...}
                        list_items_with_parent[instance] = value
                elif value and is_fk and model is ListItemValue:
                    if field_name == "valuetype":
                        valuetype = models.DValueType.objects.get(valuetype=value)
                        setattr(instance, field_name, valuetype)
                    elif field_name == "language":
                        try:
                            related_language = models.Language.objects.get(code=value)
                            setattr(instance, field_name, related_language)
                        except models.Language.DoesNotExist:
                            self.stderr.write(
                                f"Language with code {value} does not exist. Please create this language before importing these data."
                            )
                            sys.exit()
                    else:
                        related_list_item = ListItem.objects.get(id=value)
                        setattr(instance, field_name, related_list_item)
                else:
                    setattr(instance, field_name, value)

            # run validation on all non-parent fields & gather for bulk create
            instance.clean_fields(exclude={"parent"})
            instances.append(instance)
            instance_pks.append(instance.pk)

        model.objects.bulk_create(instances)

        if model is ListItem:
            # Create list item relationships after all list items have been created
            for child, parent in list_items_with_parent.items():
                child.parent = ListItem.objects.get(id=parent)
                child.clean_fields(
                    exclude={
                        field["name"] for field in fields if field["name"] != "parent"
                    }
                )
                child.save()

        return instance_pks

    def export_controlled_lists(self, data_dest, file_name, controlled_lists, format):
        if format == "xlsx":
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "List"
            self.export_model_to_sheet(ws, List)
            self.export_model_to_sheet(wb, ListItem)
            self.export_model_to_sheet(wb, ListItemValue)

            if data_dest != "" and data_dest != ".":
                wb.save(os.path.join(data_dest, f"{file_name}.xlsx"))
                self.stdout.write(f"Data exported successfully to {file_name}.xlsx")
            else:
                self.stdout.write(
                    "No destination directory specified. Please rerun this command with the '-d' parameter populated."
                )

        elif format == "skos-rdf":
            parsed_lists = [lst.strip() for lst in controlled_lists.split(",")]
            if parsed_lists != [""]:
                export_lists = List.objects.filter(
                    Q(name__in=parsed_lists) | Q(id__in=parsed_lists)
                )
                export_list_items = ListItem.objects.filter(
                    list__in=export_lists
                ).prefetch_related("list_item_values", "parent", "children")
            else:
                export_lists = List.objects.all()
                export_list_items = ListItem.objects.all().prefetch_related(
                    "list_item_values", "parent", "children"
                )
            skos = SKOSWriter()
            skos_file = skos.write_controlled_lists(
                export_lists, export_list_items, format="pretty-xml"
            )

            if data_dest != "" and data_dest != ".":
                with open(os.path.join(data_dest, f"{file_name}.xml"), "wb") as file:
                    file.write(skos_file)
                self.stdout.write(f"Data exported successfully to {file_name}.xml")

        else:
            self.stdout.write(
                f"The specified format {format} is not supported. Please rerun this command with a supported format."
            )

    def export_model_to_sheet(self, wb, model):
        # For the first sheet (List), use blank sheet that is initiallized with workbook
        # otherwise, append a new sheet
        if isinstance(wb, openpyxl.worksheet.worksheet.Worksheet):
            ws = wb
        else:
            ws = wb.create_sheet(title=model.__name__)
        fields = [
            {"name": field.name, "datatype": field.get_internal_type()}
            for field in model._meta.fields
        ]
        ws.append(field["name"] for field in fields)
        for instance in model.objects.all():
            row_data = []
            for field in fields:
                value = getattr(instance, field["name"])
                if isinstance(
                    value,
                    (
                        List,
                        ListItem,
                        ListItemValue,
                    ),
                ):
                    row_data.append(str(getattr(value, "id")) if value else "")
                elif isinstance(value, models.Language):
                    row_data.append(str(value.code))
                elif isinstance(value, models.DValueType):
                    row_data.append(str(value.valuetype))
                elif field["datatype"] == "UUIDField":
                    row_data.append(str(value) if value else "")
                elif field["datatype"] == "BooleanField":
                    row_data.append("1" if value else "0")
                elif field["datatype"] == "IntegerField":
                    row_data.append(str(value))
                else:
                    row_data.append(value if value else "")
            ws.append(row_data)
