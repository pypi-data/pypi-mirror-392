import os
import pandas as pd
from openpyxl import load_workbook

from validata.io.loader import load_file  # not used here but kept for consistency


def clear_rows(file, sheet=None):
    """
    Remove all rows except the header row.

    For CSV:
        - keeps only the header line and overwrites the file.

    For Excel:
        - keeps only the header row in the specified sheet
          (or first sheet if not specified).
        - preserves sheet names and order.
        - other sheets are unchanged.

    Parameters
    ----------
    file : str or path-like
        Path to the CSV or Excel file.
    sheet : str | int | None, optional
        Sheet name or index for Excel. Ignored for CSV.
        If None, uses the first sheet for Excel.
    """
    lower = str(file).lower()

    if lower.endswith(".csv"):
        _clear_rows_csv(file)
    elif lower.endswith(".xlsx") or lower.endswith(".xls"):
        _clear_rows_excel(file, sheet)
    else:
        raise ValueError(f"Unsupported file type: {file!r}")


def _clear_rows_csv(file):
    # load header only
    df = pd.read_csv(file, nrows=0)
    empty_df = pd.DataFrame(columns=df.columns)
    empty_df.to_csv(file, index=False)


def _clear_rows_excel(file, sheet):
    wb = load_workbook(filename=file)

    # select sheet
    if sheet is None:
        ws = wb.worksheets[0]
    elif isinstance(sheet, int):
        ws = wb.worksheets[sheet]
    else:
        ws = wb[sheet]

    max_row = ws.max_row
    if max_row > 1:
        # delete all rows below header
        ws.delete_rows(2, max_row - 1)

    wb.save(file)
