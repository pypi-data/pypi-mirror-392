"""XLSX style: writes an Excel workbook using openpyxl.

Optional dependency: openpyxl (group: xlsx)
"""

from typing import Any, IO, TYPE_CHECKING
from datetime import date, datetime
from pathlib import Path

from .table_style import TableStyle

if TYPE_CHECKING:  # avoid circular import at runtime
    from ..craftable import ColDefList


class XlsxStyle(TableStyle):
    def __init__(self, sheet_name: str = "Sheet1"):
        super().__init__()
        self.terminal_style = False
        self.string_output = False  # Binary format
        self.sheet_name = sheet_name

    ###############################################################################
    # write_table
    ###############################################################################

    def write_table(
        self,
        value_rows: list[list[Any]],
        header_row: list[str] | None,
        col_defs: "ColDefList",
        header_defs: "ColDefList | None",
        file: str | Path | IO[bytes],
    ) -> None:
        """Write table to XLSX file format."""
        try:
            from openpyxl import Workbook  # type: ignore
            from openpyxl.utils import get_column_letter  # type: ignore
            from openpyxl.styles import Alignment, Font  # type: ignore
        except Exception as e:  # pragma: no cover - only hit without optional dep
            raise ImportError(
                "openpyxl is required for XlsxStyle; install group 'xlsx'"
            ) from e

        wb = Workbook()
        ws: Any = wb.active  # type: ignore[assignment]
        ws.title = self.sheet_name

        # Write header row
        row_idx = 1
        if header_row:
            for col_idx, text in enumerate(header_row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=str(text))
                cell.font = Font(bold=True)
                # Use alignment from header_defs if available, else center
                if header_defs and (col_idx - 1) < len(header_defs):
                    hdr_def = header_defs[col_idx - 1]
                    if ">" in hdr_def.align:
                        cell.alignment = Alignment(horizontal="right")
                    elif "^" in hdr_def.align:
                        cell.alignment = Alignment(horizontal="center")
                    else:
                        cell.alignment = Alignment(horizontal="left")
                else:
                    cell.alignment = Alignment(horizontal="center")
            row_idx += 1

        # Helper for converting format specs to Excel number formats
        def _to_excel_number_format(fmt_spec_obj) -> str | None:
            """Map FormatSpec or string to Excel number format.

            Supports:
            - Float with precision and grouping: .2f, ,.2f -> 0.00, #,##0.00
            - Integer with grouping: d, ,d -> 0, #,##0
            - Percentage with precision: .2%, % -> 0.00%, 0%

            Returns None if unsupported or if fmt_spec_obj is None.
            """
            if fmt_spec_obj is None:
                return None

            # Handle FormatSpec objects by extracting relevant fields
            if hasattr(fmt_spec_obj, "type"):
                type_char = fmt_spec_obj.type
                has_group = fmt_spec_obj.grouping == ","
                precision_str = fmt_spec_obj.precision

                # Parse precision (e.g., ".2" -> 2)
                decimals = 0
                if precision_str and precision_str.startswith("."):
                    try:
                        decimals = int(precision_str[1:])
                    except (ValueError, IndexError):
                        decimals = 0

                if type_char == "f":
                    base = "#,##0" if has_group else "0"
                    if decimals:
                        base += "." + ("0" * decimals)
                    return base
                elif type_char == "d":
                    return "#,##0" if has_group else "0"
                elif type_char == "%":
                    base = "0" if decimals == 0 else "0." + ("0" * decimals)
                    return base + "%"
                return None

            # Fallback: parse string representation
            if not isinstance(fmt_spec_obj, str):
                fmt_spec_obj = str(fmt_spec_obj)

            spec = fmt_spec_obj.strip().lstrip("<>^")
            if spec.endswith("f"):
                parts = spec[:-1].split(".")
                decimals = 0
                if len(parts) == 2 and parts[1].isdigit():
                    decimals = int(parts[1])
                has_group = "," in parts[0]
                base = "#,##0" if has_group else "0"
                if decimals:
                    base += "." + ("0" * decimals)
                return base
            if spec.endswith("d"):
                has_group = "," in spec[:-1]
                return "#,##0" if has_group else "0"
            if spec.endswith("%"):
                # Extract precision before %
                decimals = 0
                if ".%" in spec:
                    # Pattern like ".2%"
                    parts = spec.split(".")
                    if len(parts) >= 2:
                        num_part = parts[1].rstrip("%")
                        if num_part.isdigit():
                            decimals = int(num_part)
                base = "0" if decimals == 0 else "0." + ("0" * decimals)
                return base + "%"
            return None

        # Helper for formatting cell values with preprocess/postprocess
        def _coerce_cell_value(
            cd, original: Any, row: list[Any], col_idx: int
        ) -> tuple[Any, str | None]:
            """Return (value_for_cell, excel_number_format).
            Falls back to formatted text when numeric coercion not safe.
            """
            v2 = cd.preprocess(original, row, col_idx)
            if v2 is None:
                return "", None
            # If prefix/suffix/postprocessor present we treat as text
            if cd.prefix or cd.suffix or cd.postprocessor is not None:
                try:
                    if cd.format_spec:
                        fmt = f"{{:{cd.format_spec}}}"
                        text = fmt.format(v2)
                    else:
                        text = str(v2)
                except Exception:
                    text = str(v2)
                text = cd.prefix + text + cd.suffix
                return cd.postprocess(original, text, row, col_idx), None
            # Direct acceptable python types
            if isinstance(v2, (bool, int, float, date, datetime)):
                nf = _to_excel_number_format(cd.format_spec) if cd.format_spec else None
                return v2, nf
            # Try to parse strings into numbers
            if isinstance(v2, str):
                txt = v2.strip().replace(",", "")
                try:
                    if "." in txt:
                        num = float(txt)
                        nf = (
                            _to_excel_number_format(cd.format_spec)
                            if cd.format_spec
                            else None
                        )
                        return num, nf
                    num = int(txt)
                    nf = (
                        _to_excel_number_format(cd.format_spec)
                        if cd.format_spec
                        else None
                    )
                    return num, nf
                except Exception:
                    pass
            # Fallback to formatted text
            try:
                if cd.format_spec:
                    fmt = f"{{:{cd.format_spec}}}"
                    text = fmt.format(v2)
                else:
                    text = str(v2)
            except Exception:
                text = str(v2)
            text = cd.prefix + text + cd.suffix
            text = cd.postprocess(original, text, row, col_idx)
            return text, None

        # Write data rows
        for r, row in enumerate(value_rows):
            for c, (val, col_def) in enumerate(zip(row, col_defs), start=1):
                cell_value, number_format = _coerce_cell_value(col_def, val, row, c - 1)
                cell = ws.cell(row=row_idx, column=c, value=cell_value)
                if number_format:
                    try:
                        cell.number_format = number_format
                        if isinstance(cell.value, int) and "." in number_format:
                            cell.value = float(cell.value)
                    except Exception:
                        pass
                if ">" in col_def.align:
                    cell.alignment = Alignment(horizontal="right")
                elif "^" in col_def.align:
                    cell.alignment = Alignment(horizontal="center")
                else:
                    cell.alignment = Alignment(horizontal="left")
            row_idx += 1

        # Set column widths based on ColDef widths
        # Excel width units are approximately character widths at default font
        # Add small padding for better visual appearance
        for c, col_def in enumerate(col_defs, start=1):
            letter = get_column_letter(c)
            char_width = max(col_def.width, 1)
            # Excel width formula: slightly wider than character count for padding
            ws.column_dimensions[letter].width = char_width + 2

        wb.save(file)
