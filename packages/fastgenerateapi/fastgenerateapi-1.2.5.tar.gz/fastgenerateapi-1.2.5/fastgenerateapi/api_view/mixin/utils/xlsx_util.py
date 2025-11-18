import importlib
import io
import operator
from tempfile import NamedTemporaryFile
from typing import List, Union, Optional, Dict, Type

import openpyxl
from fastapi import UploadFile
from openpyxl.styles import Alignment, PatternFill
from openpyxl.styles.colors import COLOR_INDEX, Color
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from starlette._utils import is_async_callable
from starlette.responses import StreamingResponse, JSONResponse, FileResponse
from tortoise import Model
from pydantic import ValidationError

from fastgenerateapi import BaseModel, BaseView
from fastgenerateapi.api_view.mixin.response_mixin import ResponseMixin
from fastgenerateapi.schemas_factory.common_schema_factory import common_schema_factory


class XlsxUtil:
    default_align = Alignment(
        horizontal='center',
        vertical='center',
        text_rotation=0,
        wrap_text=True,
        shrink_to_fit=True,
        indent=0,
    )
    default_fill = PatternFill(
        start_color=Color(COLOR_INDEX[44]),
        end_color=Color(COLOR_INDEX[44]),
        fill_type='solid'
    )

    @staticmethod
    def write_headers(sh: Worksheet, headers: List[str]) -> List[int]:
        """
        写入第一行信息
        :return:
        """
        col_max_len_list = []
        sh.row_dimensions[1].height = 26
        for col, header in enumerate(headers, 1):
            sh.cell(1, col).value = header
            sh.cell(1, col).alignment = XlsxUtil.default_align
            sh.cell(1, col).fill = XlsxUtil.default_fill
            sh.cell(1, col).alignment = XlsxUtil.default_align
            col_max_len_list.append(len(header.encode('gb18030')))

        return col_max_len_list

    @staticmethod
    def write_content(model_list: List[Model]):
        """
        填写内容部分
        :return:
        """
        # 跳过标题，从第二行开始写入
        for row, model in enumerate(model_list, 2):
            ...

        return

    @staticmethod
    def adaptive_format(sh: Worksheet, col_max_len_list: List[int], height_num: int):
        """
        自适应宽度
        :return:
        """
        # 设置自适应列宽
        for i, col_max_len in enumerate(col_max_len_list, 1):
            # 256*字符数得到excel列宽,为了不显得特别紧凑添加两个字符宽度
            max_width = col_max_len + 4
            if max_width > 256:
                max_width = 256
            sh.column_dimensions[get_column_letter(i)].width = max_width
        for y in range(2, height_num + 2):
            sh.row_dimensions[y].height = 18

        return

    async def export_xlsx(
            self,
            model_list: List[Model],
            headers: List[str],
            fields: List[str],
            fields_handler: dict,
            file_save_path: Optional[str] = None,
            # rpc_param: Union[Dict[str, Dict[str, List[str]]], Type[RPCParam], None] = None,
            title: str = None,
    ) -> StreamingResponse:
        wb = openpyxl.Workbook()
        col_max_len_list = []

        def write(sh, row, col, value):
            sh.cell(row, col).value = value
            sh.cell(row, col).alignment = XlsxUtil.default_align
            if col_max_len_list[col - 1] < len(str(value).encode('gb18030')):
                col_max_len_list[col - 1] = len(str(value).encode('gb18030'))

        start_row = 1
        try:
            sh = wb.active
            sh.title = title if title else f'{self.model_class._meta.table_description}'

            col_max_len_list = XlsxUtil.write_headers(sh, headers)

            for row, model in enumerate(model_list, start_row + 1):
                model = await BaseView.getattr_model(model=model, fields=fields)
                # model = await self.setattr_model_rpc(self.model_class, model, rpc_param)

                for col, field in enumerate(fields, 1):
                    info = getattr(model, field, "")
                    handler = fields_handler.get(field)
                    if handler and hasattr(handler, "__call__"):
                        if is_async_callable(handler):
                            info = await handler(info)
                        else:
                            info = handler(info)
                    write(sh, row, col, info)

            XlsxUtil.adaptive_format(sh, col_max_len_list, len(model_list))
        finally:
            if file_save_path:
                wb.save(file_save_path)
                return ResponseMixin.success(msg="请求成功")
            bytes_io = io.BytesIO()
            wb.save(bytes_io)
            bytes_io.seek(0)

        return ResponseMixin.stream(bytes_io, is_xlsx=True)

    async def import_xlsx(
            self,
            file: UploadFile,
            headers: List[str],
            # [
            #        "name",
            #       ("is_male", {"男"： True, "女": False} 或者 方法, {"额外字段"： 方法}, ...),
            # ]
            # 方法(默认传excel的值)
            fields: List[Union[str, dict, tuple, list]],
            combine_fields: Optional[List[Dict[str, any]]] = None,
            model_class: Optional[Type[Model]] = None,
            create_schema: Optional[Type[BaseModel]] = None,
            # storage_path: Union[str, Path],
            # rpc_param: Union[Dict[str, Dict[str, List[Union[str, tuple]]]], Type[RPCParam]] = None,
            modules: str = "openpyxl",
    ) -> JSONResponse:
        """
        fields: 方法(默认传excel的值)
        例如：
        [
            "name",    # 传入值是 name 字段的值
            ("is_male", {"男"： True, "女": False} 或者 方法, {"额外字段"： 方法}, ...),
            # 值 "男" 获取为bool值，不在字典里为None, 页可以自定义 同步或异步方法 获取值
        ]
        """
        limit_modules = ["openpyxl"]
        if modules not in limit_modules:
            return ResponseMixin.error(msg=f"export xlsx modules only import {'、'.join(limit_modules)}")

        if not file:
            return ResponseMixin.fail(msg=f"请先选择合适的文件")

        if not model_class:
            model_class = self.model_class
        if not create_schema:
            create_schema = common_schema_factory(model_class, name=f"{model_class.__name__}ExcelImportSchema")

        with NamedTemporaryFile() as tmp2:
            tmp2.write(await file.read())
            try:
                wb = importlib.import_module(modules).load_workbook(tmp2, read_only=True, data_only=True)
            except Exception:
                return ResponseMixin.error(msg=f"please pip install {modules}")

            try:
                ws = wb.active

                header_row = ws[1]
                header_list = []
                for msg in header_row:
                    header_list.append(str(msg.value).replace(" ", ''))

                if len(header_list) != len(headers):
                    return ResponseMixin.fail(message="文件首行长度校验错误")

                if not operator.eq(header_list, headers):
                    return ResponseMixin.fail(message="文件首行内容校验错误")

                # if ws.max_row < 2:
                #     return ResponseMixin.fail(msg="导入数据不能为空")

                create_list = []
                effective_row = 0
                for row in range(2, ws.max_row + 1):
                    data = {}
                    # data_schema = {}
                    row_data = ws[row]
                    if await self.excel_row_is_empty(row_data):
                        continue
                    effective_row += 1
                    for col, field_input in enumerate(fields):
                        if type(field_input) in [str, int]:
                            data[field_input] = row_data[col].value
                            # data_schema[field_input] = (type(row_data[col].value), ...)

                        if type(field_input) == tuple or type(field_input) == list:
                            key = field_input[0]
                            val = field_input[1]
                            required_doc = {}
                            if len(field_input) > 2:
                                required_doc = field_input[2]
                                if required_doc == "required":
                                    required_doc = {"required": True}
                            if type(val) == dict:
                                model_val = val.get(row_data[col].value)
                                if not model_val and required_doc.get("required"):
                                    return ResponseMixin.fail(
                                        msg=required_doc.get("error",
                                                             "") or f"第{row}行{self.get_field_description(key)}不能为空")
                                data[key] = model_val
                                # data_schema[key] = (type(model_val), ...)
                            elif hasattr(val, "__call__"):
                                if is_async_callable(val):
                                    model_val = await val(row_data[col].value)
                                else:
                                    model_val = val(row_data[col].value)
                                data[key] = model_val
                                # data_schema[key] = (type(model_val), ...)
                            else:
                                raise NotImplemented
                        else:
                            raise NotImplemented
                    for combine_field in combine_fields:
                        field = combine_field.get("field", None)
                        value = combine_field.get("value", None)
                        function = combine_field.get("function", None)
                        args = combine_field.get("args", None)
                        if not field or (not function and not value):
                            continue
                        if value:
                            data[field] = value
                        else:
                            if not args:
                                if is_async_callable(function):
                                    model_val = await function()
                                else:
                                    model_val = function()
                            else:
                                args_list = []
                                for arg in args:
                                    args_list.append(data.get(arg, ""))
                                if is_async_callable(function):
                                    model_val = await function(*args_list)
                                else:
                                    model_val = function(*args_list)
                            data[field] = model_val
                    try:
                        create_obj = model_class(**create_schema(**data).dict(exclude_unset=True))
                    except ValidationError as e:
                        error_field = e.errors()[0].get('loc')[0]
                        description = self.get_field_description(error_field)
                        if not data.get(error_field):
                            return ResponseMixin.fail(message=f"第{row}行{description}不能为空")
                        return ResponseMixin.fail(message=f"第{row}行{description}填写错误")
                    await self.check_unique_field(create_obj, model_class=model_class)
                    create_list.append(create_obj)

                await model_class.bulk_create(create_list)
            finally:
                wb.close()
            if effective_row == 0:
                return ResponseMixin.fail(message="导入数据不能为空")
        return ResponseMixin.success(msg='创建成功')

    @staticmethod
    async def excel_row_is_empty(row_list) -> bool:
        is_empty = True
        for row in row_list:
            if row.value is not None:
                return False

        return is_empty

    async def excel_model(
            self,
            headers: List[str] = None,
            model_class: Optional[Model] = None,
            excel_model_path: Optional[str] = None,
            modules: str = "openpyxl",
            title: Optional[str] = None,
    ) -> Union[FileResponse, StreamingResponse]:
        if excel_model_path:
            return FileResponse(
                path=excel_model_path,
                filename="导入模板.xlsx",
                media_type="xlsx",
            )

        limit_modules = ["openpyxl", "xlsxwriter"]
        if modules not in limit_modules:
            return ResponseMixin.error(msg=f"export xlsx modules only import {'、'.join(limit_modules)}")
        try:
            wb = importlib.import_module(modules).Workbook()
        except Exception:
            return ResponseMixin.error(msg=f"please pip install {modules}")
        if modules == "openpyxl":
            def write(sh, row, col, value):
                sh.cell(row, col).value = value

            start_col = 1
            start_row = 1
        else:
            def write(sh, row, col, value):
                sh.write(row, col, value)

            start_col = 0
            start_row = 0
        try:
            sh = wb.active
            sh.title = title if title else f'{model_class._meta.table_description}'

            for col, header in enumerate(headers, start_col):
                write(sh, start_row, col, header)

        finally:
            bytes_io = io.BytesIO()
            wb.save(bytes_io)
            bytes_io.seek(0)

        return ResponseMixin.stream(bytes_io, is_xlsx=True)
