from enum import Enum
import importlib
from importlib import resources
import json
from pathlib import Path
import re

import openpyxl

from nemo_library_etl.adapter.migman.config_models_migman import ConfigMigMan
from nemo_library_etl.adapter.migman.model.migman import MigMan
import pandas as pd

from nemo_library_etl.adapter.migman.symbols import MIGMAN_PROJECT_POSTFIX_SEPARATOR


class MigManUtils:
    @staticmethod
    def MigManDatabaseLoad() -> list[MigMan]:
        """
        Load the MigMan database from a JSON file.

        Returns:
            list[MigMan]: A list of MigMan objects loaded from the database.
        """
        with importlib.resources.open_text(
            "nemo_library.templates", "migmantemplates.json"
        ) as file:
            data = json.load(file)

        return [MigMan(**element) for element in data]

    @staticmethod
    def validate_columns(
        project: str, postfix: str, columns: list[str], missing_ok: bool = False
    ) -> bool:
        """
        Validate if the given columns match the expected columns for a specific project and postfix.
        Raises a detailed error showing which columns are missing or unexpected.

        Args:
            project (str): The project name.
            postfix (str): The postfix for the table.
            columns (list[str]): The list of columns to validate.

        Returns:
            bool: True if the columns are valid, False otherwise.
        """
        db = MigManUtils.MigManDatabaseLoad()
        expected_columns = [
            col.nemo_import_name
            for col in db
            if col.project == project and col.postfix == postfix
        ]

        expected_set = set(expected_columns)
        given_set = set(columns)

        missing = sorted(expected_set - given_set)
        additional = sorted(given_set - expected_set)

        if (missing and not missing_ok) or additional:
            error_message = [
                f"Column mismatch for project '{project}' and postfix '{postfix}':",
                f"- Expected ({len(expected_columns)}): {expected_columns}",
                f"- Found ({len(columns)}):    {columns}",
            ]
            if missing:
                error_message.append(f"- Missing columns ({len(missing)}): {missing}")
            if additional:
                error_message.append(
                    f"- Unexpected columns ({len(additional)}): {additional}"
                )

            raise ValueError("\n".join(error_message))
        return True

    @staticmethod
    def get_migman_projects(cfg: ConfigMigMan) -> list[str]:
        if cfg.setup.project_status_file:
            return MigManUtils.get_migman_projects_from_excel(
                Path(cfg.setup.project_status_file)
            )
        return cfg.setup.projects

    @staticmethod
    def split_migman_project_name(project_name: str) -> tuple[str, str]:

        # if there is no "_" in the project name, return the whole name as project and empty string as postfix
        if MIGMAN_PROJECT_POSTFIX_SEPARATOR not in project_name:
            return project_name, ""

        # otherwise, split at the first "-" and return the parts
        parts = project_name.split(MIGMAN_PROJECT_POSTFIX_SEPARATOR, 1)
        return parts[0].strip(), parts[1].strip() if len(parts) > 1 else ""

    @staticmethod
    def get_migman_projects_from_excel(file: Path) -> list[str]:
        path = Path(file)
        if not path.exists():
            raise FileNotFoundError(f"The file {path} does not exist.")
        workbook = openpyxl.load_workbook(path)
        worksheet = workbook["Status Datenübernahme"]

        data = []
        for row in worksheet.iter_rows(
            min_row=10, max_row=300, min_col=1, max_col=10, values_only=True
        ):
            data.append(row)

        # Create a DataFrame from the extracted data
        columns = [
            worksheet.cell(row=9, column=i).value for i in range(1, 11)
        ]  # Headers in row 9
        dataframe = pd.DataFrame(data, columns=columns)

        # Drop rows where "Importreihenfolge" is NaN or empty
        if "Importreihenfolge" in dataframe.columns:
            dataframe = dataframe.dropna(subset=["Importreihenfolge"])
        else:
            raise ValueError(
                "The column 'Importreihenfolge' does not exist in the data."
            )

        if "Name des Importprograms / Name der Erfassungsmaske" in dataframe.columns:
            nemosteps = dataframe[dataframe["Migrationsart"] == "NEMO"][
                "Name des Importprograms / Name der Erfassungsmaske"
            ].to_list()

            nemosteps = [x.title().strip() for x in nemosteps]
            replacements = {
                "European Article Numbers": "Global Trade Item Numbers",
                "Part-Storage Areas Relationship": "Part-Storage Areas Relationships",
                "Sales Tax Id": "Sales Tax ID",
                "Mrp Parameters": "MRP Parameters",
                "Sales Units Of Measure": "Sales Units of Measure",
                "Standard Boms (Header Data)": "Standard BOMs (Header Data)",
                "Standard Boms (Line Data)": "Standard BOMs (Line Data)",
                "Routings (Standard Boms)": "Routings (Standard BOMs)",
                "Bills Of Materials For Operations (Routings Production)": "Bills of Materials for Operations (Routings Production)",
            }

            nemosteps = [
                replacements[item] if item in replacements else item
                for item in nemosteps
            ]

            return nemosteps
        else:
            raise ValueError(
                "The column 'Name des Importprograms / Name der Erfassungsmaske' does not exist in the data."
            )

    @staticmethod
    def MigManProjectName(project: str, addon: str | None, postfix: str) -> str:
        return f"{project}{" " + addon if addon else ""}{(" (" + postfix + ")") if postfix else ""}"

    @staticmethod
    def getJoinFileName(project: str) -> str:
        return MigManUtils.slugify_filename("join_" + project) + ".sql"

    @staticmethod
    def getJoinQuery(adapter: str, project: str) -> str:
        query = None
        indiv_file = Path(f"./config/{adapter}/joins/") / MigManUtils.getJoinFileName(
            project
        )
        if indiv_file.exists():
            with indiv_file.open("r", encoding="utf-8") as f:
                query = f.read()
        else:
            file = (
                resources.files("nemo_library_etl")
                / "adapter"
                / "migman"
                / "config"
                / "joins"
                / f"{adapter}"
                / MigManUtils.getJoinFileName(project)
            )

            with resources.as_file(file) as sql_file:
                query = sql_file.read_text(encoding="utf-8")

        if not query:
            raise ValueError(
                f"Join query file {MigManUtils.getJoinFileName(project)} for adapter {adapter} not found."
            )

        return query

    @staticmethod
    def slugify_filename(name: str | Enum) -> str:
        """Make a safe lowercase file stem from a human name."""

        if isinstance(name, Enum):
            name = name.value
        s = str(name)
        s = s.strip().lower()
        s = re.sub(r"\s+", "_", s)
        s = re.sub(r"[^a-z0-9_]", "_", s)
        s = re.sub(r"_+", "_", s)
        return s.strip("_") or "table"
