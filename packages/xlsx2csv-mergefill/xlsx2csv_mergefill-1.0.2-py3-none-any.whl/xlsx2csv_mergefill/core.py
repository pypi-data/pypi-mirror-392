from __future__ import annotations

import csv
import io
from pathlib import Path
from typing import Optional, Iterable, List, Tuple, Dict

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook

# 型エイリアス
CellCoord = Tuple[int, int]


def _load_workbook(input_xlsx: Path | str) -> Workbook:
    """Excelワークブックを読み込む共通関数"""
    input_path = Path(input_xlsx)
    if not input_path.exists():
        raise FileNotFoundError(f"入力ファイルが見つかりません: {input_path}")
    
    return load_workbook(filename=str(input_path), data_only=True, read_only=False)


def _sanitize_filename(name: str) -> str:
    """
    ファイル名として使用できない文字を置換する
    """
    import re
    sanitized = re.sub(r'[<>:"/\\|?*]', '_', name)
    sanitized = re.sub(r'_+', '_', sanitized)
    sanitized = sanitized.strip(' .')
    return sanitized if sanitized else "Sheet"


def _build_merged_value_map(ws: Worksheet) -> Dict[CellCoord, Optional[object]]:
    """マージセルの値マップを構築"""
    merged_map: Dict[CellCoord, Optional[object]] = {}
    for mr in ws.merged_cells.ranges:
        min_row, min_col, max_row, max_col = mr.min_row, mr.min_col, mr.max_row, mr.max_col
        top_left_val = ws.cell(min_row, min_col).value
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                merged_map[(r, c)] = top_left_val
    return merged_map


def _iter_rows_values(ws: Worksheet) -> Iterable[List[Optional[object]]]:
    """ワークシートからマージセル展開済みの行データを取得"""
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    merged_map = _build_merged_value_map(ws)

    for r in range(1, max_row + 1):
        row_vals: List[Optional[object]] = []
        for c in range(1, max_col + 1):
            val = ws.cell(r, c).value
            if (r, c) in merged_map:
                val = merged_map[(r, c)]
            row_vals.append(val)
        yield row_vals


def convert_file(input_xlsx: Path | str, output_csv: Path | str) -> int:
    """
    ExcelファイルをCSVファイルに変換する（シンプル版）
    """
    wb = _load_workbook(input_xlsx)
    sheets = wb.worksheets
    
    if not sheets:
        raise ValueError("ワークシートが見つかりません")
    
    output_path = Path(output_csv)
    base_output = output_path.with_suffix("")
    
    if len(sheets) == 1:
        _write_csv(_iter_rows_values(sheets[0]), output_path)
    else:
        for ws in sheets:
            safe_sheet_name = _sanitize_filename(ws.title)
            target_path = base_output.parent / f"{base_output.name}_{safe_sheet_name}.csv"
            _write_csv(_iter_rows_values(ws), target_path)
    
    return 0


# Backward-compatible alias (deprecated)
def excel_to_csv(input_xlsx: Path | str, output_csv: Path | str) -> int:
    """Deprecated alias for convert_file. Will be removed in a future release."""
    return convert_file(input_xlsx, output_csv)


def read_sheet(input_xlsx: Path | str) -> List[List[Optional[object]]]:
    """Excelファイルからデータを読み込んでリスト形式で返す（シンプル版）"""
    wb = _load_workbook(input_xlsx)
    
    ws = wb.active
    if ws is None:
        sheets = wb.worksheets
        if not sheets:
            raise ValueError("ワークシートが見つかりません")
        ws = sheets[0]
    
    return list(_iter_rows_values(ws))


# Backward-compatible alias (deprecated)
def load_excel_data(input_xlsx: Path | str) -> List[List[Optional[object]]]:
    """Deprecated alias for read_sheet. Will be removed in a future release."""
    return read_sheet(input_xlsx)


def read_workbook(input_xlsx: Path | str) -> Dict[str, List[List[Optional[object]]]]:
    """Excelファイルの全シートからデータを読み込んで辞書形式で返す（シンプル版）"""
    wb = _load_workbook(input_xlsx)
    
    return {ws.title: list(_iter_rows_values(ws)) for ws in wb.worksheets}


# Backward-compatible alias (deprecated)
def load_all_sheets_data(input_xlsx: Path | str) -> Dict[str, List[List[Optional[object]]]]:
    """Deprecated alias for read_workbook. Will be removed in a future release."""
    return read_workbook(input_xlsx)


def list_sheets(input_xlsx: Path | str) -> List[str]:
    """Excelファイルのシート名一覧を取得"""
    wb = _load_workbook(input_xlsx)
    return wb.sheetnames


# Backward-compatible alias (deprecated)
def get_sheet_names(input_xlsx: Path | str) -> List[str]:
    """Deprecated alias for list_sheets. Will be removed in a future release."""
    return list_sheets(input_xlsx)


def to_csv_string(data: List[List[Optional[object]]]) -> str:
    """データをCSV文字列に変換（シンプル版）"""
    output = io.StringIO()
    writer = csv.writer(output, delimiter=",", quoting=csv.QUOTE_MINIMAL)
    for row in data:
        writer.writerow(["" if v is None else str(v) for v in row])
    return output.getvalue()


# Backward-compatible alias (deprecated)
def data_to_csv_string(data: List[List[Optional[object]]]) -> str:
    """Deprecated alias for to_csv_string. Will be removed in a future release."""
    return to_csv_string(data)


def _write_csv(rows: Iterable[List[Optional[object]]], out_path: Path) -> None:
    """内部用CSV書き込み関数（UTF-8/カンマ区切り固定）"""
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", newline="", encoding="UTF-8") as f:
        writer = csv.writer(f, delimiter=",", quoting=csv.QUOTE_MINIMAL)
        for row in rows:
            writer.writerow(["" if v is None else str(v) for v in row])
