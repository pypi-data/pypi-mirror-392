import io
from collections.abc import Callable
from pathlib import Path

from aixtools.agents.context.data_models import FileExtractionResult, FileType, TruncationInfo
from aixtools.agents.context.processors.common import (
    check_and_apply_output_limit,
    create_error_result,
    create_file_metadata,
    format_section_header,
    truncate_string,
)
from aixtools.utils import config


def _read_sheet_data(sheet, max_rows_head: int, max_rows_tail: int, max_columns: int) -> tuple[list[list], int, int]:
    """Read sheet data with row and column limits."""
    all_rows = list(sheet.iter_rows(values_only=True))
    if not all_rows:
        return [], 0, 0

    total_rows = len(all_rows) - 1  # Exclude header
    header = all_rows[0] if all_rows else []
    data_rows = all_rows[1:] if len(all_rows) > 1 else []

    # Determine columns to show
    total_columns = len(header)
    columns_to_show = min(total_columns, max_columns)

    # Truncate header
    truncated_header = list(header[:columns_to_show])

    # Select rows (head + tail pattern)
    if total_rows <= (max_rows_head + max_rows_tail):
        selected_rows = data_rows
    else:
        head_rows = data_rows[:max_rows_head]
        tail_rows = data_rows[-max_rows_tail:] if max_rows_tail > 0 else []
        selected_rows = head_rows + tail_rows

    # Truncate rows to column limit
    result_rows = [truncated_header]
    for row in selected_rows:
        truncated_row = list(row[:columns_to_show])
        result_rows.append(truncated_row)

    return result_rows, total_columns, total_rows


def _format_sheet_as_csv(rows: list[list], max_cell_length: int, max_line_length: int) -> tuple[str, int]:
    """Format sheet rows as CSV with cell and line truncation."""
    output = io.StringIO()
    cells_truncated = 0

    for row in rows:
        # Truncate each cell
        truncated_cells = []
        for cell in row:
            truncated_cell, was_truncated = truncate_string(cell, max_cell_length)
            truncated_cells.append(truncated_cell)
            if was_truncated:
                cells_truncated += 1

        # Create CSV line
        line = ",".join(f'"{cell}"' if "," in cell or '"' in cell else cell for cell in truncated_cells)

        # Truncate line if needed
        if len(line) > max_line_length:
            line = line[:max_line_length] + "..."

        output.write(line + "\n")

    return output.getvalue(), cells_truncated


def _process_excel_file(
    file_path: Path,
    max_sheets: int,
    max_rows_head: int,
    max_rows_tail: int,
    max_columns: int,
    max_cell_length: int,
    max_line_length: int,
    max_total_output: int,
) -> tuple[str, TruncationInfo]:
    """Process Excel file (.xlsx) using openpyxl."""
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required for Excel file processing. Install with: uv add openpyxl")

    workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheet_names = workbook.sheetnames
    total_sheets = len(sheet_names)
    sheets_to_process = min(total_sheets, max_sheets)

    output = io.StringIO()
    output.write(f"Spreadsheet: {file_path.name}\n")
    output.write(f"Sheets: {sheets_to_process} of {total_sheets}\n\n")

    total_cells_truncated = 0
    total_columns_truncated = 0
    total_rows_shown = 0
    total_rows_available = 0

    for i in range(sheets_to_process):
        sheet_name = sheet_names[i]
        sheet = workbook[sheet_name]

        # Add sheet header
        output.write(format_section_header(f"{sheet_name}", i, total_sheets))

        # Read and truncate sheet data
        rows, total_columns, total_rows = _read_sheet_data(sheet, max_rows_head, max_rows_tail, max_columns)

        if not rows:
            output.write("(empty sheet)\n")
            continue

        # Format as CSV
        csv_content, cells_truncated = _format_sheet_as_csv(rows, max_cell_length, max_line_length)
        total_cells_truncated += cells_truncated

        # Write metadata
        columns_shown = min(total_columns, max_columns)
        rows_shown = min(len(rows) - 1, total_rows)  # Exclude header from count
        output.write(f"Columns: {columns_shown} (of {total_columns} total)\n")
        output.write(f"Rows: {rows_shown} (of {total_rows} total)\n\n")

        # Write data
        output.write(csv_content)

        total_columns_truncated += (total_columns - columns_shown) if total_columns > columns_shown else 0
        total_rows_shown += rows_shown
        total_rows_available += total_rows

        # Check output limit
        if output.tell() > max_total_output:
            output.write("\n...\n")
            break

    workbook.close()

    # Build truncation info
    truncation_info = TruncationInfo(
        cells_truncated=total_cells_truncated, total_output_limit_reached=output.tell() > max_total_output
    )

    if sheets_to_process < total_sheets:
        truncation_info.rows_shown = f"{sheets_to_process} sheets of {total_sheets} total"

    return output.getvalue(), truncation_info


def _process_xls_file(
    file_path: Path,
    max_sheets: int,
    max_rows_head: int,
    max_rows_tail: int,
    max_columns: int,
    max_cell_length: int,
    max_line_length: int,
    max_total_output: int,
) -> tuple[str, TruncationInfo]:
    """Process legacy Excel file (.xls) using xlrd."""
    try:
        import xlrd
    except ImportError:
        raise ImportError("xlrd is required for .xls file processing. Install with: uv add xlrd")

    workbook = xlrd.open_workbook(file_path)
    sheet_names = workbook.sheet_names()
    total_sheets = len(sheet_names)
    sheets_to_process = min(total_sheets, max_sheets)

    output = io.StringIO()
    output.write(f"Spreadsheet: {file_path.name}\n")
    output.write(f"Sheets: {sheets_to_process} of {total_sheets}\n\n")

    total_cells_truncated = 0
    total_columns_truncated = 0
    total_rows_shown = 0
    total_rows_available = 0

    for i in range(sheets_to_process):
        sheet_name = sheet_names[i]
        sheet = workbook.sheet_by_name(sheet_name)

        # Add sheet header
        output.write(format_section_header(f"{sheet_name}", i, total_sheets))

        if sheet.nrows == 0:
            output.write("(empty sheet)\n")
            continue

        # Read rows
        total_rows = sheet.nrows - 1  # Exclude header
        total_columns = sheet.ncols
        columns_to_show = min(total_columns, max_columns)

        # Build header
        header = [str(sheet.cell_value(0, col)) for col in range(columns_to_show)]

        # Select rows (head + tail pattern)
        if total_rows <= (max_rows_head + max_rows_tail):
            row_indices = list(range(1, sheet.nrows))
        else:
            head_indices = list(range(1, max_rows_head + 1))
            tail_indices = list(range(sheet.nrows - max_rows_tail, sheet.nrows)) if max_rows_tail > 0 else []
            row_indices = head_indices + tail_indices

        # Build rows
        rows = [header]
        for row_idx in row_indices:
            row = [str(sheet.cell_value(row_idx, col)) for col in range(columns_to_show)]
            rows.append(row)

        # Format as CSV
        csv_content, cells_truncated = _format_sheet_as_csv(rows, max_cell_length, max_line_length)
        total_cells_truncated += cells_truncated

        # Write metadata
        rows_shown = len(row_indices)
        output.write(f"Columns: {columns_to_show} (of {total_columns} total)\n")
        output.write(f"Rows: {rows_shown} (of {total_rows} total)\n\n")

        # Write data
        output.write(csv_content)

        total_columns_truncated += (total_columns - columns_to_show) if total_columns > columns_to_show else 0
        total_rows_shown += rows_shown
        total_rows_available += total_rows

        # Check output limit
        if output.tell() > max_total_output:
            output.write("\n...\n")
            break

    # Build truncation info
    truncation_info = TruncationInfo(
        cells_truncated=total_cells_truncated, total_output_limit_reached=output.tell() > max_total_output
    )

    if sheets_to_process < total_sheets:
        truncation_info.rows_shown = f"{sheets_to_process} sheets of {total_sheets} total"

    return output.getvalue(), truncation_info


def _process_ods_file(
    file_path: Path,
    max_sheets: int,
    max_rows_head: int,
    max_rows_tail: int,
    max_columns: int,
    max_cell_length: int,
    max_line_length: int,
    max_total_output: int,
) -> tuple[str, TruncationInfo]:
    """Process OpenDocument Spreadsheet (.ods) using odfpy or pandas."""
    try:
        import pandas as pd
    except ImportError:
        raise ImportError("pandas is required for ODS file processing. Install with: uv add pandas odfpy")

    # Read all sheets
    all_sheets = pd.read_excel(file_path, sheet_name=None, engine="odf")
    sheet_names = list(all_sheets.keys())
    total_sheets = len(sheet_names)
    sheets_to_process = min(total_sheets, max_sheets)

    output = io.StringIO()
    output.write(f"Spreadsheet: {file_path.name}\n")
    output.write(f"Sheets: {sheets_to_process} of {total_sheets}\n\n")

    total_cells_truncated = 0

    for i, sheet_name in enumerate(sheet_names[:sheets_to_process]):
        df = all_sheets[sheet_name]

        # Add sheet header
        output.write(format_section_header(f"{sheet_name}", i, total_sheets))

        if df.empty:
            output.write("(empty sheet)\n")
            continue

        # Truncate columns
        total_columns = len(df.columns)
        columns_to_show = min(total_columns, max_columns)
        df_truncated = df.iloc[:, :columns_to_show]

        # Truncate rows (head + tail)
        total_rows = len(df)
        if total_rows <= (max_rows_head + max_rows_tail):
            df_selected = df_truncated
        else:
            df_head = df_truncated.head(max_rows_head)
            df_tail = df_truncated.tail(max_rows_tail) if max_rows_tail > 0 else pd.DataFrame()
            df_selected = pd.concat([df_head, df_tail])

        rows_shown = len(df_selected)

        # Write metadata
        output.write(f"Columns: {columns_to_show} (of {total_columns} total)\n")
        output.write(f"Rows: {rows_shown} (of {total_rows} total)\n\n")

        # Convert to list of lists for CSV formatting
        header = list(df_selected.columns)
        data_rows = df_selected.values.tolist()
        all_rows = [header] + data_rows

        # Format as CSV with truncation
        csv_content, cells_truncated = _format_sheet_as_csv(all_rows, max_cell_length, max_line_length)
        total_cells_truncated += cells_truncated

        output.write(csv_content)

        # Check output limit
        if output.tell() > max_total_output:
            output.write("\n...\n")
            break

    # Build truncation info
    truncation_info = TruncationInfo(
        cells_truncated=total_cells_truncated, total_output_limit_reached=output.tell() > max_total_output
    )

    if sheets_to_process < total_sheets:
        truncation_info.rows_shown = f"{sheets_to_process} sheets of {total_sheets} total"

    return output.getvalue(), truncation_info


def process_spreadsheet(
    file_path: Path,
    max_sheets: int = 3,
    max_rows_head: int = config.DEFAULT_ROWS_HEAD,
    max_rows_tail: int = config.DEFAULT_ROWS_TAIL,
    max_columns: int = config.MAX_COLUMNS,
    max_cell_length: int = config.MAX_CELL_LENGTH,
    max_line_length: int = config.MAX_LINE_LENGTH,
    max_total_output: int = config.MAX_TOTAL_OUTPUT,
    tokenizer: Callable | None = None,
) -> FileExtractionResult:
    """Process spreadsheet files (.xlsx, .xls, .ods).

    Args:
        file_path: Path to spreadsheet file
        max_sheets: Maximum number of sheets to process
        max_rows_head: Maximum rows from start of each sheet
        max_rows_tail: Maximum rows from end of each sheet
        max_columns: Maximum columns per sheet
        max_cell_length: Maximum characters per cell
        max_line_length: Maximum characters per line
        max_total_output: Maximum total output length
        tokenizer: Optional tokenizer function
    """
    try:
        file_type = FileType.XLSX
        metadata = create_file_metadata(file_path, mime_type=f"spreadsheet/{file_path.suffix[1:]}")

        # Process based on file type
        suffix = file_path.suffix.lower()
        if suffix == ".xlsx":
            content, truncation_info = _process_excel_file(
                file_path,
                max_sheets,
                max_rows_head,
                max_rows_tail,
                max_columns,
                max_cell_length,
                max_line_length,
                max_total_output,
            )
        elif suffix == ".xls":
            content, truncation_info = _process_xls_file(
                file_path,
                max_sheets,
                max_rows_head,
                max_rows_tail,
                max_columns,
                max_cell_length,
                max_line_length,
                max_total_output,
            )
        elif suffix == ".ods":
            content, truncation_info = _process_ods_file(
                file_path,
                max_sheets,
                max_rows_head,
                max_rows_tail,
                max_columns,
                max_cell_length,
                max_line_length,
                max_total_output,
            )
        else:
            return FileExtractionResult(
                content=None,
                success=False,
                error_message=f"Unsupported spreadsheet format: {suffix}",
                file_type=file_type,
                metadata=metadata,
            )

        # Apply tokenizer if provided
        if tokenizer:
            truncation_info.tokens_shown = tokenizer(content)

        # Final length check
        content = check_and_apply_output_limit(content, max_total_output, truncation_info)

        return FileExtractionResult(
            content=content,
            success=True,
            was_extracted=True,
            file_type=file_type,
            truncation_info=truncation_info,
            metadata=metadata,
        )

    except ImportError as e:
        return create_error_result(e, FileType.XLSX, file_path, "spreadsheet (missing dependencies)")
    except Exception as e:
        return create_error_result(e, FileType.XLSX, file_path, "spreadsheet")
