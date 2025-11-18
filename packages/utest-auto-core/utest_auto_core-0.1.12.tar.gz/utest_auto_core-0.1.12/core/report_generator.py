"""
简化版报告生成模块

提供Excel和JSON格式的测试报告生成功能
包含测试结果、步骤详情、截图信息和全局监控结果
"""

import os
import json
import logging
from datetime import datetime
from typing import Dict, List, Any, Optional, Union
from pathlib import Path
from dataclasses import asdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .test_case import TestResult, StepResult, TestStatus, StepStatus
from .config_manager import ReportConfig


logger = logging.getLogger(__name__)

# 简化版报告生成器，不再需要复杂的字体设置


class ReportGenerator:
    """简化版报告生成器"""
    
    def __init__(self, config: ReportConfig):
        """
        初始化报告生成器
        
        Args:
            config: 报告配置
        """
        self.config = config
        self.output_dir = Path(config.output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # 简化版报告生成器，不需要创建子目录
    
    
    def generate_report(self, test_results: List[TestResult], 
                       device_info: Dict[str, Any] = None,
                       global_monitor_result: Dict[str, Any] = None,
                       exit_code: int = None) -> Union[str, List[str]]:
        """
        生成测试报告
        
        Args:
            test_results: 测试结果列表
            device_info: 设备信息
            global_monitor_result: 全局监控结果（ANR/Crash监控等）
            
        Returns:
            Union[str, List[str]]: 报告文件路径或路径列表
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # 不再生成图表，简化报告
        
        # 根据配置生成相应格式的报告
        if self.config.report_format == "excel":
            return self._generate_excel_report(test_results, device_info, timestamp, global_monitor_result, exit_code)
        elif self.config.report_format == "json":
            return self._generate_json_report(test_results, device_info, timestamp, global_monitor_result, exit_code)
        elif self.config.report_format == "all":
            # 生成所有格式的报告
            reports = []
            reports.append(self._generate_excel_report(test_results, device_info, timestamp, global_monitor_result, exit_code))
            reports.append(self._generate_json_report(test_results, device_info, timestamp, global_monitor_result, exit_code))
            return reports
        else:
            raise ValueError(f"不支持的报告格式: {self.config.report_format}，支持的格式: excel, json, all")
    
    
    
    
    
    def _generate_excel_report(self, test_results: List[TestResult], 
                              device_info: Dict[str, Any],
                              timestamp: str,
                              global_monitor_result: Dict[str, Any] = None,
                              exit_code: int = None) -> str:
        """
        生成增强版Excel报告
        
        Args:
            test_results: 测试结果列表
            device_info: 设备信息
            timestamp: 时间戳
            global_monitor_result: 全局监控结果
            
        Returns:
            str: 报告文件路径
        """
        report_file = self.output_dir / f"test_report_{timestamp}.xlsx"
        
        wb = Workbook()
        
        # 删除默认工作表
        wb.remove(wb.active)
        
        # 创建测试概览工作表
        self._create_summary_sheet(wb, test_results, device_info, timestamp, global_monitor_result, exit_code)
        
        # 创建测试汇总工作表
        self._create_test_summary_sheet(wb, test_results)
        
        # 创建测试详情工作表（包含步骤详情和截图）
        self._create_test_details_sheet(wb, test_results)
        
        # 创建错误分析工作表
        self._create_error_analysis_sheet(wb, test_results)
        
        # 创建性能指标汇总工作表
        self._create_performance_summary_sheet(wb, test_results)
        
        # 保存工作簿
        wb.save(report_file)
        logger.info(f"增强版Excel报告已生成: {report_file}")
        
        return str(report_file)
    
    
    
    
    def _create_summary_sheet(self, wb: Workbook, test_results: List[TestResult], 
                             device_info: Dict[str, Any], timestamp: str, 
                             global_monitor_result: Dict[str, Any] = None,
                             exit_code: int = None) -> None:
        """创建测试概览工作表"""
        ws = wb.create_sheet("测试概览", 0)
        
        # 设置标题样式
        title_font = Font(name="微软雅黑", size=16, bold=True, color="FFFFFF")
        title_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title_alignment = Alignment(horizontal="center", vertical="center")
        
        # 设置表头样式
        header_font = Font(name="微软雅黑", size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 设置数据样式
        data_font = Font(name="微软雅黑", size=11)
        data_alignment = Alignment(horizontal="left", vertical="center")
        
        # 设置边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 标题
        ws.merge_cells('A1:H1')
        ws['A1'] = f"自动化测试报告 - {timestamp}"
        ws['A1'].font = title_font
        ws['A1'].fill = title_fill
        ws['A1'].alignment = title_alignment
        
        # 设备信息
        row = 3
        ws[f'A{row}'] = "设备信息"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws[f'A{row}'].alignment = header_alignment
        ws.merge_cells(f'A{row}:H{row}')
        
        row += 1
        # 初始化设备信息数据列表
        device_info_data = []
        
        if device_info:
            device_info_data = [
                ("设备UDID", device_info.get('udid', 'N/A')),
                ("平台类型", device_info.get('platform', 'N/A')),
                ("设备型号", device_info.get('model', 'N/A')),
                ("制造商", device_info.get('manufacturer', 'N/A')),
                ("系统版本", device_info.get('os_version', 'N/A')),
                ("屏幕尺寸", f"{device_info.get('screen_width', 0)}x{device_info.get('screen_height', 0)}"),
                ("测试时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
                ("报告生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            ]
            
            for i, (key, value) in enumerate(device_info_data):
                ws[f'A{row + i}'] = key
                ws[f'B{row + i}'] = str(value)
                ws[f'A{row + i}'].font = data_font
                ws[f'B{row + i}'].font = data_font
                ws[f'A{row + i}'].alignment = data_alignment
                ws[f'B{row + i}'].alignment = data_alignment
                ws[f'A{row + i}'].border = thin_border
                ws[f'B{row + i}'].border = thin_border
        else:
            # 当没有设备信息时，显示默认信息
            device_info_data = [
                ("设备信息", "未提供设备信息"),
                ("测试时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
                ("报告生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            ]
            
            for i, (key, value) in enumerate(device_info_data):
                ws[f'A{row + i}'] = key
                ws[f'B{row + i}'] = str(value)
                ws[f'A{row + i}'].font = data_font
                ws[f'B{row + i}'].font = data_font
                ws[f'A{row + i}'].alignment = data_alignment
                ws[f'B{row + i}'].alignment = data_alignment
                ws[f'A{row + i}'].border = thin_border
                ws[f'B{row + i}'].border = thin_border
        
        # 全局监控结果
        if global_monitor_result:
            row += len(device_info_data) + 2
            ws[f'A{row}'] = "全局监控结果"
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            ws[f'A{row}'].alignment = header_alignment
            ws.merge_cells(f'A{row}:H{row}')
            
            row += 1
            monitor_data = [
                ("监控状态", "成功" if global_monitor_result.get('success', False) else "失败"),
                ("运行时间", f"{global_monitor_result.get('run_time', 0):.2f}秒"),
                ("Crash次数", global_monitor_result.get('crash_count', 0)),
                ("ANR次数", global_monitor_result.get('anr_count', 0)),
            ]
            
            for i, (key, value) in enumerate(monitor_data):
                ws[f'A{row + i}'] = key
                ws[f'B{row + i}'] = str(value)
                ws[f'A{row + i}'].font = data_font
                ws[f'B{row + i}'].font = data_font
                ws[f'A{row + i}'].alignment = data_alignment
                ws[f'B{row + i}'].alignment = data_alignment
                ws[f'A{row + i}'].border = thin_border
                ws[f'B{row + i}'].border = thin_border
            
            row += len(monitor_data) + 2
        
        # 测试统计
        ws[f'A{row}'] = "测试统计"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws[f'A{row}'].alignment = header_alignment
        ws.merge_cells(f'A{row}:H{row}')
        
        # 统计测试结果
        total_tests = len(test_results)
        passed_tests = sum(1 for r in test_results if r.status == TestStatus.PASSED)
        failed_tests = sum(1 for r in test_results if r.status == TestStatus.FAILED)
        error_tests = sum(1 for r in test_results if r.status == TestStatus.ERROR)
        
        total_duration = sum(r.duration or 0 for r in test_results)
        
        row += 1
        stats_data = [
            ("总测试数", total_tests),
            ("通过", passed_tests),
            ("失败", failed_tests),
            ("错误", error_tests),
            ("通过率", f"{(passed_tests/total_tests*100):.1f}%" if total_tests > 0 else "0%"),
            ("总耗时", f"{total_duration:.2f}秒"),
            ("退出码", exit_code if exit_code is not None else "未知")
        ]
        
        for i, (key, value) in enumerate(stats_data):
            ws[f'A{row + i}'] = key
            ws[f'B{row + i}'] = str(value)
            ws[f'A{row + i}'].font = data_font
            ws[f'B{row + i}'].font = data_font
            ws[f'A{row + i}'].alignment = data_alignment
            ws[f'B{row + i}'].alignment = data_alignment
            ws[f'A{row + i}'].border = thin_border
            ws[f'B{row + i}'].border = thin_border
        
        # 调整列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 25
        
        # 设置行高
        for i in range(1, row + len(stats_data) + 1):
            ws.row_dimensions[i].height = 25

    def _create_performance_summary_sheet(self, wb: Workbook, test_results: List[TestResult]) -> None:
        """创建性能指标汇总工作表"""
        ws = wb.create_sheet("性能指标汇总")
        
        # 设置表头样式
        header_font = Font(name="微软雅黑", size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 设置数据样式
        data_font = Font(name="微软雅黑", size=11)
        data_alignment = Alignment(horizontal="center", vertical="center")
        
        # 设置边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 标题
        ws.merge_cells('A1:P1')
        ws['A1'] = "性能指标汇总（基于ubox统计指标）"
        ws['A1'].font = Font(name="微软雅黑", size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        
        # 表头 - 包含更多ubox提供的指标
        headers = [
            "测试名称", "应用CPU(%)", "总CPU(%)", "内存峰值(MB)", "内存平均(MB)",
            "平均FPS", "最高FPS", "最低FPS", "FPS-P50", "卡顿率(%)",
            "大卡顿", "小卡顿", "GPU使用率(%)", "上传流量(KB)", "下载流量(KB)", "数据来源"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 填充性能数据
        current_row = 3
        for result in test_results:
            # 获取性能数据
            perf_data = result.performance_data or {}
            
            # 获取ubox提供的丰富性能指标
            cpu_app_avg = perf_data.get('cpu_usage_avg', 0.0)
            cpu_total_avg = perf_data.get('cpu_total_avg', 0.0)
            memory_peak = perf_data.get('memory_peak_mb', 0.0)
            memory_avg = perf_data.get('memory_avg_mb', 0.0)
            fps_avg = perf_data.get('fps_avg', 0.0)
            fps_max = perf_data.get('fps_max', 0.0)
            fps_min = perf_data.get('fps_min', 0.0)
            fps_p50 = perf_data.get('fps_p50', 0.0)
            stutter_rate = perf_data.get('stutter_rate_percent', 0.0)
            big_jank = perf_data.get('big_jank_count', 0)
            small_jank = perf_data.get('small_jank_count', 0)
            gpu_avg = perf_data.get('gpu_avg', 0.0)
            net_up = perf_data.get('network_upload_total_kb', 0.0)
            net_down = perf_data.get('network_download_total_kb', 0.0)
            data_source = perf_data.get('data_source', 'unknown')
            
            # 填充数据
            ws.cell(row=current_row, column=1, value=result.test_name).font = data_font
            ws.cell(row=current_row, column=2, value=f"{cpu_app_avg:.2f}").font = data_font
            ws.cell(row=current_row, column=3, value=f"{cpu_total_avg:.2f}").font = data_font
            ws.cell(row=current_row, column=4, value=f"{memory_peak:.2f}").font = data_font
            ws.cell(row=current_row, column=5, value=f"{memory_avg:.2f}").font = data_font
            ws.cell(row=current_row, column=6, value=f"{fps_avg:.2f}").font = data_font
            ws.cell(row=current_row, column=7, value=f"{fps_max:.2f}").font = data_font
            ws.cell(row=current_row, column=8, value=f"{fps_min:.2f}").font = data_font
            ws.cell(row=current_row, column=9, value=f"{fps_p50:.2f}").font = data_font
            ws.cell(row=current_row, column=10, value=f"{stutter_rate:.2f}").font = data_font
            ws.cell(row=current_row, column=11, value=big_jank).font = data_font
            ws.cell(row=current_row, column=12, value=small_jank).font = data_font
            ws.cell(row=current_row, column=13, value=f"{gpu_avg:.2f}").font = data_font
            ws.cell(row=current_row, column=14, value=f"{net_up:.2f}").font = data_font
            ws.cell(row=current_row, column=15, value=f"{net_down:.2f}").font = data_font
            ws.cell(row=current_row, column=16, value=data_source).font = data_font
            
            # 设置所有单元格的边框和对齐
            for col in range(1, 17):
                cell = ws.cell(row=current_row, column=col)
                cell.border = thin_border
                cell.alignment = data_alignment
            
            # 根据性能指标设置颜色（可选）
            self._apply_performance_colors(ws, current_row, cpu_app_avg, fps_avg, stutter_rate)
            
            current_row += 1
        
        # 调整列宽 - 根据表头字数自动调整
        self._auto_adjust_column_widths(ws, headers)
        
        # 设置行高
        for row in range(1, current_row):
            ws.row_dimensions[row].height = 25

    def _auto_adjust_column_widths(self, ws, headers: List[str]) -> None:
        """根据表头内容自动调整列宽"""
        for col, header in enumerate(headers, 1):
            # 特殊列的处理
            if "截图" in header:
                # 截图列需要更宽的空间
                width = 80
            elif "错误信息" in header or "错误详情" in header:
                # 错误信息列需要更宽的空间
                width = 50
            elif "监控数据" in header:
                # 监控数据列需要更宽的空间
                width = 50
            elif "步骤描述" in header:
                # 步骤描述列需要更宽的空间
                width = 35
            elif "设备型号" in header:
                # 设备型号列需要更宽的空间
                width = 20
            else:
                # 普通列：根据字符数计算宽度
                chinese_chars = len([c for c in header if '\u4e00' <= c <= '\u9fff'])
                english_chars = len(header) - chinese_chars
                
                # 计算宽度：中文字符*2.5 + 英文字符*1.2 + 边距4
                width = chinese_chars * 2.5 + english_chars * 1.2 + 4
                
                # 设置最小和最大宽度
                min_width = 8
                max_width = 30
                
                # 限制宽度范围
                width = max(min_width, min(width, max_width))
            
            # 设置列宽
            ws.column_dimensions[get_column_letter(col)].width = width

    def _apply_performance_colors(self, ws, row: int, cpu_usage: float, fps_avg: float, stutter_rate: float) -> None:
        """根据性能指标设置单元格颜色"""
        try:
            # CPU使用率颜色（绿色：<50%, 黄色：50-80%, 红色：>80%）
            cpu_cell = ws.cell(row=row, column=2)
            if cpu_usage > 80:
                cpu_cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            elif cpu_usage > 50:
                cpu_cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            else:
                cpu_cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
            
            # FPS颜色（绿色：>30, 黄色：20-30, 红色：<20）
            fps_cell = ws.cell(row=row, column=4)
            if fps_avg < 20:
                fps_cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            elif fps_avg < 30:
                fps_cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            else:
                fps_cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
            
            # 卡顿率颜色（绿色：<5%, 黄色：5-10%, 红色：>10%）
            stutter_cell = ws.cell(row=row, column=5)
            if stutter_rate > 10:
                stutter_cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            elif stutter_rate > 5:
                stutter_cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            else:
                stutter_cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
                
        except Exception as e:
            logger.warning(f"设置性能指标颜色失败: {e}")


    
    def _create_test_details_sheet(self, wb: Workbook, test_results: List[TestResult]) -> None:
        """创建测试详情工作表（包含步骤详情和截图）"""
        ws = wb.create_sheet("测试详情")
        
        # 设置样式
        header_font = Font(name="微软雅黑", size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        data_font = Font(name="微软雅黑", size=11)
        data_alignment = Alignment(horizontal="left", vertical="center")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 标题
        ws.merge_cells('A1:L1')
        ws['A1'] = "测试详情（包含步骤详情和截图）"
        ws['A1'].font = Font(name="微软雅黑", size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        
        # 表头
        headers = [
            "测试名称", "测试状态", "开始时间", "结束时间", "耗时(秒)",
            "步骤名称", "步骤描述", "步骤状态", "步骤耗时(秒)", "错误信息",
            "截图", "监控数据"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 填充数据
        current_row = 3
        for result in test_results:
            status_color = self._get_status_color(result.status)
            
            if result.steps:
                # 每个步骤一行
                start_row = current_row
                for step_idx, step in enumerate(result.steps):
                    # 测试用例基本信息（只在第一步时填写）
                    if step_idx == 0:
                        ws.cell(row=current_row, column=1, value=result.test_name).font = data_font
                        ws.cell(row=current_row, column=2, value=result.status.value).font = data_font
                        ws.cell(row=current_row, column=3, value=result.start_time.strftime("%Y-%m-%d %H:%M:%S") if result.start_time else "").font = data_font
                        ws.cell(row=current_row, column=4, value=result.end_time.strftime("%Y-%m-%d %H:%M:%S") if result.end_time else "").font = data_font
                        ws.cell(row=current_row, column=5, value=f"{result.duration:.2f}" if result.duration else "").font = data_font
                        
                        # 设置测试状态颜色
                        status_cell = ws.cell(row=current_row, column=2)
                        status_cell.fill = PatternFill(start_color=status_color, end_color=status_color, fill_type="solid")
                        status_cell.font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
                    
                    # 步骤信息
                    ws.cell(row=current_row, column=6, value=step.step_name).font = data_font
                    ws.cell(row=current_row, column=7, value=step.description).font = data_font
                    ws.cell(row=current_row, column=8, value=step.status.value).font = data_font
                    ws.cell(row=current_row, column=9, value=f"{step.duration:.2f}" if step.duration else "").font = data_font
                    ws.cell(row=current_row, column=10, value=step.error_message or "").font = data_font
                    
                    # 设置步骤状态颜色
                    step_status_cell = ws.cell(row=current_row, column=8)
                    step_status_color = self._get_status_color(step.status)
                    step_status_cell.fill = PatternFill(start_color=step_status_color, end_color=step_status_color, fill_type="solid")
                    step_status_cell.font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
                    
                    # 插入步骤截图（直接嵌入图片）
                    if step.screenshots:
                        # 优先使用步骤级截图
                        self._insert_screenshots_to_cell(ws, current_row, 11, step.screenshots)
                    elif result.screenshots:
                        # 步骤无截图时，回退展示用例级截图，避免报告空白
                        self._insert_screenshots_to_cell(ws, current_row, 11, result.screenshots)
                    else:
                        ws.cell(row=current_row, column=11, value="无截图").font = data_font
                    
                    # 监控数据（只在第一步时显示）
                    if step_idx == 0:
                        # 合并性能监控数据
                        monitor_text = self._format_all_monitor_data(result)
                        ws.cell(row=current_row, column=12, value=monitor_text).font = data_font
                    
                    # 设置所有单元格的边框和对齐
                    for col in range(1, 13):
                        cell = ws.cell(row=current_row, column=col)
                        cell.border = thin_border
                        cell.alignment = data_alignment
                    
                    current_row += 1
                
                # 合并测试用例基本信息的单元格
                if len(result.steps) > 1:
                    ws.merge_cells(f'A{start_row}:A{current_row - 1}')
                    ws.merge_cells(f'B{start_row}:B{current_row - 1}')
                    ws.merge_cells(f'C{start_row}:C{current_row - 1}')
                    ws.merge_cells(f'D{start_row}:D{current_row - 1}')
                    ws.merge_cells(f'E{start_row}:E{current_row - 1}')
                    ws.merge_cells(f'L{start_row}:L{current_row - 1}')  # 监控数据
                    
                    # 设置合并单元格的对齐方式
                    for col in [1, 2, 3, 4, 5, 12]:
                        cell = ws.cell(row=start_row, column=col)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                # 没有步骤的测试用例
                ws.cell(row=current_row, column=1, value=result.test_name).font = data_font
                ws.cell(row=current_row, column=2, value=result.status.value).font = data_font
                ws.cell(row=current_row, column=3, value=result.start_time.strftime("%Y-%m-%d %H:%M:%S") if result.start_time else "").font = data_font
                ws.cell(row=current_row, column=4, value=result.end_time.strftime("%Y-%m-%d %H:%M:%S") if result.end_time else "").font = data_font
                ws.cell(row=current_row, column=5, value=f"{result.duration:.2f}" if result.duration else "").font = data_font
                ws.cell(row=current_row, column=6, value="无步骤").font = data_font
                ws.cell(row=current_row, column=7, value="").font = data_font
                ws.cell(row=current_row, column=8, value="").font = data_font
                ws.cell(row=current_row, column=9, value="").font = data_font
                ws.cell(row=current_row, column=10, value=result.error_message or "").font = data_font
                ws.cell(row=current_row, column=11, value="无截图").font = data_font
                
                # 监控数据
                monitor_text = self._format_all_monitor_data(result)
                ws.cell(row=current_row, column=12, value=monitor_text).font = data_font
                
                # 设置测试状态颜色
                status_cell = ws.cell(row=current_row, column=2)
                status_cell.fill = PatternFill(start_color=status_color, end_color=status_color, fill_type="solid")
                status_cell.font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
                
                # 设置所有单元格的边框和对齐
                for col in range(1, 13):
                    cell = ws.cell(row=current_row, column=col)
                    cell.border = thin_border
                    cell.alignment = data_alignment
                
                current_row += 1
        
        # 调整列宽
        self._auto_adjust_column_widths(ws, headers)
        
        # 设置行高
        for row in range(1, current_row):
            ws.row_dimensions[row].height = 200  # 增加行高以容纳截图
    
    def _get_status_color(self, status: Union[TestStatus, StepStatus]) -> str:
        """获取状态对应的颜色"""
        color_map = {
            TestStatus.PASSED: "00B050",
            TestStatus.FAILED: "FF0000",
            TestStatus.ERROR: "FF6600",
            TestStatus.RUNNING: "0070C0",
            StepStatus.PASSED: "00B050",
            StepStatus.FAILED: "FF0000",
            StepStatus.ERROR: "FF6600",
            StepStatus.SKIPPED: "FFC000",
            StepStatus.RUNNING: "0070C0",
        }
        return color_map.get(status, "808080")
    
    def _generate_json_report(self, test_results: List[TestResult], 
                             device_info: Dict[str, Any],
                             timestamp: str,
                             global_monitor_result: Dict[str, Any] = None,
                             exit_code: int = None) -> str:
        report_file = self.output_dir / f"test_report_{timestamp}.json"
        
        # 转换测试结果为精简字典（不包含重复的性能数据）
        results_data = []
        for result in test_results:
            result_dict = {
                "test_name": result.test_name,
                "description": result.description,
                "status": result.status.value,
                "start_time": result.start_time.isoformat() if result.start_time else None,
                "end_time": result.end_time.isoformat() if result.end_time else None,
                "duration": result.duration,
                "error_message": result.error_message,
                "screenshots": result.screenshots,
                "logs": result.logs,
                # 只保留性能数据的关键信息，避免重复
                "performance_summary": self._extract_performance_summary_from_result(result),
                "logcat_data": result.logcat_data,
                "recording_data": result.recording_data,
                "steps": self._convert_steps_to_dict(result.steps)
            }
            results_data.append(result_dict)
        
        # 构建精简的报告数据
        report_data = {
            "report_info": {
                "generated_at": datetime.now().isoformat(),
                "timestamp": timestamp,
                "format": "json",
                "version": "2.0",
                "exit_code": exit_code
            },
            "summary": {
                "total_tests": len(test_results),
                "passed": len([r for r in test_results if r.status.value == "passed"]),
                "failed": len([r for r in test_results if r.status.value == "failed"]),
                "error": len([r for r in test_results if r.status.value == "error"]),
                "skipped": len([r for r in test_results if r.status.value == "skipped"])
            },
            "device_info": device_info,
            "test_results": results_data,
            "global_monitor_result": global_monitor_result
        }
        
        # 保存JSON文件
        with open(report_file, 'w', encoding='utf-8') as f:
            json.dump(report_data, f, ensure_ascii=False, indent=2)
        
        logger.info(f"精简JSON报告已生成: {report_file}")
        return str(report_file)
    
    def _extract_performance_summary_from_result(self, result: TestResult) -> Dict[str, Any]:
        """从测试结果中提取精简的性能汇总信息"""
        if not result.performance_data:
            return {
                "data_source": "no_data",
                "metrics_count": 0,
                "file_info": {
                    "file_path": None,
                    "file_size": 0,
                    "file_exists": False
                },
                "core_metrics": {
                    "cpu_usage_avg": 0.0,
                    "memory_peak_mb": 0.0,
                    "fps_avg": 0.0,
                    "stutter_rate_percent": 0.0,
                    "network_upload_total_kb": 0.0,
                    "network_download_total_kb": 0.0
                }
            }
        
        perf_data = result.performance_data
        data_source = perf_data.get('data_source', 'unknown')
        
        # 提取文件信息
        file_path = perf_data.get('file_path', None)
        file_size = perf_data.get('file_size', 0)
        file_exists = file_path and os.path.exists(file_path) if file_path else False
        
        file_info = {
            "file_path": file_path,
            "file_size": file_size,
            "file_exists": file_exists
        }
        
        # 只保留核心指标，避免数据冗余
        core_metrics = {
            "cpu_usage_avg": perf_data.get('cpu_usage_avg', 0.0),
            "memory_peak_mb": perf_data.get('memory_peak_mb', 0.0),
            "fps_avg": perf_data.get('fps_avg', 0.0),
            "stutter_rate_percent": perf_data.get('stutter_rate_percent', 0.0),
            "network_upload_total_kb": perf_data.get('network_upload_total_kb', 0.0),
            "network_download_total_kb": perf_data.get('network_download_total_kb', 0.0)
        }
        
        # 如果是ubox数据，添加一些关键详细信息
        if data_source == 'ubox_overview':
            core_metrics.update({
                "cpu_total_avg": perf_data.get('cpu_total_avg', 0.0),
                "fps_max": perf_data.get('fps_max', 0.0),
                "fps_min": perf_data.get('fps_min', 0.0),
                "gpu_avg": perf_data.get('gpu_avg', 0.0),
                "big_jank_count": perf_data.get('big_jank_count', 0),
                "small_jank_count": perf_data.get('small_jank_count', 0)
            })
        
        return {
            "data_source": data_source,
            "metrics_count": perf_data.get('metrics_count', 0),
            "file_info": file_info,
            "core_metrics": core_metrics
        }
    
    def _convert_steps_to_dict(self, steps: List[StepResult]) -> List[Dict[str, Any]]:
        """转换步骤结果为字典格式"""
        steps_data = []
        for step in steps:
            step_dict = {
                "step_name": step.step_name,
                "status": step.status.value,
                "start_time": step.start_time.isoformat() if step.start_time else None,
                "end_time": step.end_time.isoformat() if step.end_time else None,
                "duration": step.duration,
                "error_message": step.error_message,
                "screenshots": step.screenshots,
                "logs": step.logs,
                "description": step.description
            }
            steps_data.append(step_dict)
        return steps_data
    
    def _insert_screenshots_to_cell(self, ws, row: int, col: int, screenshots: List[str]) -> None:
        """将截图插入到Excel单元格中，保持原图比例，优化多图显示"""
        try:
            from openpyxl.drawing import image as xl_image
            from PIL import Image as PILImage
            
            # 过滤存在的截图文件，并记录调试信息
            valid_screenshots = [s for s in screenshots if os.path.exists(s)]
            if not valid_screenshots and screenshots:
                logger.warning(f"截图路径不存在或无权限: {screenshots}")
                logger.warning(f"当前工作目录: {os.getcwd()}")
            
            if not valid_screenshots:
                # 显示原始路径，便于排查
                ws.cell(row=row, column=col, value="\n".join(screenshots) if screenshots else "截图文件不存在")
                return
            
            # 如果只有一个截图，直接插入
            if len(valid_screenshots) == 1:
                screenshot_path = valid_screenshots[0]
                # 获取原图尺寸
                with PILImage.open(screenshot_path) as pil_img:
                    original_width, original_height = pil_img.size
                
                # 计算缩放比例，保持宽高比
                max_width = 300  # 最大宽度
                max_height = 200  # 最大高度
                
                # 计算缩放比例
                width_ratio = max_width / original_width
                height_ratio = max_height / original_height
                scale_ratio = min(width_ratio, height_ratio, 1.0)  # 不放大，只缩小
                
                # 计算新尺寸
                new_width = int(original_width * scale_ratio)
                new_height = int(original_height * scale_ratio)
                
                img = xl_image.Image(screenshot_path)
                img.width = new_width
                img.height = new_height
                img.anchor = f"{get_column_letter(col)}{row}"
                ws.add_image(img)
                
            elif len(valid_screenshots) <= 4:
                # 2-4个截图，创建缩略图网格
                self._insert_multiple_screenshots(ws, row, col, valid_screenshots)
                
            else:
                # 超过4个截图，显示前4个，其余显示提示
                self._insert_multiple_screenshots(ws, row, col, valid_screenshots[:4])
                # 在下方显示剩余截图数量
                remaining_count = len(valid_screenshots) - 4
                ws.cell(row=row + 2, column=col, value=f"... 还有 {remaining_count} 个截图")
                
        except Exception as e:
            logger.warning(f"插入截图失败: {e}")
            # 如果插入失败，显示文件路径
            screenshot_text = "\n".join([f"{i+1}. {screenshot}" for i, screenshot in enumerate(screenshots)])
            ws.cell(row=row, column=col, value=screenshot_text)
    
    def _format_monitor_data(self, monitor_data: Dict[str, Any], monitor_type: str) -> str:
        """格式化监控数据为可读文本"""
        if not monitor_data:
            return f"{monitor_type}: 未启用"
        
        # 构建信息
        info_lines = [f"{monitor_type}: 已启用"]
        
        # 添加文件路径
        if 'file_path' in monitor_data:
            info_lines.append(f"文件路径: {monitor_data['file_path']}")
        elif 'log_file' in monitor_data:
            info_lines.append(f"日志文件: {monitor_data['log_file']}")
        elif 'video_file' in monitor_data:
            info_lines.append(f"视频文件: {monitor_data['video_file']}")
        
        # 添加文件大小（如果存在）
        if 'file_size' in monitor_data:
            size_mb = monitor_data['file_size'] / (1024 * 1024)
            info_lines.append(f"文件大小: {size_mb:.2f}MB")
        
        # 性能监控特殊处理
        if monitor_type == "性能监控":
            if 'app_version' in monitor_data:
                info_lines.append(f"应用版本: {monitor_data['app_version']}")
            if 'device_model' in monitor_data:
                info_lines.append(f"设备型号: {monitor_data['device_model']}")
            if 'data_count' in monitor_data:
                info_lines.append(f"数据点数: {monitor_data['data_count']}")
            
            # 客户需要的核心性能指标
            if 'cpu_usage_avg' in monitor_data:
                info_lines.append(f"CPU使用率: {monitor_data['cpu_usage_avg']:.2f}%")
            if 'memory_peak_mb' in monitor_data:
                info_lines.append(f"内存峰值: {monitor_data['memory_peak_mb']:.2f}MB")
            if 'fps_avg' in monitor_data:
                info_lines.append(f"平均FPS: {monitor_data['fps_avg']:.2f}")
            if 'stutter_rate_percent' in monitor_data:
                info_lines.append(f"卡顿率: {monitor_data['stutter_rate_percent']:.2f}%")
            if 'network_upload_total_kb' in monitor_data:
                info_lines.append(f"上传流量: {monitor_data['network_upload_total_kb']:.2f}KB")
            if 'network_download_total_kb' in monitor_data:
                info_lines.append(f"下载流量: {monitor_data['network_download_total_kb']:.2f}KB")
            
            # 兼容旧的字段名
            if 'avg_fps' in monitor_data and 'fps_avg' not in monitor_data:
                info_lines.append(f"平均FPS: {monitor_data['avg_fps']:.1f}")
            if 'avg_cpu_usage' in monitor_data and 'cpu_usage_avg' not in monitor_data:
                info_lines.append(f"平均CPU: {monitor_data['avg_cpu_usage']:.1f}%")
            if 'avg_memory_usage' in monitor_data and 'memory_peak_mb' not in monitor_data:
                info_lines.append(f"平均内存: {monitor_data['avg_memory_usage']:.1f}MB")
        
        # 添加持续时间
        if 'duration' in monitor_data and monitor_data['duration'] > 0:
            info_lines.append(f"持续时间: {monitor_data['duration']:.2f}秒")
        
        return "\n".join(info_lines)
    
    def _format_all_monitor_data(self, result: TestResult) -> str:
        """格式化所有监控数据为可读文本 - 使用ubox提供的丰富统计指标"""
        monitor_lines = []
        
        # 性能监控数据
        if result.performance_data:
            perf_data = result.performance_data
            data_source = perf_data.get('data_source', 'unknown')
            
            if data_source == 'ubox_overview':
                perf_lines = ["性能监控: ubox统计指标"]
                
                # 客户需要的核心性能指标
                if 'cpu_usage_avg' in perf_data:
                    perf_lines.append(f"应用CPU: {perf_data['cpu_usage_avg']:.2f}%")
                if 'cpu_total_avg' in perf_data:
                    perf_lines.append(f"总CPU: {perf_data['cpu_total_avg']:.2f}%")
                if 'memory_peak_mb' in perf_data:
                    perf_lines.append(f"内存峰值: {perf_data['memory_peak_mb']:.2f}MB")
                if 'memory_avg_mb' in perf_data:
                    perf_lines.append(f"内存平均: {perf_data['memory_avg_mb']:.2f}MB")
                if 'fps_avg' in perf_data:
                    perf_lines.append(f"平均FPS: {perf_data['fps_avg']:.2f}")
                if 'fps_max' in perf_data:
                    perf_lines.append(f"最高FPS: {perf_data['fps_max']:.2f}")
                if 'fps_min' in perf_data:
                    perf_lines.append(f"最低FPS: {perf_data['fps_min']:.2f}")
                if 'fps_p50' in perf_data:
                    perf_lines.append(f"FPS-P50: {perf_data['fps_p50']:.2f}")
                if 'stutter_rate_percent' in perf_data:
                    perf_lines.append(f"卡顿率: {perf_data['stutter_rate_percent']:.2f}%")
                if 'big_jank_count' in perf_data:
                    perf_lines.append(f"大卡顿: {perf_data['big_jank_count']}")
                if 'small_jank_count' in perf_data:
                    perf_lines.append(f"小卡顿: {perf_data['small_jank_count']}")
                if 'gpu_avg' in perf_data:
                    perf_lines.append(f"GPU使用率: {perf_data['gpu_avg']:.2f}%")
                if 'network_upload_total_kb' in perf_data:
                    perf_lines.append(f"上传流量: {perf_data['network_upload_total_kb']:.2f}KB")
                if 'network_download_total_kb' in perf_data:
                    perf_lines.append(f"下载流量: {perf_data['network_download_total_kb']:.2f}KB")
                
                # 温度信息
                if 'cpu_temp_avg' in perf_data:
                    perf_lines.append(f"CPU温度: {perf_data['cpu_temp_avg']:.1f}°C")
                if 'battery_temp_avg' in perf_data:
                    perf_lines.append(f"电池温度: {perf_data['battery_temp_avg']:.1f}°C")
                
                # 功耗信息
                if 'power_avg' in perf_data:
                    perf_lines.append(f"平均功耗: {perf_data['power_avg']:.2f}mW")
                
                # 指标数量
                metrics_count = perf_data.get('metrics_count', 0)
                perf_lines.append(f"统计指标: {metrics_count}个")
                
            else:
                # 兼容旧格式
                perf_lines = ["性能监控: 已启用"]
                if 'cpu_usage_avg' in perf_data:
                    perf_lines.append(f"CPU使用率: {perf_data['cpu_usage_avg']:.2f}%")
                if 'memory_peak_mb' in perf_data:
                    perf_lines.append(f"内存峰值: {perf_data['memory_peak_mb']:.2f}MB")
                if 'fps_avg' in perf_data:
                    perf_lines.append(f"平均FPS: {perf_data['fps_avg']:.2f}")
                if 'stutter_rate_percent' in perf_data:
                    perf_lines.append(f"卡顿率: {perf_data['stutter_rate_percent']:.2f}%")
                if 'network_upload_total_kb' in perf_data:
                    perf_lines.append(f"上传流量: {perf_data['network_upload_total_kb']:.2f}KB")
                if 'network_download_total_kb' in perf_data:
                    perf_lines.append(f"下载流量: {perf_data['network_download_total_kb']:.2f}KB")
            
            monitor_lines.extend(perf_lines)
        else:
            monitor_lines.append("性能监控: 未启用")
        
        # Logcat数据
        if result.logcat_data:
            logcat_lines = ["Logcat: 已启用"]
            if 'file_path' in result.logcat_data:
                logcat_lines.append(f"文件: {result.logcat_data['file_path']}")
            monitor_lines.extend(logcat_lines)
        else:
            monitor_lines.append("Logcat: 未启用")
        
        # 录制数据
        if result.recording_data:
            recording_lines = ["录制: 已启用"]
            if 'file_path' in result.recording_data:
                recording_lines.append(f"文件: {result.recording_data['file_path']}")
            monitor_lines.extend(recording_lines)
        else:
            monitor_lines.append("录制: 未启用")
        
        return "\n".join(monitor_lines)
    
    def _insert_multiple_screenshots(self, ws, row: int, col: int, screenshots: List[str]) -> None:
        """插入多个截图到单元格中，保持原图比例，智能布局"""
        try:
            from openpyxl.drawing import image as xl_image
            from PIL import Image as PILImage
            import io
            
            # 根据截图数量决定布局
            total_images = len(screenshots)
            
            if total_images == 2:
                # 2个截图：水平排列
                max_images_per_row = 2
                max_width = 140
                max_height = 100
            elif total_images == 3:
                # 3个截图：2+1布局
                max_images_per_row = 2
                max_width = 120
                max_height = 90
            else:  # 4个截图
                # 4个截图：2x2网格
                max_images_per_row = 2
                max_width = 120
                max_height = 90
            
            # 计算网格布局
            rows = (total_images + max_images_per_row - 1) // max_images_per_row
            
            for i, screenshot_path in enumerate(screenshots):
                if not os.path.exists(screenshot_path):
                    continue
                    
                # 计算位置
                img_row = i // max_images_per_row
                img_col = i % max_images_per_row
                
                # 创建缩略图，保持原图比例
                try:
                    with PILImage.open(screenshot_path) as img:
                        original_width, original_height = img.size
                        
                        # 计算缩放比例，保持宽高比
                        width_ratio = max_width / original_width
                        height_ratio = max_height / original_height
                        scale_ratio = min(width_ratio, height_ratio, 1.0)  # 不放大，只缩小
                        
                        # 计算新尺寸
                        new_width = int(original_width * scale_ratio)
                        new_height = int(original_height * scale_ratio)
                        
                        # 创建缩略图
                        img.thumbnail((new_width, new_height), PILImage.Resampling.LANCZOS)
                        
                        # 保存到内存
                        img_buffer = io.BytesIO()
                        img.save(img_buffer, format='PNG')
                        img_buffer.seek(0)
                        
                        # 插入到Excel
                        xl_img = xl_image.Image(img_buffer)
                        xl_img.width = new_width
                        xl_img.height = new_height
                        
                        # 计算锚点位置
                        anchor_col = col + img_col
                        anchor_row = row + img_row
                        xl_img.anchor = f"{get_column_letter(anchor_col)}{anchor_row}"
                        
                        ws.add_image(xl_img)
                        
                except Exception as e:
                    logger.warning(f"处理截图 {screenshot_path} 失败: {e}")
                    continue
                
        except Exception as e:
            logger.warning(f"插入多个截图失败: {e}")
            # 如果失败，显示文件路径列表
            screenshot_text = "\n".join([f"{i+1}. {screenshot}" for i, screenshot in enumerate(screenshots)])
            ws.cell(row=row, column=col, value=screenshot_text)

    def _create_test_summary_sheet(self, wb: Workbook, test_results: List[TestResult]) -> None:
        """创建测试汇总工作表"""
        ws = wb.create_sheet("测试汇总")
        
        # 设置样式
        header_font = Font(name="微软雅黑", size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        data_font = Font(name="微软雅黑", size=11)
        data_alignment = Alignment(horizontal="center", vertical="center")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 标题
        ws.merge_cells('A1:H1')
        ws['A1'] = "测试用例汇总"
        ws['A1'].font = Font(name="微软雅黑", size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        
        # 表头
        headers = ["测试名称", "状态", "开始时间", "结束时间", "耗时(秒)", "步骤数", "失败步骤", "错误信息"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 填充数据
        current_row = 3
        for result in test_results:
            # 统计步骤信息
            total_steps = len(result.steps)
            failed_steps = len([s for s in result.steps if s.status == StepStatus.FAILED])
            
            # 测试用例基本信息
            ws.cell(row=current_row, column=1, value=result.test_name).font = data_font
            ws.cell(row=current_row, column=2, value=result.status.value).font = data_font
            ws.cell(row=current_row, column=3, value=result.start_time.strftime("%Y-%m-%d %H:%M:%S") if result.start_time else "").font = data_font
            ws.cell(row=current_row, column=4, value=result.end_time.strftime("%Y-%m-%d %H:%M:%S") if result.end_time else "").font = data_font
            ws.cell(row=current_row, column=5, value=f"{result.duration:.2f}" if result.duration else "").font = data_font
            ws.cell(row=current_row, column=6, value=total_steps).font = data_font
            ws.cell(row=current_row, column=7, value=failed_steps).font = data_font
            ws.cell(row=current_row, column=8, value=result.error_message or "").font = data_font
            
            # 设置状态颜色
            status_color = self._get_status_color(result.status)
            status_cell = ws.cell(row=current_row, column=2)
            status_cell.fill = PatternFill(start_color=status_color, end_color=status_color, fill_type="solid")
            status_cell.font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
            
            # 设置所有单元格的边框和对齐
            for col in range(1, 9):
                cell = ws.cell(row=current_row, column=col)
                cell.border = thin_border
                cell.alignment = data_alignment
            
            current_row += 1
        
        # 调整列宽
        self._auto_adjust_column_widths(ws, headers)
        
        # 设置行高
        for row in range(1, current_row):
            ws.row_dimensions[row].height = 25

    def _create_error_analysis_sheet(self, wb: Workbook, test_results: List[TestResult]) -> None:
        """创建错误分析工作表"""
        ws = wb.create_sheet("错误分析")
        
        # 设置样式
        header_font = Font(name="微软雅黑", size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        data_font = Font(name="微软雅黑", size=11)
        data_alignment = Alignment(horizontal="left", vertical="center")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 标题
        ws.merge_cells('A1:F1')
        ws['A1'] = "错误分析汇总"
        ws['A1'].font = Font(name="微软雅黑", size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        
        # 表头
        headers = ["测试名称", "错误类型", "错误位置", "错误信息", "发生时间", "错误详情"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 收集所有错误信息
        current_row = 3
        error_count = 0
        
        for result in test_results:
            # 测试用例级别的错误
            if result.status in [TestStatus.FAILED, TestStatus.ERROR] and result.error_message:
                ws.cell(row=current_row, column=1, value=result.test_name).font = data_font
                ws.cell(row=current_row, column=2, value="测试用例错误").font = data_font
                ws.cell(row=current_row, column=3, value="测试用例级别").font = data_font
                ws.cell(row=current_row, column=4, value=result.error_message).font = data_font
                ws.cell(row=current_row, column=5, value=result.start_time.strftime("%Y-%m-%d %H:%M:%S") if result.start_time else "").font = data_font
                ws.cell(row=current_row, column=6, value="").font = data_font  # 错误详情列留空
                
                # 设置所有单元格的边框和对齐
                for col in range(1, 7):
                    cell = ws.cell(row=current_row, column=col)
                    cell.border = thin_border
                    cell.alignment = data_alignment
                
                current_row += 1
                error_count += 1
            
            # 步骤级别的错误
            for step in result.steps:
                if step.status in [StepStatus.FAILED, StepStatus.ERROR] and step.error_message:
                    ws.cell(row=current_row, column=1, value=result.test_name).font = data_font
                    ws.cell(row=current_row, column=2, value="步骤错误").font = data_font
                    ws.cell(row=current_row, column=3, value=step.step_name).font = data_font
                    ws.cell(row=current_row, column=4, value=step.error_message).font = data_font
                    ws.cell(row=current_row, column=5, value=step.start_time.strftime("%Y-%m-%d %H:%M:%S") if step.start_time else "").font = data_font
                    ws.cell(row=current_row, column=6, value="").font = data_font  # 错误详情列留空
                    
                    # 设置所有单元格的边框和对齐
                    for col in range(1, 7):
                        cell = ws.cell(row=current_row, column=col)
                        cell.border = thin_border
                        cell.alignment = data_alignment
                    
                    current_row += 1
                    error_count += 1
        
        # 如果没有错误，显示提示信息
        if error_count == 0:
            ws.merge_cells('A3:F3')
            ws['A3'] = "本次测试未发现错误"
            ws['A3'].font = data_font
            ws['A3'].alignment = Alignment(horizontal="center", vertical="center")
            ws['A3'].border = thin_border
        
        # 调整列宽
        self._auto_adjust_column_widths(ws, headers)
        
        # 设置行高
        for row in range(1, current_row):
            ws.row_dimensions[row].height = 25

