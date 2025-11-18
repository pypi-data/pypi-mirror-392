from datetime import UTC
from datetime import date
from datetime import datetime
from decimal import Decimal
from typing import TYPE_CHECKING
from typing import Final
from typing import overload
from warnings import filterwarnings

from openpyxl import load_workbook

from movslib.model import KV
from movslib.model import ZERO
from movslib.model import Row
from movslib.model import Rows

if TYPE_CHECKING:
    from collections.abc import Iterable

    from openpyxl.worksheet.worksheet import Worksheet

SHEET_NAME: Final = 'ListaMovimenti'
MIN_ROW: Final = 7
MAX_ROW: Final = 999


def _load_sheet(fn: str) -> 'Worksheet':
    # ignore openpyxl warning about libretto .xlsx quality
    filterwarnings('ignore', module='openpyxl.styles.stylesheet')
    wb = load_workbook(fn, read_only=True, data_only=True, keep_links=False)
    return wb[SHEET_NAME]


def _parse_date(dt: str) -> date:
    return datetime.strptime(dt[-10:], '%d/%m/%Y').replace(tzinfo=UTC).date()


def _parse_decimal(d: str) -> Decimal:
    return Decimal(d.replace('.', '').replace(',', '.'))


def _importos(sheet: 'Worksheet') -> 'Iterable[Decimal]':
    for (importo_raw,) in sheet.iter_rows(
        MIN_ROW, MAX_ROW, min_col=4, max_col=4, values_only=True
    ):
        if not isinstance(importo_raw, str):
            raise TypeError(importo_raw)
        yield _parse_decimal(importo_raw)


def _read_kv(sheet: 'Worksheet') -> KV:
    libretto_n_raw = sheet['A1'].value
    periodo_da_raw = sheet['A3'].value
    periodo_a_raw = sheet['A4'].value
    periodo_da = _parse_date(periodo_da_raw)
    periodo_a = _parse_date(periodo_a_raw)
    libretto_n = libretto_n_raw[-12:]
    sum_importos = sum(_importos(sheet), ZERO)

    return KV(
        da=periodo_da,
        a=periodo_a,
        tipo='',
        conto_bancoposta=libretto_n,
        intestato_a='',
        saldo_al=periodo_a,
        saldo_contabile=sum_importos,
        saldo_disponibile=sum_importos,
    )


def read_kv(fn_sheet: 'str | Worksheet') -> KV:
    sheet = _load_sheet(fn_sheet) if isinstance(fn_sheet, str) else fn_sheet
    return _read_kv(sheet)


def _read_csv(sheet: 'Worksheet') -> 'Iterable[Row]':
    for contabile_raw, valuta_raw, descrizione, importo_raw in sheet.iter_rows(
        MIN_ROW, MAX_ROW, min_col=1, max_col=4, values_only=True
    ):
        if not isinstance(contabile_raw, str):
            raise TypeError(contabile_raw)
        if not isinstance(valuta_raw, str):
            raise TypeError(valuta_raw)
        if not isinstance(descrizione, str):
            raise TypeError(descrizione)
        if not isinstance(importo_raw, str):
            raise TypeError(importo_raw)
        importo = _parse_decimal(importo_raw)
        yield Row(
            data_contabile=_parse_date(contabile_raw),
            data_valuta=_parse_date(valuta_raw),
            addebiti=-importo if importo < 0 else None,
            accrediti=importo if importo > 0 else None,
            descrizione_operazioni=descrizione,
        )


def read_csv(fn_sheet: 'str | Worksheet') -> 'Iterable[Row]':
    sheet = _load_sheet(fn_sheet) if isinstance(fn_sheet, str) else fn_sheet
    return _read_csv(sheet)


@overload
def read_libretto(fn: str) -> tuple[KV, list[Row]]: ...


@overload
def read_libretto(fn: str, name: str) -> tuple[KV, Rows]: ...


def read_libretto(
    fn: str, name: str | None = None
) -> tuple[KV, list[Row] | Rows]:
    sheet = _load_sheet(fn)
    kv = read_kv(sheet)
    csv = read_csv(sheet)

    return kv, (list(csv) if name is None else Rows(name, csv))
