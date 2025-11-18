from datetime import UTC
from datetime import date
from datetime import datetime
from decimal import Decimal
from typing import TYPE_CHECKING
from typing import Final
from typing import overload
from warnings import filterwarnings

from openpyxl import load_workbook

from movslib.model import KV
from movslib.model import Row
from movslib.model import Rows

if TYPE_CHECKING:
    from collections.abc import Iterable

    from openpyxl.worksheet.worksheet import Worksheet

SHEET_NAME: Final = 'ListaMovimenti'
MIN_ROW: Final = 13
MAX_ROW: Final = 999


def _load_sheet(fn: str) -> 'Worksheet':
    # ignore openpyxl warning about libretto .xlsx quality
    filterwarnings('ignore', module='openpyxl.styles.stylesheet')
    wb = load_workbook(fn, read_only=True, data_only=True, keep_links=False)
    return wb[SHEET_NAME]


def _parse_date(dt: str) -> date:
    return datetime.strptime(dt[-10:], '%d/%m/%Y').replace(tzinfo=UTC).date()


def _parse_decimal(d: str) -> Decimal:
    return Decimal(d.replace('.', '').replace(',', '.'))


def _importos(sheet: 'Worksheet') -> 'Iterable[Decimal]':
    for (importo_raw,) in sheet.iter_rows(
        MIN_ROW, MAX_ROW, min_col=4, max_col=4, values_only=True
    ):
        if not isinstance(importo_raw, str):
            raise TypeError(importo_raw)
        yield _parse_decimal(importo_raw)


def _read_kv(sheet: 'Worksheet') -> KV:
    def saldo(raw: str) -> str:
        return raw.split(': ', 1)[-1].split(' ', 1)[0]

    return KV(
        da=_parse_date(sheet['A2'].value),
        a=_parse_date(sheet['A3'].value),
        tipo=sheet['A4'].value.split(': ', 1)[-1],
        conto_bancoposta=sheet['A6'].value.split(': ', 1)[-1],
        intestato_a=sheet['A7'].value.split(': ', 1)[-1],
        saldo_al=_parse_date(sheet['A8'].value),
        saldo_contabile=_parse_decimal(saldo(sheet['A9'].value)),
        saldo_disponibile=_parse_decimal(saldo(sheet['A10'].value)),
    )


def read_kv(fn_sheet: 'str | Worksheet') -> KV:
    sheet = _load_sheet(fn_sheet) if isinstance(fn_sheet, str) else fn_sheet
    return _read_kv(sheet)


def _read_csv(sheet: 'Worksheet') -> 'Iterable[Row]':
    def deci(raw: float | None) -> Decimal | None:
        if raw is None:
            return None
        return round(Decimal(raw), 2)

    for (
        contabile_raw,
        valuta_raw,
        addebiti_raw,
        accrediti_raw,
        descrizione,
    ) in sheet.iter_rows(
        MIN_ROW, MAX_ROW, min_col=1, max_col=5, values_only=True
    ):
        if not isinstance(contabile_raw, datetime):
            raise TypeError(contabile_raw)
        if not isinstance(valuta_raw, datetime):
            raise TypeError(valuta_raw)
        if not isinstance(addebiti_raw, float | None):
            raise TypeError(addebiti_raw)
        if not isinstance(accrediti_raw, float | None):
            raise TypeError(accrediti_raw)
        if not isinstance(descrizione, str):
            raise TypeError(descrizione)

        yield Row(
            data_contabile=contabile_raw.date(),
            data_valuta=valuta_raw.date(),
            addebiti=deci(addebiti_raw),
            accrediti=deci(accrediti_raw),
            descrizione_operazioni=descrizione,
        )


def read_csv(fn_sheet: 'str | Worksheet') -> 'Iterable[Row]':
    sheet = _load_sheet(fn_sheet) if isinstance(fn_sheet, str) else fn_sheet
    return _read_csv(sheet)


@overload
def read_lista_movimenti_xlsx(fn: str) -> tuple[KV, list[Row]]: ...


@overload
def read_lista_movimenti_xlsx(fn: str, name: str) -> tuple[KV, Rows]: ...


def read_lista_movimenti_xlsx(
    fn: str, name: str | None = None
) -> tuple[KV, list[Row] | Rows]:
    sheet = _load_sheet(fn)
    kv = read_kv(sheet)
    csv = read_csv(sheet)

    return kv, (list(csv) if name is None else Rows(name, csv))
