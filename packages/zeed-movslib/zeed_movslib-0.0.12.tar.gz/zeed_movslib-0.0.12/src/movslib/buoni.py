from datetime import UTC
from datetime import datetime
from decimal import Decimal
from operator import attrgetter
from typing import TYPE_CHECKING
from typing import Final
from typing import overload

from openpyxl import load_workbook

from movslib.model import KV
from movslib.model import ZERO
from movslib.model import Row
from movslib.model import Rows

if TYPE_CHECKING:
    from collections.abc import Iterable

    from openpyxl.worksheet.worksheet import Worksheet


SHEET_NAME: Final = 'RPOL_PatrimonioBuoni'
MIN_ROW: Final = 2
MAX_ROW: Final = 999


def _load_sheet(fn: str) -> 'Worksheet':
    wb = load_workbook(fn, read_only=True, data_only=True, keep_links=False)
    return wb[SHEET_NAME]


def _read_csv(sheet: 'Worksheet') -> 'Iterable[Row]':
    for (
        _,
        tipologia,
        _,
        _,
        _,
        _,
        _,
        _,
        netto_a_scadenza_raw,
        data_sottoscrizione_raw,
        valore_nominale_raw,
        scadenza_raw,
        *_,
    ) in sheet.iter_rows(
        MIN_ROW, MAX_ROW, min_col=1, max_col=19, values_only=True
    ):
        if not isinstance(tipologia, str):
            raise TypeError(type(tipologia))
        if not isinstance(netto_a_scadenza_raw, str):
            raise TypeError(type(netto_a_scadenza_raw))
        if not isinstance(data_sottoscrizione_raw, str):
            raise TypeError(type(data_sottoscrizione_raw))
        if not isinstance(valore_nominale_raw, str):
            raise TypeError(type(valore_nominale_raw))
        if not isinstance(scadenza_raw, str):
            raise TypeError(type(scadenza_raw))

        netto_a_scadenza = Decimal(
            netto_a_scadenza_raw[1:].replace('.', '').replace(',', '.')
        )
        data_sottoscrizione = datetime.strptime(
            data_sottoscrizione_raw[:10], '%d/%m/%Y'
        ).replace(tzinfo=UTC)
        valore_nominale = Decimal(
            valore_nominale_raw[1:].replace('.', '').replace(',', '.')
        )
        scadenza = datetime.strptime(scadenza_raw, '%d.%m.%Y').replace(
            tzinfo=UTC
        )

        yield Row(
            data_contabile=data_sottoscrizione.date(),
            data_valuta=data_sottoscrizione.date(),
            addebiti=valore_nominale,
            accrediti=None,
            descrizione_operazioni=tipologia,
        )
        yield Row(
            data_contabile=scadenza.date(),
            data_valuta=scadenza.date(),
            addebiti=None,
            accrediti=netto_a_scadenza,
            descrizione_operazioni=tipologia,
        )


def read_csv(fn_sheet: 'str | Worksheet') -> 'Iterable[Row]':
    sheet = _load_sheet(fn_sheet) if isinstance(fn_sheet, str) else fn_sheet
    return sorted(_read_csv(sheet), key=attrgetter('date'))


@overload
def read_buoni(fn: str) -> 'tuple[KV, list[Row]]': ...


@overload
def read_buoni(fn: str, name: str) -> 'tuple[KV, Rows]': ...


def read_buoni(
    fn: str, name: str | None = None
) -> 'tuple[KV, list[Row] | Rows]':
    sheet = _load_sheet(fn)
    kv = KV(
        da=None,
        a=None,
        tipo='buoni postali',
        conto_bancoposta='',
        intestato_a='',
        saldo_al=None,
        saldo_contabile=ZERO,
        saldo_disponibile=ZERO,
    )
    csv = read_csv(sheet)

    return kv, (list(csv) if name is None else Rows(name, csv))
