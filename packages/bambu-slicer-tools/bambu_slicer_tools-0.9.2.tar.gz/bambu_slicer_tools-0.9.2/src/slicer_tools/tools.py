#!/usr/bin/env python
# cspell:ignore Bambu Fiberon Prusa
# TODO: Add special types to formatting (name, inherits, etc.) - also make version
# xx.xx.xx.xx. Document all of this.
# TODO: INCLUDE HUGE WARNING ABOUT INHERITING FROM USER DEFINED BASE TYPES (I
# create these as system types - they are very much not - see _user_base_fix).
# TODO: Document that differencing is only reliable for system group. Others are
# untested.
"""Quick and dirty  tools for reading and comparing Bambu Studio slicer presets.

Summary: This package includes classes for loading system and user presets, and
presets/overrides specified in model .3mf files. The main feature of the package is
comparing presets in one or more .3mf files (the default mode summarises differences
relative to system presets). Differences are provided in .xlsx format. In addition, all
settings in a single preset can be dumped as a json file (from walking the inheritance
tree).

Notes:
    - In this package, I focus on filament, machine and process as the core preset
    types.
    - Preset types may be system, user and project (these three follow the BambuStudio
    naming convention) and **override**, which is a preset generated from project file
    overrides of one of three other groups. Override is not BambuStudio terminology,
    but provides a convenient naming convention for handling project file oddities.

Cautions:
    - This implementation has only been tested on  Windows. But if the Bambu Studio file
    structure for macos mimics Windows, it should work OK on there too (or at least
    can be made to work with minimal updates).
    - I have only implemented an english version, so this package may break when using
    other languages.
    - Right now the package is built around Bambu Studio files. However, I expect that
    it should be easy to extend to Orca Slicer files, and possibly also Prusa Slicer.
"""
import json
import string
import random
from argparse import ArgumentParser, Namespace
from enum import Enum, StrEnum
from typing import cast, NamedTuple, Generator
from zipfile import ZipFile
from zipfile import Path as zPath
from pathlib import Path

from openpyxl import Workbook  # type: ignore[import-untyped]

from slicer_tools.common import AllNodeSettings, NodeMetadata, PresetType, PresetGroup
from slicer_tools.common import SettingsDict, SettingValue
from slicer_tools.common import FROM, INHERITS, NAME, DEFAULT_ENCODING
from slicer_tools.common import choose
from slicer_tools.presets import ProjectPresets

# Navigation and file name components
THREE_MF_FOLDER = "./3MF"
THREE_MF_EXT = ".3mf"
CONFIG_EXT = ".config"
METADATA = "Metadata/"

# Constants for manipulating .configs
# Can build all json file names from these two constants and the
# PresetType enum below.
SETTINGS = "_settings"
PROJECT_CONFIG_NAME = "project" + SETTINGS
EXCLUDE_CONFIGS = ["model_settings", "slice_info"]

# more json key elements
# annoyingly, in config files, settings id uses a different terminology to
# the file naming - specifically print in the json is equivalent to process and
# printer is equivalent to machine. Yay!
PRINT = "print"
PRINTER = "printer"
DIFFS_TO_SYSTEM = "different_settings_to_system"
ID = "_id"
INHERITS_GROUP = "inherits_group"  # Because why do one thing one way ...
VERSION = "version"
SPECIAL_VERSION = "XX.XX.XX.XX"
# Special settings that will require .
SPECIAL_TYPES = [
    "from",
    "filament_extruder_variant",
    "filament_settings_id",
    "inherits",
    "name",
    "print_extruder_id",
    "print_extruder_variant",
    "print_settings_id",
    "printer_extruder_id",
    "printer_extruder_variant",
    "printer_settings_id",
    "version",
]


class CellFormat(StrEnum):
    """Builtin Excel cell formats."""

    NORMAL = "Normal"
    GOOD = "Good"
    BAD = "Bad"
    INPUT = "Input"
    NOTE = "Note"
    NEUTRAL = "Neutral"
    HEADING4 = "Headline 4"


class DiffType(Enum):
    """Difference types."""

    # No difference in values.
    NO_DIFF = 0
    # Reference value (the value differenced to!).
    REFERENCE = 1
    # Normal difference to reference value.
    DIFFERENCE = 2
    # Override system difference.
    OVERRIDE = 3
    # Unset reference value.
    UNSET = 4
    # Too complex to diff (for now at least).
    COMPLEX = 5
    # Value requires special/manual handling
    SPECIAL = 6


class CellInfo(NamedTuple):
    """Summary data for writing to an Excel cell."""

    row: int
    column: int
    value: str
    format: DiffType | CellFormat


class DiffValue(NamedTuple):
    """Container for difference value and difference type information."""

    value: str
    type: DiffType


class DiffValuePath(NamedTuple):
    """Provide a unique path/key for difference sets.

    This tuple eliminates the requirement for sparse nested dictionaries or minimal
    classes emulating nested dicts, at the cost of using tuples as dict keys.
    """

    # This is pretty similar to PresetPath. However, a diff set may contain multiple
    # preset types and multiple .3mf files, so we need to add  the filename to ensure
    # uniqueness - as this info is already available in node metadata, we're just using
    # that. In addition, the value (row) name will also be sparse. As adding row name
    # to the key makes iteration over differences easier, I've done that as well.
    row_name: str
    metadata: NodeMetadata


class DiffMatrix:
    """Difference data in a sparse matrix form."""

    # This started as a dataclass, but as all members are now private, converted to
    # full fledged class.
    _rows: dict[str, int]
    _cols: dict[NodeMetadata, int]
    _values: dict[DiffValuePath, DiffValue]
    # Flag indicating indices need recalculating. Set to true every time a new
    # value is added.
    _reset_required: bool
    # Group, filename, preset name.
    _header_count: int = 3

    def __init__(self) -> None:
        """Create instance variables."""
        self._rows = {}
        self._cols = {}
        self._values = {}
        self._reset_required = True

    def column_exits(self, metadata: NodeMetadata) -> bool:
        """Test if column is defined."""
        if metadata in self._cols:
            return True
        return False

    def add_value(
        self, row_name: str, column_id: NodeMetadata, value: str, value_type: DiffType
    ) -> None:
        """Add or overwrite a value to the diff matrix.

        Automatically triggers an instance reset and will make any active generator
        invalid (values/indices have been modified in ways that are unpredictable for
        the generator).
        """
        self._reset_required = True
        self._rows[row_name] = -1
        self._cols[column_id] = -1
        self._values[DiffValuePath(row_name, column_id)] = DiffValue(value, value_type)

    def _reset_lookups(self) -> None:
        """Prepare row and column lookup indices for writing tables.

        All indices are 0 based.
        """
        key: str | NodeMetadata

        keys = self._rows.keys()
        for i, key in enumerate(sorted(keys), start=0):
            self._rows[key] = i

        # Column sort is not worth doing, as we want a rough sort along the lines:
        #   Reference, Override, Project and User.
        # While it would be possible to write a compare function, it's nearly as
        # simple to just run over the column list for each group.
        i = 0
        for group in [
            PresetGroup.SYSTEM,
            PresetGroup.OVERRIDE,
            PresetGroup.PROJECT,
            PresetGroup.USER,
        ]:
            for key, _ in self._cols.items():
                if key.group == group:
                    self._cols[key] = i
                    i += 1

    def table_cells(self) -> Generator[CellInfo]:
        """Generate table and associated row names and column headers.

        Return values are row offset (0 based), column number, value, value type or
        format.
        """
        if self._reset_required:
            self._reset_lookups()

        # Header names. All 0 based.
        yield CellInfo(0, 0, "Group", CellFormat.NORMAL)
        yield CellInfo(1, 0, "Filename", CellFormat.NORMAL)
        yield CellInfo(2, 0, "Preset", CellFormat.NORMAL)

        for row_name, row_offset in self._rows.items():
            yield CellInfo(
                row_offset + self._header_count, 0, row_name, CellFormat.NORMAL
            )

        for metadata, col_offset in self._cols.items():
            yield CellInfo(0, col_offset + 1, metadata.group.value, CellFormat.NORMAL)
            yield CellInfo(1, col_offset + 1, metadata.filename, CellFormat.NORMAL)
            yield CellInfo(2, col_offset + 1, metadata.name, CellFormat.NORMAL)

        for key, value in self._values.items():
            row = self._rows[key.row_name] + self._header_count
            col = self._cols[key.metadata] + 1
            yield CellInfo(row, col, value.value, value.type)

    def row_count(self) -> int:
        """Row count in the table include header lines."""
        return len(self._rows) + self._header_count


class ThreeMFPresets:
    """Container for all presets in a 3mf file.

    Currently skips per object and slice data. And will probably never implement these.
    """

    filename: str
    # Project settings extracted from project_settings.config.
    project_config: SettingsDict

    # Collection of the following presets:
    # - All project presets and overrides in the project file.
    # - All user presets that the project presets/overrider might inherit from.
    # - All system presets that the project presets/overrides might inherit from.
    # ThreeMFPresets creates the first group, the other two are loaded automatically.
    # After all project presets have been created, the project_presets member
    # can generate settings and differences as required.
    # The caveat is that calls to generate settings/differences must be deferred until
    # all Presets in the 3mf have been instantiated, as otherwise there is a risk that
    # the project_presets object will not hold the data required for a complete roll up.
    # (In effect this is creating a half way house to a class variable that is only
    # visible within each ThreeMFSettings instance and the Settings instances contained
    # by it. Yuck!)
    presets: ProjectPresets

    def __init__(
        self, project_file: Path, appdata_path: Path | None = None, user_id: str = ""
    ) -> None:
        """Create project metadata container.

        appdata_path and user_id override the default locations for find the system
        and user presets.
        """
        self.filename = project_file.name
        self.presets = ProjectPresets(appdata_path=appdata_path, bbl_user_id=user_id)
        with ZipFile(project_file, mode="r") as archive:
            for zip_path in zPath(archive, at=METADATA).iterdir():
                # Iterate over metadata folder to find presets (.config files).
                if (
                    zip_path.name.endswith(CONFIG_EXT)
                    and zip_path.stem not in EXCLUDE_CONFIGS
                ):
                    # Excluding xml format configs for now.
                    # We'll need the config data no matter what.
                    # By rights, this should be a json file at this point.
                    settings = json.loads(zip_path.read_text(encoding=DEFAULT_ENCODING))
                    if zip_path.stem == PROJECT_CONFIG_NAME:
                        # Project config. Will contain multiple overrides and
                        # requires special handling.
                        self.project_config = settings
                        self._process_project_settings()
                    else:
                        # Find out what type we are working with and create preset.
                        for prefix in PresetType:
                            if zip_path.stem.startswith(prefix + SETTINGS):
                                self.presets.add_project_node(
                                    NodeMetadata(
                                        name=settings[NAME],
                                        filename=project_file.name,
                                        preset_type=prefix,
                                        group=PresetGroup.PROJECT,
                                    ),
                                    settings=settings,
                                )
                                break

    def _process_project_settings(self) -> None:
        """Extract presets from project_settings.config and add to project presets."""
        # process the differences by type.
        # Guaranteed there is a better way to this, but patience is gone.
        group = PresetGroup.OVERRIDE
        filament_count = len(self.project_config[PresetType.FILAMENT + SETTINGS + ID])
        filament_idx = -1
        for diff_idx, diff in enumerate(self.project_config[DIFFS_TO_SYSTEM]):
            if diff_idx == 0:
                # Process.
                name = self.project_config[PRINT + SETTINGS + ID]
                value_idx = 0
                preset_type = PresetType.PROCESS

            elif diff_idx == filament_count + 1:
                # Machine.
                name = self.project_config[PRINTER + SETTINGS + ID]
                value_idx = 0
                preset_type = PresetType.MACHINE

            else:
                # Filament
                filament_idx += 1
                name = self.project_config[PresetType.FILAMENT + SETTINGS + ID][
                    filament_idx
                ]
                value_idx = filament_idx
                preset_type = PresetType.FILAMENT

            # Set up the special unicorn tears preset.
            # Create settings and add key/values missing from the override differences.
            # Unfortunately, the project override name is not unique, as it is also the
            # the name of the immediate parent that the override inherits from (see
            # PresetMetadata for more on this, and why we are setting up the fix we
            # do here).
            # The fix is to create the preset metadata with the parent preset inherit
            # specified and the preset name made unique by adding 4 random characters
            # to the end of the name.
            name = cast(str, name)
            metadata = NodeMetadata(
                name=name + "_" + "".join(random.sample(string.ascii_lowercase, 4)),
                filename=self.filename,
                preset_type=preset_type,
                group=group,
                # This is the immediate parent, which may be different from the
                # inherits_group value in project_settings. Yay!
                override_inherits=name,
            )

            # And now we pre-populate settings with values that project_settings does
            # not provide. As another project_settings wrinkle: if the project does not
            # override the system values at all, "inherits" is set to "".
            # We do some munging here to fix this as well. Yay.
            inherits = self.project_config[INHERITS_GROUP][diff_idx]
            if not inherits:
                inherits = name

            settings = {
                # Use unique name for override.
                NAME: metadata.name,
                INHERITS: inherits,
                FROM: self.project_config[FROM],
                # special value here, as versions can be all over the place.
                VERSION: SPECIAL_VERSION,
            }

            # Finally collect the actual differences.
            self._differences_to_settings(
                value_idx=value_idx,
                filament_count=filament_count,
                diff=diff,
                settings=settings,
                metadata=metadata,
            )

            self.presets.add_project_node(
                metadata=metadata,
                settings=settings,
            )

    def _differences_to_settings(
        self,
        value_idx: int,
        diff: str,
        settings: SettingsDict,
        metadata: NodeMetadata,
        filament_count: int = 0,
    ) -> None:
        """Create override settings from difference list.

        expect_count is the number of values expected in the value lists. Used for
        list length validations. List length validation tries to be smart, but may get
        things wrong. Check errors and warnings to resolve.
        """
        diff_keys = diff.split(";")
        if not diff_keys[0]:
            return

        # If I was confident about value counts, could do this with a comprehension.
        # But instead we'll do it slow with checks. And we need to, because of
        # inconsistency.
        for key in diff_keys:
            values = self.project_config[key]

            # Deal with filament first.
            if metadata.preset_type == PresetType.FILAMENT:
                if key == "filament_notes":
                    print(
                        f"Warning: key `filament_notes' appears in override"
                        f" {metadata.name}. Key ignored."
                        f"\n  (BambuStudio currently retains the note value for one"
                        f"  filament only in the config file and I have no idea which"
                        f"  one is important.)"
                    )
                    continue

                if not isinstance(values, list):
                    raise TypeError(
                        f"Expected list of values for setting '{key}'"
                        f"\n  of filament override {metadata.name}."
                        f"\n  Got {type(values)}. (Values = {values})."
                    )

                if len(values) != filament_count:
                    # Will raise a warning or error. Prep info string accordingly.
                    raise IndexError(
                        f"Unexpected override list size."
                        f"\n  List length is expected to match override/project"
                        f" filament count."
                        f"\n  Override `{metadata.name}` setting `{key}`:"
                        f"\n     Expected list length count {filament_count}, got"
                        f" {len(values)}."
                        f"\n    (Values = {values})"
                    )

                # If we're here, I think it's real.
                settings[key] = values[value_idx]

            else:
                # Now dealing with machine and process values. These could be anything,
                # but typically only expect either a string or single value list.
                # Warn about anything unexpected and grab the value anyway.
                if isinstance(values, str) or (
                    isinstance(values, list) and len(values) == 1
                ):
                    # What we expected.
                    settings[key] = values

                else:
                    print(
                        f"Warning: Unexpected type for machine/process override."
                        f" This is probably OK, but should be checked to make sure."
                        f"\n  In override `{metadata.name}` for setting `{key}`:"
                        f"\n  Expected either a single value list or a string."
                        f"\n  Got (Values = {values})"
                    )
                    # Not what we expected, but still probably fine.
                    settings[key] = values


class ProjectDiffSet:
    """Group of project/3mf settings differences/comparisons.

    All differences/comparisons in a ProjectDiffSet instance are relative to a reference
    PresetGroup specified when the instance is initialised. The default reference
    PresetGroup is "system", which follows the BambuStudio model for project configs
    (as implemented by "different_settings_to_system" in the project_settings.config
    file).

    ProjectDiffSet uses the following information to generate differences:
    - Setting values for the project presets.
    - Setting values for the reference presets that project presets inherit from (may
    be unset).
    - If the the project preset is an override, the setting values the project preset
    inherits from its parent (may be unset).

    The ProjectDiffSet presents the following data when differences exist:
    - The setting value and the reference value (may be unset).
    - A reference value flag of REFERENCE, COMPLEX, SPECIAL or UNSET.
    - A settings value flag of DIFFERENT, COMPLEX, SPECIAL or OVERRIDE.
    - OVERRIDE represents values that Bambu Studio highlights in orange.
    A more long winded explanation (because I need it):
            - The project preset is created by inheriting a user or project preset, but
            overrides the value inherited from its parent (orange in BS GUI).
            - The parent value is different to the reference value (for a regular
            difference, this would be a black value in the BS GUI).
            - The system reference value from the system preset is different to both
            of the above values.
    The OVERRIDE parent values are not recorded.
    - COMPLEX means that type of the setting is too complicated for this difference
    tool to evaluate (e.g. a list of lists or other similar nested or mixed type
    objects.)
    - SPECIAL means that these values require special handling or may have been hacked
    to deliver package functionality - refer to the package documentation on the best
    way to handle these setting/value pairs.

    Note: this class is intended to allow comparison of multiple settings in
    multiple files. A separate class should be implemented to compare two items (e.g.
    filaments). I haven't implemented this for now, as this is largely available in the
    BS GUI (albeit with no export function).
    """

    _reference_group: PresetGroup
    # Data is a dict of sparse matrices storing values and index info.
    _data: dict[PresetType, DiffMatrix]
    # The current target row for writing out Excel data.
    _write_row: int = 1

    def __init__(self, reference_group: PresetGroup = PresetGroup.SYSTEM) -> None:
        """Create preset diff."""
        # The default preset group is SYSTEM, which more or less follows the BS
        # representation.
        self._reference_group = reference_group
        self._data = {}
        for preset_type in PresetType:
            self._data[preset_type] = DiffMatrix()

    @staticmethod
    def _extract_str(value: SettingValue, diff_type: DiffType) -> tuple[str, DiffType]:
        """Very simple type checking and string extraction.

        Returns (simple string value or repr of complex value, diff_type if string value
        found | DiffType.Complex if complex value)

        Value is simple if it is a string or list with one string member.
        """
        if isinstance(value, str):
            return value, diff_type
        if isinstance(value, list) and len(value) == 1 and isinstance(value[0], str):
            return value[0], diff_type

        return repr(value), DiffType.COMPLEX

    def _add_diff(
        self, settings: AllNodeSettings, inherited: AllNodeSettings | None = None
    ) -> None:
        """Add any differences in preset to the DiffSet."""
        # Easy checks first.

        # Validate that this is a unique project preset.
        metadata = settings.metadata
        if self._data[metadata.preset_type].column_exits(metadata):
            raise IndexError(f"Repeated difference for preset '{metadata}'")

        # Terminate if there is no reference data (should not happen).
        if settings.ref_metadata is None:
            raise ValueError(
                f"Preset '{metadata}' is missing difference reference data"
                f"\nThis should not be possible!"
            )

        # Comparison process, which is bit messy. I think the following truth table
        # covers it:
        # v1 != v2 != v3
        # Preset  Ref.    Over.   Over.   Preset   Ref.
        # Value   Value   type?   Value   Flag     Flag  Record?  Notes
        #   v1      v1     Any      NA    NO_DIFF  REF   No       Shouldn't happen,
        #                                                         check anyway.
        #   v1      Unset  No       NA    DIFF     UNSET Yes
        #   v1      Unset  Yes      Unset DIFF     UNSET Yes
        #   v1      Unset  Yes      v1    DIFF     UNSET Yes
        #   v1      Unset  Yes      v3    OVER.    UNSET Yes
        #
        #   v1      v2     No       NA    DIFF     REF   Yes
        #   v1      v2     Yes      Unset DIFF     REF   Yes
        #   v1      v2     Yes      v1    DIFF     REF   Yes
        #   v1      v2     Yes      v2    DIFF     REF   Yes      Shouldn't happen!
        #   v1      v2     Yes      v3    OVER.    REF   Yes
        #
        # To really do this properly, we should cross check that value types match,
        # which will get ugly for nest types (probably needs pydantic), and then
        # for differences in the values themselves.
        # HOWEVER, that's a bit much work for the simple difference tool that I'm aiming
        # for in the first pass. So for the first pass, we take the following shortcuts.
        # A) Assume the inheritance tree is definitive for differences. So we take it
        # on faith the differences exist if the tree says so and don't check.
        # B) We do check that value types are either a string or a single string value
        # in a list, which we will need to do make sure override types match the
        # presets they inherit from (see implementation).
        # C) Anything more complex than (B) is recorded as TOO_COMPLEX for the
        # simple difference. Up to the user to resolve. (For most cases this probably
        # won't be an issue).
        # Revisit this if complex types happen more often than I expect.

        for setting_name, value in settings.source_subtree.items():
            if setting_name in settings.reference_subtree:
                ref_value = settings.reference_subtree[setting_name]
                ref_value, ref_type = self._extract_str(ref_value, DiffType.REFERENCE)

            else:
                ref_value = ""
                ref_type = DiffType.UNSET

            value, value_type = self._extract_str(value, DiffType.DIFFERENCE)

            if setting_name in SPECIAL_TYPES:
                # Special type, don't bother checking override details, just update
                # types.
                value_type = DiffType.SPECIAL
                ref_type = DiffType.SPECIAL

            elif (
                value_type == DiffType.DIFFERENCE
                and inherited
                and setting_name in inherited.source_subtree
            ):
                # Handle the odd duck if we have a difference.
                # No point checking overrides for the complex case.
                inherit_value = inherited.source_subtree[setting_name]
                inherit_value, inherit_type = self._extract_str(
                    inherit_value, DiffType.DIFFERENCE
                )
                # now a long string of checks to see if we have a meaningful
                # override.
                if inherit_type == DiffType.DIFFERENCE and inherit_value != value:
                    if ref_type == DiffType.UNSET or (
                        ref_type == DiffType.REFERENCE and ref_value != inherit_value
                    ):
                        # value != inherit != (unset | ref_value).
                        value_type = DiffType.OVERRIDE

            # Record value, add setting name to row indices (likely to be duplicate
            # calls, not worth chasing efficiency here).
            self._data[settings.metadata.preset_type].add_value(
                setting_name, settings.metadata, value, value_type
            )
            self._data[settings.metadata.preset_type].add_value(
                setting_name, settings.ref_metadata, ref_value, ref_type
            )

    def add_project_presets(self, project: ThreeMFPresets) -> None:
        """Create diffs for all project presets in a 3mf file."""
        for settings in project.presets.project_presets(self._reference_group):

            # If settings are for an OVERRIDE preset, then the diff set needs to
            # know the parent settings for preset the that the override preset is built
            # from. Gather them here if needed.
            inherited: AllNodeSettings | None = None
            if settings.metadata.override_inherits:
                inherited = project.presets.all_node_settings(
                    settings.metadata.preset_type,
                    node_name=settings.metadata.override_inherits,
                    ref_group=self._reference_group,
                )
                # Previously tried to do a basic check for consistency of reference
                # values, but this falls down if trying to compare anything other
                # than a system ref group, so I've dropped it and just take it on
                # faith that BS is doing the right thing.

            self._add_diff(settings, inherited)

    @staticmethod
    def _xlsx_cell(  # type: ignore[no-untyped-def]
        ws, row: int, column: int, value: str, cell_format: DiffType | CellFormat
    ) -> int:
        """Set the value and format for an xlsx cell.

        returns row + 1 (useful if moving down after each cell).
        """
        cell = ws.cell(row=row, column=column, value=value)
        # can add Neutral if I want a special value.
        if isinstance(cell_format, DiffType):
            match cell_format:
                case DiffType.DIFFERENCE:
                    cell.style = CellFormat.NORMAL
                case DiffType.REFERENCE:
                    cell.style = CellFormat.GOOD
                case DiffType.UNSET:
                    cell.style = CellFormat.BAD
                case DiffType.OVERRIDE:
                    cell.style = CellFormat.INPUT
                case DiffType.COMPLEX:
                    cell.style = CellFormat.NOTE
                case DiffType.SPECIAL:
                    cell.style = CellFormat.NEUTRAL
                case _:
                    cell.style = "Normal"
        else:
            cell.style = cell_format

        return row + 1

    def _xlsx_empty_line(self) -> None:
        """Add an empty line to current sheet."""
        self._write_row += 1

    def _xlsx_legend(self, ws) -> None:  # type: ignore[no-untyped-def]
        """Write legend information into worksheet."""
        self._write_row = self._xlsx_cell(
            ws,
            row=self._write_row,
            column=1,
            value="Normal/difference value",
            cell_format=DiffType.DIFFERENCE,
        )
        self._write_row = self._xlsx_cell(
            ws,
            row=self._write_row,
            column=1,
            value="Reference value",
            cell_format=DiffType.REFERENCE,
        )
        self._write_row = self._xlsx_cell(
            ws,
            row=self._write_row,
            column=1,
            value="Unset (problem) value",
            cell_format=DiffType.UNSET,
        )
        self._write_row = self._xlsx_cell(
            ws,
            row=self._write_row,
            column=1,
            value="Value too complex to difference",
            cell_format=DiffType.COMPLEX,
        )
        self._write_row = self._xlsx_cell(
            ws,
            row=self._write_row,
            column=1,
            value="Override value",
            cell_format=DiffType.OVERRIDE,
        )
        self._write_row = self._xlsx_cell(
            ws,
            row=self._write_row,
            column=1,
            value=(
                "Special. Value type needs special/alternative handling. See package"
                " documentation for details."
            ),
            cell_format=DiffType.SPECIAL,
        )

    def _xls_diff_table(  # type: ignore[no-untyped-def]
        self, ws, preset: PresetType
    ) -> None:
        """Write diff table to ws."""
        self._write_row = self._xlsx_cell(
            ws, self._write_row, 1, preset.value, CellFormat.HEADING4
        )

        data = self._data[preset]
        for cell in data.table_cells():
            self._xlsx_cell(
                ws, self._write_row + cell.row, cell.column + 1, cell.value, cell.format
            )

        self._write_row += data.row_count()

    def save_xlsx(self, xlsx_file: Path | None = None) -> None:
        """Write the DiffSet to xlsx file using openpyxl."""
        if not xlsx_file:
            xlsx_file = Path().cwd() / "Preset diffs.xlsx"

        if xlsx_file.exists():
            c = input("Difference file exists. y/Y to overwrite> ")

            if c[0] != "y" and c[0] != "Y":
                print("Differences not saved.")
                return

        # Create workbook and worksheets and populate. We could also modify this to dump
        # the tables to a single worksheet, but I think the per worksheet is better for
        # separating machine/process/filament.
        wb = Workbook()
        for preset in PresetType:
            ws = wb.create_sheet(title=preset.value)
            self._write_row = 1
            self._xlsx_legend(ws)
            self._xlsx_empty_line()
            self._xls_diff_table(ws, preset)

        # Clean up - delete empty ws[0]
        wb.remove(wb.worksheets[0])
        wb.save(filename=str(xlsx_file.resolve()))


def presets_to_xlsx(cl_args: Namespace) -> None:
    """Create differences from arguments."""
    if cl_args.source:
        working = Path(cl_args.source).resolve()
    else:
        working = Path.cwd()

    # For non-default system or user presets, create a StandardPreset instance and use
    # it when creating the first instance of ThreeMFConfig (it will be ignored on
    # second and later calls).
    #  standard_presets = StandardPresets(appdata_path, user_id)
    #  this_config = ThreeMFConfig(...., standard_presets=standard_presets)
    # Could add this to command line arguments if there is a real call for it.
    # For now, I'm assuming that the vast majority of users won't need this.

    # Create a diff set that mimics the standard BambuStudio diff to system.
    diffs = ProjectDiffSet(cl_args.ref)

    if working.is_file():
        this_3mf = ThreeMFPresets(working)
        diffs.add_project_presets(this_3mf)
        working = working.parent
    else:
        for filepath in working.iterdir():
            if filepath.is_file() and filepath.suffix.lower() == THREE_MF_EXT:
                # Grab metadata, discard after generating differences.
                this_3mf = ThreeMFPresets(filepath)

                # This should have loaded all key configuration data.
                # Generate differences from system settings.
                diffs.add_project_presets(this_3mf)

    if cast(str, cl_args.output)[-5:] != ".xlsx":
        xlsx_path = working / (cl_args.output + ".xlsx")
    else:
        xlsx_path = working / cl_args.output

    diffs.save_xlsx(xlsx_path)


def dump_json(cl_args: Namespace) -> None:
    """Export settings from preset(s) in .3mf file."""
    filepath = Path(cl_args.source).resolve()
    this_3mf = ThreeMFPresets(filepath)

    selection = choose(
        options=[preset_type.value for preset_type in PresetType],
        header=(
            "Select the preset type you want to export."
        ),
        allow_multi=False,
        no_action="Return without exporting.",
    )

    if not selection:
        return

    preset_type = PresetType(selection[0])
    names = this_3mf.presets.preset_names(preset_type)

    selection = choose(
        names,
        header=(
            "\nEnter one or more comma separated numbers to choose the presets"
            "\nthat you want to export. E.g. 1, 3, 5<Enter>. Filenames will be"
            "\nthe preset name with '.json' suffix. Refer package documentation"
            "\nfor explanation of override naming (random letters on end of names)."
        ),
        allow_multi=True,
        no_action="Return without exporting.",
    )

    if not selection:
        return

    working = filepath.parent
    for name in selection:
        settings = this_3mf.presets.all_node_settings(preset_type, name)
        filepath = working / (name + ".json")

        with open(filepath, "w", encoding=DEFAULT_ENCODING) as fp:
            json.dump(settings.source_subtree, fp)


def get_parser() -> ArgumentParser:
    """Create argument parser for command line."""
    this_parser = ArgumentParser(
        description="Simple difference/dumper utility for BambuStudio .3mf files."
    )

    subparsers = this_parser.add_subparsers(
        help="<command> --help, -h for subcommand help.", required=True
    )

    parser_diff = subparsers.add_parser(
        "diff",
        description="Generate differences from file or folder.",
    )

    parser_diff.add_argument(
        "--source",
        help=(
            "Specify a source .3mf file or folder containing .3mf files."
            " Defaults to cwd()."
        ),
    )

    parser_diff.add_argument(
        "--output",
        help="Filename for output without suffix (.xlsx or .json files will"
        " be created based on mode). Defaults to `preset_diffs`",
        default="preset_diffs",
    )

    # Ended up deciding that I don't trust anything other than system, so removed
    # this capability altogether.
    # parser_diff.add_argument(
    #    "--ref",
    #    help=(
    #        "Name of the reference group for differencing"
    #        " (`system` | `user` | `project`)."
    #        " Defaults to `system`. 'system' is robust, others are untested and should"
    #        " be used with extreme caution (I recommend only using system)."
    #    ),
    #    type=PresetGroup,
    #    default=PresetGroup.SYSTEM,
    # )
    parser_diff.set_defaults(ref=PresetGroup.SYSTEM)

    parser_diff.set_defaults(func=presets_to_xlsx)

    parser_dump = subparsers.add_parser(
        "export",
        description=(
            "Export json for one or more presets from a file. An interactive menu"
            " system will guide you through selecting preset type and preset names"
            " for export."
        ),
    )

    parser_dump.add_argument(
        "source",
        help=("Specify a source .3mf file."),
    )

    parser_dump.set_defaults(func=dump_json)

    return this_parser


if __name__ == "__main__":
    parser = get_parser()
    args = parser.parse_args()
    args.func(args)
