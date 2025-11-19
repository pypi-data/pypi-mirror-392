"""Export functionality for DNS benchmark results."""

import os
import tempfile
from typing import Any, Dict, List, Optional

import matplotlib
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from weasyprint import HTML

from dns_benchmark.analysis import BenchmarkAnalyzer
from dns_benchmark.core import DNSQueryResult

matplotlib.use("Agg")  # Use non-interactive backend


class ExportBundle:
    @staticmethod
    def export_json(
        results: List[DNSQueryResult],
        analyzer: BenchmarkAnalyzer,
        domain_stats: Optional[List[Dict[str, Any]]],
        record_type_stats: Optional[List[Dict[str, Any]]],
        error_stats: Optional[Dict[str, int]],
        output_path: str,
    ) -> None:
        payload = {
            "overall": analyzer.get_overall_statistics(),
            "resolver_stats": [vars(s) for s in analyzer.get_resolver_statistics()],
            "raw_results": [
                {
                    "resolver_name": r.resolver_name,
                    "resolver_ip": r.resolver_ip,
                    "domain": r.domain,
                    "record_type": r.record_type,
                    "latency_ms": r.latency_ms,
                    "status": r.status.value,
                    "answers_count": len(r.answers),
                    "ttl": r.ttl,
                    "error_message": r.error_message,
                    "start_time": r.start_time,
                    "end_time": r.end_time,
                    "attempt_number": r.attempt_number,
                    "cache_hit": r.cache_hit,
                    "iteration": r.iteration,
                    "query_id": r.query_id,
                }
                for r in results
            ],
            "domain_stats": domain_stats,
            "record_type_stats": record_type_stats,
            "error_stats": error_stats,
        }
        with open(output_path, "w") as f:
            import json

            json.dump(payload, f, indent=2)


class CSVExporter:
    """Export DNS benchmark results to CSV format."""

    @staticmethod
    def export_raw_results(results: List[DNSQueryResult], output_path: str) -> None:
        """Export raw query results to CSV."""
        data = []
        for result in results:
            data.append(
                {
                    "timestamp": result.start_time,
                    "resolver_name": result.resolver_name,
                    "resolver_ip": result.resolver_ip,
                    "domain": result.domain,
                    "record_type": result.record_type,
                    "latency_ms": result.latency_ms,
                    "status": result.status.value,
                    "answers_count": len(result.answers),
                    "ttl": result.ttl or "",
                    "error_message": result.error_message or "",
                    "cache_hit": result.cache_hit,
                    "iteration": result.iteration,
                    "query_id": result.query_id,
                }
            )

        df = pd.DataFrame(data)
        df.to_csv(output_path, index=False)

    @staticmethod
    def export_summary_statistics(
        analyzer: BenchmarkAnalyzer, output_path: str
    ) -> None:
        """Export summary statistics to CSV."""
        resolver_stats = analyzer.get_resolver_statistics()

        data = []
        for stats in resolver_stats:
            data.append(
                {
                    "resolver_name": stats.resolver_name,
                    "resolver_ip": stats.resolver_ip,
                    "total_queries": stats.total_queries,
                    "successful_queries": stats.successful_queries,
                    "success_rate": stats.success_rate,
                    "min_latency_ms": stats.min_latency,
                    "avg_latency_ms": stats.avg_latency,
                    "median_latency_ms": stats.median_latency,
                    "max_latency_ms": stats.max_latency,
                    "std_latency_ms": stats.std_latency,
                    "p95_latency_ms": stats.p95_latency,
                    "p99_latency_ms": stats.p99_latency,
                    "jitter_ms": stats.jitter,
                    "consistency_score": stats.consistency_score,
                }
            )

        df = pd.DataFrame(data)
        df.to_csv(output_path, index=False)

    @staticmethod
    def export_domain_statistics(
        domain_stats: List[Dict[str, Any]], output_path: str
    ) -> None:
        df = pd.DataFrame(domain_stats)
        df.to_csv(output_path, index=False)

    @staticmethod
    def export_record_type_statistics(
        rt_stats: List[Dict[str, Any]], output_path: str
    ) -> None:
        df = pd.DataFrame(rt_stats)
        df.to_csv(output_path, index=False)

    @staticmethod
    def export_error_statistics(error_stats: Dict[str, int], output_path: str) -> None:
        df = pd.DataFrame(
            [{"error_message": k, "count": v} for k, v in error_stats.items()]
        )
        df.to_csv(output_path, index=False)


class ExcelExporter:
    """Export DNS benchmark results to Excel format."""

    @staticmethod
    def export_results(
        results: List[DNSQueryResult],
        analyzer: BenchmarkAnalyzer,
        output_path: str,
        domain_stats: Optional[List[Dict[str, Any]]] = None,
        record_type_stats: Optional[List[Dict[str, Any]]] = None,
        error_stats: Optional[Dict[str, int]] = None,
        include_charts: bool = False,
    ) -> None:
        wb = Workbook()
        wb.remove(wb.active)

        # Temporary directory for chart images
        temp_dir = None
        chart_paths = []
        try:
            # Add standard sheets
            ExcelExporter._add_raw_data_sheet(wb, results)
            ExcelExporter._add_resolver_summary_sheet(wb, analyzer)

            if domain_stats:
                ExcelExporter._add_simple_table_sheet(
                    wb, "Domain Stats", pd.DataFrame(domain_stats)
                )
            if record_type_stats:
                ExcelExporter._add_simple_table_sheet(
                    wb, "Record Type Stats", pd.DataFrame(record_type_stats)
                )
            if error_stats:
                df = pd.DataFrame(
                    [{"Error": k, "Count": v} for k, v in error_stats.items()]
                )
                ExcelExporter._add_simple_table_sheet(wb, "Error Breakdown", df)

            # Add charts if requested
            if include_charts:
                temp_dir = tempfile.mkdtemp()
                chart_paths = ExcelExporter._add_charts_sheet(wb, analyzer, temp_dir)

            # Save workbook before cleaning up temp files
            wb.save(output_path)
        finally:
            # Clean up temporary files after workbook is saved
            for chart_path in chart_paths:
                if os.path.exists(chart_path):
                    try:
                        os.remove(chart_path)
                    except OSError:
                        pass

            # Remove temp directory if it exists
            if temp_dir and os.path.exists(temp_dir):
                try:
                    os.rmdir(temp_dir)
                except OSError:
                    pass

    @staticmethod
    def _add_simple_table_sheet(wb: Workbook, title: str, df: pd.DataFrame) -> None:
        ws = wb.create_sheet(title)
        headers = list(df.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"
            )
        for row_idx, row in enumerate(
            dataframe_to_rows(df, index=False, header=False), 2
        ):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        for column in ws.columns:
            max_length = 0
            letter = column[0].column_letter
            for cell in column:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:  # noqa: E722
                    pass
            ws.column_dimensions[letter].width = min(max_length + 2, 50)

    @staticmethod
    def _add_raw_data_sheet(wb: Workbook, results: List[DNSQueryResult]) -> None:
        """Add raw query results sheet."""
        ws = wb.create_sheet("Raw Data")

        data = []
        for result in results:
            data.append(
                {
                    "Resolver Name": result.resolver_name,
                    "Resolver IP": result.resolver_ip,
                    "Domain": result.domain,
                    "Record Type": result.record_type,
                    "Latency (ms)": result.latency_ms,
                    "Status": result.status.value,
                    "Answers Count": len(result.answers),
                    "TTL": result.ttl or "",
                    "Error Message": result.error_message or "",
                    "Attempts": result.attempt_number,
                    "Cached": result.cache_hit,
                    "Iteration": result.iteration,
                }
            )

        df = pd.DataFrame(data)

        # Add headers with formatting
        headers = list(df.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"
            )

        # Add data
        for row_idx, row in enumerate(
            dataframe_to_rows(df, index=False, header=False), 2
        ):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:  # noqa: E722
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    @staticmethod
    def _add_resolver_summary_sheet(wb: Workbook, analyzer: BenchmarkAnalyzer) -> None:
        """Add resolver statistics sheet."""
        ws = wb.create_sheet("Resolver Summary")

        resolver_stats = analyzer.get_resolver_statistics()

        data = []
        for stats in resolver_stats:
            data.append(
                {
                    "Resolver Name": stats.resolver_name,
                    "Resolver IP": stats.resolver_ip,
                    "Total Queries": stats.total_queries,
                    "Successful Queries": stats.successful_queries,
                    "Success Rate (%)": round(stats.success_rate, 2),
                    "Min Latency (ms)": round(stats.min_latency, 2),
                    "Avg Latency (ms)": round(stats.avg_latency, 2),
                    "Median Latency (ms)": round(stats.median_latency, 2),
                    "Max Latency (ms)": round(stats.max_latency, 2),
                    "Std Dev (ms)": round(stats.std_latency, 2),
                    "P95 Latency (ms)": round(stats.p95_latency, 2),
                    "P99 Latency (ms)": round(stats.p99_latency, 2),
                    "Jitter": round(stats.jitter, 2),
                    "Consistency": round(stats.consistency_score, 2),
                }
            )

        df = pd.DataFrame(data)

        # Add headers
        headers = list(df.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"
            )

        # Add data
        for row_idx, row in enumerate(
            dataframe_to_rows(df, index=False, header=False), 2
        ):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:  # noqa: E722
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    @staticmethod
    def _add_charts_sheet(
        wb: Workbook, analyzer: BenchmarkAnalyzer, temp_dir: str
    ) -> List[str]:
        """Add a sheet with embedded chart images.

        Returns:
            List of chart file paths that need to be cleaned up later.
        """
        from openpyxl.drawing.image import Image as XLImage

        ws = wb.create_sheet("Charts")

        # Generate charts
        latency_chart_path = ExcelExporter._generate_latency_chart_for_excel(
            analyzer, temp_dir
        )
        success_chart_path = ExcelExporter._generate_success_chart_for_excel(
            analyzer, temp_dir
        )

        # Add title
        ws["A1"] = "DNS Resolver Performance Charts"
        ws["A1"].font = Font(bold=True, size=14)

        # Add latency chart
        ws["A3"] = "Average Latency Comparison"
        ws["A3"].font = Font(bold=True, size=12)
        img1 = XLImage(latency_chart_path)
        img1.width = 600
        img1.height = 360
        ws.add_image(img1, "A4")

        # Add success rate chart
        ws["A24"] = "Success Rate Comparison"
        ws["A24"].font = Font(bold=True, size=12)
        img2 = XLImage(success_chart_path)
        img2.width = 600
        img2.height = 360
        ws.add_image(img2, "A25")

        # Return paths for cleanup after workbook is saved
        return [latency_chart_path, success_chart_path]

    @staticmethod
    def _generate_latency_chart_for_excel(
        analyzer: BenchmarkAnalyzer, output_dir: str
    ) -> str:
        """Generate latency chart for Excel embedding."""
        resolver_stats = analyzer.get_resolver_statistics()
        valid_resolvers = [s for s in resolver_stats if s.successful_queries > 0]

        fig, ax = plt.subplots(figsize=(10, 6))

        if not valid_resolvers:
            ax.text(
                0.5, 0.5, "No successful queries", ha="center", va="center", fontsize=14
            )
            ax.axis("off")
        else:
            names = [s.resolver_name for s in valid_resolvers]
            avg_latencies = [s.avg_latency for s in valid_resolvers]
            colors = [
                "#2ecc71" if lat < 50 else "#f39c12" if lat < 100 else "#e74c3c"
                for lat in avg_latencies
            ]

            bars = ax.bar(range(len(names)), avg_latencies, color=colors)
            ax.set_xticks(range(len(names)))
            ax.set_xticklabels(names, rotation=45, ha="right")
            ax.set_ylabel("Average Latency (ms)")
            ax.set_title("DNS Resolver Performance Comparison")

            # Add value labels on bars
            for bar in bars:
                height = bar.get_height()
                ax.annotate(
                    f"{height:.1f}",
                    xy=(bar.get_x() + bar.get_width() / 2, height),
                    xytext=(0, 3),
                    textcoords="offset points",
                    ha="center",
                    va="bottom",
                )

            ax.grid(True, alpha=0.3, axis="y")

        plt.tight_layout()
        chart_path = os.path.join(output_dir, "excel_latency_chart.png")
        plt.savefig(chart_path, dpi=150, bbox_inches="tight")
        plt.close()

        return chart_path

    @staticmethod
    def _generate_success_chart_for_excel(
        analyzer: BenchmarkAnalyzer, output_dir: str
    ) -> str:
        """Generate success rate chart for Excel embedding."""
        resolver_stats = analyzer.get_resolver_statistics()
        names = [s.resolver_name for s in resolver_stats]
        rates = [s.success_rate for s in resolver_stats]
        colors = [
            "#2ecc71" if r > 95 else "#f39c12" if r > 80 else "#e74c3c" for r in rates
        ]

        fig, ax = plt.subplots(figsize=(10, 6))
        bars = ax.bar(range(len(names)), rates, color=colors)
        ax.set_xticks(range(len(names)))
        ax.set_xticklabels(names, rotation=45, ha="right")
        ax.set_ylabel("Success Rate (%)")
        ax.set_title("DNS Resolver Success Rates")
        # ax.set_ylim(0, 100)

        # Add value labels
        for bar in bars:
            height = bar.get_height()
            ax.annotate(
                f"{height:.1f}%",
                xy=(bar.get_x() + bar.get_width() / 2, height),
                xytext=(0, 3),
                textcoords="offset points",
                ha="center",
                va="bottom",
            )

        ax.grid(True, alpha=0.3, axis="y")

        plt.tight_layout()
        chart_path = os.path.join(output_dir, "excel_success_chart.png")
        plt.savefig(chart_path, dpi=150, bbox_inches="tight")
        plt.close()

        return chart_path


class PDFExporter:
    """Export DNS benchmark results to PDF format."""

    @staticmethod
    def export_results(
        results: List[DNSQueryResult],
        analyzer: BenchmarkAnalyzer,
        output_path: str,
        include_success_chart: bool = False,
    ) -> None:
        charts_dir = tempfile.mkdtemp()

        try:
            # Generate charts
            latency_chart_path = PDFExporter._generate_latency_chart(
                analyzer, charts_dir
            )
            success_chart_path = (
                PDFExporter._generate_success_rate_chart(analyzer, charts_dir)
                if include_success_chart
                else None
            )

            # Convert images to base64 for embedding
            import base64

            with open(latency_chart_path, "rb") as f:
                latency_b64 = base64.b64encode(f.read()).decode("utf-8")
            success_b64 = None
            if success_chart_path:
                with open(success_chart_path, "rb") as f:
                    success_b64 = base64.b64encode(f.read()).decode("utf-8")

            # Generate HTML content with base64 images
            html_content = PDFExporter._generate_html_content(
                analyzer, latency_b64, success_b64
            )
            # Write PDF before cleaning up temp files
            HTML(string=html_content).write_pdf(output_path)
        finally:
            # Clean up temporary files after PDF is written
            for p in [latency_chart_path, success_chart_path]:
                if p and os.path.exists(p):
                    try:
                        os.remove(p)
                    except OSError:
                        pass

            if os.path.exists(charts_dir):
                try:
                    os.rmdir(charts_dir)
                except OSError:
                    pass

    @staticmethod
    def _generate_latency_chart(analyzer: BenchmarkAnalyzer, output_dir: str) -> str:
        """Generate latency comparison bar chart."""
        resolver_stats = analyzer.get_resolver_statistics()
        valid_resolvers = [s for s in resolver_stats if s.successful_queries > 0]

        fig, ax = plt.subplots(figsize=(10, 6))
        if not valid_resolvers:
            ax.text(
                0.5, 0.5, "No successful queries", ha="center", va="center", fontsize=14
            )
            ax.axis("off")
        else:
            names = [s.resolver_name for s in valid_resolvers]
            avg_latencies = [s.avg_latency for s in valid_resolvers]
            colors = [
                "#2ecc71" if latency < 50 else "#f39c12" if latency < 100 else "#e74c3c"
                for latency in avg_latencies
            ]
            bars = ax.bar(range(len(names)), avg_latencies, color=colors)
            ax.set_xticks(range(len(names)))
            ax.set_xticklabels(names, rotation=45, ha="right")
            ax.set_ylabel("Average Latency (ms)")
            ax.set_title("DNS Resolver Performance Comparison")
            for bar in bars:
                h = bar.get_height()
                ax.annotate(
                    f"{h:.1f}",
                    xy=(bar.get_x() + bar.get_width() / 2, h),
                    xytext=(0, 3),
                    textcoords="offset points",
                    ha="center",
                    va="bottom",
                )
            ax.grid(True, alpha=0.3, axis="y")

        plt.tight_layout()
        chart_path = os.path.join(output_dir, "latency_comparison.png")
        plt.savefig(chart_path, dpi=150, bbox_inches="tight")
        plt.close()
        return chart_path

    @staticmethod
    def _generate_success_rate_chart(
        analyzer: BenchmarkAnalyzer, output_dir: str
    ) -> str:
        """Generate success rate chart."""
        resolver_stats = analyzer.get_resolver_statistics()
        names = [s.resolver_name for s in resolver_stats]
        rates = [s.success_rate for s in resolver_stats]
        colors = [
            "#2ecc71" if r > 95 else "#f39c12" if r > 80 else "#e74c3c" for r in rates
        ]

        fig, ax = plt.subplots(figsize=(10, 6))
        bars = ax.bar(range(len(names)), rates, color=colors)
        ax.set_xticks(range(len(names)))
        ax.set_xticklabels(names, rotation=45, ha="right")
        ax.set_ylabel("Success Rate (%)")
        ax.set_title("DNS Resolver Success Rates")
        # ax.set_ylim(0, 100)
        for bar in bars:
            h = bar.get_height()
            ax.annotate(
                f"{h:.1f}%",
                xy=(bar.get_x() + bar.get_width() / 2, h),
                xytext=(0, 3),
                textcoords="offset points",
                ha="center",
                va="bottom",
            )
        ax.grid(True, alpha=0.3, axis="y")

        plt.tight_layout()
        chart_path = os.path.join(output_dir, "success_rates.png")
        plt.savefig(chart_path, dpi=150, bbox_inches="tight")
        plt.close()
        return chart_path

    @staticmethod
    def _generate_html_content(
        analyzer: BenchmarkAnalyzer,
        latency_chart_b64: str,
        success_chart_b64: Optional[str] = None,
    ) -> str:
        """Generate HTML content for PDF report."""
        resolver_stats = analyzer.get_resolver_statistics()
        overall_stats = analyzer.get_overall_statistics()
        ranked_resolvers = sorted(
            [s for s in resolver_stats if s.successful_queries > 0],
            key=lambda x: x.avg_latency,
        )

        from datetime import datetime

        current_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Optional success chart block
        success_block = ""
        if success_chart_b64:
            success_block = f"""
            <div class="section">
                <h2>Success Rates</h2>
                <div class="chart">
                    <img src="data:image/png;base64,{success_chart_b64}" alt="DNS Resolver Success Rates">
                    <p><em>Success rate comparison across DNS resolvers (higher is better)</em></p>
                </div>
            </div>
            """

        template_str = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <title>DNS Benchmark Report</title>
            <style>
                body {{
                    font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                    margin: 0;
                    padding: 20px;
                    color: #333;
                    line-height: 1.6;
                }}
                .header {{
                    text-align: center;
                    border-bottom: 3px solid #2c3e50;
                    padding-bottom: 20px;
                    margin-bottom: 30px;
                }}
                .section {{
                    margin-bottom: 40px;
                }}
                table {{
                    width: 100%;
                    border-collapse: collapse;
                    margin: 20px 0;
                    font-size: 0.9em;
                }}
                th, td {{
                    padding: 12px 15px;
                    text-align: left;
                    border-bottom: 1px solid #ddd;
                }}
                th {{
                    background-color: #34495e;
                    color: white;
                }}
                tr:nth-child(even) {{
                    background-color: #f8f9fa;
                }}
                .chart {{
                    text-align: center;
                    margin: 30px 0;
                }}
                .chart img {{
                    max-width: 100%;
                    height: auto;
                    border: 1px solid #ddd;
                    border-radius: 5px;
                }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>DNS Benchmark Report</h1>
                <p>Generated on: {current_date}</p>
            </div>

            <div class="section">
                <h2>Executive Summary</h2>
                <p><strong>Total queries:</strong> {overall_stats['total_queries']}</p>
                <p><strong>Successful:</strong> {overall_stats['successful_queries']} ({overall_stats['overall_success_rate']:.1f}%)</p>
                <p><strong>Average latency:</strong> {overall_stats['overall_avg_latency']:.1f} ms</p>
                <p><strong>Median latency:</strong> {overall_stats['overall_median_latency']:.1f} ms</p>
                <p><strong>Resolvers tested:</strong> {overall_stats['resolver_count']}</p>
                <p><strong>Domains tested:</strong> {overall_stats['domain_count']}</p>
                <p><strong>Fastest resolver:</strong> {overall_stats['fastest_resolver']}</p>
                <p><strong>Slowest resolver:</strong> {overall_stats['slowest_resolver']}</p>
            </div>

            <div class="section">
                <h2>Latency Comparison</h2>
                <div class="chart">
                    <img src="data:image/png;base64,{latency_chart_b64}" alt="Latency Comparison">
                </div>
            </div>
            {success_block}
            <div class="section">
            <h2>Resolver Rankings</h2>
            <table>
                <tr>
                    <th>Rank</th>
                    <th>Resolver</th>
                    <th>Avg Latency (ms)</th>
                    <th>Success Rate (%)</th>
                    <th>Queries</th>
                </tr>
                {''.join(
                    f"<tr><td>{i + 1}</td><td>{r.resolver_name}</td>"
                    f"<td>{r.avg_latency:.1f}</td><td>{r.success_rate:.1f}%</td>"
                    f"<td>{r.successful_queries}/{r.total_queries}</td></tr>"
                    for i, r in enumerate(ranked_resolvers)
                )}
            </table>
            </div>

        <div class="section">
            <h2>Detailed Statistics</h2>
            <table>
                <tr>
                    <th>Resolver</th>
                    <th>IP Address</th>
                    <th>Min (ms)</th>
                    <th>Avg (ms)</th>
                    <th>Max (ms)</th>
                    <th>Std Dev</th>
                    <th>P95 (ms)</th>
                    <th>Success Rate (%)</th>
                </tr>
                {''.join(
                    f"<tr><td>{r.resolver_name}</td><td>{r.resolver_ip}</td>"
                    f"<td>{r.min_latency:.1f}</td><td>{r.avg_latency:.1f}</td>"
                    f"<td>{r.max_latency:.1f}</td><td>{r.std_latency:.1f}</td>"
                    f"<td>{r.p95_latency:.1f}</td><td>{r.success_rate:.1f}%</td></tr>"
                    for r in resolver_stats
                )}
            </table>
        </div>
        </body>
        </html>
        """
        return template_str
