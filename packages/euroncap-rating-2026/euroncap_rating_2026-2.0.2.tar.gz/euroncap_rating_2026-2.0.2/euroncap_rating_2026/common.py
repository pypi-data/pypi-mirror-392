from functools import wraps
from euroncap_rating_2026.version import VERSION
import logging
import pandas as pd
import random
import numpy as np
from collections import Counter
import openpyxl
from openpyxl.utils import get_column_letter
import copy

logger = logging.getLogger(__name__)


def with_footer(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        result = func(*args, **kwargs)

        # Print footer after all command output
        footer_lines = [
            " " * 40,
            f"Generated with version {VERSION}",
            " " * 40,
            "Copyright 2025, Euro NCAP IVZW\nCreated by IVEX NV (https://ivex.ai)",
        ]
        for line in footer_lines:
            logger.info(line)
            print(line)
        return result

    return wrapper


def read_excel_file_to_dfs(input_file):
    # Load the Excel file
    xls = pd.ExcelFile(input_file)

    sheet_names = xls.sheet_names
    if not sheet_names:
        raise ValueError("The Excel file does not contain any sheets.")
    dfs = {}

    for sheet_name in sheet_names:
        dfs[sheet_name] = pd.read_excel(xls, sheet_name=sheet_name, header=0)

    xls.close()

    return dfs


def frequency_proportional_sample(data, N, seed=None):
    """
    Sample N (row, col, color) tuples from data, using color frequency for proportional sampling.
    Keeps (row, col) in the result.
    """
    if seed is not None:
        random.seed(seed)

    if N >= len(data):
        return data.copy()
    # Count color frequencies
    color_counts = Counter([color for _, _, color in data])
    total = sum(color_counts.values())

    # Compute ideal proportions
    proportions = {color: count / total for color, count in color_counts.items()}
    logger.info(f"proportions: {proportions}")

    # Initial rounded down count
    raw_counts = {}
    for color, p in proportions.items():
        count = int(np.floor(p * N))
        raw_counts[color] = count
    logger.debug(f"raw_counts pre remaining distribution: {raw_counts}")
    # Distribute remaining samples
    remaining = N - sum(raw_counts.values())
    residuals = {
        color: (proportions[color] * N - raw_counts[color]) for color in color_counts
    }
    for color in sorted(residuals, key=residuals.get, reverse=True)[:remaining]:
        raw_counts[color] += 1
    logger.debug(f"raw_counts post remaining distribution: {raw_counts}")

    # Sample actual items, keeping (row, col, color)
    sampled = []
    for color, count in raw_counts.items():
        choices = [x for x in data if x[2] == color]
        sampled += random.sample(choices, min(len(choices), count))

    return sampled


def get_initials(text):
    if pd.isna(text):
        return ""
    return "".join([word[0].upper() for word in str(text).split()])


def get_first_word(text):
    if pd.isna(text):
        return ""
    return str(text).split()[0]


def get_param_df(dfs):

    # Read the "Input parameters" sheet into a DataFrame and remove it from dfs
    if "Input parameters" in dfs:
        input_parameters_df = dfs.pop("Input parameters")
    else:
        input_parameters_df = None

    if input_parameters_df is None:
        raise ValueError(
            "The Excel file does not contain the required 'Input parameters' sheet."
        )
    # Extract input parameters with their corresponding stage, stage element, and stage subelement
    input_parameters = []
    current_stage = None
    current_stage_element = None
    current_stage_subelement = None

    for _, row in input_parameters_df.iterrows():
        if not pd.isna(row.get("Stage")):
            current_stage = row["Stage"]
        if not pd.isna(row.get("Stage element")):
            current_stage_element = row["Stage element"]
        if not pd.isna(row.get("Stage Subelement")):
            current_stage_subelement = row["Stage Subelement"]
        input_param = row.get("Input parameter")
        input_value = row.get("Value")
        if not pd.isna(input_param) and not pd.isna(input_value):
            input_parameters.append(
                {
                    "Input parameter": input_param,
                    "Stage": current_stage,
                    "Stage element": current_stage_element,
                    "Stage Subelement": current_stage_subelement,
                    "Value": input_value,
                }
            )
    # Create a DataFrame of input parameters
    param_df = pd.DataFrame(input_parameters)

    param_df["param_code"] = (
        param_df["Stage"].apply(get_initials)
        + " - "
        + param_df["Stage element"].apply(get_first_word)
        + " "
        + param_df["Stage Subelement"].astype(str)
    )
    param_df = param_df[["param_code", "Input parameter", "Value"]]
    logger.debug(f"Extracted param_df: {param_df}")
    return param_df


def hard_copy_sheet(input_file: str, sheet_name: str, output_file: str):
    # Load input and output workbooks
    in_wb = openpyxl.load_workbook(input_file)
    out_wb = openpyxl.load_workbook(output_file)

    if sheet_name not in in_wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in {input_file}")

    in_ws = in_wb[sheet_name]
    # Remove sheet if it already exists in output
    if sheet_name in out_wb.sheetnames:
        std = out_wb[sheet_name]
        out_wb.remove(std)
    out_ws = out_wb.create_sheet(sheet_name)

    # Copy row heights
    for row in in_ws.row_dimensions:
        out_ws.row_dimensions[row].height = in_ws.row_dimensions[row].height

    # Copy cell values, styles, and fills
    for row in in_ws.iter_rows():
        for cell in row:
            new_cell = out_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy.copy(cell.font)
                new_cell.border = copy.copy(cell.border)
                new_cell.fill = copy.copy(cell.fill)
                new_cell.number_format = copy.copy(cell.number_format)
                new_cell.protection = copy.copy(cell.protection)
                new_cell.alignment = copy.copy(cell.alignment)

    # Copy merged cells
    for merged_range in in_ws.merged_cells.ranges:
        out_ws.merge_cells(str(merged_range))

    # Copy column widths
    # Set column widths for columns 0 to 25 (i.e., columns A to Z)
    for col_idx in range(0, 26):
        col_letter = get_column_letter(col_idx + 1)  # openpyxl is 1-based
        if (
            col_letter in in_ws.column_dimensions
            and in_ws.column_dimensions[col_letter].width is not None
        ):
            out_ws.column_dimensions[col_letter].width = in_ws.column_dimensions[
                col_letter
            ].width
    out_wb.save(output_file)
