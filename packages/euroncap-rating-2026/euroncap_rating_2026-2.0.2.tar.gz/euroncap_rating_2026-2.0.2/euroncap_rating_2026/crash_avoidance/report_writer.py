import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from euroncap_rating_2026.crash_avoidance.matrix_processing import (
    PREDICTION_COLOR_MAP,
    PredictionColor,
)
from euroncap_rating_2026.crash_avoidance import data_model

from euroncap_rating_2026.crash_avoidance.test_info import (
    get_sheet_prefix,
    get_stage_subelement_key,
)

import logging
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter

header_align_right = [
    "Score",
    "Max Score",
    "OEM.Prediction",
    "Value",
    "Prediction.Check",
    "Colour",
]
header_align_left = [
    "Test",
    "Seat position",
    "Body Region",
    "Criteria",
    "Stage",
    "Stage element",
    "Stage subelement",
    "Dummy",
]


CM_TO_EXCEL_WIDTH = 3.78


logger = logging.getLogger(__name__)


def is_empty_cell(cell_value):
    """
    Checks if a cell value is considered empty.

    Args:
        cell_value: The value of the cell to check.

    Returns:
        bool: True if the cell is empty, False otherwise.
    """
    return (
        pd.isna(cell_value)
        or cell_value is None
        or cell_value == "None"
        or cell_value == "nan"
    )


def adjust_column_widths(columns):
    """
    Adjusts the width of columns in an Excel worksheet based on the maximum length of cell values.

    Args:
        columns: Iterable of worksheet columns (e.g., ws.columns).
    """
    for column in columns:
        max_length = 0
        if not hasattr(column[0], "column_letter"):
            continue
        column_letter = column[0].column_letter  # Get the column letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except (TypeError, AttributeError) as e:
                logger.warning(f"Error processing cell value '{cell.value}': {e}")
        adjusted_width = max_length + 2  # Add some padding
        column[0].parent.column_dimensions[column_letter].width = adjusted_width


def write_report(
    score_df_dict, selected_points_dict, output_file, format_prediction_cells=False
):
    # Write the updated copied sheets back to the output file
    for sheet_name, updated_df in score_df_dict.items():
        try:
            with pd.ExcelWriter(
                output_file, engine="openpyxl", mode="a", if_sheet_exists="replace"
            ) as writer:
                updated_df.to_excel(writer, sheet_name=sheet_name, index=False)
        except Exception as e:
            logger.error(f"Failed to write updated sheet {sheet_name}: {e}")

    wb = load_workbook(output_file)

    calculation_color = PatternFill(
        start_color="FFDD04", end_color="FFDD04", fill_type="solid"
    )
    input_color = PatternFill(
        start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"
    )

    sheetnames = wb.sheetnames

    for sheet_name in sheetnames:
        ws = wb[sheet_name]

        # Move "Value" column to the end if sheet_name contains "Verification"
        if "Verification" in sheet_name:
            headers = [cell.value for cell in ws[1]]
            if "Value" in headers:
                value_col_idx = headers.index("Value") + 1  # openpyxl is 1-based
                max_col = ws.max_column
                # If "Value" is not already at the end
                if value_col_idx != max_col:
                    # Extract "Value" column values
                    value_col = [
                        ws.cell(row=row, column=value_col_idx).value
                        for row in range(1, ws.max_row + 1)
                    ]
                    # Shift columns left to fill the gap
                    for col in range(value_col_idx, max_col):
                        for row in range(1, ws.max_row + 1):
                            ws.cell(row=row, column=col).value = ws.cell(
                                row=row, column=col + 1
                            ).value
                        # Place "Value" column at the end
                        for row in range(1, ws.max_row + 1):
                            ws.cell(row=row, column=max_col).value = value_col[row - 1]
        # Remove duplicate "Value" columns if present
        headers = [cell.value for cell in ws[1]]
        value_indices = [
            i + 1 for i, h in enumerate(headers) if h == "Value"
        ]  # 1-based
        if len(value_indices) > 1:
            # Keep the first "Value" column, remove others
            for idx in reversed(value_indices[1:]):
                ws.delete_cols(idx)
        if "Robustness" in sheet_name:
            for col in ws.iter_cols():
                for cell in col[1:]:
                    if isinstance(cell.value, str):
                        if cell.value.strip().upper() == "YES":
                            cell.fill = PatternFill(
                                start_color="E8F2A1",
                                end_color="E8F2A1",
                                fill_type="solid",
                            )  # Light green
                        elif cell.value.strip().upper() == "NO":
                            cell.fill = PatternFill(
                                start_color="FFA6A6",
                                end_color="FFA6A6",
                                fill_type="solid",
                            )  # Light red

        ###############################################################
        # Align text
        ###############################################################
        for col in ws.iter_cols():
            header = col[0].value
            align = Alignment(horizontal="center")
            if header in header_align_right:
                align = Alignment(horizontal="right")
            elif header in header_align_left:
                align = Alignment(horizontal="left")

            for cell in col:
                cell.alignment = align

            if col[0].value in ["Colour", "OEM.Prediction"]:
                for cell in col[1:]:
                    if (
                        isinstance(cell.value, str)
                        and cell.value
                        and str(cell.value).lower() not in ["nan", "none"]
                    ):
                        cell.value = cell.value.capitalize()

        ###############################################################
        # Set number precision for specific columns
        ###############################################################
        for col in ws.iter_cols():
            header = col[0].value
            if header in [
                "Inspection [%]",
                "Value",
                "Capping",
                "Body regionscore",
                "Modifiers",
            ]:
                for cell in col[1:]:
                    if isinstance(cell.value, (int, float)) and not is_empty_cell(
                        cell.value
                    ):
                        cell.number_format = "0.00"
            elif header in ["Score", "Max Score", "Max score"]:
                for cell in col[1:]:
                    if isinstance(cell.value, (int, float)) and not is_empty_cell(
                        cell.value
                    ):
                        if ws.title == "Scenario Scores":
                            cell.number_format = "0.000"
                        elif ws.title not in score_df_dict.keys():
                            cell.number_format = "0.00"
                        elif ws.title == "Test Scores":
                            cell.number_format = "0.000"
                        else:
                            cell.number_format = "0.0000"

        ###############################################################
        # Color columns based on their headers
        ###############################################################
        for col in ws.iter_cols():
            if col[0].value in [
                "Score",
                "Points",
                "Prediction.Check",
                "Body regionscore",
                "Modifiers",
                "Modifier",
                "Colour",
                "Capping?",
            ]:
                for cell in col[1:]:  # Skip the first row (header)
                    cell.fill = calculation_color
                    if is_empty_cell(cell.value):
                        cell.value = ""
            elif col[0].value in ["Inspection [%]", "Value", "OEM.Prediction"]:
                for cell in col[1:]:  # Skip the first row (header)
                    if (
                        sheet_name
                        in [
                            "FC - Car & PTW Verification",
                            "FC - Ped & Cyc Verification",
                            "LDC - Single Veh Verification",
                            "LDC - Car & PTW Verification",
                            "LSC - Car & PTW Verification",
                            "LSC - Ped & Cyc Verification",
                        ]
                        and col[0].value == "OEM.Prediction"
                    ):
                        logger.debug(
                            f"Skipping OEM.Prediction color fill for sheet {sheet_name} and column {col[0].value}"
                        )
                    else:
                        cell.fill = input_color
                    if is_empty_cell(cell.value):
                        cell.value = ""

    ###############################################################
    # Adjust column widths based on content
    ###############################################################
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        if "Prediction" in sheet_name:
            for i, column in enumerate(ws.columns):
                if i < 4:
                    adjust_column_widths([column])
                else:
                    col_letter = get_column_letter(i)
                    ws.column_dimensions[col_letter].width = 2.25 * CM_TO_EXCEL_WIDTH
        else:
            adjust_column_widths(ws.columns)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        logger.debug(f"Adding borders to sheet: {sheet_name}")
        # Add thick border to the last column for all rows including header
        for row in range(1, ws.max_row + 1):  # Start from 1 to include header
            cell = ws.cell(row=row, column=ws.max_column)
            cell.border = cell.border + Border(
                right=Side(border_style="thick", color="000000")
            )
        # Add thick border to the last row for all columns
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.border = cell.border + Border(
                bottom=Side(border_style="thick", color="000000")
            )

    # Set font size to 11 and center alignment for all cells in the VRU Prediction sheet
    font_large = Font(name="Calibri", size=11)
    center_align = Alignment(horizontal="center", vertical="center")

    if format_prediction_cells:
        for index, stage_info in enumerate(data_model.STAGE_SUBELEMENTS):
            stage_element = stage_info["Stage element"]
            stage_subelement = stage_info["Stage subelement"]
            sheet_prefix = get_sheet_prefix(stage_element, stage_subelement)
            stage_subelement_key = get_stage_subelement_key(stage_element)
            prediction_sheet_name = f"{sheet_prefix} Prediction"

            if prediction_sheet_name not in [
                "FC - Car & PTW Prediction",
                "FC - Ped & Cyc Prediction",
                "LDC - Single Veh Prediction",
                "LDC - Car & PTW Prediction",
                "LSC - Car & PTW Prediction",
                "LSC - Ped & Cyc Prediction",
            ]:
                continue

            pred_ws = wb[prediction_sheet_name]
            # Remove the value of stage_subelement_key + " - " from sheet_prefix
            prefix_to_remove = f"{stage_subelement_key.value} - "
            if sheet_prefix.startswith(prefix_to_remove):
                subdict_key = sheet_prefix[len(prefix_to_remove) :]

            loadcase_list = data_model.STAGE_SUBELEMENT_TO_LOADCASES[
                stage_subelement_key
            ][subdict_key]
            for scenario in loadcase_list:
                for point in selected_points_dict.get(scenario, []):
                    logger.debug(
                        f"Selected point - color: {PREDICTION_COLOR_MAP.get(str(point.color).lower(), None)}, "
                        f"coordinates: (row={point.row}, col={point.col})"
                    )
                matrix_indices = data_model.MATRIX_INDICES.get(scenario, None)

                for row_index, row in enumerate(pred_ws.iter_rows()):
                    for col_index, cell in enumerate(row):
                        key = str(cell.value).lower()
                        if (
                            matrix_indices
                            and matrix_indices["start_row"] + 1
                            <= row_index
                            <= matrix_indices["start_row"] + matrix_indices["n_rows"]
                            and matrix_indices["start_col"]
                            <= col_index
                            < matrix_indices["start_col"] + matrix_indices["n_cols"]
                        ):
                            logger.debug(
                                f"Processing cell at row {row_index + 1}, col {col_index + 1} key: {key}"
                            )
                            if key in PREDICTION_COLOR_MAP:
                                color = PREDICTION_COLOR_MAP[key]
                            else:
                                color = None
                            logger.debug(
                                f"Setting cell at row {row_index + 1}, col {col_index + 1} to color: {color}"
                            )
                            if color:
                                # Convert float RGB (0-1) to int (0-255) and then to hex
                                color_int = tuple(int(round(c * 255)) for c in color)
                                color_hex = "{:02X}{:02X}{:02X}".format(*color_int)
                                fill = PatternFill(
                                    start_color=color_hex,
                                    end_color=color_hex,
                                    fill_type="solid",
                                )
                                cell.fill = fill
                            if not isinstance(cell, MergedCell):
                                cell.value = ""
                            # Iterate over selected_points and mark them using data_model with offsets

                if scenario not in selected_points_dict:
                    logger.info(
                        f"No selected points found for scenario {scenario} in selected_points_dict."
                    )
                    continue
                for point in selected_points_dict[scenario]:
                    cell_to_mark = pred_ws.cell(
                        row=point.row + matrix_indices["start_row"] + 2,
                        column=point.col + matrix_indices["start_col"] + 1,
                    )
                    logger.debug(f"Marking point at row={point.row}, col={point.col}")
                    logger.debug(
                        f"Updated coordinates = row={cell_to_mark.row}, col={cell_to_mark.column}"
                    )
                    cell_to_mark.value = "X"
                    cell_to_mark.alignment = Alignment(
                        horizontal="center", vertical="center"
                    )

                    cell_to_mark.font = font_large
                    cell_to_mark.alignment = center_align

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ###############################################################
        # Set first row (header) to black background and white text
        ###############################################################
        header_fill = PatternFill(
            start_color="000000", end_color="000000", fill_type="solid"
        )
        header_font = Font(name="Calibri", color="FFFFFF", bold=True)
        # Original header formatting for first row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Remove everything until the first " - " occurrence, included
        stage_subelement = sheet_name.replace(" Prediction", "")
        if " - " in stage_subelement:
            parts = stage_subelement.split(" - ", 1)
            stage_subelement = parts[1]
            prefix = parts[0]
        else:
            stage_subelement = None
            prefix = None
        scenario_list = data_model.STAGE_SUBELEMENT_TO_LOADCASES.get(prefix, {}).get(
            stage_subelement, []
        )

        if len(scenario_list) == 0:
            continue

        # Apply header formatting to row-1 and row-2 of every start_row in MATRIX_INDICES (if sheet matches)
        for scenario, indices in data_model.MATRIX_INDICES.items():
            # Only apply if this sheet is relevant (sheet_name contains scenario or is a prediction sheet)
            if any(scenario in s for s in scenario_list):
                for offset in [0, 1]:  # row-1 and row-2
                    row_num = indices["start_row"] - offset + 1  # openpyxl is 1-based
                    if row_num > 0 and row_num <= ws.max_row:
                        for cell in ws[row_num]:
                            cell.fill = header_fill
                            cell.font = header_font
                last_num = indices["start_row"] - 1  # openpyxl is 1-based
                if last_num > 0 and last_num <= ws.max_row:
                    for cell in ws[last_num]:
                        cell.value = ""
                        cell.border = cell.border + Border(
                            top=Side(border_style="thick", color="000000")
                        )
                        cell.border = Border(
                            left=cell.border.left,
                            top=cell.border.top,
                            bottom=cell.border.bottom,
                            right=None,
                        )

    # Reorder sheets according to the specified order
    def reorder_sheets(workbook):
        # Define the main order
        main_order = [
            "Input parameters",
            "Test Scores",
            "Category Scores",
            "Scenario Scores",
        ]
        # Define the prefixes for the remaining groups
        prefixes = [
            "FC - Car & PTW",
            "FC - Ped & Cyc",
            "LDC - Single Veh",
            "LDC - Car & PTW",
            "LSC - Car & PTW",
            "LSC - Ped & Cyc",
        ]
        suffixes = ["Prediction", "Robustness", "Verification"]

        # Build the desired order
        desired_order = []
        # Add main sheets if present
        for sheet in main_order:
            if sheet in workbook.sheetnames:
                desired_order.append(sheet)
        # Add grouped sheets in the specified order
        for prefix in prefixes:
            for suffix in suffixes:
                sheet_name = f"{prefix} {suffix}"
                if sheet_name in workbook.sheetnames:
                    desired_order.append(sheet_name)
        # Insert "LDC - Robustness" after all LDC sheets if present
        ldc_robustness = "LDC - Robustness"
        if ldc_robustness in workbook.sheetnames:
            # Find the last LDC sheet in desired_order
            ldc_indices = [
                i for i, name in enumerate(desired_order) if name.startswith("LDC - ")
            ]
            if ldc_indices:
                insert_pos = max(ldc_indices) + 1
            else:
                insert_pos = len(desired_order)
            # Remove if already present
            if ldc_robustness in desired_order:
                desired_order.remove(ldc_robustness)
            desired_order.insert(insert_pos, ldc_robustness)
        # Add any remaining sheets that were not specified
        for sheet in workbook.sheetnames:
            if sheet not in desired_order:
                desired_order.append(sheet)

        # Reorder sheets in workbook
        for idx, sheet_name in enumerate(desired_order):
            if workbook.sheetnames[idx] != sheet_name:
                sheet = workbook[sheet_name]
                workbook._sheets.insert(
                    idx, workbook._sheets.pop(workbook._sheets.index(sheet))
                )

    reorder_sheets(wb)
    wb.save(output_file)
