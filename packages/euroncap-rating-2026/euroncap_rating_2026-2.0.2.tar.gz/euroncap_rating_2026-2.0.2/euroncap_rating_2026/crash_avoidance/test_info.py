import random
import pandas as pd
import logging
from dataclasses import dataclass
import openpyxl
from euroncap_rating_2026.crash_avoidance import matrix_processing
from euroncap_rating_2026.crash_avoidance import data_model
from euroncap_rating_2026.crash_avoidance import robustness_layer
import euroncap_rating_2026.crash_avoidance.matrix_processing as matrix_processing


logger = logging.getLogger(__name__)


@dataclass
class StageSubelement:
    name: str
    selected_points_dict: dict
    selected_robustness_dict: dict
    test_points_df: None


@dataclass
class SelectionInfo:
    test_name: str
    standard_points: list
    extended_points: list
    selected_robustness_layer: robustness_layer.RobustnessLayer = None
    assessment_criteria: robustness_layer.AssessmentCriteria = None
    verification_condition: str = None


@dataclass
class LoadcaseInfo:
    test_name: str
    test_points: list
    robustness_layers: dict
    num_extended_tests: int = 2
    num_robustness_tests: int = 1
    score_extended_range: float = 0.0
    score_standard_range: float = 0.0
    robustness_layer_score: float = 0.0

    @property
    def num_standard_tests(self):
        return data_model.STANDARD_RANGE_VERIFICATION_TEST_NUM.get(self.test_name, 0)

    def select_test_points(self):
        """
        Select test points based on the defined criteria.
        Randomly selects:
        - num_standard_tests from non-extended range test points
        - num_extended_tests from extended range test points
        - num_robustness_tests from robustness layers (if available)
        """

        # Handle subtest case: if self.test_name is in SUBTEST_TO_TEST_DICT, update num_standard_tests using _day and _night
        is_subtest = False
        if f"{self.test_name}_day" in data_model.SUBTEST_TO_TEST_DICT:
            is_subtest = True
            day_test = f"{self.test_name}_day"
            night_test = f"{self.test_name}_night"
            num_day_test = data_model.STANDARD_RANGE_VERIFICATION_TEST_NUM.get(
                day_test, 0
            )
            num_night_test = data_model.STANDARD_RANGE_VERIFICATION_TEST_NUM.get(
                night_test, 0
            )

        # Filter out test points that are red or grey
        filtered_test_points = [
            tp
            for tp in self.test_points
            if tp.color
            not in (
                matrix_processing.PredictionColor.RED,
                matrix_processing.PredictionColor.GREY,
            )
        ]

        # Select standard range test points
        if is_subtest:
            # Split standard points into day and night based on row index
            day_points = [
                tp
                for tp in filtered_test_points
                if tp.test_range == matrix_processing.TestRange.STANDARD
                and 0 <= tp.row <= 5
            ]
            night_points = [
                tp
                for tp in filtered_test_points
                if tp.test_range == matrix_processing.TestRange.STANDARD
                and 6 <= tp.row <= 11
            ]
            selected_day = random.sample(day_points, min(num_day_test, len(day_points)))
            selected_night = random.sample(
                night_points, min(num_night_test, len(night_points))
            )
            selected_standard = selected_day + selected_night
        else:
            standard_points = [
                tp
                for tp in filtered_test_points
                if tp.test_range == matrix_processing.TestRange.STANDARD
            ]
            selected_standard = random.sample(
                standard_points, min(self.num_standard_tests, len(standard_points))
            )

        # Select extended range test points
        extended_points = [
            tp
            for tp in filtered_test_points
            if tp.test_range == matrix_processing.TestRange.EXTENDED
        ]
        selected_extended = random.sample(
            extended_points, min(self.num_extended_tests, len(extended_points))
        )

        # Select one random robustness layer with value "YES"
        yes_layers = [
            layer for layer, val in self.robustness_layers.items() if val == "YES"
        ]
        selected_robustness = random.sample(
            yes_layers, min(self.num_robustness_tests, len(yes_layers))
        )

        if selected_robustness:
            # Get the verification condition for the selected robustness layer
            selected_robustness = selected_robustness[0]
            verification_condition, assessment_criteria = (
                robustness_layer.get_robustness_layer_verification_condition(
                    selected_robustness, self.test_name
                )
            )
        else:
            selected_robustness = None
            verification_condition = None
            assessment_criteria = None

        logger.debug(
            f"Selected robustness layer for {self.test_name}: {selected_robustness}, "
            f"Verification condition: {verification_condition}, "
            f"Assessment criteria: {assessment_criteria}"
        )
        return SelectionInfo(
            test_name=self.test_name,
            standard_points=selected_standard,
            extended_points=selected_extended,
            selected_robustness_layer=selected_robustness,
            assessment_criteria=assessment_criteria,
            verification_condition=verification_condition,
        )


def get_lsc_selected_points(loadcase_info):
    """
    Low speed collision point selection criteria:

    • Car & PTW Scenarios: test lowest and highest target speed, and 1 random target speed in between. In case of impact, test adjacent cases and keep testing in +10km/h increments of target speed until prediction is met.
    • CPMRCm: Test 1.00 and 2.00m gaps for all EPTc speeds. In case of impact, test 1.50m gap.
    • CPMRCs, CPMFC: Test all cases for the 25 and 75% impact location. In case of impact, test the 50% case.
    • CBNAO - SfS: Test the highest and lowest target speed, in all ‘d’ cases. In case of impact, test the mid target speed (if applicable).
    • CBDA: Test the largest and shortest gap in combination with the highest and lowest target speed. In case of the prediction not being met, test adjacent grid cells in all directions until prediction is met.
    """
    standard_points = []

    test_name = loadcase_info.test_name
    test_points = loadcase_info.test_points

    filtered_test_points = [
        tp
        for tp in test_points
        if tp.color
        not in (
            matrix_processing.PredictionColor.RED,
            matrix_processing.PredictionColor.GREY,
        )
    ]

    # CPMRCm: Test 1.00 and 2.00m gaps for all EPTc speeds.
    if test_name.startswith("CPMRCm"):
        # Test 1.00 and 2.00m gaps for all EPTc speeds.
        # For each EPTc speed, select points with Gap 1.00 and 2.00
        eptc_speeds = sorted(
            set(
                tp.attributes.get("EPTc speed")
                for tp in filtered_test_points
                if tp.test_range == matrix_processing.TestRange.STANDARD
                and tp.attributes.get("EPTc speed") is not None
            )
        )
        for speed in eptc_speeds:
            for gap in ["1.00 m", "2.00 m"]:
                for tp in filtered_test_points:
                    if (
                        tp.test_range == matrix_processing.TestRange.STANDARD
                        and tp.attributes.get("EPTc speed") == speed
                        and tp.attributes.get("Gap") == gap
                    ):
                        standard_points.append(tp)
                        break  # Only one point per speed-gap combination

    # CPMRCs, CPMFC: Test all cases for the 25 and 75% impact location.
    elif test_name.startswith("CPMRCs") or test_name.startswith("CPMFC"):
        # Test all cases for 25% and 75% impact location.
        for tp in filtered_test_points:
            if (
                tp.test_range == matrix_processing.TestRange.STANDARD
                and tp.attributes.get("Impact Location") in ["25%", "75%"]
            ):
                standard_points.append(tp)

    # CBNAO: Test the highest and lowest target speed, in all ‘d’ cases.
    elif test_name.startswith("CBNAO - SfS"):
        # Test highest, lowest, and all available target speeds, in all 'd' cases.
        speeds = sorted(
            set(
                tp.attributes.get("EBT speed")
                for tp in filtered_test_points
                if tp.test_range == matrix_processing.TestRange.STANDARD
                and tp.attributes.get("d") is not None
            )
        )
        if speeds:
            lowest, highest = speeds[0], speeds[-1]
            for tp in filtered_test_points:
                if (
                    tp.test_range == matrix_processing.TestRange.STANDARD
                    and tp.attributes.get("d") is not None
                    and tp.attributes.get("EBT speed") in [lowest, highest]
                ):
                    standard_points.append(tp)

    # CBDA: Test the largest and shortest gap in combination with the highest and lowest target speed.
    elif test_name.startswith("CBDA"):
        # Test largest and shortest gap with highest and lowest target speed.
        gaps = sorted(
            set(
                tp.attributes.get("Gap")
                for tp in filtered_test_points
                if tp.test_range == matrix_processing.TestRange.STANDARD
            )
        )
        speeds = sorted(
            set(
                tp.attributes.get("EBT speed")
                for tp in filtered_test_points
                if tp.test_range == matrix_processing.TestRange.STANDARD
            )
        )
        if gaps and speeds:
            # Select combinations of shortest/longest gap and lowest/highest EBT speed
            for gap in (gaps[0], gaps[-1]):
                for speed in (speeds[0], speeds[-1]):
                    for tp in filtered_test_points:
                        if (
                            tp.test_range == matrix_processing.TestRange.STANDARD
                            and tp.attributes.get("Gap") == gap
                            and tp.attributes.get("EBT speed") == speed
                        ):
                            standard_points.append(tp)
                            break  # Stop after first match

    else:
        # Car & PTW Scenarios: test lowest and highest target speed, and 1 random target speed in between.
        # Handle cases where the speed attribute can be "GVT speed", or "EMT speed"
        speed_keys = ["GVT speed", "EMT speed"]
        found_speed_key = None
        for key in speed_keys:
            if any(
                tp.test_range == matrix_processing.TestRange.STANDARD
                and tp.attributes.get(key) is not None
                for tp in filtered_test_points
            ):
                found_speed_key = key
                break
        if found_speed_key:
            speeds = sorted(
                set(
                    tp.attributes.get(found_speed_key)
                    for tp in filtered_test_points
                    if tp.test_range == matrix_processing.TestRange.STANDARD
                    and tp.attributes.get(found_speed_key) is not None
                )
            )
        else:
            speeds = []
        if speeds:
            lowest, highest = speeds[0], speeds[-1]
            mid_speeds = [s for s in speeds if s != lowest and s != highest]
            mid = random.choice(mid_speeds) if mid_speeds else None
            for tp in filtered_test_points:
                if (
                    tp.test_range == matrix_processing.TestRange.STANDARD
                    and tp.attributes.get(found_speed_key) in [lowest, highest]
                ):
                    standard_points.append(tp)
            if mid is not None:
                for tp in filtered_test_points:
                    if (
                        tp.test_range == matrix_processing.TestRange.STANDARD
                        and tp.attributes.get(found_speed_key) == mid
                    ):
                        standard_points.append(tp)

    return SelectionInfo(
        test_name=loadcase_info.test_name,
        standard_points=standard_points,
        extended_points=[],
        selected_robustness_layer=None,
        assessment_criteria=None,
        verification_condition=None,
    )


def compute_robustness_layer_score(robustness_layers_df, loadcase_key, subtest_name):
    if robustness_layers_df is None:
        logger.debug("No Robustness Layers sheet provided.")
        return 0.0

    if loadcase_key not in robustness_layers_df.columns:
        raise ValueError(
            f"Column '{loadcase_key}' not found in Robustness Layers sheet."
        )

    test_robustness_df = robustness_layers_df[[loadcase_key]]

    number_of_claimed_robustness_layers = float(
        (test_robustness_df[loadcase_key] == "YES").sum()
    )
    number_of_applicable_robustness_layers = float(
        test_robustness_df[loadcase_key].notna().sum()
    )
    logger.debug(
        f"Number of claimed robustness layers: {number_of_claimed_robustness_layers}"
    )
    logger.debug(
        f"Number of applicable robustness layers: {number_of_applicable_robustness_layers}"
    )

    total_robustness_score = data_model.TOTAL_SCORES.get(subtest_name, {}).get(
        "Robustness", 0
    )

    if total_robustness_score is None:
        total_robustness_score = 0.0

    logger.debug(
        f"total_robustness_score: {total_robustness_score}, number_of_applicable_robustness_layers: {number_of_applicable_robustness_layers}"
    )
    robustness_layer_tested_score = (
        total_robustness_score / number_of_applicable_robustness_layers * 1
    )

    logger.debug(
        f"total_robustness_score: {total_robustness_score}, number_of_applicable_robustness_layers: {number_of_applicable_robustness_layers}, not_tested_claimed_robustness_layers_num: {number_of_claimed_robustness_layers - 1}"
    )
    # Number of claimed robustness layers is decreased by 1 because 1 has been selected for testing
    not_tested_claimed_robustness_layers_num = number_of_claimed_robustness_layers - 1
    robustness_layer_contant_score = (
        total_robustness_score
        / number_of_applicable_robustness_layers
        * not_tested_claimed_robustness_layers_num
    )

    logger.debug(
        f"Calculated robustness score for {subtest_name}: {robustness_layer_contant_score}"
    )
    robustness_layer_score = robustness_layer.RobustnessLayerScore(
        robustness_layer_tested_score, robustness_layer_contant_score
    )
    logger.info(
        f"Robustness Layer total: {robustness_layer_score.total_score}, "
        f"constant part: {robustness_layer_score.constant_score}, "
        f"tested part: {robustness_layer_score.tested_score}, "
    )
    return robustness_layer_score


def compute_predicted_score(test_name, test_points):

    extended_score_sum = sum(
        tp.predicted_score
        for tp in test_points
        if tp.test_range == matrix_processing.TestRange.EXTENDED
    )
    standard_score_sum = sum(
        tp.predicted_score
        for tp in test_points
        if tp.test_range == matrix_processing.TestRange.STANDARD
    )
    extended_count = sum(
        1 for tp in test_points if tp.test_range == matrix_processing.TestRange.EXTENDED
    )
    standard_count = sum(
        1 for tp in test_points if tp.test_range == matrix_processing.TestRange.STANDARD
    )
    logger.debug(f"Number of test points (is_extended_range=True): {extended_count}")
    logger.debug(f"Number of test points (is_extended_range=False): {standard_count}")
    logger.debug(
        f"Sum of predicted scores (is_extended_range=True): {extended_score_sum}"
    )
    logger.debug(
        f"Sum of predicted scores (is_extended_range=False): {standard_score_sum}"
    )

    total_standard_score = data_model.TOTAL_SCORES[test_name]["Standard"]
    if "Extended" in data_model.TOTAL_SCORES[test_name]:
        total_extended_score = data_model.TOTAL_SCORES[test_name]["Extended"]
    else:
        total_extended_score = 0.0

    logger.debug(
        f"standard_score_sum: {standard_score_sum}, standard_count: {standard_count}, total_standard_score: {total_standard_score}, "
    )
    score_standard_range = (
        standard_score_sum / standard_count * total_standard_score
        if standard_count > 0
        else 0
    )

    logger.debug(
        f"extended_score_sum: {extended_score_sum}, extended_count: {extended_count}, total_extended_score: {total_extended_score}, "
    )
    score_extended_range = (
        extended_score_sum / extended_count * total_extended_score
        if extended_count > 0
        else 0
    )

    logger.debug(f"Prefiltered extended score: {score_extended_range}")

    # Adjust scores based on the defined ranges
    if score_extended_range < 0.5 * total_extended_score:
        score_extended_range = 0
    elif (
        0.5 * total_extended_score <= score_extended_range < 0.75 * total_extended_score
    ):
        score_extended_range = 0.5 * total_extended_score
    elif (
        0.75 * total_extended_score <= score_extended_range < 1.0 * total_extended_score
    ):
        score_extended_range = 0.75 * total_extended_score
    elif score_extended_range == 1.0 * total_extended_score:
        score_extended_range = 1.0 * total_extended_score

    logger.debug(f"Calculated score for standard range: {score_standard_range}")
    logger.debug(f"Calculated score for extended range: {score_extended_range}")

    return score_standard_range, score_extended_range


def strip_suffix(test_name):
    suffixes = ("fs", "ns", "fo", "no")
    for suffix in suffixes:
        if test_name.endswith(suffix):
            return test_name[: -len(suffix)]
    return test_name


def read_loadcase_info(
    prediction_df, robustness_layers_df, loadcase_name, stage_subelement_key
):

    test_points = matrix_processing.get_test_matrix(
        prediction_df, loadcase_name, stage_subelement_key
    )
    subtest_name = None
    if loadcase_name in data_model.SUBTEST_TO_TEST_DICT.keys():
        logger.debug(
            f"Scenario {loadcase_name} is a subtest, mapping to main test: {data_model.SUBTEST_TO_TEST_DICT[loadcase_name]}"
        )
        subtest_name = loadcase_name
        loadcase_name = data_model.SUBTEST_TO_TEST_DICT[loadcase_name]

    if robustness_layers_df is not None:
        robustness_key = loadcase_name

        if robustness_key not in robustness_layers_df.columns:
            raise ValueError(
                f"Column '{robustness_key}' not found in Robustness Layers sheet."
            )
        # Create a dictionary with robustness layer names as keys and their values for the given loadcase_name
        robustness_dict = {
            robustness_layers_df.at[idx, "Robustness Layer"]: val
            for idx, val in robustness_layers_df[robustness_key].dropna().items()
        }
        logger.debug(f"Robustness layers for {robustness_key}: {robustness_dict}")
    else:
        robustness_dict = {}
    return LoadcaseInfo(
        test_name=subtest_name if subtest_name else loadcase_name,
        test_points=test_points,
        robustness_layers=robustness_dict,
    )


def get_sheet_prefix(stage_element_name, subelement_name):
    logger.debug(
        f"get_sheet_prefix called with stage_element_name='{stage_element_name}', subelement_name='{subelement_name}'"
    )
    element_initials = get_stage_subelement_key(stage_element_name).value
    if subelement_name == "Pedestrian & Cyclist":
        subelement_name = "Ped & Cyc"
    if subelement_name == "Single Vehicle":
        subelement_name = "Single Veh"
    return f"{element_initials} - {subelement_name}"


def check_robustness_scoring_enabled(subtest_name, test_points):
    predicted_standard_score, _ = compute_predicted_score(subtest_name, test_points)
    total_standard_score = data_model.TOTAL_SCORES[subtest_name]["Standard"]

    return predicted_standard_score >= 0.5 * total_standard_score


def get_stage_subelement_key(stage_element_name):
    element_initials = "".join([word[0].upper() for word in stage_element_name.split()])
    if element_initials not in data_model.StageSubelementKey.__members__:
        raise ValueError(f"Invalid stage element name: {stage_element_name}")
    return data_model.StageSubelementKey[element_initials]


def find_duplicate_rows(test_points):
    """
    Detects duplicate test points based on their attributes.
    Returns a set of tuples, each containing the row indices of duplicate test points.
    """
    attr_to_rows = {}
    duplicate_rows = set()
    for tp in test_points:
        # Convert attributes dict to a tuple of sorted items for hashability
        attr_tuple = tuple(sorted(tp.attributes.items()))
        if attr_tuple in attr_to_rows:
            prev_row = attr_to_rows[attr_tuple]["row"]
            duplicate_rows.add((prev_row, tp.row))
            attr_to_rows[attr_tuple]["rows"].append(tp.row)
        else:
            attr_to_rows[attr_tuple] = {"rows": [tp.row], "row": tp.row}
    return duplicate_rows


def check_cpla_cbla_coherence(test_points):
    """
    Checks CPLA/CBLA coherence by ensuring that duplicate test points (with same attributes)
    only have allowed colors (GREEN or RED, skipping GREY).
    Returns True if coherent, False otherwise.
    """
    duplicate_rows = find_duplicate_rows(test_points)

    if duplicate_rows:
        logger.debug(
            f"Duplicate rows found for CPLA/CBLA coherence check: {duplicate_rows}"
        )
    else:
        logger.debug("No duplicate rows found for CPLA/CBLA coherence check.")
        return True

    for prev_idx, idx in duplicate_rows:
        tp_list = [tp for tp in test_points if tp.row == prev_idx]

        for tp in tp_list:
            if tp.color == matrix_processing.PredictionColor.GREY:
                continue
            if tp.color not in (
                matrix_processing.PredictionColor.GREEN,
                matrix_processing.PredictionColor.RED,
            ):
                logger.error(
                    f"Duplicate test point at row {tp.row} (index {prev_idx}) has color {tp.color}, "
                    "but only GREEN or RED are allowed for CPLA/CBLA coherence check."
                )
                return False

    return True


def preprocess_stage_subelement(dfs, stage_element_name, subelement_name):
    sheet_prefix = get_sheet_prefix(stage_element_name, subelement_name)
    robustness_sheet_name = f"{sheet_prefix} Robustness"
    prediction_sheet_name = f"{sheet_prefix} Prediction"
    # Key is needed to access STAGE_SUBELEMENT_TO_LOADCASES with the right prefix (LDC, FC, ..)
    stage_subelement_key = get_stage_subelement_key(stage_element_name)

    if stage_subelement_key == data_model.StageSubelementKey.LDC:
        robustness_sheet_name = f"{stage_subelement_key.value} - Robustness"
    elif stage_subelement_key == data_model.StageSubelementKey.LSC:
        robustness_sheet_name = None

    if robustness_sheet_name is not None and robustness_sheet_name not in dfs:
        raise ValueError(f"{robustness_sheet_name} sheet not found in the input data.")

    test_points_df = pd.DataFrame()

    test_point_rows = []

    if subelement_name == "Pedestrian & Cyclist":
        subelement_name = "Ped & Cyc"
    if subelement_name == "Single Vehicle":
        subelement_name = "Single Veh"
    stage_subelement_dict = data_model.STAGE_SUBELEMENT_TO_LOADCASES[
        stage_subelement_key
    ]

    if subelement_name not in stage_subelement_dict:
        raise ValueError(
            f"Test prefix '{subelement_name}' not found in STAGE_SUBELEMENT_TO_LOADCASES."
        )

    selected_points_dict = {}
    selected_robustness_dict = {}
    selected_day_robustness = None
    # stage_subelement_dict contains matching test names for the given test prefix
    for loadcase_name in stage_subelement_dict[subelement_name]:
        prediction_df = dfs[prediction_sheet_name]
        if robustness_sheet_name is not None:
            robustness_layers_df = dfs[robustness_sheet_name]
        else:
            robustness_layers_df = None

        if loadcase_name in ["Driveability", "Driver State Link"]:
            logger.debug(
                f"Skipping scenario {loadcase_name} as it does not require test point selection."
            )
            continue
        logger.debug(f"Preprocessing scenario: {loadcase_name}")

        loadcase_info = read_loadcase_info(
            prediction_df, robustness_layers_df, loadcase_name, stage_subelement_key
        )

        if stage_subelement_key == data_model.StageSubelementKey.LSC:
            selected_info = get_lsc_selected_points(loadcase_info)
            selected_standard_points = selected_info.standard_points
            selected_extended_points = selected_info.extended_points
            is_robustness_scoring_enabled = False
        else:
            selected_info = loadcase_info.select_test_points()
            selected_standard_points = selected_info.standard_points
            selected_extended_points = selected_info.extended_points
            is_robustness_scoring_enabled = check_robustness_scoring_enabled(
                loadcase_info.test_name, selected_standard_points
            )

        if stage_subelement_key == data_model.StageSubelementKey.FC and (
            "CPLA" in loadcase_name or "CBLA" in loadcase_name
        ):
            coherence_check = check_cpla_cbla_coherence(loadcase_info.test_points)
            if not coherence_check:
                raise ValueError(
                    f"CPLA and CBLA coherence check failed for scenario {loadcase_name}."
                )
            else:
                logger.info(
                    f"CPLA and CBLA coherence check passed for scenario {loadcase_name}."
                )

        if not is_robustness_scoring_enabled:
            selected_robustness = "N/A"
            selected_info.verification_condition = "N/A"
        else:
            if "_day" in loadcase_info.test_name:
                selected_day_robustness = selected_info.selected_robustness_layer
                selected_robustness = selected_info.selected_robustness_layer
            if "_night" in loadcase_info.test_name:
                selected_robustness = selected_day_robustness
                selected_day_robustness = None
            else:
                selected_robustness = selected_info.selected_robustness_layer

        selected_robustness_dict[loadcase_info.test_name] = selected_robustness

        for point in selected_standard_points:
            row = {
                "Scenario": loadcase_info.test_name,
                "Test Point": f"({point.row}, {point.col})",
                "Range": point.test_range.value,
                "Robustness Layer": selected_robustness,
                "Verification Condition": (
                    selected_info.verification_condition
                    if selected_info.verification_condition
                    else None
                ),
            }

            for attr_name, attr_value in point.attributes.items():
                row[attr_name] = attr_value
            # Add OEM.Prediction and Value at the end
            row["OEM.Prediction"] = str(point.color.value).capitalize()
            row["Value"] = 0.0

            test_point_rows.append(row)

        for point in selected_extended_points:
            row = {
                "Scenario": loadcase_info.test_name,
                "Test Point": f"({point.row}, {point.col})",
                "Range": point.test_range.value,
            }

            for attr_name, attr_value in point.attributes.items():
                row[attr_name] = attr_value
            # Add OEM.Prediction and Value at the end
            row["OEM.Prediction"] = str(point.color.value).capitalize()
            row["Value"] = 0.0

            test_point_rows.append(row)

        # Handle verification conditions and robustness layers for test point rows
        for row in test_point_rows:
            # Remove Verification Condition if it's NaN or "N/A"
            if row.get("Verification Condition") in [None, "N/A"] or pd.isna(
                row.get("Verification Condition")
            ):
                row.pop("Verification Condition", None)

            # Remove Robustness Layer if it's NaN or "N/A"
            if row.get("Robustness Layer") in [None, "N/A"] or pd.isna(
                row.get("Robustness Layer")
            ):
                row.pop("Robustness Layer", None)

        selected_points_dict[loadcase_info.test_name] = (
            selected_standard_points + selected_extended_points
        )

    test_points_df = pd.DataFrame(test_point_rows)
    # Remove "Verification Condition" columns if all values are NaN
    for col in ["Verification Condition"]:
        if col in test_points_df.columns and test_points_df[col].isna().all():
            test_points_df = test_points_df.drop(columns=[col])
    return StageSubelement(
        name=subelement_name,
        selected_points_dict=selected_points_dict,
        selected_robustness_dict=selected_robustness_dict,
        test_points_df=test_points_df,
    )


def read_test_points(df):
    test_points = []
    computed_info_dict = {}
    for _, row in df.iterrows():
        # Parse "Test Point" from string like "(4, 2)" to row and col integers
        row_str = row["Test Point"]
        row_idx, col_idx = map(int, row_str[1:-1].split(","))

        current_verification_condition = row.get("Verification Condition", None)
        value = row.get("Value", None)
        computed_info_dict[f"{row['Scenario']}_{row['Test Point']}"] = {
            "value": value,
            "verification_condition": current_verification_condition,
        }
        test_point = matrix_processing.TestPoint(
            row=row_idx,
            col=col_idx,
            test_range=matrix_processing.TestRange(row["Range"]),
            color=matrix_processing.PredictionColor(row["OEM.Prediction"].lower()),
            robustness_layer=row.get("Robustness Layer", None),
            attributes={
                k: v
                for k, v in row.items()
                if k
                not in [
                    "Scenario",
                    "Robustness Layer",
                    "Test Point",
                    "Range",
                    "OEM.Prediction",
                    "Value",
                    "Verification Condition",
                    "Test Run",
                ]
                and pd.notna(v)
            },
        )
        test_points.append(test_point)

    return test_points, computed_info_dict


@dataclass
class LoadcaseScore:
    loadcase_name: str
    standard_score: float
    extended_score: float
    robustness_layer_score: float
    total_score: float


def compute_loadcase_score(
    score_test_name,
    subtest_name,
    computed_test_points,
    standard_oem_prediction_method,
    extended_oem_prediction_method,
    predicted_standard_score,
    predicted_extended_score,
    robustness_layer_score,
):
    # Define color ranking: lower is better (green < yellow < grey < red)
    color_rank = {
        matrix_processing.PredictionColor.GREEN: 0,
        matrix_processing.PredictionColor.YELLOW: 1,
        matrix_processing.PredictionColor.ORANGE: 2,
        matrix_processing.PredictionColor.BROWN: 2,
        matrix_processing.PredictionColor.RED: 3,
    }

    correct_count_dict = {"Standard": 0, "Extended": 0}
    for ctp in computed_test_points:

        pred_color = ctp.test_point.color
        comp_color = ctp.computed_color

        logger.debug(
            "Assessing test point: %s | pred_color: %s | comp_color: %s | assessment_criteria: %s | prediction_result: %s",
            f"{ctp.test_point.row}, {ctp.test_point.col}",
            pred_color,
            comp_color,
            ctp.assessment_criteria,
            ctp.prediction_result,
        )

        if (
            comp_color == matrix_processing.PredictionColor.GREY
            or pred_color == matrix_processing.PredictionColor.GREY
        ):
            logger.debug(f"---> Skip grey test point")
            continue

        if ctp.assessment_criteria == robustness_layer.AssessmentCriteria.NOT_RED:
            if ctp.computed_color != matrix_processing.PredictionColor.RED:
                logger.debug(f"---> Correct! Not red following assessment criteria")
                if ctp.test_point.test_range == matrix_processing.TestRange.STANDARD:
                    correct_count_dict["Standard"] += 1
                else:
                    correct_count_dict["Extended"] += 1
        elif (
            not ctp.assessment_criteria
            or ctp.assessment_criteria
            == robustness_layer.AssessmentCriteria.SAME_OR_BETTER_THAN_PREDICTED
        ):
            # Correct if computed == predicted or in tolerance
            if (
                ctp.prediction_result == matrix_processing.PredictionResult.CORRECT
                or ctp.prediction_result
                == matrix_processing.PredictionResult.IN_TOLERANCE
            ):
                logger.debug(f"---> Correct! Computed matches predicted")
                if ctp.test_point.test_range == matrix_processing.TestRange.STANDARD:
                    correct_count_dict["Standard"] += 1
                else:
                    correct_count_dict["Extended"] += 1
            # Correct if computed is better (lower rank) than predicted
            elif color_rank[comp_color] < color_rank[pred_color]:
                logger.debug(f"---> Correct! Computed is better than predicted")
                if ctp.test_point.test_range == matrix_processing.TestRange.STANDARD:
                    correct_count_dict["Standard"] += 1
                else:
                    correct_count_dict["Extended"] += 1

    logger.info(
        f"Number of correct or better-than-predicted test points for {subtest_name}: {correct_count_dict}"
    )

    standard_range_verification_test_num = (
        data_model.STANDARD_RANGE_VERIFICATION_TEST_NUM[subtest_name]
    )
    extended_range_verification_test_num = 2

    standard_discount_factor_thresholds = (
        data_model.SCORE_FROM_VERIFICATION_TEST_OUTCOME["Standard"][
            standard_oem_prediction_method.value
        ][standard_range_verification_test_num]
    )
    extended_discount_factor_thresholds = (
        data_model.SCORE_FROM_VERIFICATION_TEST_OUTCOME["Extended"][
            extended_oem_prediction_method.value
        ][extended_range_verification_test_num]
    )

    standard_discount_factor = (
        standard_discount_factor_thresholds[correct_count_dict["Standard"]] / 100
    )
    extended_discount_factor = (
        extended_discount_factor_thresholds[correct_count_dict["Extended"]] / 100
    )

    some_standard_failed = (
        correct_count_dict["Standard"] < standard_range_verification_test_num
    )

    if some_standard_failed:
        robustness_layer_score.total_score = robustness_layer_score.constant_score
        logger.info(
            (
                f"[!!!] Some standard range verification tests failed for {score_test_name}-{subtest_name}. "
                f"Removing testing part of robustness layer score {robustness_layer_score.tested_score}. "
                f"New robustness layer score: {robustness_layer_score.total_score}"
            )
        )
    standard_final_score = predicted_standard_score * standard_discount_factor
    extended_final_score = predicted_extended_score * extended_discount_factor
    logger.debug(
        f"Final scores for {score_test_name}-{subtest_name}: "
        f"Standard Range: {standard_final_score} "
        f"(initial: {predicted_standard_score}, discount: {standard_discount_factor}), "
        f"Extended Range: {extended_final_score} "
        f"(initial: {predicted_extended_score}, discount: {extended_discount_factor})"
    )

    # Thresholding robustness layer score
    # It is counted only if predicted_standard_score is ≥50% of the total available score in
    # the Standard Range of that category
    total_standard_score = data_model.TOTAL_SCORES[subtest_name]["Standard"]

    if standard_final_score < 0.5 * total_standard_score:
        robustness_layer_score.total_score = 0.0
        logger.info(
            f"[!!!] Standard final score {standard_final_score} is below the threshold of 50% of the total standard score {total_standard_score}. Setting the robustness layer score to 0."
        )

    total_test_score = (
        standard_final_score + extended_final_score + robustness_layer_score.total_score
    )

    logger.info(
        f"Total score for {score_test_name}-{subtest_name}: {total_test_score} "
        f"(Standard: {standard_final_score}, Extended: {extended_final_score}, Robustness Layer: {robustness_layer_score.total_score})"
    )
    return LoadcaseScore(
        loadcase_name=score_test_name,
        standard_score=standard_final_score,
        extended_score=extended_final_score,
        robustness_layer_score=robustness_layer_score.total_score,
        total_score=total_test_score,
    )


def compute_lsc_loadcase_score(test_points_df_dict, loadcase_name, input_parameters_df):

    # Read "Doors selected" and "Vehicle response" from input_parameters_df
    doors_selected = input_parameters_df.loc[
        input_parameters_df["Input parameter"] == "Doors selected", "Value"
    ]
    vehicle_response = input_parameters_df.loc[
        input_parameters_df["Input parameter"] == "Vehicle response", "Value"
    ]

    doors_selected_value = doors_selected.iloc[0] if not doors_selected.empty else None
    vehicle_response_value = (
        vehicle_response.iloc[0] if not vehicle_response.empty else None
    )
    # Map string values to enums
    doors_selected_enum = None
    vehicle_response_enum = None

    doors_selected_enum = data_model.DoorSelected(doors_selected_value)
    vehicle_response_enum = data_model.VehicleResponse(vehicle_response_value)

    test_points, computed_info_dict = read_test_points(
        test_points_df_dict[loadcase_name]
    )

    oem_colors = []
    computed_colors = []
    for test_point in test_points:
        row_key = f"{loadcase_name}_({test_point.row}, {test_point.col})"
        current_computed_dict = computed_info_dict.get(row_key, {})
        computed_test_point = matrix_processing.ComputedTestPoint(
            test_name=loadcase_name,
            test_point=test_point,
            value=round(current_computed_dict.get("value", 0.0), 2),
        )

        computed_test_point.assessment_criteria = None
        updated_color = computed_test_point.get_lsc_computed_color(
            vehicle_response_enum, doors_selected_enum
        )

        oem_colors.append(test_point.color)
        computed_colors.append(updated_color)

    lsc_final_score = 0.0
    color_score_map = {
        matrix_processing.PredictionColor.GREEN: 1.0,
        matrix_processing.PredictionColor.YELLOW: 0.75,
        matrix_processing.PredictionColor.ORANGE: 0.5,
        matrix_processing.PredictionColor.BROWN: 0.25,
        matrix_processing.PredictionColor.RED: 0.0,
    }
    for idx in range(len(computed_colors)):
        pred_color = oem_colors[idx]
        comp_color = computed_colors[idx]

        logger.debug(
            "Assessing test point: %s | pred_color: %s | comp_color: %s",
            idx,
            pred_color,
            comp_color,
        )

        if (
            comp_color == matrix_processing.PredictionColor.GREY
            or pred_color == matrix_processing.PredictionColor.GREY
        ):
            logger.debug("---> Skip grey test point")
            continue

        if comp_color == pred_color:
            logger.debug("---> Correct! Computed matches predicted")
            # Add score based on predicted color using a mapping dict
            lsc_final_score += color_score_map.get(pred_color, 0.0)

    total_standard_score = data_model.TOTAL_SCORES[loadcase_name]["Standard"]

    standard_count = sum(
        1 for tp in test_points if tp.test_range == matrix_processing.TestRange.STANDARD
    )

    if standard_count > 0:
        lsc_final_score = lsc_final_score / standard_count * total_standard_score
    else:
        logger.warning(
            "No STANDARD test points found for loadcase '%s'. Setting lsc_final_score to 0.0.",
            loadcase_name,
        )
        lsc_final_score = 0.0

    return LoadcaseScore(
        loadcase_name=loadcase_name,
        standard_score=lsc_final_score,
        extended_score=0.0,
        robustness_layer_score=0.0,
        total_score=lsc_final_score,
    )


# Convert bg_color to PredictionColor
def rgb_tuple_from_hex(hex_str):
    if hex_str is None or len(hex_str) != 8:
        return None
    r = int(hex_str[2:4], 16) / 255
    g = int(hex_str[4:6], 16) / 255
    b = int(hex_str[6:8], 16) / 255
    return (r, g, b)


def get_prediction_color_from_bgcolor(bg_color_hex):
    rgb = rgb_tuple_from_hex(bg_color_hex)
    if rgb is None:
        return None
    for color, rgb_val in matrix_processing.PREDICTION_COLOR_MAP.items():
        if all(abs(a - b) < 0.02 for a, b in zip(rgb, rgb_val)):
            return color
    return None


def load_verification_sheet_with_bg_colors(input_file, sheet_name):
    wb = openpyxl.load_workbook(input_file, data_only=True)
    ws = wb[sheet_name]

    # Get columns and rows
    columns = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    num_cols = len(columns)
    data_bg = []
    data_pred_color = []

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        bg_row = []
        pred_color_row = []
        for cell in row:
            fill = cell.fill
            bg_color = None
            if fill and fill.fgColor and fill.fgColor.type == "rgb":
                bg_color = fill.fgColor.rgb
            bg_row.append(bg_color)

            pred_color = get_prediction_color_from_bgcolor(bg_color)
            pred_color_row.append(pred_color)
        data_bg.append(bg_row)
        data_pred_color.append(pred_color_row)

    # Create DataFrame for prediction colors only
    df_pred_color = pd.DataFrame(data_pred_color, columns=columns)
    # The combined DataFrame only needs prediction colors
    df_combined = pd.DataFrame(
        [
            [df_pred_color.iloc[i, j] for j in range(num_cols)]
            for i in range(len(df_pred_color))
        ],
        columns=columns,
    )
    wb.close()
    return df_combined


def compute_stage_score(dfs, stage_info, loadcase_score_dict, input_file):
    stage_element = stage_info["Stage element"]
    stage_subelement = stage_info["Stage subelement"]

    sheet_prefix = get_sheet_prefix(stage_element, stage_subelement)
    robustness_sheet_name = f"{sheet_prefix} Robustness"
    verification_sheet_name = f"{sheet_prefix} Verification"
    prediction_sheet_name = f"{sheet_prefix} Prediction"
    stage_subelement_key = get_stage_subelement_key(stage_element)

    # Load the verification sheet as a DataFrame, including cell background colors
    prediction_df_bg_pred = load_verification_sheet_with_bg_colors(
        input_file, prediction_sheet_name
    )
    logger.debug(f"Prediction sheet with bg and color: {prediction_df_bg_pred}")

    if stage_subelement_key == data_model.StageSubelementKey.LDC:
        robustness_sheet_name = f"{stage_subelement_key.value} - Robustness"
    elif stage_subelement_key == data_model.StageSubelementKey.LSC:
        robustness_sheet_name = None

    if robustness_sheet_name is not None and robustness_sheet_name not in dfs:
        raise ValueError(f"{robustness_sheet_name} sheet not found in the input data.")

    if prediction_sheet_name not in dfs:
        raise ValueError(f"{prediction_sheet_name} sheet not found in the input data.")

    if stage_subelement == "Pedestrian & Cyclist":
        stage_subelement = "Ped & Cyc"
    if stage_subelement == "Single Vehicle":
        stage_subelement = "Single Veh"

    stage_subelement_dict = data_model.STAGE_SUBELEMENT_TO_LOADCASES[
        stage_subelement_key
    ]

    if stage_subelement not in stage_subelement_dict:
        raise ValueError(
            f"Test prefix '{stage_subelement}' not found in STAGE_SUBELEMENT_TO_LOADCASES."
        )

    verification_tests_df = dfs.get(verification_sheet_name, None)
    if verification_tests_df is None:
        raise ValueError(
            f"Verification sheet for '{stage_subelement}' not found in the input data."
        )

    input_parameters_df = dfs["Input parameters"]

    # Split test_points_df into a dictionary of DataFrames based on the "Scenario" column
    test_points_df_dict = {}
    for loadcase_name in verification_tests_df["Scenario"].unique():
        test_points_df_dict[loadcase_name] = verification_tests_df[
            verification_tests_df["Scenario"] == loadcase_name
        ].copy()

    score_set = False
    for loadcase_name in stage_subelement_dict[stage_subelement]:

        if stage_subelement_key == data_model.StageSubelementKey.LSC:
            robustness_layers_df = None

            loadcase_score = compute_lsc_loadcase_score(
                test_points_df_dict,
                loadcase_name,
                input_parameters_df,
            )
            score_test_name = loadcase_name
            score_set = True

        if stage_subelement_key == data_model.StageSubelementKey.LDC:
            if loadcase_name == "Driveability":
                logger.info("Handling non-loadcase scenario: Driveability")
                # Extract relevant input parameters and assign scores
                input_params = [
                    "Overriding torque",
                    "Continuous Intervention",
                    "Steering Response",
                    "Returning Vlat",
                ]
                score = 0.0
                for param in input_params:
                    value = input_parameters_df.loc[
                        input_parameters_df["Input parameter"] == param, "Value"
                    ]
                    if (
                        not value.empty
                        and str(value.iloc[0]).strip().lower() == "included"
                    ):
                        score += 0.5

                # Store the score in loadcase_score_dict
                if stage_element not in loadcase_score_dict:
                    loadcase_score_dict[stage_element] = {}
                if stage_subelement not in loadcase_score_dict[stage_element]:
                    loadcase_score_dict[stage_element][stage_subelement] = {}

                loadcase_score_dict[stage_element][stage_subelement][loadcase_name] = (
                    LoadcaseScore(
                        loadcase_name=loadcase_name,
                        standard_score=score,
                        extended_score=0.0,
                        robustness_layer_score=0.0,
                        total_score=score,
                    )
                )
                continue
            elif loadcase_name == "Driver State Link":
                logger.info("Handling non-loadcase scenario: Driver State Link")
                # Only one row "Driver state link", if Included then score is 3 otherwise 0
                value = input_parameters_df.loc[
                    input_parameters_df["Input parameter"] == "Driver state link",
                    "Value",
                ]
                score = (
                    3.0
                    if (
                        not value.empty
                        and str(value.iloc[0]).strip().lower() == "included"
                    )
                    else 0.0
                )

                if stage_element not in loadcase_score_dict:
                    loadcase_score_dict[stage_element] = {}
                if stage_subelement not in loadcase_score_dict[stage_element]:
                    loadcase_score_dict[stage_element][stage_subelement] = {}

                loadcase_score_dict[stage_element][stage_subelement][loadcase_name] = (
                    LoadcaseScore(
                        loadcase_name=loadcase_name,
                        standard_score=score,
                        extended_score=0.0,
                        robustness_layer_score=0.0,
                        total_score=score,
                    )
                )
                continue

        if not score_set:
            robustness_layers_df = dfs[robustness_sheet_name]
            verification_test_df = dfs.get(verification_sheet_name, None)

            if verification_test_df is None:
                raise ValueError(
                    f"Verification sheet not found for prefix '{sheet_prefix}'."
                )

            if loadcase_name in data_model.SUBTEST_TO_TEST_DICT.keys():
                logger.debug(
                    f"Scenario {loadcase_name} is a subtest, mapping to main test: {data_model.SUBTEST_TO_TEST_DICT[loadcase_name]}"
                )
                subtest_name = loadcase_name
                score_test_name = data_model.SUBTEST_TO_TEST_DICT[loadcase_name]
            else:
                subtest_name = loadcase_name
                score_test_name = loadcase_name

            standard_oem_prediction_method, extended_oem_prediction_method = (
                matrix_processing.get_prediction_methods(
                    input_parameters_df, score_test_name
                )
            )
            test_points, computed_info_dict = read_test_points(
                test_points_df_dict[subtest_name]
            )
            all_test_points = matrix_processing.get_test_matrix(
                prediction_df_bg_pred, loadcase_name, stage_subelement_key
            )
            # Remove PredictionColor.GREY from all_test_points
            all_test_points = [
                tp
                for tp in all_test_points
                if tp.color != matrix_processing.PredictionColor.GREY
            ]
            robustness_layer_score = compute_robustness_layer_score(
                robustness_layers_df, score_test_name, subtest_name
            )

            predicted_standard_score, predicted_extended_score = (
                compute_predicted_score(subtest_name, all_test_points)
            )

            logger.debug(
                f"Scenario: {score_test_name}-{subtest_name}, "
                f"Predicted Standard Score: {predicted_standard_score}, "
                f"Predicted Extended Score: {predicted_extended_score}, "
                f"Predicted Robustness Score: {robustness_layer_score}, "
            )

            computed_test_points = []
            for test_point in test_points:
                row_key = f"{loadcase_name}_({test_point.row}, {test_point.col})"
                current_computed_dict = computed_info_dict.get(row_key, {})
                current_verification_condition = current_computed_dict.get(
                    "verification_condition", None
                )
                logger.debug(f"test_point.attributes: {test_point.attributes}")
                computed_test_point = matrix_processing.ComputedTestPoint(
                    test_name=loadcase_name,
                    test_point=test_point,
                    value=round(current_computed_dict.get("value", 0.0), 2),
                )
                if current_verification_condition is None:
                    computed_test_point.assessment_criteria = None
                else:
                    computed_test_point.assessment_criteria = (
                        robustness_layer.get_assessment_criteria(
                            computed_test_point.test_point.robustness_layer,
                            loadcase_name,
                            current_verification_condition,
                        )
                    )
                logger.debug(f"ComputedTestPoint: {computed_test_point}")
                logger.debug(
                    f"--> computed color: {computed_test_point.computed_color}"
                )
                logger.debug(
                    f"--> prediction result: {computed_test_point.prediction_result}"
                )
                logger.debug(
                    f"--> assessment criteria: {computed_test_point.assessment_criteria}"
                )
                computed_test_points.append(computed_test_point)

            if (
                stage_subelement_key == data_model.StageSubelementKey.LDC
                and loadcase_name == "ELK-RE"
            ):
                extended_range_performance = input_parameters_df.loc[
                    input_parameters_df["Input parameter"]
                    == "Extended range performance",
                    "Value",
                ]
                extended_range_performance = (
                    str(extended_range_performance.iloc[0]).strip().lower()
                )
                if extended_range_performance == "ldw":
                    predicted_extended_score = 0.5 * predicted_extended_score

            loadcase_score = compute_loadcase_score(
                score_test_name,
                subtest_name,
                computed_test_points,
                standard_oem_prediction_method,
                extended_oem_prediction_method,
                predicted_standard_score,
                predicted_extended_score,
                robustness_layer_score,
            )
        # If already present, update the score fields; otherwise, set the entry
        if (
            stage_element in loadcase_score_dict
            and stage_subelement in loadcase_score_dict[stage_element]
            and score_test_name in loadcase_score_dict[stage_element][stage_subelement]
        ):
            logger.debug(
                f"Updating existing scenario score for {stage_element} - {stage_subelement} - {score_test_name}"
            )
            logger.debug(
                f"Previous score: {loadcase_score_dict[stage_element][stage_subelement][score_test_name]}"
            )
            loadcase_score_dict[stage_element][stage_subelement][
                score_test_name
            ].standard_score += loadcase_score.standard_score
            loadcase_score_dict[stage_element][stage_subelement][
                score_test_name
            ].extended_score += loadcase_score.extended_score
            loadcase_score_dict[stage_element][stage_subelement][
                score_test_name
            ].robustness_layer_score += loadcase_score.robustness_layer_score
            logger.debug(
                f"Updated score: {loadcase_score_dict[stage_element][stage_subelement][score_test_name]}"
            )
        else:
            logger.debug(
                f"Adding new scenario score for {stage_element} - {stage_subelement} - {score_test_name}"
            )
            logger.debug(f"New score: {loadcase_score}")
            loadcase_score_dict[stage_element][stage_subelement][
                score_test_name
            ] = loadcase_score
