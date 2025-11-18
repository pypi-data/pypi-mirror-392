import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from euroncap_rating_2026.crash_protection.vru_processing import (
    VRU_PREDICTION_COLOR_MAP,
    VruPredictionColor,
)


header_align_right = [
    "Body regionscore",
    "Modifiers",
    "Inspection [%]",
    "Score",
    "Max Score",
    "HPL",
    "LPL",
    "Capping",
    "OEM.Prediction",
    "Value",
    "Prediction.Check",
    "Colour",
    "Points",
    "Modifier",
]
header_align_left = [
    "Loadcase",
    "Seat position",
    "Body Region",
    "Criteria",
    "Stage",
    "Stage element",
    "Stage subelement",
    "Dummy",
]


CM_TO_EXCEL_WIDTH = 3.78

import logging

logger = logging.getLogger(__name__)


def is_empty_cell(cell_value):
    """
    Checks if a cell value is considered empty.

    Args:
        cell_value: The value of the cell to check.

    Returns:
        bool: True if the cell is empty, False otherwise.
    """
    return (
        pd.isna(cell_value)
        or cell_value is None
        or cell_value == "None"
        or cell_value == "nan"
    )


def write_report(
    score_df_dict, vru_x_cell_coords, output_file, format_vru_prediction=False
):
    # Write the updated copied sheets back to the output file
    for sheet_name, updated_df in score_df_dict.items():
        try:
            with pd.ExcelWriter(
                output_file, engine="openpyxl", mode="a", if_sheet_exists="replace"
            ) as writer:
                updated_df.to_excel(writer, sheet_name=sheet_name, index=False)
        except Exception as e:
            logger.error(f"Failed to write updated sheet {sheet_name}: {e}")

    wb = load_workbook(output_file)

    calculation_color = PatternFill(
        start_color="FFDD04", end_color="FFDD04", fill_type="solid"
    )
    input_color = PatternFill(
        start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"
    )
    # If not formatting VRU prediction, keep a list of sheetnames without "CP - VRU Prediction"
    if not format_vru_prediction and "CP - VRU Prediction" in wb.sheetnames:
        sheetnames = [name for name in wb.sheetnames if name != "CP - VRU Prediction"]
        # Move "CP - VRU Prediction" to the second last position in wb._sheets if present
        sheets = wb._sheets
        for i, ws in enumerate(sheets):
            if ws.title == "CP - VRU Prediction":
                vru_ws = sheets.pop(i)
                sheets.insert(len(sheets) - 2, vru_ws)
                break
        sheetnames = [name for name in wb.sheetnames if name != "CP - VRU Prediction"]

    else:
        sheetnames = wb.sheetnames

    for sheet_name in sheetnames:
        ws = wb[sheet_name]

        ###############################################################
        # Set first row (header) to black background and white text
        ###############################################################
        header_fill = PatternFill(
            start_color="000000", end_color="000000", fill_type="solid"
        )
        header_font = Font(name="Calibri", color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        ###############################################################
        # Align text
        ###############################################################
        for col in ws.iter_cols():
            header = col[0].value
            if header in header_align_right:
                align = Alignment(horizontal="right")
            elif header in header_align_left:
                align = Alignment(horizontal="left")
            for cell in col:
                cell.alignment = align

            if col[0].value in ["Colour", "OEM.Prediction"]:
                for cell in col[1:]:
                    if (
                        isinstance(cell.value, str)
                        and cell.value
                        and str(cell.value).lower() not in ["nan", "none"]
                    ):
                        cell.value = cell.value.capitalize()

        ###############################################################
        # Set number precision for specific columns
        ###############################################################
        for col in ws.iter_cols():
            header = col[0].value
            if header in [
                "Inspection [%]",
                "Value",
                "Capping",
                "Body regionscore",
                "Modifiers",
            ]:
                for cell in col[1:]:
                    if isinstance(cell.value, (int, float)) and not is_empty_cell(
                        cell.value
                    ):
                        cell.number_format = "0.00"
            elif header in ["Score", "Max Score", "Max score"]:
                for cell in col[1:]:
                    if isinstance(cell.value, (int, float)) and not is_empty_cell(
                        cell.value
                    ):
                        if ws.title not in score_df_dict.keys():
                            cell.number_format = "0.00"
                        elif ws.title == "Test Scores":
                            cell.number_format = "0.000"
                        else:
                            cell.number_format = "0.0000"

        ###############################################################
        # Color columns based on their headers
        ###############################################################
        for col in ws.iter_cols():
            if col[0].value in [
                "Score",
                "Points",
                "Prediction.Check",
                "Body regionscore",
                "Modifiers",
                "Modifier",
                "Colour",
                "Capping?",
            ]:
                for cell in col[1:]:  # Skip the first row (header)
                    cell.fill = calculation_color
                    if is_empty_cell(cell.value):
                        cell.value = ""
            elif col[0].value in ["Inspection [%]", "Value", "OEM.Prediction"]:
                for cell in col[1:]:  # Skip the first row (header)
                    if (
                        sheet_name
                        in ["CP - VRU Head Impact", "CP - VRU Pelvis & Leg Impact"]
                        and col[0].value == "OEM.Prediction"
                    ):
                        logger.debug(
                            f"Skipping OEM.Prediction color fill for sheet {sheet_name} and column {col[0].value}"
                        )
                    else:
                        cell.fill = input_color
                    if is_empty_cell(cell.value):
                        cell.value = ""

    ###############################################################
    # Adjust column widths based on content
    ###############################################################
    for sheet_name in wb.sheetnames:
        if sheet_name == "CP - VRU Prediction":
            continue
        ws = wb[sheet_name]
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter  # Get the column letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except (TypeError, AttributeError) as e:
                    logger.warning(f"Error processing cell value '{cell.value}': {e}")
            adjusted_width = max_length + 2  # Add some padding
            ws.column_dimensions[column_letter].width = adjusted_width

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        prev_value = None
        # Add thick border to the last column for all rows including header
        for row in range(1, ws.max_row + 1):  # Start from 1 to include header
            cell = ws.cell(row=row, column=ws.max_column)
            cell.border = cell.border + Border(
                right=Side(border_style="thick", color="000000")
            )
        # Add thick border to the last row for all columns
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.border = cell.border + Border(
                bottom=Side(border_style="thick", color="000000")
            )
        for row in range(1, ws.max_row + 1):  # Skip header row
            cell = ws.cell(row=row, column=1)
            # Set top border when previous value is empty (nan/None) and current is not empty
            if (
                (prev_value is None or str(prev_value).lower() in ["nan", "none", ""])
                and (
                    cell.value is not None
                    and str(cell.value).lower() not in ["nan", "none", ""]
                )
                and row != 2
            ):
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).border = ws.cell(
                        row=row, column=col
                    ).border + Border(top=Side(border_style="thick", color="000000"))
            prev_value = cell.value

    if format_vru_prediction:
        vru_ws = wb["CP - VRU Prediction"]
        for row_index, row in enumerate(vru_ws.iter_rows()):
            for col_index, cell in enumerate(row):
                key = str(cell.value).lower()
                if (
                    key in ["nan", "none", ""]
                    and col_index > 3
                    and row_index in list(range(3, 22)) + list(range(26, 29))
                ):
                    key = VruPredictionColor.GREY
                if key in VRU_PREDICTION_COLOR_MAP:
                    color = VRU_PREDICTION_COLOR_MAP[key]

                    # Convert float RGB (0-1) to int (0-255) and then to hex
                    color_int = tuple(int(round(c * 255)) for c in color)
                    color_hex = "{:02X}{:02X}{:02X}".format(*color_int)
                    fill = PatternFill(
                        start_color=color_hex,
                        end_color=color_hex,
                        fill_type="solid",
                    )
                    cell.fill = fill
                    if (cell.row - 2, cell.column - 1) in vru_x_cell_coords:
                        cell.value = "X"
                        cell.alignment = Alignment(
                            horizontal="center", vertical="center"
                        )
                    else:
                        cell.value = ""

        # Set thick black border for all cells
        for row in list(range(4, 23)) + list(
            range(27, 30)
        ):  # Excel rows 4-21 and 27-29 inclusive
            for col in range(5, 26):  # Excel columns 3 to 25 inclusive
                cell = vru_ws.cell(row=row, column=col)
                cell.border = Border(
                    left=Side(border_style="thin", color="000000"),
                    right=Side(border_style="thin", color="000000"),
                    top=Side(border_style="thin", color="000000"),
                    bottom=Side(border_style="thin", color="000000"),
                )

        for idx, col in enumerate(vru_ws.columns):
            col_letter = col[0].column_letter
            if idx < 3:
                vru_ws.column_dimensions[col_letter].width = 2 * CM_TO_EXCEL_WIDTH
            else:
                vru_ws.column_dimensions[col_letter].width = CM_TO_EXCEL_WIDTH

        # Set font size to 25 and center alignment for all cells in the VRU Prediction sheet
        font_large = Font(name="Calibri", size=25)
        center_align = Alignment(horizontal="center", vertical="center")

        # Merge cells in the first row for the "Headform" header and keep the original text
        headform_text = vru_ws.cell(
            row=1, column=1
        ).value  # Preserve existing text if any
        vru_ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=25)
        vru_ws.cell(row=1, column=1).value = headform_text
        vru_ws.cell(row=1, column=1).alignment = center_align
        vru_ws.cell(row=1, column=1).fill = header_fill
        vru_ws.cell(row=1, column=1).font = header_font

        # Merge cells in the first row for the "Headform" header and keep the original text
        centerline_text = vru_ws.cell(
            row=4, column=1
        ).value  # Preserve existing text if any
        vru_ws.merge_cells(start_row=4, start_column=1, end_row=22, end_column=1)
        vru_ws.cell(row=4, column=1).value = centerline_text
        vru_ws.cell(row=4, column=1).alignment = center_align
        vru_ws.cell(row=4, column=1).fill = header_fill
        vru_ws.cell(row=4, column=1).font = header_font
        # Set orientation to 90 degrees (vertical text) for the merged center cell (centerline)
        vru_ws.cell(row=4, column=1).alignment = Alignment(
            horizontal="center", vertical="center", text_rotation=90
        )

        # Merge cells in the 24th row for the "Legform" header and keep the original text
        legform_text = vru_ws.cell(
            row=24, column=1
        ).value  # Preserve existing text if any
        vru_ws.merge_cells(start_row=24, start_column=1, end_row=25, end_column=25)
        vru_ws.cell(row=24, column=1).value = legform_text
        vru_ws.cell(row=24, column=1).alignment = center_align
        vru_ws.cell(row=24, column=1).fill = header_fill
        vru_ws.cell(row=24, column=1).font = header_font

        cell.font = font_large
        cell.alignment = center_align

        # Merge cells in rows 24, 27, 28, and 29 for the headers and keep the original text
        for row_num in [27, 28, 29]:
            header_text = vru_ws.cell(
                row=row_num, column=1
            ).value  # Preserve existing text if any
            vru_ws.merge_cells(
                start_row=row_num, start_column=1, end_row=row_num, end_column=4
            )
            vru_ws.cell(row=row_num, column=1).value = header_text
            vru_ws.row_dimensions[row_num].height = 0.55 * 28.35

            vru_ws.cell(row=row_num, column=1).fill = header_fill
            vru_ws.cell(row=row_num, column=1).font = header_font

    wb.save(output_file)
