import os
import sys
import openpyxl
import pandas as pd
from ..i18n import *


__all__ = [
    "read_table_item",
    "print_table",
    "new_table",
    "write_table",
    "write_table_item",
    "save_table",
]


class TableReader:

    def __init__(
        self, 
        cache_size: int,
    ):

        self.cache_size = cache_size  
        self.cache = {}
        self.current_cache_size = 0


    def _load_table(
        self, 
        table_path: str,
    ):

        if table_path in self.cache: return
        
        if not os.path.exists(table_path):
            raise FileNotFoundError(
                translate("未找到表格文件 %s") % (table_path)
            )

        try:
            
            workbook = openpyxl.load_workbook(table_path)
            sheet = workbook.active
            
            assert sheet is not None
            
            headers = [sheet.cell(row=1, column=col).value for col in range(2, sheet.max_column + 1)]
            row_names = [sheet.cell(row=row, column=1).value for row in range(2, sheet.max_row + 1)]
            
            table_data = {
                row_name: {header: sheet.cell(row=row, column=col).value 
                            for col, header in enumerate(headers, start=2)}
                for row, row_name in enumerate(row_names, start=2)
            }
            
            table_size = sys.getsizeof(table_data)

            self.cache[table_path] = table_data
            self.current_cache_size += table_size

            while self.current_cache_size > self.cache_size:
                oldest_table = next(iter(self.cache))
                self.current_cache_size -= sys.getsizeof(self.cache[oldest_table])
                del self.cache[oldest_table]
                
        except Exception as error:
            
            raise Exception(
                translate("读取表格时发生错误: %s") % (error)
            )

    def _read_table(
        self, 
        table_path: str,
    ):
        
        self._load_table(table_path)
        return self.cache[table_path]

    def _get_value(
        self, 
        table_path: str, 
        row_name: str, 
        column_name: str,
    ):
        
        table_info = self._read_table(table_path)
        row_data = table_info.get(row_name)
        return row_data.get(column_name) if row_data else None


    def _print_table(
        self, 
        table_path: str,
    ):
        
        table_info = self._read_table(table_path)
        
        for row_name, row_data in table_info.items():
            print(f"{row_name}: {row_data}")
            
            
table_reader_cache_size = 1024 * 1024 * 1

table_reader = TableReader(
    cache_size = table_reader_cache_size
)


def read_table_item(
    table_path: str, 
    row_name: str, 
    column_name: str,
):
    
    return table_reader._get_value(table_path, row_name, column_name)


def print_table(
    table_path: str,
):
    
    return table_reader._print_table(table_path)


def new_table(
    column_name_list,
):

    return pd.DataFrame(columns=column_name_list)


def write_table(
    table, 
    row_index: int, 
    column_index: int, 
    value,
):
    
    row_index = row_index - 1
    
    num_rows, num_columns = table.shape

    if row_index >= num_rows:

        additional_rows = row_index - num_rows + 1

        table = pd.concat([table, pd.DataFrame([[None]*num_columns] * additional_rows, columns=table.columns)], ignore_index=True)

    if column_index >= num_columns:

        additional_columns = column_index - num_columns + 1

        for i in range(additional_columns):
            table[table.columns[-1] + f'_new_{i}'] = [None] * table.shape[0]

    table.iloc[row_index, column_index] = value
    
    return table


def write_table_item(
    table, 
    row_name: str, 
    column_name: str, 
    value,
):
    """Write a value to a specific cell in a table identified by row and column names.

    Args:
        table: The target table (DataFrame)
        row_name (str): Name of the row to locate (must be in first column)
        column_name (str): Name of the column to locate
        value: Value to write to the specified cell

    Returns:
        The modified table

    Raises:
        ValueError: If row or column name is not found
    """
    if row_name not in table.iloc[:, 0].values:
        raise ValueError(
            translate("行名 '%s' 未在第一列中找到") % (row_name)
        )

    if column_name not in table.columns:
        raise ValueError(
            translate("列名 '%s' 未在表格中找到") % (column_name)
        )

    row_index = table[table.iloc[:, 0] == row_name].index[0]
    column_index = table.columns.get_loc(column_name)
    table.iloc[row_index, column_index] = value
    
    return table


def save_table(
    table, 
    table_path: str,
):
    
    table.to_excel(table_path, index=False)