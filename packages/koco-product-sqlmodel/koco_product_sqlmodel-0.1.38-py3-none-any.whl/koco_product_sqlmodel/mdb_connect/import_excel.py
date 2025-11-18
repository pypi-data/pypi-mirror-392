from fastapi import HTTPException
import openpyxl as opyx
import openpyxl.worksheet.worksheet as opyx_ws
import koco_product_sqlmodel.dbmodels.definition as sql_def
import koco_product_sqlmodel.dbmodels.models_enums as sql_enum
import koco_product_sqlmodel.mdb_connect.select as mdb_select
import koco_product_sqlmodel.mdb_connect.mdb_connector as mdb_con
import koco_product_sqlmodel.dbmodels.changelog as sql_change
import dataclasses as dc

IS_NONE_STR = lambda in_str: None if in_str == "None" else in_str


@dc.dataclass
class TableData:
    ws: opyx_ws.Worksheet
    spectable_type: sql_enum.SpectableTypeEnum
    start_row: int | None = None  # first index = 1!
    n_rows: int | None = None
    start_col: int | None = None  # first index = 1!
    n_cols: int | None = None
    has_unit: bool = False
    unit_col: int | None = None
    max_col: int | None = None
    min_col: int | None = None
    header: list[str] = dc.field(default_factory=lambda: [])
    data: list[list[str]] = dc.field(default_factory=lambda: [])
    units: list[str] = dc.field(default_factory=lambda: [])
    min_vals: list[str] = dc.field(default_factory=lambda: [])
    max_vals: list[str] = dc.field(default_factory=lambda: [])

    def __post_init__(self):
        self._get_starting_row_col()
        self._get_table_dims()
        self._get_special_cols()
        self._get_header()
        self._get_data()

    def _get_starting_row_col(self):
        for irows, line in enumerate(self.ws):
            for icols, cell in enumerate(line):
                if cell.value:
                    self.start_row = irows + 1
                    self.start_col = icols + 1
                    return

    def _get_table_dims(self):
        n_rows = 0
        n_cols = 0
        for line in self.ws:
            n_rows += 1
            if n_rows == 1:
                n_cols = len(line)
        self.n_rows = n_rows - self.start_row + 1
        self.n_cols = n_cols - self.start_col + 1

    def _get_special_cols(self):
        self.has_unit = False
        if self.spectable_type.value == "free":
            self.has_unit = False
            self.unit_col = None
            self.min_col = None
            self.max_col = None
            return
        for col in range(self.start_col, self.start_col + self.n_cols):
            if self.ws.cell(row=self.start_row, column=col).value.lower() == "unit":
                self.has_unit = True
                self.unit_col = col
            if self.ws.cell(row=self.start_row, column=col).value.lower() == "max":
                self.max_col = col
            if self.ws.cell(row=self.start_row, column=col).value.lower() == "min":
                self.min_col = col

    def _get_header(self):
        self.header = []
        for col in range(self.start_col, self.start_col + self.n_cols):
            if col == self.unit_col:
                continue
            self.header.append(str(self.ws.cell(row=self.start_row, column=col).value))

    def _get_data(self):
        self.data = []
        self.units = []
        for row in range(self.start_row + 1, self.start_row + self.n_rows):
            data = []
            for col in range(self.start_col, self.start_col + self.n_cols):
                if col == self.unit_col:
                    self.units.append(
                        IS_NONE_STR(str(self.ws.cell(row=row, column=col).value))
                    )
                    continue
                if col == self.min_col:
                    self.min_vals.append(
                        IS_NONE_STR(str(self.ws.cell(row=row, column=col).value))
                    )
                if col == self.max_col:
                    self.max_vals.append(
                        IS_NONE_STR(str(self.ws.cell(row=row, column=col).value))
                    )

                data.append(IS_NONE_STR(str(self.ws.cell(row=row, column=col).value)))
            self.data.append(data)


def import_spectable_from_excel(
    file_content: bytes,
    entity_id: int,
    entity_type: str,
    spectable_name: str,
    spectable_type: sql_enum.SpectableTypeEnum,
    user_id: int,
) -> sql_def.CSpecTable:
    check_args(entity_type=entity_type, spectable_type=spectable_type)
    wb = opyx.load_workbook(filename=file_content, data_only=True)
    ws = get_table_worksheet(wb=wb, spectable_name=spectable_name)
    table_data = TableData(ws=ws, spectable_type=spectable_type)
    spectable = mdb_select.select_spectable(
        parent_id=entity_id,
        parent_type=sql_enum.SpectableParentEnum(adapt_parent_type(entity_type)),
        st_type=spectable_type,
        name=spectable_name,
    )
    if spectable != None:
        mdb_con.delete_spectable_by_id(
            log_func=spectable,
            spectable_id=spectable.id,
            delete_connected_items=True,
            user_id=user_id,
        )
    spectable = add_spectable_to_db(
        parent=sql_enum.SpectableParentEnum(adapt_parent_type(entity_type)),
        parent_id=entity_id,
        type=spectable_type,
        name=spectable_name,
        user_id=user_id,
        has_unit=table_data.has_unit,
    )
    add_spectable_items(
        spectable_id=spectable.id,
        table_data=table_data,
        spectable_type=spectable_type,
        user_id=user_id,
    )
    return spectable


def check_args(entity_type: str, spectable_type: sql_enum.SpectableTypeEnum) -> None:
    if (
        entity_type.lower() not in ("carticle", "article")
        and spectable_type == sql_enum.SpectableTypeEnum.overview
    ):
        raise HTTPException(
            status_code=501,
            detail="overview-tables only possible for entity-type carticle",
        )


def add_spectable_items(
    spectable_id: int,
    table_data: TableData,
    spectable_type: sql_enum.SpectableTypeEnum,
    user_id: int,
):
    if spectable_type == sql_enum.SpectableTypeEnum.multicol:
        add_spectable_items_multicol(
            spectable_id=spectable_id, table_data=table_data, user_id=user_id
        )
        return
    if spectable_type == sql_enum.SpectableTypeEnum.free:
        add_spectable_items_free(
            spectable_id=spectable_id, table_data=table_data, user_id=user_id
        )
        return
    if spectable_type in (
        sql_enum.SpectableTypeEnum.singlecol,
        sql_enum.SpectableTypeEnum.overview,
    ):
        add_spectable_items_singlecol(
            spectable_id=spectable_id, table_data=table_data, user_id=user_id
        )
        return
    raise HTTPException(
        status_code=501,
        detail="Only multicol, free, singlecol, and overview tables implemented so far",
    )


def add_spectable_items_singlecol(
    spectable_id: int, table_data: TableData, user_id: int
):
    data = []
    for iy, line in enumerate(table_data.data):
        unit, min_val, max_val = None, None, None
        if table_data.unit_col:
            unit = table_data.units[iy]
        if table_data.max_col:
            max_val = table_data.max_vals[iy]
        if table_data.min_col:
            min_val = table_data.min_vals[iy]
        for ix, d in enumerate(line):
            if ix in (
                table_data.max_col - table_data.start_col - 1,
                table_data.min_col - table_data.start_col - 1,
                table_data.unit_col - table_data.start_col - 1,
            ):
                continue
            data.append(
                sql_def.CSpecTableItem(
                    pos=f"{iy+1}",
                    name=line[0],
                    value=d,
                    spec_table_id=spectable_id,
                    unit=unit,
                    max_value=max_val,
                    min_value=min_val,
                    user_id=user_id,
                )
            )
    add_spectableitems_to_db(spec_items=data)


def add_spectable_items_free(spectable_id: int, table_data: TableData, user_id: int):
    headers = []
    for ix, header in enumerate(table_data.header):
        unit = None
        headers.append(
            sql_def.CSpecTableItem(
                pos=f"0;{ix+1}",
                name="free_header",
                value=header,
                spec_table_id=spectable_id,
                user_id=user_id,
            )
        )
    data = []
    for iy, line in enumerate(table_data.data):
        for ix, d in enumerate(line):
            unit = None
            data.append(
                sql_def.CSpecTableItem(
                    pos=f"{iy+1};{ix+1}",
                    name="free_cell",
                    value=d,
                    spec_table_id=spectable_id,
                    unit=unit,
                )
            )
    spec_items = headers + data
    add_spectableitems_to_db(spec_items=spec_items)


def add_spectable_items_multicol(
    spectable_id: int, table_data: TableData, user_id: int
):
    headers = []
    for ix, header in enumerate(table_data.header[1:]):
        unit = None
        headers.append(
            sql_def.CSpecTableItem(
                pos=f"0;{ix+1}",
                name=header,
                value=header,
                spec_table_id=spectable_id,
                user_id=user_id,
            )
        )
    data = []
    for iy, line in enumerate(table_data.data):
        for ix, d in enumerate(line[1:]):
            unit = None
            if table_data.has_unit:
                unit = table_data.units[iy]
            data.append(
                sql_def.CSpecTableItem(
                    pos=f"{iy+1};{ix+1}",
                    name=table_data.data[iy][0],
                    value=d,
                    spec_table_id=spectable_id,
                    unit=unit,
                )
            )
    spec_items = headers + data
    add_spectableitems_to_db(spec_items=spec_items)


def add_spectableitems_to_db(spec_items: list[sql_def.CSpecTableItem]) -> None:
    with mdb_con.Session(mdb_con.mdb_engine) as session:
        for sti in spec_items:
            session.add(sti)
            session.commit()
            session.refresh(sti)
            log_data = sql_change.CChangelog(
                entity_id=sti.id,
                entity_type="cspectableitem",
                user_id=sti.user_id,
                action="POST",
                new_values=None,
            )
            session.add(log_data)
            session.commit()
            # print(sti)


def get_table_worksheet(wb: opyx.Workbook, spectable_name: str) -> opyx_ws.Worksheet:
    if spectable_name in wb:
        return wb[spectable_name]
    return wb.active


def add_spectable_to_db(
    parent: sql_enum.SpectableParentEnum,
    parent_id: int,
    type: sql_enum.SpectableTypeEnum,
    has_unit: bool,
    name: str,
    user_id: int,
) -> sql_def.CSpecTable:
    spectable = sql_def.CSpecTable(
        name=name, type=type, has_unit=has_unit, parent=parent, parent_id=parent_id
    )
    with mdb_con.Session(mdb_con.mdb_engine) as session:
        session.add(spectable)
        session.commit()
        session.refresh(spectable)
        log_changes(
            entity_type="cspectable",
            entity_id=spectable.id,
            action="POST",
            user_id=user_id,
            new_values=None,
        )
    return spectable


def adapt_parent_type(pt: str):
    # Remove leading letter "c" as long es spectables use "article" and "family" as parent_type
    if pt[0] == "c":
        return pt[1:]
    return pt


def log_changes(
    entity_type: str, entity_id: int, action: str, user_id: int, new_values: str
) -> None:
    with mdb_con.Session(mdb_con.mdb_engine) as session:
        change = sql_change.CChangelog(
            entity_id=entity_id,
            entity_type=entity_type,
            user_id=user_id,
            action=action,
            new_values=new_values,
        )
        session.add(change)
        session.commit()


def main():
    pass


if __name__ == "__main__":
    main()
