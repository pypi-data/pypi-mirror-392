from __future__ import annotations
import os
import re
from typing import Dict, Any, List, Set
import pandas as pd  # pip install pandas openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

_INVALID_SHEET_CHARS = r'[:\\/*?\[\]]'  # Excel-invalid characters
_MAX_SHEET_LEN = 31

# Header style: solid green background + white bold font
_HEADER_FILL = PatternFill(fill_type="solid", start_color="FF2E7D32", end_color="FF2E7D32")  # #2E7D32
_HEADER_FONT = Font(color="FFFFFFFF", bold=True)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _safe_sheet_name(raw_title: str, taken: Set[str]) -> str:
    """
    Create a safe Excel sheet name from a raw title.

    Ensures the sheet name is valid for Excel by removing invalid characters,
    handling length limits, and avoiding duplicates.
    :param raw_title: Original title to convert to sheet name
    :param taken: Set of already used sheet names to avoid conflicts
    :return: Safe Excel sheet name that doesn't conflict with existing names
    """
    name = (raw_title or "Untitled").strip()
    name = re.sub(_INVALID_SHEET_CHARS, "_", name)
    name = re.sub(r"\s+", " ", name)
    name = name[:_MAX_SHEET_LEN] if name else "Sheet"

    base = name or "Sheet"
    candidate = base
    i = 1
    while candidate in taken or not candidate:
        suffix = f"_{i}"
        candidate = (base[:_MAX_SHEET_LEN - len(suffix)] + suffix) if len(base) + len(
            suffix) > _MAX_SHEET_LEN else base + suffix
        i += 1

    taken.add(candidate)
    return candidate


def _style_header(ws, ncols: int) -> None:
    """
    Apply styling to the header row of an Excel worksheet.

    Styles the first row with green background, white bold font, and center alignment.
    Also freezes the panes below the header row.
    :param ws: OpenPyXL worksheet object to style
    :param ncols: Number of columns in the worksheet
    :return: None
    """
    if ncols > 0:
        ws.freeze_panes = "A2"
        for idx in range(1, ncols + 1):
            cell = ws.cell(row=1, column=idx)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = _HEADER_ALIGN


def _autosize_columns(ws, df: pd.DataFrame) -> None:
    """
    Automatically size columns in an Excel worksheet based on content.

    Calculates optimal column widths based on header text and sample data
    from the first 200 rows for performance.
    :param ws: OpenPyXL worksheet object to resize
    :param df: Pandas DataFrame containing the data
    :return: None
    """
    # Basic autosize based on header + sample of values
    for i, col in enumerate(df.columns, start=1):
        header = str(col) if col is not None else ""
        max_len = len(header)
        # sample first ~200 rows for performance
        if not df.empty and i <= len(df.columns):
            for val in df.iloc[:min(200, len(df)), i - 1].astype(str).values:
                if len(val) > max_len:
                    max_len = len(val)
        ws.column_dimensions[get_column_letter(i)].width = min(max(10, max_len + 2), 60)


def _style_summary_sheet(ws, df: pd.DataFrame, sheet_mapping: dict = None) -> None:
    """
    Apply special styling to the summary sheet with text wrapping for descriptions.
    Add hyperlinks to table titles that link to their corresponding sheets.
    
    :param ws: OpenPyXL worksheet object to style
    :param df: Pandas DataFrame containing the summary data
    :param sheet_mapping: Dictionary mapping table titles to their sheet names
    :return: None
    """
    _style_header(ws, ncols=df.shape[1])
    
    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    
    for row_idx in range(2, len(df) + 2):
        for col_idx in range(1, df.shape[1] + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = wrap_alignment
            
            if col_idx == 1 and sheet_mapping:
                table_title = cell.value
                if table_title and table_title in sheet_mapping:
                    sheet_name = sheet_mapping[table_title]
                    
                    if ' ' in sheet_name or any(char in sheet_name for char in ['[', ']', '*', '?', ':', '\\', '/']):
                        hyperlink_ref = f"#'{sheet_name}'!A1"
                    else:
                        hyperlink_ref = f"#{sheet_name}!A1"
                    
                    cell.hyperlink = Hyperlink(ref=hyperlink_ref, target=hyperlink_ref)
                    cell.font = Font(color="0000FF", underline="single")
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12
    
    for row_idx in range(2, len(df) + 2):
        ws.row_dimensions[row_idx].height = 60  # Allow for multiple lines


def _normalize_data(headers: List[str], rows: List[List]) -> tuple[List[str], List[List]]:
    """
    Normalize headers and rows to ensure consistent dimensions.

    :param headers: List of column headers
    :param rows: List of data rows
    :return: Tuple of (normalized_headers, normalized_rows)
    """
    if not rows:
        return headers, []

    max_cols = max(len(row) for row in rows) if rows else 0

    if headers:
        target_cols = max(len(headers), max_cols)
    else:
        target_cols = max_cols
        headers = [f"Column_{i + 1}" for i in range(target_cols)]

    normalized_headers = list(headers)
    while len(normalized_headers) < target_cols:
        normalized_headers.append(f"Column_{len(normalized_headers) + 1}")
    normalized_headers = normalized_headers[:target_cols]

    normalized_rows = []
    for row in rows:
        normalized_row = list(row)
        while len(normalized_row) < target_cols:
            normalized_row.append(None)
        normalized_rows.append(normalized_row[:target_cols])

    return normalized_headers, normalized_rows


def write_structured_excel(excel_path: str, items: List[Dict[str, Any]]) -> str | None:
    """
    Write a list of structured data items into an Excel workbook.

    Each item becomes a separate worksheet with styled headers. The function
    handles sheet name sanitization, header styling, and column autosizing.
    Automatically handles mismatched headers and data columns.

    :param excel_path: Path where the Excel file will be saved
    :param items: List of dictionaries, each containing:
                 - 'title': Sheet title (optional)
                 - 'headers': List of column headers (optional)
                 - 'rows': List of data rows (optional)
    :return: Path to the written Excel file if successful, None if no items provided
    """
    if not items:
        return None

    valid_items = []
    for item in items:
        headers = item.get("headers") or []
        rows = item.get("rows") or []
        if headers or (rows and any(
                row for row in rows if any(cell for cell in row if cell is not None and str(cell).strip()))):
            valid_items.append(item)

    if not valid_items:
        print("Warning: No valid items to write to Excel")
        return None

    os.makedirs(os.path.dirname(excel_path) or ".", exist_ok=True)
    taken: Set[str] = set()

    with pd.ExcelWriter(excel_path, engine="openpyxl", mode="w") as writer:
        summary_data = []
        sheet_mapping = {}
        
        for item in valid_items:
            title = item.get("title") or "Untitled"
            description = item.get("description") or "No description available"
            page_number = item.get("page", "Unknown")
            item_type = item.get("type", "Table")  # Default to "Table" if not specified
            
            
            summary_data.append({
                "Table Title": title,
                "Description": description,
                "Page": page_number,
                "Type": item_type
            })
        
        if summary_data:
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name="Table Summary", index=False)
            taken.add("Table Summary")

        for item in valid_items:
            try:
                title = item.get("title") or "Untitled"
                headers = item.get("headers") or []
                rows = item.get("rows") or []

                sheet_name = _safe_sheet_name(title, taken)
                
                sheet_mapping[title] = sheet_name

                normalized_headers, normalized_rows = _normalize_data(headers, rows)

                if not normalized_rows and not normalized_headers:
                    print(f"Skipping empty item: {title}")
                    continue

                try:
                    df = pd.DataFrame(normalized_rows, columns=normalized_headers)
                except Exception as e:
                    print(f"Error creating DataFrame for '{title}': {e}")
                    df = pd.DataFrame([["Error processing data"]], columns=["Message"])

                df.to_excel(writer, sheet_name=sheet_name, index=False)

                ws = writer.sheets[sheet_name]
                _style_header(ws, ncols=df.shape[1])
                _autosize_columns(ws, df)

            except Exception as e:
                print(f"Error processing item '{item.get('title', 'Unknown')}': {e}")
                continue

        if summary_data and sheet_mapping:
            summary_ws = writer.sheets["Table Summary"]
            _style_summary_sheet(summary_ws, summary_df, sheet_mapping)

    return excel_path