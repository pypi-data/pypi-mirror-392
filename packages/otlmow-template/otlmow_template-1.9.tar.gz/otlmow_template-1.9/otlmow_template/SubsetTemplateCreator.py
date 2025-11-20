import asyncio
import concurrent
import contextlib
import csv
import logging
import os
from asyncio import sleep
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed, ALL_COMPLETED
from pathlib import Path

from openpyxl.cell import Cell
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.dimensions import DimensionHolder, ColumnDimension
from openpyxl.worksheet.worksheet import Worksheet
from otlmow_converter.DotnotationHelper import DotnotationHelper
from otlmow_converter.Exceptions.UnknownExcelError import UnknownExcelError
from otlmow_converter.OtlmowConverter import OtlmowConverter
from otlmow_model.OtlmowModel.BaseClasses.BooleanField import BooleanField
from otlmow_model.OtlmowModel.BaseClasses.KeuzelijstField import KeuzelijstField
from otlmow_model.OtlmowModel.BaseClasses.OTLObject import dynamic_create_instance_from_uri, OTLObject, \
    get_attribute_by_name, dynamic_create_type_from_uri
from otlmow_model.OtlmowModel.Helpers.GenericHelper import get_ns_and_name_from_uri
from otlmow_model.OtlmowModel.Helpers.RelationCreator import create_betrokkenerelation, create_relation
from otlmow_model.OtlmowModel.Helpers.generated_lists import get_hardcoded_relation_dict
from otlmow_modelbuilder.OSLOCollector import OSLOCollector
from otlmow_modelbuilder.SQLDataClasses.OSLOClass import OSLOClass

ROOT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))


short_to_long_ns = {
    'ond': 'https://wegenenverkeer.data.vlaanderen.be/ns/onderdeel#',
    'onderdeel': 'https://wegenenverkeer.data.vlaanderen.be/ns/onderdeel#',
    'ins': 'https://wegenenverkeer.data.vlaanderen.be/ns/installatie#',
    'installatie': 'https://wegenenverkeer.data.vlaanderen.be/ns/installatie#',
    'imp': 'https://wegenenverkeer.data.vlaanderen.be/ns/implementatieelement#',
    'implementatieelement': 'https://wegenenverkeer.data.vlaanderen.be/ns/implementatieelement#',
    'proefenmeting': 'https://wegenenverkeer.data.vlaanderen.be/ns/proefenmeting#',
    'pro': 'https://wegenenverkeer.data.vlaanderen.be/ns/proefenmeting#',
    'lev': 'https://wegenenverkeer.data.vlaanderen.be/ns/levenscyclus#',
    'levenscyclus': 'https://wegenenverkeer.data.vlaanderen.be/ns/levenscyclus#'
}


class SubsetTemplateCreator:
    green_fill = PatternFill(start_color="90EE90", fill_type="solid")

    @classmethod
    def _load_collector_from_subset_path(cls, subset_path: Path) -> OSLOCollector:
        collector = OSLOCollector(subset_path)
        collector.collect_all(include_abstract=True)
        return collector

    @classmethod
    async def generate_template_from_subset_async(
            cls,
            subset_path: Path,
            template_file_path: Path,
            ignore_relations: bool = True,
            filter_attributes_by_subset: bool = True,
            class_uris_filter: [str] = None,
            dummy_data_rows: int = 1,
            add_geometry: bool = True,
            add_attribute_info: bool = False,
            add_deprecated: bool = False,
            generate_choice_list: bool = True,
            split_per_type: bool = True,
            model_directory: Path = None, **kwargs):
        """
        Generate a template from a subset file, async version.
        Await this function!

        :param subset_path: Path to the subset file
        :param template_file_path: Path to where the template file should be created
        :param ignore_relations: Whether to ignore relations when creating the template, defaults to True
        :param filter_attributes_by_subset: Whether to filter by the attributes in the subset, defaults to True
        :param class_uris_filter: List of class URIs to filter by. If not None, only classes with these URIs will be included, defaults to None
        :param dummy_data_rows: Amount of dummy data rows to add to the template, defaults to 1
        :param add_geometry: Whether to include the geometry attribute in the template, defaults to True
        :param add_attribute_info: Whether to add attribute information to the template (colored grey in Excel), defaults to False
        :param add_deprecated: Whether to add a deprecated row to the template (colored red in Excel), defaults to False
        :param generate_choice_list: Whether to generate a choice list in the template (only for Excel), defaults to True
        :param split_per_type: Whether to split the template into a file per type (only for CSV), defaults to True
        :param model_directory: Path to the model directory, defaults to None

        :return: None
        """
        # generate objects to write to file
        objects = await cls.generate_objects_for_template_async(
            subset_path=subset_path, ignore_relations=ignore_relations, class_uris_filter=class_uris_filter,
            add_geometry=add_geometry, filter_attributes_by_subset=filter_attributes_by_subset,
            dummy_data_rows=dummy_data_rows, model_directory=model_directory)

        abbreviate_excel_after = False
        if kwargs.get('abbreviate_excel_sheettitles') == True:
            kwargs['abbreviate_excel_sheettitles'] = False
            abbreviate_excel_after = True

        # write the file
        await OtlmowConverter.from_objects_to_file_async(
            file_path=template_file_path, sequence_of_objects=objects, split_per_type=split_per_type, **kwargs)

        # alter the file if needed
        extension = template_file_path.suffix.lower()
        if extension == '.xlsx':
            await cls.alter_excel_template_async(
                generate_choice_list=generate_choice_list, file_path=template_file_path, dummy_data_rows=dummy_data_rows,
                instances=objects, add_deprecated=add_deprecated, add_attribute_info=add_attribute_info,
                abbreviate_excel_sheettitles=abbreviate_excel_after)

        elif extension == '.csv':
            await cls.alter_csv_template_async(
                split_per_type=split_per_type, file_path=template_file_path, dummy_data_rows=dummy_data_rows,
                instances=objects, add_deprecated=add_deprecated, add_attribute_info=add_attribute_info)

    @classmethod
    def generate_template_from_subset(
            cls,
            subset_path: Path,
            template_file_path: Path,
            ignore_relations: bool = True,
            filter_attributes_by_subset: bool = True,
            class_uris_filter: [str] = None,
            dummy_data_rows: int = 1,
            add_geometry: bool = True,
            add_attribute_info: bool = False,
            add_deprecated: bool = False,
            generate_choice_list: bool = True,
            split_per_type: bool = True,
            model_directory: Path = None, **kwargs):
        """
         Generate a template from a subset file.

         :param subset_path: Path to the subset file
         :param template_file_path: Path to where the template file should be created
         :param ignore_relations: Whether to ignore relations when creating the template, defaults to True
         :param filter_attributes_by_subset: Whether to filter by the attributes in the subset, defaults to True
         :param class_uris_filter: List of class URIs to filter by. If not None, only classes with these URIs will be included, defaults to None
         :param dummy_data_rows: Amount of dummy data rows to add to the template, defaults to 1
         :param add_geometry: Whether to include the geometry attribute in the template, defaults to True
         :param add_attribute_info: Whether to add attribute information to the template (colored grey in Excel), defaults to False
         :param add_deprecated: Whether to tag deprecated attributes in the template, defaults to False
         :param generate_choice_list: Whether to generate a choice list in the template (only for Excel), defaults to True
         :param split_per_type: Whether to split the template into a file per type (only for CSV), defaults to True
         :param model_directory: Path to the model directory, defaults to None

         :return: None
         """
        # generate objects to write to file
        objects = cls.generate_objects_for_template(
            subset_path=subset_path, ignore_relations=ignore_relations, class_uris_filter=class_uris_filter,
            add_geometry=add_geometry, filter_attributes_by_subset=filter_attributes_by_subset,
            dummy_data_rows=dummy_data_rows, model_directory=model_directory)

        abbreviate_excel_after = False
        if kwargs.get('abbreviate_excel_sheettitles') == True:
            kwargs['abbreviate_excel_sheettitles'] = False
            abbreviate_excel_after = True

        # write the file
        OtlmowConverter.from_objects_to_file(
            file_path=template_file_path, sequence_of_objects=objects, split_per_type=split_per_type,
            model_directory=model_directory, **kwargs)

        # alter the file if needed
        extension = template_file_path.suffix.lower()
        if extension == '.xlsx':
            cls.alter_excel_template(
                generate_choice_list=generate_choice_list, file_path=template_file_path, dummy_data_rows=dummy_data_rows,
                instances=objects, add_deprecated=add_deprecated, add_attribute_info=add_attribute_info,
                abbreviate_excel_sheettitles=abbreviate_excel_after)
        elif extension == '.csv':
            cls.alter_csv_template(
                split_per_type=split_per_type, file_path=template_file_path, dummy_data_rows=dummy_data_rows,
                instances=objects, add_deprecated=add_deprecated, add_attribute_info=add_attribute_info)


    @classmethod
    def create_x_objects(cls, oslo_class, add_geometry, collector, filter_attributes_by_subset, model_directory,
                         amount_objects_to_create):
        if oslo_class.objectUri in cls.relation_dict:
            return []

        otl_objects = []
        for _ in range(amount_objects_to_create):
            otl_object = cls.generate_object_from_oslo_class(
                oslo_class=oslo_class, add_geometry=add_geometry, collector=collector,
                filter_attributes_by_subset=filter_attributes_by_subset, model_directory=model_directory)
            if otl_object is not None:
                otl_objects.append(otl_object)
        return otl_objects

    @classmethod
    def generate_objects_for_template(
            cls, subset_path: Path, class_uris_filter: [str], filter_attributes_by_subset: bool,
            dummy_data_rows: int, add_geometry: bool, ignore_relations: bool, model_directory: Path = None
    ) -> [OTLObject]:
        """
        This method is used to generate objects for the template. It will generate objects based on the subset file
        """
        collector = cls._load_collector_from_subset_path(subset_path=subset_path)
        filtered_class_list = cls.filters_classes_by_subset(collector=collector, class_uris_filter=class_uris_filter)
        if filtered_class_list == []:
            raise ValueError('Something went wrong, as the class_uri filter list is empty')
        cls.relation_dict = get_hardcoded_relation_dict(model_directory=model_directory)

        amount_objects_to_create = max(1, dummy_data_rows)
        otl_objects = []

        while True:
            with ThreadPoolExecutor() as executor:
                futures = [
                    executor.submit(cls.create_x_objects, oslo_class, add_geometry, collector,
                                    filter_attributes_by_subset, model_directory, amount_objects_to_create)
                    for oslo_class in [cl for cl in filtered_class_list if cl.abstract == 0]
                ]
                while futures:
                    done, not_done = concurrent.futures.wait(futures, return_when=ALL_COMPLETED, timeout=60)
                    for future in as_completed(done):
                        otl_objects.extend(future.result())
                    futures = not_done
                    print(f'when using multiprocessing: {len(done)} done, {len(not_done)} not done')
                    if len(not_done) > 0:
                        print('restarting the loop with the not done ones')

            created = len(otl_objects)
            unique_ids = len({obj.assetId.identificator if obj.typeURI != 'http://purl.org/dc/terms/Agent' else obj.agentId.identificator
                              for obj in otl_objects})
            if created == unique_ids:
                break
            otl_objects = []

        if not ignore_relations:
            non_relations_class_uris = [cl.objectUri for cl in filtered_class_list
                                        if cl.abstract == 0 and cl.objectUri not in cls.relation_dict]
            cls.append_relations_to_objects(otl_objects=otl_objects, collector=collector,
                                            class_uris_filter=non_relations_class_uris, model_directory=model_directory)

        return otl_objects

    @classmethod
    async def create_x_objects_async(cls, oslo_class, add_geometry, collector, filter_attributes_by_subset,
                                  model_directory,
                         amount_objects_to_create):
        if oslo_class.objectUri in cls.relation_dict:
            return []

        otl_objects = []
        for _ in range(amount_objects_to_create):
            await sleep(0)
            otl_object = cls.generate_object_from_oslo_class(
                oslo_class=oslo_class, add_geometry=add_geometry, collector=collector,
                filter_attributes_by_subset=filter_attributes_by_subset, model_directory=model_directory)
            if otl_object is not None:
                otl_objects.append(otl_object)
        return otl_objects


    @classmethod
    async def generate_objects_for_template_async(
            cls, subset_path: Path, class_uris_filter: [str], filter_attributes_by_subset: bool,
            dummy_data_rows: int, add_geometry: bool, ignore_relations: bool, model_directory: Path = None
    ) -> [OTLObject]:
        """
        This method is used to generate objects for the template. It will generate objects based on the subset file
        """
        await sleep(0)
        collector = cls._load_collector_from_subset_path(subset_path=subset_path)
        await sleep(0)
        filtered_class_list = cls.filters_classes_by_subset(collector=collector, class_uris_filter=class_uris_filter)
        if filtered_class_list == []:
            raise ValueError('Something went wrong, as the class_uri filter list is empty')
        await sleep(0)
        cls.relation_dict = get_hardcoded_relation_dict(model_directory=model_directory)

        amount_objects_to_create = max(1, dummy_data_rows)
        otl_objects = []

        while True:
            tasks = [
                cls.create_x_objects_async(oslo_class, add_geometry, collector,
                                           filter_attributes_by_subset, model_directory, amount_objects_to_create)
                for oslo_class in [cl for cl in filtered_class_list if cl.abstract == 0]
            ]
            results = await asyncio.gather(*tasks)
            otl_objects.extend([item for sublist in results for item in sublist])

            created = len(otl_objects)
            unique_ids = len({
                                 obj.assetId.identificator if obj.typeURI != 'http://purl.org/dc/terms/Agent' else obj.agentId.identificator
                                 for obj in otl_objects})
            if created == unique_ids:
                break
            otl_objects = []

        if not ignore_relations:
            non_relations_class_uris = [cl.objectUri for cl in filtered_class_list
                                        if cl.abstract == 0 and cl.objectUri not in cls.relation_dict]
            cls.append_relations_to_objects(otl_objects=otl_objects, collector=collector,
                                            class_uris_filter=non_relations_class_uris, model_directory=model_directory)

        return otl_objects

    @classmethod
    def generate_object_from_oslo_class(
            cls, oslo_class: OSLOClass, add_geometry: bool,
            filter_attributes_by_subset: bool, collector: OSLOCollector, model_directory: Path = None) -> [OTLObject]:
        """
        Generate an object from a given OSLO class
        """
        instance = dynamic_create_instance_from_uri(oslo_class.objectUri, model_directory=model_directory)
        if instance is None:
            return

        if filter_attributes_by_subset:
            for attribute_object in collector.find_attributes_by_class(oslo_class):
                attr = get_attribute_by_name(instance, attribute_object.name)
                if attr is None:
                    logging.warning(f'Attribute {attribute_object.name} not found in class {oslo_class.objectUri}')
                elif attr.naam == 'isActief':
                    attr.set_waarde(True)
                else:
                    attr.fill_with_dummy_data()
        else:
            for attr in instance:
                if attr.naam == 'isActief':
                    attr.set_waarde(True)
                elif attr.naam != 'geometry':
                    attr.fill_with_dummy_data()

        with contextlib.suppress(AttributeError):
            assetId_attr = get_attribute_by_name(instance, 'assetId')
            if assetId_attr is not None:
                assetId_attr.fill_with_dummy_data()
            if add_geometry:
                geo_attr = get_attribute_by_name(instance, 'geometry')
                if geo_attr is not None:
                    geo_attr.fill_with_dummy_data()

        asset_versie = get_attribute_by_name(instance, 'assetVersie')
        if asset_versie is not None:
            asset_versie.set_waarde(None)

        DotnotationHelper.clear_list_of_list_attributes(instance)

        return instance

    @classmethod
    def alter_excel_template(cls, instances: list, file_path: Path, add_attribute_info: bool,
                             generate_choice_list: bool, dummy_data_rows: int, add_deprecated: bool,
                             abbreviate_excel_sheettitles: bool):
        wb = load_workbook(file_path)
        wb.create_sheet('Keuzelijsten')

        choice_list_dict = {}
        for sheet in wb:
            if sheet.title == 'Keuzelijsten':
                break

            cls.alter_excel_sheet(add_attribute_info=add_attribute_info, choice_list_dict=choice_list_dict,
                                  generate_choice_list=generate_choice_list, dummy_data_rows=dummy_data_rows,
                                  instances=instances, sheet=sheet, add_deprecated=add_deprecated, workbook=wb,
                                  abbreviate_excel_sheettitle=abbreviate_excel_sheettitles)

        wb.save(file_path)
        wb.close()

    @classmethod
    def fill_class_dict(cls, instances: list) -> dict:
        class_dict = defaultdict(list)
        for instance in instances:
            class_dict[instance.typeURI].append(instance)
        return class_dict


    @classmethod
    def alter_csv_template(cls, instances: list, file_path: Path, add_attribute_info: bool,
                             split_per_type: bool, dummy_data_rows: int, add_deprecated: bool):
        classes_dict = cls.fill_class_dict(instances)
        if split_per_type:
            for type_uri, typed_instances in classes_dict.items():
                ns, name = get_ns_and_name_from_uri(type_uri)
                ns_name = f'{ns}_{name}'
                if name == 'Agent':
                    ns_name = 'Agent'
                class_file_path = file_path.parent / f'{file_path.stem}_{ns_name}.csv'
                cls.alter_csv_file(add_attribute_info=add_attribute_info, add_deprecated=add_deprecated,
                                   dummy_data_rows=dummy_data_rows, instances=typed_instances, file_path=class_file_path)
        else:
            cls.alter_csv_file(add_attribute_info=add_attribute_info, add_deprecated=add_deprecated,
                               dummy_data_rows=dummy_data_rows, instances=instances, file_path=file_path)
    
    @classmethod
    async def alter_csv_template_async(cls, instances: list, file_path: Path, add_deprecated: bool,
                                       add_attribute_info: bool, split_per_type: bool, dummy_data_rows: int):
        classes_dict = cls.fill_class_dict(instances)
        if split_per_type:
            for type_uri, typed_instances in classes_dict.items():
                await sleep(0)
                ns, name = get_ns_and_name_from_uri(type_uri)
                ns_name = f'{ns}_{name}'
                if name == 'Agent':
                    ns_name = 'Agent'
                class_file_path = file_path.parent / f'{file_path.stem}_{ns_name}.csv'
                cls.alter_csv_file(add_attribute_info=add_attribute_info,
                                   dummy_data_rows=dummy_data_rows, instances=typed_instances, add_deprecated=add_deprecated,
                                   file_path=class_file_path)
        else:
            await sleep(0)
            cls.alter_csv_file(add_attribute_info=add_attribute_info,
                               dummy_data_rows=dummy_data_rows, instances=instances, add_deprecated=add_deprecated,
                               file_path=file_path)

    @classmethod
    def alter_csv_file(cls, add_attribute_info: bool, instances: [OTLObject], add_deprecated: bool, file_path: Path,
                       dummy_data_rows: int):
        collected_attribute_info_row = []
        deprecated_attributes_row = []
        instance = instances[0]
        quote_char = '"'

        with open(file_path, encoding='utf-8') as file:
            csv_reader = csv.reader(file, delimiter=';', quotechar=quote_char)
            header_row = next(csv_reader)
            csv_data = list(csv_reader)

        for index, header in enumerate(header_row):
            if header is None or header == '':
                continue

            if header == 'typeURI':
                if add_attribute_info:
                    collected_attribute_info_row.append(
                        'De URI van het object volgens https://www.w3.org/2001/XMLSchema#anyURI .')
                if add_deprecated:
                    deprecated_attributes_row.append('')
                continue

            try:
                attribute = DotnotationHelper.get_attribute_by_dotnotation(instance, header)
            except AttributeError:
                if add_attribute_info:
                    collected_attribute_info_row.append('')
                if add_deprecated:
                    deprecated_attributes_row.append('')
                continue

            if add_attribute_info:
                collected_attribute_info_row.append(attribute.definition)

            if add_deprecated:
                deprecated_attributes_row.append('DEPRECATED' if attribute.deprecated_version else '')

        with open(file_path, 'w', newline='', encoding='utf-8') as file:
            csv_writer = csv.writer(file, delimiter=';', quotechar=quote_char, quoting=csv.QUOTE_MINIMAL)
            if add_attribute_info:
                csv_writer.writerow(collected_attribute_info_row)
            if add_deprecated and any(deprecated_attributes_row):
                csv_writer.writerow(deprecated_attributes_row)
            csv_writer.writerow(header_row)
            if dummy_data_rows != 0:
                for line in csv_data:
                    csv_writer.writerow(line)

    @classmethod
    async def alter_excel_template_async(cls, instances: list, file_path: Path, add_attribute_info: bool,
                                         generate_choice_list: bool, dummy_data_rows: int, add_deprecated: bool,
                                         abbreviate_excel_sheettitles: bool):
        wb = load_workbook(file_path)
        wb.create_sheet('Keuzelijsten')

        choice_list_dict = {}
        for sheet in wb:
            if sheet.title == 'Keuzelijsten':
                break

            cls.alter_excel_sheet(add_attribute_info=add_attribute_info, choice_list_dict=choice_list_dict,
                                  generate_choice_list=generate_choice_list, dummy_data_rows=dummy_data_rows,
                                  instances=instances, sheet=sheet, add_deprecated=add_deprecated, workbook=wb,
                                  abbreviate_excel_sheettitle=abbreviate_excel_sheettitles)
            await sleep(0)

        wb.save(file_path)
        wb.close()

    @classmethod
    def alter_excel_sheet(cls, add_attribute_info: bool, choice_list_dict: dict, generate_choice_list: bool,
                          instances: [OTLObject], sheet: Worksheet, add_deprecated: bool, workbook: Workbook,
                          dummy_data_rows: int, abbreviate_excel_sheettitle: bool = True):
        type_uri = cls.get_uri_from_sheet_name(sheet.title)
        instance = next((x for x in instances if x.typeURI == type_uri), None)
        if instance is None:
            instance = next((x for x in instances if x.typeURI.startswith(type_uri)), None)
        if instance is None:
            raise UnknownExcelError(f'When creating a template, no instance could be created for {type_uri}')

        boolean_validation = DataValidation(type="list", formula1='"TRUE,FALSE,-"', allow_blank=True)
        sheet.add_data_validation(boolean_validation)
        collected_attribute_info = []
        deprecated_attributes_row = []
        header_row = next(sheet.iter_rows(min_row=1, max_row=1))
        for header_cell in header_row:
            header = header_cell.value
            if header is None or header == '':
                continue

            if header == 'typeURI':
                data_validation = DataValidation(type="list", formula1=f'"{type_uri}"', allow_blank=True)
                sheet.add_data_validation(data_validation)
                data_validation.add(f'{header_cell.column_letter}2:{header_cell.column_letter}1000')
                if add_attribute_info:
                    collected_attribute_info.append('De URI van het object volgens https://www.w3.org/2001/XMLSchema#anyURI .')
                if add_deprecated:
                    deprecated_attributes_row.append('')
                continue

            if type_uri == 'http://purl.org/dc/terms/Agent' and header.startswith('assetId.'):
                continue

            attribute = DotnotationHelper.get_attribute_by_dotnotation(instance, header)

            if add_attribute_info:
                collected_attribute_info.append(attribute.definition)

            if add_deprecated:
                deprecated_attributes_row.append('DEPRECATED' if attribute.deprecated_version else '')

            if generate_choice_list:
                if issubclass(attribute.field, BooleanField):
                    boolean_validation.add(f'{header_cell.column_letter}2:{header_cell.column_letter}1000')
                    cls.color_choice_lists_green(sheet=sheet, header_cell_column=header_cell)

                elif issubclass(attribute.field, KeuzelijstField):
                    cls.generate_choice_list_in_excel(
                        attribute=attribute, choice_list_dict=choice_list_dict, column=header_cell.column,
                        row_nr=1, sheet=sheet, workbook=workbook)
                    cls.color_choice_lists_green(sheet=sheet, header_cell_column=header_cell)

        if dummy_data_rows == 0:
            instance_count = len([x for x in instances if x.typeURI == type_uri])
            if instance_count > 0:
                sheet.delete_rows(idx=2, amount=instance_count)

        if add_deprecated and any(deprecated_attributes_row):
            cls.add_deprecated_row_to_sheet(deprecated_attributes_row, sheet)

        if add_attribute_info:
            cls.add_attribute_info_to_sheet(collected_attribute_info, sheet)

        cls.set_fixed_column_width(sheet=sheet, width=25)

        if abbreviate_excel_sheettitle:
            cls.abbreviate_excel_sheettitle(sheet=sheet)

    @classmethod
    def color_choice_lists_green(cls, sheet: Worksheet, header_cell_column: Cell):
        for cell in sheet.iter_rows(min_col=header_cell_column.column, max_col=header_cell_column.column,
                                    min_row=2, max_row=1000):
            cell[0].fill = cls.green_fill

    @classmethod
    def add_deprecated_row_to_sheet(cls, deprecated_attributes_row, sheet):
        sheet.insert_rows(idx=1)
        for index, depr_info in enumerate(deprecated_attributes_row, start=1):
            if depr_info != '':
                cell = sheet.cell(row=1, column=index)
                cell.value = depr_info
                cell.fill = PatternFill(start_color="FF7276", end_color="FF7276", fill_type="solid")

    @classmethod
    def add_attribute_info_to_sheet(cls, collected_attribute_info: list[str], sheet: Worksheet):
        sheet.insert_rows(idx=1)
        fill = PatternFill(start_color="808080", fill_type="solid")
        alignment = Alignment(wrapText=True, vertical='top')
        for index, attr_info in enumerate(collected_attribute_info, start=1):
            cell = sheet.cell(row=1, column=index, value=attr_info)
            cell.alignment = alignment
            cell.fill = fill

    @classmethod
    def generate_choice_list_in_excel(cls, attribute, choice_list_dict, column, row_nr, sheet: Worksheet,
                                      workbook: Workbook):
        choice_list_values = [cv for cv in attribute.field.options.values()
                              if cv.status != 'verwijderd']
        if attribute.field.naam not in choice_list_dict:
            cls.add_choice_list_to_sheet(workbook=workbook, name=attribute.field.naam,
                                         options=choice_list_values, choice_list_dict=choice_list_dict)
        column_in_choice_sheet = choice_list_dict[attribute.field.naam]
        start_range = f"${column_in_choice_sheet}$2"
        end_range = f"${column_in_choice_sheet}${len(choice_list_values) + 2}"
        data_val = DataValidation(type="list", formula1=f"Keuzelijsten!{start_range}:{end_range}",
                                  allowBlank=True)
        sheet.add_data_validation(data_val)
        data_val.add(f'{get_column_letter(column)}{row_nr + 1}:'
                     f'{get_column_letter(column)}1000')

    @classmethod
    def filters_classes_by_subset(cls, collector: OSLOCollector,
                                  class_uris_filter: [str] = None) -> list[OSLOClass]:
        if class_uris_filter is None:
            return collector.classes
        return [x for x in collector.classes if x.objectUri in class_uris_filter]

    @classmethod
    def set_fixed_column_width(cls, sheet, width: int):
        dim_holder = DimensionHolder(worksheet=sheet)
        for col in range(sheet.min_column, sheet.max_column + 1):
            dim_holder[get_column_letter(col)] = ColumnDimension(sheet, min=col, max=col, width=width)
        sheet.column_dimensions = dim_holder

    @classmethod
    def add_choice_list_to_sheet(cls, workbook, name, options, choice_list_dict):
        active_sheet = workbook['Keuzelijsten']
        column_nr = choice_list_dict.keys().__len__() + 1
        row_nr = 1
        new_header = active_sheet.cell(row=row_nr, column=column_nr)
        if new_header.value is not None:
            raise ValueError(f'Header already exists at column {column_nr}: {new_header.value}')
        new_header.value = name
        cell = active_sheet.cell(row=row_nr + 1, column=column_nr)
        cell.value = "-"
        for index, option in enumerate(options, start=2):
            cell = active_sheet.cell(row=row_nr + index, column=column_nr)
            cell.value = option.invulwaarde

        choice_list_dict[name] = new_header.column_letter

    @classmethod
    def get_uri_from_sheet_name(cls, title: str) -> str:
        if title == 'Agent':
            return 'http://purl.org/dc/terms/Agent'
        if '#' not in title:
            raise ValueError('Sheet title does not contain a #')
        class_ns, class_name = title.split('#', maxsplit=1)
        return short_to_long_ns.get(class_ns, class_ns) + class_name

    @classmethod
    def append_relations_to_objects(cls, otl_objects: [OTLObject], collector: OSLOCollector, class_uris_filter: [str],
                                    model_directory: Path):
        class_dict = defaultdict(list)
        for instance in otl_objects:
            class_dict[instance.typeURI].append(instance)

        for class_uri in class_uris_filter:
            for relation in collector.find_all_concrete_relations(objectUri=class_uri):
                if class_uri != relation.bron_uri:
                    continue
                for i, bron_instance in enumerate(class_dict[relation.bron_uri]):
                    if relation.doel_uri not in class_dict:
                        continue
                    doel_instance = class_dict[relation.doel_uri][i]
                    if relation.objectUri == 'https://wegenenverkeer.data.vlaanderen.be/ns/onderdeel#HeeftBetrokkene':
                        relation_instance = create_betrokkenerelation(rol='toezichter', source=bron_instance,
                                                            target=doel_instance, model_directory=model_directory)
                    else:
                        if relation.richting == 'Unspecified' and bron_instance.assetId.identificator > doel_instance.assetId.identificator:
                            continue
                        relation_type = dynamic_create_type_from_uri(class_uri=relation.objectUri, model_directory=model_directory)
                        relation_instance = create_relation(relation_type=relation_type, source=bron_instance,
                                                            target=doel_instance, model_directory=model_directory)

                    otl_objects.append(relation_instance)

    @classmethod
    def abbreviate_excel_sheettitle(cls, sheet):
        class_name = sheet.title
        # abbreviates the class_name so it doesn't exceed the 31 character limit of sheet titles in excel
        if '#' not in class_name:
            sheet.title = class_name[:31]
            return
        split_name = class_name.split("#")
        namespace_name = split_name[0]
        subclass_name = split_name[1]
        class_name = f"{namespace_name[:3]}#{subclass_name}"
        sheet.title = class_name[:31]
