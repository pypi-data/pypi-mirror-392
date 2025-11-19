from openpyxl import Workbook
from ..models import Model, ModelSet, ModelID, Import, Attribute, Constraint, Actor, Location, ActorMember, ActorGroup, InformationSpace, Task, TaskDefinition, TaskInstance

__all__ = ('excel_full_model',
           )

def excel_full_model(model: Model) -> Workbook:
    wb = Workbook()

    #Sheets
    actor_sheet = wb.create_sheet('Actors')
    actor_sheet['A1'].value = 'ID'
    actor_sheet['B1'].value = 'Title'
    actor_sheet['C1'].value = 'Description'
    actor_sheet['D1'].value = 'Groups'
    actor_sheet['E1'].value = 'Members'

    locations_sheet = wb.create_sheet('Locations')
    locations_sheet['A1'].value = 'ID'
    locations_sheet['B1'].value = 'Title'
    locations_sheet['C1'].value = 'Description'
    locations_sheet['D1'].value = 'Groups'
    locations_sheet['E1'].value = 'Members'

    infospaces_sheet = wb.create_sheet('Information Spaces')
    infospaces_sheet['A1'].value = 'ID'
    infospaces_sheet['B1'].value = 'Title'
    infospaces_sheet['C1'].value = 'Description'

    tasks_sheet = wb.create_sheet('Tasks')
    tasks_sheet['A1'].value = 'ID'
    tasks_sheet['B1'].value = 'Title'
    tasks_sheet['C1'].value = 'Description'
    tasks_sheet['D1'].value = 'Triggers'
    tasks_sheet['E1'].value = 'Info requirements'
    tasks_sheet['F1'].value = 'Read'
    tasks_sheet['G1'].value = 'Write'
    tasks_sheet['H1'].value = 'Binding'
                     
    rows = {'actors': 2, 'locations': 2, 'information_spaces':2,'tasks': 2}

    def add_node(node, parent_node):
        match node:
            case Actor():
                row = rows['actors']
                actor_sheet[f'A{row}'].value = node.name
                actor_sheet[f'B{row}'].value = node.title
                actor_sheet[f'C{row}'].value = node.description
                actor_sheet[f'D{row}'].value= ",".join(node.groups)
                actor_sheet[f'E{row}'].value= ",".join(node.members)
                rows['actors'] += 1
            case Location(name):
                row = rows['locations']
                locations_sheet[f'A{row}'].value = node.name
                locations_sheet[f'B{row}'].value = node.title
                locations_sheet[f'C{row}'].value = node.description
                locations_sheet[f'D{row}'].value= ",".join(node.groups)
                locations_sheet[f'E{row}'].value= ",".join(node.members)
                rows['locations'] += 1
            case InformationSpace(name):
                row = rows['information_spaces']
                infospaces_sheet[f'A{row}'].value = node.node_id
                infospaces_sheet[f'B{row}'].value = node.title
                infospaces_sheet[f'C{row}'].value = node.description
                rows['information_spaces'] += 1
                for sub_node in node.nodes:
                    add_node(sub_node, node)
            case Task(name):
                row = rows['tasks']
                tasks_sheet[f'A{row}'].value = node.node_id
                tasks_sheet[f'B{row}'].value = node.title
                tasks_sheet[f'C{row}'].value = node.description
                
                #Add triggers
                trigger_row = row
                for trigger in node.triggers:
                    tasks_sheet[f'D{trigger_row}'].value = trigger
                    trigger_row += 1
                req_row = row
                for req in node.info_space_requirements:
                    tasks_sheet[f'E{req_row}'].value = req.name
                    tasks_sheet[f'F{req_row}'].value = 'Yes' if req.read else 'No'
                    tasks_sheet[f'G{req_row}'].value = 'Yes' if req.write else 'No'
                    tasks_sheet[f'H{req_row}'].value = req.binding
                    req_row += 1

                rows['tasks'] = max((row + 1,trigger_row,req_row))

                for sub_node in node.nodes:
                    add_node(sub_node, node)
            #Add TaskDef and Instance
    for node in model.nodes:
        add_node(node, model)
    
    #Remove the default sheet
    wb.remove(wb.active)
    return wb
